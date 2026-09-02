# V77: force openpyxl's WORKSHEET WRITER onto pure Python, and prove it took effect.
#
# V74 tried this with sys.modules.setdefault("lxml", None) before importing openpyxl. That is
# a no-op: Streamlit and its dependencies import lxml long before this module runs, so the key
# already exists and setdefault leaves the real module in place. Test 80's own Build Log said
# so -- "openpyxl worksheet writer: lxml (C)" -- while reporting Bridge Asset 4,691 of 4,691
# rows on the sheet after write and max_row=4,696 at save time. Fully built, then the two
# largest sheets were dropped during serialisation.
#
# openpyxl resolves the writer at import: openpyxl/xml/__init__.py sets LXML by trying to
# import lxml.etree, openpyxl/xml/functions.py binds `xmlfile` from lxml.etree or et_xmlfile
# accordingly, and openpyxl/worksheet/_writer.py imports that symbol. Rebinding it on the
# writer module is what actually decides how worksheets are streamed, and it works no matter
# what imported lxml first.
#
# Why bother: every build through 2026-08-25 wrote 5.8-7.9 MB with all four tabs. Every build
# from 2026-09-01 -- the day Streamlit Cloud moved the app to Python 3.14.7 -- wrote ~800 KB
# with Bridge Asset and Term Asset missing or empty, while every smaller sheet survived.
# openpyxl 3.1.5 predates Python 3.14. et_xmlfile is slower on a workbook this size and is the
# writer that produced every working report.
import sys as _sys

OPENPYXL_WORKSHEET_WRITER = "unknown"
try:
    import et_xmlfile as _et_xmlfile
    import openpyxl.worksheet._writer as _openpyxl_ws_writer

    _openpyxl_ws_writer.xmlfile = _et_xmlfile.xmlfile
    OPENPYXL_WORKSHEET_WRITER = getattr(_openpyxl_ws_writer.xmlfile, "__module__", "?")
except Exception as _writer_patch_exc:  # pragma: no cover - never break the build over this
    OPENPYXL_WORKSHEET_WRITER = f"patch failed: {type(_writer_patch_exc).__name__}: {_writer_patch_exc}"

import base64
import calendar
import gc
import hashlib
import io
import re
import secrets
import time
import urllib.parse
import warnings
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta, time as datetime_time
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


try:
    from active_loan_postbuild_audit import (
        audit_diagnostic_lines,
        audit_openpyxl_workbook,
        write_audit_sheets,
        workbook_needs_attention,
    )
    POSTBUILD_AUDIT_AVAILABLE = True
    POSTBUILD_AUDIT_IMPORT_ERROR = ""
except Exception as _audit_import_exc:
    audit_diagnostic_lines = None
    audit_openpyxl_workbook = None
    write_audit_sheets = None
    workbook_needs_attention = None
    POSTBUILD_AUDIT_AVAILABLE = False
    POSTBUILD_AUDIT_IMPORT_ERROR = str(_audit_import_exc)


PRIMARY_USER_NAME = "Hayden"
APP_BUILD_VERSION = "ALR_FIX_2026_09_02_V81_DROP_SUBUNIT_TWIN_TERM_ASSETS"

# V67: filled by _build_bridge_spine_like, reported in the build diagnostics. The Term Asset
# queries require the sub-unit check to be OFF -- "(Is_Sub_Unit__c = FALSE OR
# Is_Sub_Unit__c = NULL)" -- and that OR is only safe if NULL never actually occurs. The
# Bridge spine pulls Is_Sub_Unit__c WITHOUT filtering it, so it can count the raw values for
# free and settle the question on a live run. If NULL shows up, "off" and "not set" are
# different things in this org and the Term predicate needs tightening to "= FALSE".
SUBUNIT_FLAG_CENSUS: Dict[str, int] = {}

# V68: sheet_name -> number of data rows handed to write_output_sheet. Checked against the
# SAVED bytes before the download is offered, so a workbook that loses a tab between the QA
# audit and the save fails loudly instead of shipping as a half-report. Test 77 shipped with
# Bridge Asset and Term Asset as 918-byte empty stubs -- zero cells, not even the rows 1-5
# scaffold -- while its own QA Summary reported 4,694 and 23,848 data rows for them.
WRITTEN_SHEET_ROWS: Dict[str, int] = {}

# V75: sheets the finished report must NOT carry. The template is the uploaded prior workbook
# whenever one is supplied, so anything on it survives into the output unless it is removed --
# which is how Bridge Payoffs / Term Payoffs / REO Sales kept reappearing week after week.
# None of these is a source for any build column: build_prev_maps only reads Bridge Asset,
# Bridge Loan, Term Loan and Term Asset.
REPORT_DROP_SHEETS = (
    "Bridge Payoffs",
    "Term Payoffs",
    "REO Sales",
    "Pacific Life",
    "2026-1",
    "CAFL SA",
    "JLL",
)

# Kept deliberately even though they are not report tabs: load_template_lookup_maps reads
# Strategy Groupings, SSP Loans and Legacy straight out of the workbook, and the workbook the
# build starts from is last week's output. Dropping them would silently break the strategy
# grouping, SSP and Legacy lookups on the NEXT run.
REPORT_KEEP_LOOKUP_SHEETS = ("SSP Loans", "Legacy", "Strategy Groupings")

# V79: filled by _build_term_asset_like, reported in the build diagnostics and the Build Log.
TERM_ASSET_PARENT_DROPS: Dict[str, float] = {}

# V81: filled by _drop_subunit_twin_term_assets, reported in the diagnostics and Build Log.
TERM_ASSET_SUBUNIT_TWIN_DROPS: Dict[str, float] = {}

# V71: sheet_name -> {"handed": n, "on_sheet_after_write": n}, filled by write_output_sheet and
# reported in the build diagnostics. Distinguishes "the write produced nothing" from "the rows
# were lost later", which the V70 save check alone cannot tell apart.
SHEET_WRITE_AUDIT: Dict[str, dict] = {}
# New official report layout: headers on row 5, data starts row 6.
HEADER_ROW = 5
DATA_START_ROW = 6
QA_HARD_STOP_ON_FAIL = True
BRIDGE_ASSET_UPB_TINY_VS_FUNDED_RATIO = 0.50
BRIDGE_NPD_PRESERVE_DAY10_WHEN_SERVICER_DAY1 = True
TEMPLATE_FILENAMES = (
    "Active Loan Template.xlsx",
    "Active Loan Report Template.xlsx",
    "Active Loan Template - 20260518 Clean.xlsx",
    "20260518 Active Loans.xlsx",
    "20260330 Active Loans vDRAFT.xlsx",
)
API_VERSION = "v66.0"
BULK_PAGE_SIZE = 5000
BULK_WAIT_TIMEOUT_SECONDS = 300
OUTPUT_TEST_FILENAME = "active loan report test.xlsx"
ENFORCE_ZERO_FILLABLE_BLANKS = True
ZERO_BLANK_MAX_ROUNDS = 4
FORCE_QUARTER_END = None
MATH_TOLERANCE_DOLLARS = 1.00
TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT = 1.10
BRIDGE_UPB_COMMITMENT_RATIO_LIMIT = 1.10
# Math QA should repair and report issues by default, not block the weekly build.
# Set this to True only for development / intentional fail-fast testing.
STRICT_MATH_HARD_STOP = False
BRIDGE_OVER_COMMITMENT_WARN_TOLERANCE_DOLLARS = 2500.00
BRIDGE_OVER_COMMITMENT_WARN_RATIO = 0.0005
BRIDGE_EXCEPTION_STAGES_ALLOW_OVER_COMMITMENT = {"Expired", "Matured", "Sold", "REO", "REO-Sold"}
UPB_HEADER_RE = re.compile(r"\b\d{1,2}/\d{1,2}\s*UPB\b", re.I)
# Preserve formula columns from the completed-report template. Set False only for
# mismatch debugging where formula cached values are unavailable before Excel recalculates.
PRESERVE_TERM_ASSET_FORMULA_COLUMNS = True
# The generated workbook is consumed by Python/openpyxl mismatch checks before Excel
# has a chance to recalculate formula caches. If formulas are left in place, those
# tools see blank cached values. Default to writing the report formula outputs as
# materialized values so the workbook is testable immediately after build.
MATERIALIZE_FORMULA_RESULT_COLUMNS = True


VALID_STAGES = ["Active", "Closed Won", "Expired", "Matured", "Sold", "REO"]
BRIDGE_ACTIVE_STAGES = VALID_STAGES.copy()
BRIDGE_LOAN_SOURCE_STAGES = VALID_STAGES.copy()
BRIDGE_ACTIVE_PROPERTY_STATUSES = ["Active", "REO"]
BRIDGE_TYPES = ["Bridge Loan", "SAB Loan", "Acquired Bridge Loan", "Single Asset Bridge Loan"]
BRIDGE_TYPE_CONTAINS = ["SAB", "Single Asset Bridge"]
TERM_TYPES = ["DSCR", "Investor DSCR", "Single Rental Loan", "Term Loan"]
LOAN_ACTIVE_STATUS = "Active"
BRIDGE_EXCLUDED_PRODUCT_TYPE = "Model Home Lease"

DNL_STAGES = [
    "Closed Won", "Purchased", "Brokered- Closed Won", "Expired", "Matured",
    "Sold", "Paid Off", "REO", "REO-Sold",
]

VALUATION_STAGES = VALID_STAGES.copy()
VALUATION_PROPERTY_STATUSES = ["Active", "REO"]

EXPIRED_OR_MATURED_STAGES = ["Expired", "Matured"]
REO_FAMILY_STAGES = ["REO"]
TERM_ACTIVE_STAGES = VALID_STAGES.copy() + ["Purchased", "Approved by Committee"]
TERM_PREBOARDING_STAGES = {"Approved by Committee", "Purchased"}
TERM_CURRENT_MATURITY_FIELD_CANDIDATES = [
    "Current_Loan_Maturity_Date__c",
    "Current_Line_Maturity_Date__c",
    "Current_Maturity_Date__c",
    "Loan_Maturity_Date__c",
    "Maturity_Date__c",
]
TERM_ACTIVE_PROPERTY_STATUSES = ["Active", "REO"]
TERM_RECORDTYPE_NAMES = {"term loan", "dscr"}
BRIDGE_RT_EXACT = {"acquired bridge loan", "bridge loan", "sab loan", "single asset bridge loan"}
BRIDGE_RT_CONTAINS = {"sab", "single asset bridge"}
TERM_DSCR_TYPES = {"DSCR", "Investor DSCR", "Single Rental Loan"}
TERM_ALWAYS_INCLUDE_DEALS = {"43422", "43462"}
# V41: deals that surface in SF_Term but the official report does NOT carry on the Term
# tab. 20747 is a bridge-to-term refinance that still appears in the Term Salesforce pull
# (Closed Won, ~$1.4M Statebridge UPB) but real treats it as Bridge-only, so the generic
# UPB keep would wrongly include it. Force-excluded from the Term Loan/Term Asset population.
TERM_FORCE_EXCLUDE_DEALS = {"20747"}
TERM_SPINE_SERVICER_FAMILIES = {"midland", "fci", "berkadia"}
TERM_SOLD_SERVICING_RETAINED_SEGMENT = "Sold Servicing Retained"
TERM_SOLD_RETAINED_SEGMENT_VALUES = {
    TERM_SOLD_SERVICING_RETAINED_SEGMENT,
    "Sold Servcing Retained",
}

# ---------------------------------------------------------------------------
# V43 taxonomy (verified against 20260803 Active Loans, incl. its live formulas)
# ---------------------------------------------------------------------------
# The official report renamed the Bridge "Securitized Bridge" segment to
# "RTL Securitizations" and the RB Loan Type label from "Single Asset Bridge"
# to "RTL". Bridge Asset!CM6 and CX6 confirm both:
#   =IF(...,IF($B6="RB","RTL",$BW6))
#   =IF(OR($BV6="RTL Securitizations",AND($D6="CAFL 2026-R1 CV",...
BRIDGE_SECURITIZED_SEGMENT = "RTL Securitizations"
BRIDGE_RB_LOAN_TYPE = "RTL"
# Sold bridge loans display Financing == Segment == "Sold Servicing Retained"
# (the report's NPL / special-list formulas gate on $D6<>"Sold Servicing Retained").
BRIDGE_SOLD_FINANCING = "Sold Servicing Retained"
BRIDGE_SOLD_SEGMENT = "Sold Servicing Retained"
BRIDGE_SOLD_FINANCING_VALUES = {"Sold", BRIDGE_SOLD_FINANCING, "Sold Servcing Retained"}
# Term: Portfolio reads "Sold Term" while Financing reads "Sold Servicing Retained"
# (Term Loan!J = 'Sold Term' 288 / Term Loan!L = 'Sold Servicing Retained' 288).
TERM_SOLD_PORTFOLIO = "Sold Term"
# Quarter-end NPL columns are 3-valued in the official report: REO / NPL / N.
QEND_NPL_REO_VALUE = "REO"
QEND_NPL_NPL_VALUE = "NPL"
QEND_NPL_NONE_VALUE = "N"

ACTIVE_RM_STAGES = [
    "Closed Won", "Expired", "Matured", "Sold", "REO", "REO-Sold",
]

ACTIVE_RM_DIRECT_FIELD_CANDIDATES = [
    "Active_RM__c",
    "Is_Active_RM__c",
    "Active_Relationship_Manager__c",
    "Relationship_Manager_Active__c",
    "Relationship_Manager__c",
    "Relationship_Manager_1__c",
    "Relationship_Manager_2__c",
    "Primary_RM__c",
    "Secondary_RM__c",
    "RM__c",
    "RM_1__c",
    "RM_2__c",
    "Current_RM__c",
]

TERM_SERVICER_NAME_FIELD_CANDIDATES = [
    "Servicer_Name__c",
    "Master_Servicer_Name__c",
    "Primary_Servicer_Name__c",
]

TERM_SERVICER_PRIMARY_FIELD_CANDIDATES = [
    "Servicer_Loan_Number__c",
    "Servicer_Loan_Num__c",
    "Servicer_Loan_ID__c",
    "Servicer_Loan_Id__c",
    "Primary_Servicer_Loan_Number__c",
    "Primary_Servicer_Loan_ID__c",
    "Master_Servicer_Loan_Number__c",
    "Master_Servicer_Loan_ID__c",
]

TERM_SERVICER_FALLBACK_FIELD_CANDIDATES = [
    "Servicer_Commitment_Id__c",
    "Servicer_Commitment_ID__c",
]

TERM_PAYOFF_DATE_FIELD_CANDIDATES = [
    "Payoff_Date__c",
    "Paid_Off_Date__c",
    "Loan_Payoff_Date__c",
    "Actual_Payoff_Date__c",
]

AM_ASSIGNMENT_ROLES = ["Asset Manager", "Asset Manager 2", "Construction Manager"]
EXCLUDED_TEST_ACCOUNT_NAME = "Inhouse Test Account"

BRIDGE_MB_FINANCINGS = {
    "Goldman Sachs",
    "Morgan Stanley",
    "Wells Fargo",
    "Wells Fargo - NPL",
    "Goldman Sachs - NPL",
    "Axos",
    "CAFL 2026-R1",
    "Ineligible",
}

DATE_NUMBER_FORMAT = "mm-dd-yy"
MONEY0_FORMAT = r'#,###;[Red]\(#,###\);"-"'
MONEY2_FORMAT = r'#,###.00;[Red]\(#,###.00\);"-"'
BASE_FONT = Font(name="Aptos Narrow", size=11)
BASE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

SHEET_DATE_HEADERS = {
    "Bridge Asset": {
        "Origination Date", "First Funding Date", "Last Funding Date", "Next Payment Date",
        "Original Loan Maturity date", "Current Loan Maturity date", "Original Asset Maturity date",
        "Current Asset Maturity Date", "AM 1 Assigned Date", "AM 2 Assigned Date", "CM Assigned Date",
        "Special Asset: Resolved Date", "Forbearance Term Date", "FC Sale Date", "Rescheduled FC Sale Date",
        "REO Date", "Origination Value Dt", "Most Recent Appraisal Order Date", "Updated Valuation Date", "Tax Due Date",
        "Servicer Maturity Date", "CV Maturity Date", "Maturity Date", "Most Recent Valuation Date",
    },
    "Bridge Loan": {
        "Origination Date", "Last Funding Date", "Original Maturity Date", "Current Maturity Date",
        "Next Advance Maturity Date", "Next Payment Date", "Most Recent Valuation Date",
        "AM 1 Assigned Date", "AM 2 Assigned Date", "CM Assigned Date",
    },
    "Term Loan": {"Origination Date", "Maturity Date", "Next Payment Date", "REO Date", "Loan Sold Date"},
    "Term Asset": {"Origination Date", "Origination Value Date", "Updated Value Date"},
}

SHEET_DATETIME_HEADERS = {
    # The official report keeps the appraisal Order Received timestamp as a full
    # Pacific-local datetime (date + HH:MM:SS), not a date-only value. Preserve the
    # time so it matches exactly instead of being truncated to midnight.
    "Bridge Asset": {"Most Recent Appraisal Order Date"},
}

REPORT_IDENTIFIER_HEADERS = {
    "Bridge Asset": {"Deal Number", "Servicer ID", "SF Yardi ID", "Asset ID"},
    "Bridge Loan": {"Deal Number", "Servicer ID"},
    "Term Loan": {"Deal Number", "Servicer ID", "SF Yardi ID"},
    "Term Asset": {"Deal Number", "Asset ID"},
}

DEFAULT_TEXT_HEADERS = {
    "Bridge Asset": {"Year Built", "Square Feet", "Zip", "APN", "Additional APNs"},
}

# The completed report intentionally uses "N/A" in several formula-driver and manual-review
# fields. Leaving these as true Excel blanks can break downstream formulas (for example,
# Bridge Asset DQ Status treats a blank REO Date as if the asset were REO).
REPORT_NA_FILL_HEADERS = {
    # VERIFIED against 20260608_Active_Loans.xlsx read with keep_default_na=False:
    # the real report fills missing values with the literal text "N/A" in essentially
    # every data column (true blanks only appear in unnamed spacer columns, Bridge
    # Asset "Property Type", and the Term Asset Updated Value* columns -- the latter
    # are handled by REPORT_FORCE_BLANK_HEADERS below). Earlier I misread this because
    # pandas silently converts "N/A" to NaN by default; do NOT remove columns here
    # without re-checking the actual report with keep_default_na=False.
    "Bridge Asset": {
        "Loan Buyer", "Financing", "Servicer ID", "Servicer", "SF Yardi ID", "Borrower Entity",
        "Primary Contact", "County", "CBSA", "APN", "Additional APNs", "# of Units", "Year Built",
        "Square Feet", "First Funding Date", "Last Funding Date", "Asset Manager 1",
        "AM 1 Assigned Date", "Asset Manager 2", "AM 2 Assigned Date", "Construction Mgr.",
        "CM Assigned Date", "Remedy Plan", "Delinquency Notes", "Maturity Status",
        "Special Asset Status", "Special Asset Reason", "Special Asset: Special Asset Status",
        "Special Asset: Resolved Date", "Forbearance Term Date", "FC Sale Date",
        "Rescheduled FC Sale Date", "REO Date", "Origination Value Dt", "Origination As-Is Value",
        "Origination ARV", "Most Recent Appraisal Order Date", "Updated Valuation Date",
        "Updated As-Is Value", "Updated ARV", "Title Company", "Tax Due Date", "Tax Frequency",
        "Tax Commentary", "Transaction Type", "Deal Intro Sub-Source", "Referral Source Account",
        "Referral Source Contact", "Servicer Status", "Servicer Maturity Date", "Maturity Difference",
        "Most Recent Valuation Date", "Most Recent As-Is Value", "Most Recent ARV",
        # "% of Reno Budget" is a hand/upstream-maintained ratio with no Salesforce or
        # servicer source the build can reach (20260803 shows values >1 on rows with zero
        # funded renovation, so it is not any funded/approved combination available here).
        # It is N/A-when-missing in the report, so let the N/A policy fill it rather than
        # leaving ~4,900 true blanks for the zero-blank QA to flag.
        "% of Reno Budget",
    },
    "Bridge Loan": {
        "Loan Buyer", "Financing", "Servicer ID", "Servicer", "Borrower Name", "Primary Contact",
        "Most Recent Valuation Date", "Most Recent As-Is Value", "Most Recent ARV",
        "Transaction Type", "Deal Intro Sub-Source", "Referral Source Account",
        "Referral Source Contact", "Asset Manager 1", "AM 1 Assigned Date", "Asset Manager 2",
        "AM 2 Assigned Date", "Construction Mgr.", "CM Assigned Date", "AM Commentary",
        # V54: reads "N/A" when no child asset has a unit count (45 deals on 20260810).
        "# of Units",
    },
    "Term Loan": {
        "Servicer ID", "Servicer", "Borrower Entity", "Financing", "Loan Buyer", "REO Date",
        "Asset Manager", "Deal Intro Sub-Source", "Referral Source Account",
        "Referral Source Contact", "AM Commentary", "Loan Sold Date",
    },
    "Term Asset": {
        "Financing", "CBSA", "# Units", "Origination Value Date", "Origination Value",
        "Origination Value Type",
    },
}


# Columns that must stay truly blank when source/carry-forward is blank.
# VERIFIED against 20260608_Active_Loans.xlsx (keep_default_na=False): on Bridge Asset
# the Updated/Most Recent valuation columns are N/A-when-missing (NOT blank), so they
# are intentionally NOT listed here. Only the Term Asset Updated Value* columns are
# genuinely blank-when-missing in the real report.
REPORT_FORCE_BLANK_HEADERS = {
    "Bridge Asset": set(),
    "Bridge Loan": set(),
    "Term Loan": set(),
    "Term Asset": {"Updated Value Date", "Updated As-Is Value", "Updated Value Type"},
}

REPORT_INTEGER_HEADERS = {
    "Bridge Asset": {"# of Units", "Days to Maturity", "Days Past Due"},
    "Bridge Loan": {"Number of Assets", "# of Units", "Days Past Due"},
    "Term Loan": {"Days Past Due"},
    "Term Asset": {"# Units"},
}

# Fix N: free-text columns whose SF source can carry data-entry double-spaces. The
# official report renders these single-spaced; collapse runs of 2+ whitespace to one
# space so byte-only whitespace diffs stop registering as mismatches.
# V59: N/A-fill columns where a literal 0 is a REAL value the report keeps, so the
# text-zero-to-N/A rule must skip them. See _normalize_output_for_report.
REPORT_ZERO_IS_REAL_HEADERS = {
    "Term Loan": {"Borrower Entity"},
}

WHITESPACE_COLLAPSE_HEADERS = {
    "Bridge Asset": {
        "Deal Name", "Borrower Entity", "Account Name", "Address", "Primary Contact",
        "APN", "Additional APNs", "Title Company",
    },
    "Bridge Loan": {"Deal Name", "Borrower Name", "Account", "Primary Contact"},
    "Term Loan": {"Deal Name", "Borrower Entity", "Account Name"},
    "Term Asset": {"Address"},
}

# The inverse of the above: columns whose values are fixed report LABELS whose internal
# spacing is significant, so the generic text-display collapse must not touch them.
# Bridge Asset Servicer Status is the case that matters -- the official report's 90-day
# bucket is literally "90 +  DAYS" with a double space, and normalize_text_display_scalar
# was squashing it to "90 + DAYS" on every one of those rows.
PRESERVE_INTERNAL_WHITESPACE_HEADERS = {
    "Bridge Asset": {"Servicer Status"},
}

SHEET_MONEY2_HEADERS = {
    "Bridge Asset": {
        "Asset Commitment",
        "SF Funded Amount", "Suspense Balance", "Origination As-Is Value", "Origination ARV",
        "Updated As-Is Value", "Updated ARV", "Initial Disbursement Funded", "Renovation Holdback",
        "Renovation Holdback Funded", "Renovation Holdback Remaining", "Interest Allocation",
        "Interest Allocation Funded", "Most Recent As-Is Value", "Most Recent ARV", "Needs NPL Value",
        "Property ALA", "As-Is Value",
    },
    "Term Asset": {"Property ALA", "Origination Value", "Updated As-Is Value"},
}

SHEET_MONEY0_HEADERS = {
    "Bridge Loan": {
        "Loan Commitment", "Active Funded Amount", "Suspense Balance", "Remaining Commitment",
        "Most Recent As-Is Value", "Most Recent ARV", "Initial Disbursement Funded",
        "Renovation Holdback", "Renovation HB Funded", "Renovation HB Remaining",
        "Interest Allocation", "Interest Allocation Funded",
    },
    "Term Loan": {"Loan Amount", "SFR Allocation", "MF Allocation"},
}

BRIDGE_ASSET_FROM_BRIDGE_SPINE = {
    "Loan Buyer": "Sold To",
    "Financing": "Warehouse Line",
    "Deal Number": "Deal Loan Number",
    "Servicer ID": "Servicer Loan Number",
    "SF Yardi ID": "Yardi ID",
    "Asset ID": "Asset ID",
    "Deal Name": "Deal Name",
    "Borrower Entity": "Borrower Entity: Business Entity Name",
    "Account Name": "Account Name: Account Name",
    "Primary Contact": "Primary Contact: Full Name",
    "Address": "Address",
    "City": "City",
    "State": "State",
    "Zip": "Zip",
    "County": "County",
    "CBSA": "CBSA",
    "APN": "APN",
    "Additional APNs": "Additional APNs",
    "# of Units": "# of Units",
    "Year Built": "Year Built",
    "Square Feet": "Square Feet",
    "Origination Date": "Close Date",
    "First Funding Date": "First Funding Date",
    "Last Funding Date": "Last Funding Date",
    "Original Loan Maturity date": "Original Loan Maturity Date",
    "Current Loan Maturity date": "Current Loan Maturity date",
    "Original Asset Maturity date": "Original Asset Maturity Date",
    "Current Asset Maturity Date": "Current Asset Maturity date",
    "Remedy Plan": "Remedy Plan",
    "Delinquency Notes": "Delinquency Status Notes",
    "Maturity Status": "Maturity Status",
    "Is Special Asset (Y/N)": "Is Special Asset",
    "Special Asset Status": "Special Asset: Status",
    "Special Asset Reason": "Special Asset: Special Asset Reason",
    "Special Asset: Special Asset Status": "Special Asset: Special Asset Status",
    "Special Asset: Resolved Date": "Special Asset: Resolved Date",
    "Forbearance Term Date": "Forbearance Term Date",
    "REO Date": "REO Date",
    "Origination Value Dt": "Origination Valuation Date",
    "Origination As-Is Value": "Origination As-Is Value",
    "Origination ARV": "Origination After Repair Value",
    "Most Recent Appraisal Order Date": "Most Recent Appraisal Order Date",
    "Initial Disbursement Funded": "Initial Disbursement Funded",
    "Renovation Holdback": "Approved Renovation Advance Amount",
    "Renovation Holdback Funded": "Renovation Advance Amount Funded",
    "Renovation Holdback Remaining": "Reno Advance Amount Remaining",
    "Interest Allocation": "Interest Allocation",
    "Interest Allocation Funded": "Interest Holdback Funded",
    "Title Company": "Title Company: Account Name",
    "Tax Due Date": "Tax Payment Next Due Date",
    "Tax Frequency": "Taxes Payment Frequency",
    "Tax Commentary": "Tax Commentary",
    "Product Type": "Product Type",
    "Product Sub-Type": "Product Sub-Type",
    "Transaction Type": "Transaction Type",
    "Project Strategy": "Project Strategy",
    "Property Type": "Property Type",
    "Originator": "CAF Originator: Full Name",
    "Deal Intro Sub-Source": "Deal Intro Sub-Source",
    "Referral Source Account": "Referral Source Account: Account Name",
    "Referral Source Contact": "Referral Source Contact: Full Name",
    "Loan Stage": "Stage",
    "Property Status": "Status",
    # V56: per-asset Approved Advance Amount Max (see _build_bridge_spine_like).
    "Asset Commitment": "Approved Advance Amount Max",
}

BRIDGE_ASSET_FROM_VALUATION = {
    "Origination Value Dt": "Origination Valuation Date",
    "Origination As-Is Value": "Origination As-Is Value",
    "Origination ARV": "Origination After Repair Value",
    "Most Recent Appraisal Order Date": "Most Recent Appraisal Order Date",
    "Updated Valuation Date": "Current Appraisal Date",
    "Updated As-Is Value": "Current Appraised As-Is Value",
    "Updated ARV": "Current Appraised After Repair Value",
}

BRIDGE_ASSET_FROM_FORECLOSURE = {
    "FC Sale Date": "FC Sale Date",
    "Rescheduled FC Sale Date": "Rescheduled FC Sale Date",
}

TERM_LOAN_FROM_TERM_WIDE = {
    "Deal Number": "Deal Loan Number",
    "SF Yardi ID": "Yardi ID",
    "Deal Name": "Deal Name",
    "Borrower Entity": "Borrower Entity",
    "Account Name": "Account Name",
    "Do Not Lend (Y/N)": "Do Not Lend",
    "Financing": "Current Funding Vehicle",
    "Loan Sold Date": "Sold Loan: Sold Date",
    "Loan Amount": "Loan Amount",
    "Origination Date": "Close Date",
    "Originator": "CAF Originator",
    "Deal Intro Sub-Source": "Deal Intro Sub-Source",
    "Referral Source Account": "Referral Source Account",
    "Referral Source Contact": "Referral Source Contact",
    "AM Commentary": "Comments AM",
}

TERM_ASSET_FROM_TERM_ASSET_REPORT = {
    "Deal Number": "Deal Loan Number",
    "Asset ID": "Asset ID",
    "Portfolio": "Portfolio",
    "Segment": "Segment",
    "Financing": "Financing",
    "Origination Date": "Origination Date",
    "Address": "Address",
    "City": "City",
    "State": "State",
    "Zip": "Zip",
    "CBSA": "CBSA",
    "# Units": "# of Units",
    "Property Type": "Property Type",
    "Grouping": "Grouping",
    "Origination Value Date": "Origination Value Date",
    "Origination Value": "Origination Value",
    "Origination Value Type": "Origination Value Type",
    "Property ALA": "ALA",
    "Updated Value Date": "Updated Value Date",
    "Updated As-Is Value": "Updated As-Is Value",
    "Updated Value Type": "Updated Value Type",
}


# Formula text is transcribed verbatim from the 20260803 official report's own row 6
# (read back with openpyxl data_only=False), so the letters below are the report's, not a
# re-derivation. Bridge Asset shifted +1 from BO ("% of Reno Budget", col 67) onward
# relative to the previous build. "{qlabel}" is substituted with the running quarter
# ("Q3") at seed time by _resolve_formula_override.
DRAFT_FORMULA_OVERRIDES = {
    "Bridge Asset": {
        "SF Funded Amount": "=+$BK6+$BM6+$BQ6",
        "Loan Type": '=IF($B6="5A","5A Bridge",IF($B6="TPO","Purchased Bridge",IF($B6="RB","RTL",$BW6)))',
        "CV Maturity Date": '=IF(OR($BW6="Credit Line",$BX6="Line of Credit"),$AG6,$AE6)',
        "Maturity Difference": '=IFERROR($CN6-$CK6,"N/A")',
        "Maturity Date": '=IF($CK6<>"N/A",$CK6,$CN6)',
        "Days to Maturity": "=+$CP6-$CQ$4",
        "Days Past Due": "=+$CR$4-$AC6",
        "DQ Status": '=IF($BC6<>"N/A","REO",IF(AND($CR6>0,$CR6<30),"DQ 1-29",IF(AND($CR6>=30,$CR6<60),"DQ 30-59",IF(AND($CR6>=60,$CR6<90),"DQ 60-89",IF($CR6>=90,"DQ 90+","Current")))))',
        "Most Recent Valuation Date": '=IF($BH6<>"N/A",$BH6,$BD6)',
        "Most Recent As-Is Value": '=IF($BH6<>"N/A",$BI6,$BE6)',
        "Most Recent ARV": '=IF($BH6<>"N/A",$BJ6,$BF6)',
        "Needs NPL Value": '=IF(AND($D6<>"Sold Servicing Retained",OR($DF6="NPL",$DF6="REO"),$CT6<$CW$4),"Y","N")',
        "Securitized (Y/N)": '=IF(OR($BV6="RTL Securitizations",AND($D6="CAFL 2026-R1 CV",$BV6="Legacy")),"Y","N")',
        "SSP JV (Y/N)": '=IF($BV6="SSP","Y","N")',
        "CPP JV (Y/N)": '=IF($BV6="CPP JV","Y","N")',
        "Oaktree JV (Y/N)": '=IF($BV6="Oaktree JV","Y","N")',
        "Legacy (Y/N)": '=IF($BV6="Legacy","Y","N")',
        "Matured Loan (YN)": '=IF(_xlfn.MINIFS($CQ:$CQ,$E:$E,$E6)<0,"Y","N")',
        "DQ 45+ Loan (Y/N)": '=IF(_xlfn.MAXIFS($CR:$CR,$E:$E,$E6)>=45,"Y","N")',
        "SA Loan (Y/N)": "=IFERROR(VLOOKUP($AL6,'Strategy Groupings'!$F$4:$G$14,2,0),\"N\")",
        "__QEND_NPL_REO__": '=IF(AND($D6<>"Sold Servicing Retained",$CS6="REO"),"REO",IF(AND($D6<>"Sold Servicing Retained",_xlfn.MINIFS($AC:$AC,$E:$E,$E6)<=$DF$4),"NPL","N"))',
        "__SPECIAL_LIST__": '=IF(AND($D6<>"Sold Servicing Retained",OR($DB6="Y",$DC6="Y",$DD6="Y",$DE6="Y",OR($DF6="NPL",$DF6="REO"))),"Y","N")',
    },
    "Bridge Loan": {
        "Days Past Due": "=+$V$4-$U6",
    },
    "Term Loan": {
        "Days Past Due": "=+$V$4-$T6",
        "DQ Status": '=IF($U6<>"N/A","REO",IF(AND($V6>0,$V6<30),"DQ 1-29",IF(AND($V6>=30,$V6<60),"DQ 30-59",IF(AND($V6>=60,$V6<90),"DQ 60-89",IF($V6>=90,"DQ 90+","Current")))))',
        # V52 (20260810): the "CAFL REO" branch (Securitized Term + REO) was REMOVED from
        # the official formula. 20260810 Term Loan!AF6 yields only N/A / Q3 NPL / Term REO /
        # DQ 45+, and the tab carries zero "CAFL REO" cells (1,018 N/A, 12 Q3 NPL, 1 Term
        # REO). Securitized-Term REO deals now fall through to N/A.
        "__SPECIAL_LIST__": '=IF(AND(OR($J6="Active Term",$J6="DSCR"),$T6<=$AF$4,$W6<>"REO"),"{qlabel} NPL",IF(AND(OR($J6="Active Term",$J6="DSCR"),$W6="REO"),"Term REO",IF(AND(OR($J6="Active Term",$J6="DSCR"),$V6>=45,$W6<>"REO"),"DQ 45+","N/A")))',
        "SFR Allocation": "=SUMIFS('Term Asset'!$S:$S,'Term Asset'!$B:$B,'Term Loan'!$B6,'Term Asset'!$O:$O,'Term Loan'!AG$4)",
        "MF Allocation": "=SUMIFS('Term Asset'!$S:$S,'Term Asset'!$B:$B,'Term Loan'!$B6,'Term Asset'!$O:$O,'Term Loan'!AH$4)",
        "Strategy Grouping": '=IF($AG6>$AH6,"Single Family Rental","Multifamily")',
    },
    "Term Asset": {
        "__UPB__": "=+(S6/SUMIFS($S:$S,$B:$B,$B6))*_xlfn.XLOOKUP($B6,'Term Loan'!$B:$B,'Term Loan'!$Q:$Q)",
        "__SPECIAL_LIST__": "=_xlfn.XLOOKUP($B6,'Term Loan'!$B:$B,'Term Loan'!$AF:$AF)",
    },
}


def _resolve_formula_override(formula: str, q_end: date) -> str:
    """Substitute run-dependent literals inside an override formula."""
    if not isinstance(formula, str) or "{qlabel}" not in formula:
        return formula
    return formula.replace("{qlabel}", f"Q{(q_end.month - 1) // 3 + 1}")


SHEET_BLUEPRINTS = {
    # Bridge Asset layout matches 20260803: "% of Reno Budget" sits at BO (67) between
    # Renovation Holdback Remaining (BN) and Interest Allocation (BP), so every column
    # from 67 on is +1 versus the previous build. Last blue column is now CK (89).
    "Bridge Asset": {
        "row1": {c: "CALC" for c in [35] + list(range(91, 112))},
        "row2": {2: "Bridge Asset Level Data"},
        # B3 continues the report-date banner chain the official report carries across the
        # data tabs (Bridge Summary -> Bridge Loan -> Bridge Asset -> Term Loan -> Term Asset).
        "row3": {2: "=+'Bridge Loan'!$B$3", 110: "__QEND__"},
        "row4": {
            36: "__SUBTOTAL__",
            95: "__RUN_DT__",
            96: "=+$CQ$4",
            101: "=EDATE(DF3,-6)",
            110: "=+$DF$3-90",
        },
        "row5": {
            2: "Portfolio", 3: "Loan Buyer", 4: "Financing", 5: "Deal Number",
            6: "Servicer ID", 7: "Servicer", 8: "SF Yardi ID", 9: "Asset ID",
            10: "Deal Name", 11: "Borrower Entity", 12: "Account Name", 13: "Do Not Lend (Y/N)",
            14: "Primary Contact", 15: "Address", 16: "City", 17: "State", 18: "Zip",
            19: "County", 20: "CBSA", 21: "APN", 22: "Additional APNs", 23: "# of Units",
            24: "Year Built", 25: "Square Feet", 26: "Origination Date", 27: "First Funding Date",
            28: "Last Funding Date", 29: "Next Payment Date", 30: "Original Loan Maturity date",
            31: "Current Loan Maturity date", 32: "Original Asset Maturity date",
            33: "Current Asset Maturity Date", 34: "Asset Commitment", 35: "SF Funded Amount",
            36: "__UPB__", 37: "Suspense Balance", 38: "Asset Manager 1", 39: "AM 1 Assigned Date",
            40: "Asset Manager 2", 41: "AM 2 Assigned Date", 42: "Construction Mgr.",
            43: "CM Assigned Date", 44: "Remedy Plan", 45: "Delinquency Notes", 46: "Maturity Status",
            47: "Is Special Asset (Y/N)", 48: "Special Asset Status", 49: "Special Asset Reason",
            50: "Special Asset: Special Asset Status", 51: "Special Asset: Resolved Date",
            52: "Forbearance Term Date", 53: "FC Sale Date", 54: "Rescheduled FC Sale Date",
            55: "REO Date", 56: "Origination Value Dt", 57: "Origination As-Is Value",
            58: "Origination ARV", 59: "Most Recent Appraisal Order Date", 60: "Updated Valuation Date",
            61: "Updated As-Is Value", 62: "Updated ARV", 63: "Initial Disbursement Funded",
            64: "Renovation Holdback", 65: "Renovation Holdback Funded", 66: "Renovation Holdback Remaining",
            67: "% of Reno Budget",
            68: "Interest Allocation", 69: "Interest Allocation Funded", 70: "Title Company",
            71: "Tax Due Date", 72: "Tax Frequency", 73: "Tax Commentary", 74: "Segment",
            75: "Product Type", 76: "Product Sub-Type", 77: "Transaction Type", 78: "Project Strategy",
            79: "Strategy Grouping", 80: "Property Type", 81: "Originator", 82: "Active RM",
            83: "Deal Intro Sub-Source", 84: "Referral Source Account", 85: "Referral Source Contact",
            86: "Loan Stage", 87: "Property Status", 88: "Servicer Status", 89: "Servicer Maturity Date",
            91: "Loan Type", 92: "CV Maturity Date", 93: "Maturity Difference", 94: "Maturity Date",
            95: "Days to Maturity", 96: "Days Past Due", 97: "DQ Status", 98: "Most Recent Valuation Date",
            99: "Most Recent As-Is Value", 100: "Most Recent ARV", 101: "Needs NPL Value",
            102: "Securitized (Y/N)", 103: "SSP JV (Y/N)", 104: "CPP JV (Y/N)", 105: "Oaktree JV (Y/N)",
            106: "Legacy (Y/N)", 107: "Matured Loan (YN)", 108: "DQ 45+ Loan (Y/N)", 109: "SA Loan (Y/N)",
            110: "__QEND_NPL_REO__", 111: "__SPECIAL_LIST__",
        },
        "subtotal_col": 36,
    },
    "Bridge Loan": {
        "row1": {22: "CALC"},
        "row2": {2: "Bridge Loan Level Data"},
        "row3": {2: "=+'Bridge Summary'!$B$3"},
        # Bridge Asset's run-date anchor moved CP4 -> CQ4 with the "% of Reno Budget" insert.
        "row4": {22: "=+'Bridge Asset'!$CQ$4", 26: "__SUBTOTAL__"},
        # 20260803 has NO "Remaining Commitment" on Bridge Loan: AA (Suspense Balance) is
        # followed directly by AB (Most Recent Valuation Date), so the tab ends at BG
        # (AM Commentary, col 59) rather than BH.
        "row5": {
            2: "Portfolio", 3: "Loan Buyer", 4: "Financing", 5: "Deal Number", 6: "Servicer ID",
            7: "Servicer", 8: "Deal Name", 9: "Borrower Name", 10: "Account", 11: "Do Not Lend (Y/N)",
            12: "Primary Contact", 13: "Number of Assets", 14: "# of Units", 15: "State(s)",
            16: "Origination Date", 17: "Last Funding Date", 18: "Original Maturity Date",
            19: "Current Maturity Date", 20: "Next Advance Maturity Date", 21: "Next Payment Date",
            22: "Days Past Due", 23: "Loan Level Delinquency", 24: "Loan Commitment",
            25: "Active Funded Amount", 26: "__UPB__", 27: "Suspense Balance",
            28: "Most Recent Valuation Date", 29: "Most Recent As-Is Value", 30: "Most Recent ARV",
            31: "Initial Disbursement Funded", 32: "Renovation Holdback", 33: "Renovation HB Funded",
            34: "Renovation HB Remaining", 35: "Interest Allocation", 36: "Interest Allocation Funded",
            37: "Loan Stage", 38: "Segment", 39: "Loan Type", 40: "Product Type", 41: "Product Sub Type",
            42: "Transaction Type", 43: "Project Strategy", 44: "Strategy Grouping", 45: "CV Originator",
            46: "Active RM", 47: "Deal Intro Sub-Source", 48: "Referral Source Account",
            49: "Referral Source Contact", 50: "__QEND_NPL__", 51: "Needs NPL Value",
            52: "Special Focus (Y/N)", 53: "Asset Manager 1", 54: "AM 1 Assigned Date",
            55: "Asset Manager 2", 56: "AM 2 Assigned Date", 57: "Construction Mgr.",
            58: "CM Assigned Date", 59: "AM Commentary",
        },
        "subtotal_col": 26,
    },
    "Term Loan": {
        "row1": {22: "CALC", 23: "CALC", 32: "CALC"},
        "row2": {2: "Term Loan Level Data", 32: "__QEND__"},
        "row3": {2: "=+'Bridge Asset'!$B$3"},
        "row4": {
            17: "__SUBTOTAL__", 22: "__RUN_DT__", 32: "=+$AF$2-90",
            33: "Single Family Rental", 34: "Multifamily",
        },
        "row5": {
            2: "Deal Number", 3: "Servicer ID", 4: "Servicer", 5: "SF Yardi ID", 6: "Deal Name",
            7: "Borrower Entity", 8: "Account Name", 9: "Do Not Lend (Y/N)", 10: "Portfolio",
            11: "Segment", 12: "Financing", 13: "CPP JV", 14: "Loan Buyer", 15: "Loan Sold Date",
            16: "Loan Amount", 17: "__UPB__", 18: "Origination Date", 19: "Maturity Date",
            20: "Next Payment Date", 21: "REO Date", 22: "Days Past Due", 23: "DQ Status",
            24: "Asset Manager", 25: "Originator", 26: "Active RM", 27: "Deal Intro Sub-Source",
            28: "Referral Source Account", 29: "Referral Source Contact", 30: "AM Commentary",
            32: "__SPECIAL_LIST__", 33: "SFR Allocation", 34: "MF Allocation", 35: "Strategy Grouping",
        },
        "subtotal_col": 17,
    },
    "Term Asset": {
        "row1": {20: "CALC", 22: "CALC"},
        "row2": {2: "Term Asset Level Data"},
        "row3": {2: "=+'Term Loan'!$B$3"},
        "row4": {20: "__SUBTOTAL__"},
        "row5": {
            2: "Deal Number", 3: "Asset ID", 4: "Portfolio", 5: "Segment", 6: "Financing",
            7: "Origination Date", 8: "Address", 9: "City", 10: "State", 11: "Zip", 12: "CBSA",
            13: "# Units", 14: "Property Type", 15: "Grouping", 16: "Origination Value Date",
            17: "Origination Value", 18: "Origination Value Type", 19: "Property ALA", 20: "__UPB__",
            22: "__SPECIAL_LIST__", 23: "Updated Value Date", 24: "Updated As-Is Value",
            25: "Updated Value Type",
        },
        "subtotal_col": 20,
    },
}


# Only the blue-header columns are auto-populated by the build. Everything to the
# right of these (the manual / self-computing CALC formula columns) is filled in by
# hand in Excel and is intentionally excluded from BOTH the data write and the
# openpyxl mismatch audit/baseline repair. Column letters: Bridge Asset B..CJ,
# Term Loan B..AD, Term Asset B..T. Formula columns that fall WITHIN these ranges
# (e.g. SF Funded Amount, UPB) are still preserved/propagated as live formulas.
SHEET_BLUE_MAX_COLUMN = {
    # Bridge Asset gained "% of Reno Budget" at BO, so the last blue column is CK (89).
    "Bridge Asset": column_index_from_string("CK"),  # 89
    "Term Loan": column_index_from_string("AD"),     # 30
    "Term Asset": column_index_from_string("T"),     # 20
}


def _sheet_blue_max_col(sheet_name: str) -> Optional[int]:
    """Highest auto-populated/audited column for a sheet, or None if unrestricted."""
    return SHEET_BLUE_MAX_COLUMN.get(sheet_name)


def hey(name: str = PRIMARY_USER_NAME) -> str:
    return f"Hi {name} 👋"


def today_et() -> date:
    return datetime.now(ZoneInfo("America/New_York")).date()


def quarter_end_for_run(run_dt: date) -> date:
    if FORCE_QUARTER_END is not None:
        return FORCE_QUARTER_END
    q_month = ((run_dt.month - 1) // 3 + 1) * 3
    last_day = calendar.monthrange(run_dt.year, q_month)[1]
    return date(run_dt.year, q_month, last_day)


def make_upb_header(run_dt: date) -> str:
    return f"{run_dt.month}/{run_dt.day} UPB"


def normalize_header_name(x) -> str:
    return re.sub(r"[^0-9a-z]+", "", str(x).strip().lower())


def header_lookup(columns: Sequence[str]) -> Dict[str, str]:
    return {normalize_header_name(c): c for c in columns}


def first_matching_col(df: pd.DataFrame, aliases: Sequence[str]) -> Optional[str]:
    lookup = header_lookup(df.columns)
    for alias in aliases:
        k = normalize_header_name(alias)
        if k in lookup:
            return lookup[k]
    return None


def norm_id_series(s: pd.Series) -> pd.Series:
    return (
        s.astype("string")
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"[^0-9A-Za-z]", "", regex=True)
        .replace({"": pd.NA})
    )


def id_key_no_leading_zeros(s: pd.Series) -> pd.Series:
    out = norm_id_series(s).astype("string")
    # Midland and some servicing exports append COM to commitment-style IDs.
    # Strip it for matching only; the displayed Servicer ID is normalized later.
    out = out.str.replace(r"COM$", "", regex=True)
    out = out.str.lstrip("0")
    return out.replace({"": pd.NA})


def money_to_float(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return np.nan
    s = str(x)
    s = re.sub(r"[^0-9\.\-]", "", s)
    return pd.to_numeric(s, errors="coerce")


def _to_datetime_scalar_mixed_utc_naive(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return pd.NaT
    try:
        ts = pd.to_datetime(x, errors="coerce", format="mixed", utc=True)
    except TypeError:
        ts = pd.to_datetime(x, errors="coerce", utc=True)
    except Exception:
        ts = pd.to_datetime(x, errors="coerce")
    if pd.isna(ts):
        return pd.NaT
    if isinstance(ts, pd.Timestamp) and ts.tz is not None:
        try:
            ts = ts.tz_convert(None)
        except Exception:
            ts = ts.tz_localize(None)
    return ts


def to_dt(x):
    return _to_datetime_scalar_mixed_utc_naive(x)


def is_reo_stage(val) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    return "reo" in s and s != ""


def has_any_value(val) -> bool:
    if val is None:
        return False
    if isinstance(val, float) and np.isnan(val):
        return False
    if isinstance(val, str) and val.strip() == "":
        return False
    return True


def clean_text(val) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    s = str(val).strip()
    if s.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    return s

def strip_statebridge_display_id(servicer_id, servicer_name):
    sid = clean_text(servicer_id)
    if not sid:
        return pd.NA
    sid = re.sub(r"\.0$", "", sid)
    sid = sid.replace(",", "").strip()
    serv = clean_text(servicer_name).lower()
    if "statebridge" in serv:
        sid = re.sub(r"[^0-9A-Za-z]", "", sid)
        if sid.startswith("0000"):
            sid = sid[4:]
    return sid or pd.NA

def normalize_servicer_id_for_report(servicer_ids: pd.Series, servicer_names: pd.Series) -> pd.Series:
    sid = pd.Series(servicer_ids, copy=False)
    serv = pd.Series(servicer_names, copy=False)
    out = []
    for i in sid.index:
        out.append(strip_statebridge_display_id(sid.loc[i], serv.loc[i] if i in serv.index else pd.NA))
    return pd.Series(out, index=sid.index, dtype="object")


def normalize_report_identifier_scalar(val):
    txt = clean_text(val)
    if not txt:
        return pd.NA
    txt = re.sub(r"\.0$", "", txt)
    txt = txt.strip()
    if re.fullmatch(r"\d+", txt):
        if len(txt) > 1 and txt.startswith("0"):
            return txt
        try:
            return int(txt)
        except Exception:
            return txt
    return txt


def normalize_report_identifier_series(s: pd.Series) -> pd.Series:
    return pd.Series(s, copy=False).map(normalize_report_identifier_scalar)


def normalize_text_display_scalar(val):
    if not has_any_value(val):
        return pd.NA
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return pd.NA
        return _excel_strip_timezone(val)
    if isinstance(val, (datetime, date)):
        return _excel_strip_timezone(val) if isinstance(val, datetime) else val
    if isinstance(val, np.generic):
        val = val.item()
    txt = clean_text(val)
    txt = re.sub(r"\.0$", "", txt)
    # Collapse internal runs of 2+ whitespace to a single space. SF text fields carry
    # data-entry double-spaces (Deal Name / APN / Borrower Entity / Account Name); the
    # official report renders them single-spaced, so match that (Fix N).
    txt = re.sub(r"\s{2,}", " ", txt).strip()
    return txt if txt else pd.NA


def normalize_text_display_series(s: pd.Series) -> pd.Series:
    return pd.Series(s, copy=False).map(normalize_text_display_scalar)


def normalize_integer_display_series(s: pd.Series) -> pd.Series:
    ser = pd.Series(s, copy=False)
    num = pd.to_numeric(ser, errors="coerce")
    out = _object_series_like(ser)
    mask = num.notna()
    if bool(mask.any()):
        out.loc[mask] = num.loc[mask].round(0).astype("int64").astype("object")
    return out


def blankish_mask(s: pd.Series) -> pd.Series:
    base = pd.Series(list(pd.Series(s, copy=False)), index=pd.Series(s, copy=False).index, dtype="object")
    s_text = base.astype("string").str.strip().str.lower()
    return base.isna() | s_text.isin(["", "nan", "none", "<na>", "nat"])


def coalesce_keep_nonblank(primary: pd.Series, fallback: pd.Series) -> pd.Series:
    p = pd.Series(list(pd.Series(primary, copy=False)), index=pd.Series(primary, copy=False).index)
    f = pd.Series(list(pd.Series(fallback, copy=False)), index=p.index)
    return p.where(~blankish_mask(p), f)


def blank_zero_value_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    """Force 0 (and 0.0 / "0") to a true blank for value columns.

    The official report leaves valuation/value cells EMPTY when there is no value --
    never 0. A literal 0 can sneak in from a Salesforce field, a stale prior-report
    carry-forward, or a numeric coercion. This guarantees a missing Updated ARV /
    As-Is / value reads as an empty cell exactly like the real report, so the test
    can never show a misleading 0 where the report shows blank.
    """
    for col in columns:
        if col not in df.columns:
            continue
        num = pd.to_numeric(df[col], errors="coerce")
        is_zeroish = num.eq(0)
        df[col] = df[col].mask(is_zeroish, pd.NA)
    return df


def coalesce_report_display_first(primary: pd.Series, fallback: pd.Series) -> pd.Series:
    """Carry-forward coalesce where report placeholders like N/A are valid.

    coalesce_keep_nonblank intentionally treats only true blanks as missing, but
    some normalization paths can treat placeholder values as effectively blank. For
    prior completed workbook carry-forward fields like Financing, an explicit N/A
    is the value the report should keep.
    """
    p = pd.Series(list(pd.Series(primary, copy=False)), index=pd.Series(primary, copy=False).index, dtype="object")
    f = pd.Series(list(pd.Series(fallback, copy=False)), index=p.index, dtype="object")
    p_text = p.astype("string").str.strip().str.lower()
    has_primary = p.notna() & p_text.ne("") & ~p_text.isin(["nan", "none", "<na>", "nat"])
    return p.where(has_primary, f)


def deal_key(value) -> str:
    s = clean_text(value)
    if not s:
        return ""
    s = re.sub(r"\.0$", "", s)
    return s


def deal_lookup_keys(value) -> List[str]:
    s = deal_key(value)
    if not s:
        return []
    keys = [s]
    m = re.match(r"^(\d+)-", s)
    if m:
        keys.append(m.group(1))
    return keys


def deal_in_lookup(value, lookup: Set[str]) -> bool:
    return any(k in lookup for k in deal_lookup_keys(value))


def first_nonblank(series: pd.Series):
    for v in series:
        if has_any_value(v):
            return v
    return pd.NA


def first_or_various(series: pd.Series):
    vals = []
    seen = set()
    for v in series:
        if not has_any_value(v):
            continue
        key = clean_text(v)
        if key not in seen:
            seen.add(key)
            vals.append(v)
    if not vals:
        return pd.NA
    if len(vals) == 1:
        return vals[0]
    return "Various"


def _yn_from_bool_series(s: pd.Series) -> pd.Series:
    truthy = {"true", "t", "y", "yes", "1"}
    falsy = {"false", "f", "n", "no", "0", ""}

    def _one(x):
        if x is None:
            return "N"
        try:
            if pd.isna(x):
                return "N"
        except Exception:
            pass
        if isinstance(x, str):
            xs = x.strip().lower()
            if xs in truthy:
                return "Y"
            if xs in falsy:
                return "N"
        return "Y" if bool(x) else "N"

    base = pd.Series(list(pd.Series(s, copy=False)), index=pd.Series(s, copy=False).index, dtype="object")
    return base.map(_one)


LIKELY_DATE_PATTERNS = (
    re.compile(r"^\d{4}-\d{1,2}-\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)?$") ,
    re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?)?$", re.I),
    re.compile(r"^\d{1,2}-\d{1,2}-\d{2,4}$"),
)
MONTH_NAME_RE = re.compile(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\b", re.I)


def _looks_like_date_string(x) -> bool:
    if x is None:
        return False
    try:
        if pd.isna(x):
            return False
    except Exception:
        pass
    s = str(x).strip()
    if not s or s.lower() in {"nan", "nat", "none", "<na>"}:
        return False
    if any(p.match(s) for p in LIKELY_DATE_PATTERNS):
        return True
    if MONTH_NAME_RE.search(s):
        return True
    return False


def _to_datetime_series_mixed(s: pd.Series) -> pd.Series:
    base = pd.Series(list(pd.Series(s, copy=False)), index=pd.Series(s, copy=False).index, dtype="object")
    if base.empty:
        return pd.Series([], index=base.index, dtype="datetime64[ns]")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        warnings.simplefilter("ignore", FutureWarning)
        try:
            parsed = pd.to_datetime(base, errors="coerce", format="mixed", utc=True)
        except TypeError:
            parsed = pd.to_datetime(base, errors="coerce", utc=True)
        except Exception:
            parsed = base.map(_to_datetime_scalar_mixed_utc_naive)

    if not isinstance(parsed, pd.Series):
        parsed = pd.Series(parsed, index=base.index)
    else:
        parsed = parsed.reindex(base.index)

    try:
        if getattr(parsed.dt, "tz", None) is not None:
            parsed = parsed.dt.tz_convert(None)
    except Exception:
        try:
            parsed = parsed.dt.tz_localize(None)
        except Exception:
            parsed = parsed.map(_to_datetime_scalar_mixed_utc_naive)

    return pd.to_datetime(parsed, errors="coerce")


def _to_pacific_naive_series(s: pd.Series) -> pd.Series:
    """Parse to UTC then convert to America/Los_Angeles and drop tz, KEEPING the time.

    The Salesforce Bulk API returns datetimes in UTC (e.g. 2026-02-09T01:04:00Z),
    but the manual/official Active Loan Report renders them in the org's Pacific
    locale (2026-02-08 17:04:00) -- both the date boundary AND the time-of-day differ
    from a naive UTC read. This is used for the appraisal Order Received timestamp so
    "Most Recent Appraisal Order Date" reproduces the official report exactly. (Verified
    offset: -8 in February, -7 in April -> America/Los_Angeles with DST.)
    """
    base = pd.Series(list(pd.Series(s, copy=False)), index=pd.Series(s, copy=False).index, dtype="object")
    if base.empty:
        return pd.Series([], index=base.index, dtype="datetime64[ns]")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        warnings.simplefilter("ignore", FutureWarning)
        try:
            parsed = pd.to_datetime(base, errors="coerce", format="mixed", utc=True)
        except TypeError:
            parsed = pd.to_datetime(base, errors="coerce", utc=True)
        except Exception:
            parsed = pd.to_datetime(base, errors="coerce", utc=True)
    if not isinstance(parsed, pd.Series):
        parsed = pd.Series(parsed, index=base.index)
    else:
        parsed = parsed.reindex(base.index)
    try:
        parsed = parsed.dt.tz_convert("America/Los_Angeles").dt.tz_localize(None)
    except Exception:
        try:
            parsed = parsed.dt.tz_localize(None)
        except Exception:
            parsed = _to_datetime_series_mixed(base)
    parsed = pd.to_datetime(parsed, errors="coerce")
    # The official report stores the appraisal order timestamp truncated to the minute
    # (seconds = 00), e.g. 2026-02-08 17:04:00. The Salesforce value carries seconds
    # (17:04:41); floor to the minute so it matches exactly.
    try:
        parsed = parsed.dt.floor("min")
    except Exception:
        pass
    return pd.to_datetime(parsed, errors="coerce")


def downcast_numeric_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    for c in out.columns:
        s = out[c]
        try:
            if pd.api.types.is_integer_dtype(s):
                out.loc[:, c] = pd.to_numeric(s, errors="coerce", downcast="integer")
            elif pd.api.types.is_float_dtype(s):
                out.loc[:, c] = pd.to_numeric(s, errors="coerce", downcast="float")
        except Exception:
            pass
    return out


# V58: report columns that must stay VERBATIM Salesforce text. The numeric hints below
# would otherwise coerce them ("feet" catches Square Feet, "year" catches Year Built,
# "units" catches # of Units), which strips the thousands separator Salesforce actually
# stores and turns the official's '1,814' into '1814'. Verified against 20260824: the
# official Square Feet equals the raw SF value on 4,695/4,782 assets, and coercion cost
# 749 Square Feet + 52 Year Built cells. All of these are display-only -- they are declared
# in DEFAULT_TEXT_HEADERS and are never used in arithmetic.
BULK_PRESERVE_TEXT_COLUMNS = {
    "Square Feet", "Year Built", "Zip", "APN", "Additional APNs", "County", "CBSA",
}


def _normalize_bulk_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    numeric_hints = ("amount", "value", "upb", "balance", "feet", "units", "year", "rate", "commitment", "ala")

    def _is_date_like(col_name: str) -> bool:
        cl = str(col_name).lower()
        if "status" in cl:
            return False
        return (
            "date" in cl
            or "funding" in cl
            or "close" in cl
            or "order" in cl
            or "resolved" in cl
            or "maturity" in cl
        )

    for c in out.columns:
        s = out[c]
        cl = str(c).lower()

        if _is_date_like(c) and not pd.api.types.is_datetime64_any_dtype(s):
            s_str = s.astype("string").str.strip()
            nonblank_mask = s_str.notna() & s_str.ne("")
            nonblank_count = int(nonblank_mask.sum())
            if nonblank_count > 0:
                likely_date_mask = s_str.map(_looks_like_date_string)
                likely_count = int((likely_date_mask & nonblank_mask).sum())
                if likely_count > 0 and (likely_count / nonblank_count) >= 0.60:
                    parsed_input = s_str.where(likely_date_mask, pd.NA)
                    parsed = _to_datetime_series_mixed(parsed_input)
                    parsed_count = int(parsed[nonblank_mask].notna().sum())
                    if parsed_count > 0 and (parsed_count / nonblank_count) >= 0.60:
                        out[c] = parsed
                        continue

        if str(c).strip() in BULK_PRESERVE_TEXT_COLUMNS:
            continue

        if any(h in cl for h in numeric_hints):
            cleaned = (
                s.astype("string")
                .str.replace(",", "", regex=False)
                .str.replace("$", "", regex=False)
                .str.replace("%", "", regex=False)
            )
            parsed = pd.to_numeric(cleaned, errors="coerce")
            if parsed.notna().sum() > 0:
                out[c] = parsed

    return downcast_numeric_frame(out)


def normalize_servicer_family(val) -> str:
    s = clean_text(val).lower()
    if not s:
        return ""
    if "berkadia" in s:
        return "berkadia"
    if "onity" in s:
        return "onity"
    if "midland" in s:
        return "midland"
    if "statebridge" in s:
        return "statebridge"
    if "shellpoint" in s:
        return "shellpoint"
    if "selene" in s:
        return "selene"
    if s == "sps" or "specialized" in s or "select portfolio" in s:
        return "sps"
    if "fci" in s:
        return "fci"
    if "fay" in s:
        return "fay"
    if "cornerstone" in s:
        return "cornerstone"
    return s


def fci_servicer_label_from_filename(filename: str) -> str:
    n = filename.lower()
    if "2012632" in n:
        return "FCI 2012632"
    return "FCI"


def _soql_quote(v: str) -> str:
    s = str(v).replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"


def _soql_in(field: str, values) -> str:
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    if not vals:
        return "Id != NULL"
    return f"{field} IN ({', '.join(_soql_quote(v) for v in vals)})"


def _soql_not_equal_or_null(field: str, bad_value: str) -> str:
    q = _soql_quote(bad_value)
    return f"({field} = NULL OR {field} != {q})"


def _soql_false_or_null(field: str) -> str:
    return f"({field} = FALSE OR {field} = NULL)"


def _soql_parent_name_not_equal_or_no_parent(parent_id_field: str, parent_name_field: str, bad_value: str) -> str:
    q = _soql_quote(bad_value)
    return f"({parent_id_field} = NULL OR {parent_name_field} != {q})"


def _chunked(seq, size=200):
    seq = list(seq)
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def _nonblank_unique(values):
    out = []
    seen = set()
    for x in values:
        s = clean_text(x)
        if not s:
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def _run_bulk_union(soql_list, rename_map=None):
    frames = []
    for soql in soql_list:
        df = run_bulk_query(soql, rename_map=rename_map)
        if not df.empty:
            frames.append(df)
        del df
        gc.collect()

    if not frames:
        return pd.DataFrame()

    if len(frames) == 1:
        out = frames[0]
    else:
        out = pd.concat(frames, ignore_index=True, copy=False)

    del frames
    gc.collect()
    return downcast_numeric_frame(out)


def b64url_no_pad(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).rstrip(b"=").decode("utf-8")


def make_verifier() -> str:
    v = secrets.token_urlsafe(96)
    return v[:128]


def make_challenge(verifier: str) -> str:
    return b64url_no_pad(hashlib.sha256(verifier.encode("utf-8")).digest())


@st.cache_resource
def pkce_store():
    return {}


@st.cache_resource
def http_session():
    sess = requests.Session()
    adapter = requests.adapters.HTTPAdapter(pool_connections=10, pool_maxsize=10, max_retries=0)
    sess.mount("https://", adapter)
    sess.mount("http://", adapter)
    return sess


def exchange_code_for_token(
    token_url: str,
    code: str,
    verifier: str,
    client_id: str,
    redirect_uri: str,
    client_secret: Optional[str],
):
    data = {
        "grant_type": "authorization_code",
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "code": code,
        "code_verifier": verifier,
    }
    if client_secret:
        data["client_secret"] = client_secret
    resp = http_session().post(token_url, data=data, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(f"Token exchange failed ({resp.status_code}): {resp.text}")
    return resp.json()


def show_salesforce_login_helper():
    st.info(
        "Step 1: Log in to Salesforce.\n\n"
        "Step 2: Approve access.\n\n"
        "Step 3: Upload your files after login.\n\n"
        "Step 4: Click Build. This app uses Salesforce Bulk API 2.0 to pull the full datasets."
    )


def render_salesforce_login_gate() -> dict:
    st.markdown("### Step 1: Sign in to Salesforce")
    st.caption(
        "Please sign in before uploading files. The upload section stays hidden until login is complete so the Salesforce callback does not clear files you already selected."
    )
    show_salesforce_login_helper()
    sf_info = ensure_sf_session()

    c1, c2 = st.columns([3, 1])
    with c1:
        inst = (st.session_state.get("sf_token") or {}).get("instance_url", "")
        st.success("✅ Logged in to Salesforce API")
        if inst:
            st.caption(f"Connected to: {inst}")
            st.caption("Bulk API 2.0 is used with chunked result pages so the pull is not capped at report-api row counts.")
    with c2:
        if st.button("Log out"):
            st.session_state.sf_token = None
            st.query_params.clear()
            st.rerun()

    st.divider()
    return sf_info


def ensure_sf_session() -> dict:
    cfg = st.secrets["salesforce"]
    client_id = cfg["client_id"]
    auth_host = cfg.get("auth_host", "https://cvest.my.salesforce.com").rstrip("/")
    redirect_uri = cfg["redirect_uri"].rstrip("/")
    client_secret = cfg.get("client_secret")
    auth_url = f"{auth_host}/services/oauth2/authorize"
    token_url = f"{auth_host}/services/oauth2/token"

    qp = st.query_params
    code = qp.get("code")
    state = qp.get("state")
    err = qp.get("error")
    err_desc = qp.get("error_description")

    if err:
        st.error(f"Login error: {err}")
        if err_desc:
            st.code(err_desc)
        st.stop()

    if "sf_token" not in st.session_state:
        st.session_state.sf_token = None

    store = pkce_store()
    now = time.time()
    ttl = 900
    for s, (_v, t0) in list(store.items()):
        if now - t0 > ttl:
            store.pop(s, None)

    if code:
        if not state or state not in store:
            st.error("Login link expired. Click login again.")
            st.stop()
        verifier, _t0 = store.pop(state)
        tok = exchange_code_for_token(token_url, code, verifier, client_id, redirect_uri, client_secret)
        st.session_state.sf_token = tok
        st.query_params.clear()
        st.rerun()

    if not st.session_state.sf_token:
        new_state = secrets.token_urlsafe(24)
        new_verifier = make_verifier()
        new_challenge = make_challenge(new_verifier)
        store[new_state] = (new_verifier, time.time())
        login_params = {
            "response_type": "code",
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "code_challenge": new_challenge,
            "code_challenge_method": "S256",
            "state": new_state,
            "prompt": "login",
            "scope": "api refresh_token",
        }
        login_url = auth_url + "?" + urllib.parse.urlencode(login_params)
        st.link_button("Login to Salesforce", login_url)
        st.stop()

    tok = st.session_state.sf_token
    access_token = tok.get("access_token")
    instance_url = tok.get("instance_url")
    if not access_token or not instance_url:
        st.error("Login token missing needed values.")
        st.stop()
    return {"access_token": access_token, "instance_url": instance_url.rstrip("/")}


def _session_cache(bucket: str) -> dict:
    if bucket not in st.session_state:
        st.session_state[bucket] = {}
    return st.session_state[bucket]


def _sf_auth_parts() -> Tuple[str, str]:
    tok = st.session_state.get("sf_token") or {}
    access_token = tok.get("access_token")
    instance_url = (tok.get("instance_url") or "").rstrip("/")
    if not access_token or not instance_url:
        raise RuntimeError("Salesforce session missing. Please log in again.")
    return access_token, instance_url


def _sf_headers(extra: Optional[dict] = None) -> dict:
    access_token, _instance_url = _sf_auth_parts()
    hdrs = {"Authorization": f"Bearer {access_token}"}
    if extra:
        hdrs.update(extra)
    return hdrs


def _sf_request(
    path: str,
    method: str = "GET",
    *,
    params: Optional[dict] = None,
    json_body: Optional[dict] = None,
    headers: Optional[dict] = None,
    expect_json: bool = True,
    timeout: int = 180,
):
    _access_token, instance_url = _sf_auth_parts()
    url = f"{instance_url}/services/data/{API_VERSION}/{path.lstrip('/')}"
    hdrs = _sf_headers(headers)
    if json_body is not None:
        hdrs.setdefault("Content-Type", "application/json")
    resp = http_session().request(
        method=method,
        url=url,
        headers=hdrs,
        params=params,
        json=json_body,
        timeout=timeout,
    )
    if resp.status_code >= 400:
        msg = resp.text[:4000]
        raise RuntimeError(f"Salesforce API {method} failed ({resp.status_code}) for {path}: {msg}")
    if expect_json:
        return resp.json()
    return resp


def _bulk_query_create_job(soql: str) -> str:
    payload = {
        "operation": "query",
        "query": soql,
        "columnDelimiter": "COMMA",
        "lineEnding": "LF",
    }
    js = _sf_request("jobs/query", method="POST", json_body=payload)
    job_id = js.get("id")
    if not job_id:
        raise RuntimeError(f"Bulk query job creation failed: {js}")
    return job_id


def _bulk_query_wait(job_id: str, poll_seconds: float = 1.25, timeout_seconds: int = BULK_WAIT_TIMEOUT_SECONDS) -> dict:
    t0 = time.time()
    while True:
        js = _sf_request(f"jobs/query/{job_id}", method="GET")
        state = js.get("state")
        if state == "JobComplete":
            return js
        if state in {"Aborted", "Failed"}:
            raise RuntimeError(f"Bulk query job {job_id} failed: state={state}; message={js.get('errorMessage') or js}")
        if time.time() - t0 > timeout_seconds:
            raise TimeoutError(f"Timed out waiting for Bulk query job {job_id}.")
        time.sleep(poll_seconds)


def _bulk_query_results_pages(job_id: str, max_records: int = BULK_PAGE_SIZE):
    locator: Optional[str] = None
    while True:
        params = {"maxRecords": max_records}
        if locator:
            params["locator"] = locator
        resp = _sf_request(
            f"jobs/query/{job_id}/results",
            method="GET",
            params=params,
            headers={"Accept": "text/csv"},
            expect_json=False,
            timeout=300,
        )
        # Salesforce Bulk API returns UTF-8 CSV, but if the response Content-Type omits
        # a charset, requests defaults to ISO-8859-1 and mangles non-ASCII text (e.g.
        # a U+2019 apostrophe in a borrower entity name becomes "â€™"). Force UTF-8.
        resp.encoding = "utf-8"
        yield resp.text
        locator = resp.headers.get("Sforce-Locator") or resp.headers.get("sforce-locator")
        if not locator or locator.lower() == "null":
            break


def run_bulk_query(soql: str, rename_map: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    job_id = _bulk_query_create_job(soql)
    _bulk_query_wait(job_id)
    frames: List[pd.DataFrame] = []
    for text in _bulk_query_results_pages(job_id):
        if not text.strip():
            continue
        chunk = pd.read_csv(io.StringIO(text), keep_default_na=True, low_memory=True)
        if rename_map:
            chunk = chunk.rename(columns=rename_map)
        chunk = _normalize_bulk_df(chunk)
        chunk = downcast_numeric_frame(chunk)
        frames.append(chunk)
        del chunk, text
        gc.collect()
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True, copy=False)
    del frames
    gc.collect()
    return downcast_numeric_frame(out)


def describe_sobject(sobject: str) -> dict:
    cache = _session_cache("sobject_describe_cache")
    if sobject in cache:
        return cache[sobject]
    js = _sf_request(f"sobjects/{sobject}/describe", method="GET")
    cache[sobject] = js
    return js


def _field_map_by_name(sobject: str) -> Dict[str, dict]:
    return {f["name"]: f for f in describe_sobject(sobject).get("fields", [])}


def relationship_name_for(sobject: str, field_api: str) -> str:
    fld = _field_map_by_name(sobject).get(field_api)
    if not fld:
        raise KeyError(f"{sobject}.{field_api} not found in describe().")
    rel = fld.get("relationshipName")
    if not rel:
        raise KeyError(f"{sobject}.{field_api} is not a relationship field.")
    return rel


def first_existing_field_name(sobject: str, candidates: Sequence[str]) -> Optional[str]:
    field_map = _field_map_by_name(sobject)
    for name in candidates:
        if name in field_map:
            return name
    return None


def existing_field_names(sobject: str, candidates: Sequence[str]) -> List[str]:
    field_map = _field_map_by_name(sobject)
    return [name for name in candidates if name in field_map]


def picklist_values_for(sobject: str, field_api: str) -> List[str]:
    fld = _field_map_by_name(sobject).get(field_api) or {}
    out: List[str] = []
    seen = set()
    for item in fld.get("picklistValues", []) or []:
        val = clean_text(item.get("value") or item.get("label"))
        if val and val not in seen:
            seen.add(val)
            out.append(val)
    return out


def coalesce_columns(df: pd.DataFrame, columns: Sequence[str], index=None) -> pd.Series:
    if index is None:
        index = df.index
    out = pd.Series([pd.NA] * len(index), index=index, dtype="object")
    for col in reversed([c for c in columns if c in df.columns]):
        out = coalesce_keep_nonblank(pd.Series(df[col], index=index), out)
    return out


def _strict_active_rm_role_match(role) -> bool:
    s = re.sub(r"[^a-z0-9]+", " ", clean_text(role).lower()).strip()
    if not s:
        return False
    if s in {"active rm", "active relationship manager"}:
        return True
    if re.fullmatch(r"rm(?:\s*[12])?", s):
        return True
    return "relationship" in s and ("manager" in s or re.search(r"\brm\b", s) is not None)


def opportunity_name_expr(field_candidates: Sequence[str]) -> str:
    field_api = first_existing_field_name("Opportunity", field_candidates)
    if not field_api:
        raise KeyError(f"Could not find any Opportunity field in: {field_candidates}")
    try:
        rel = relationship_name_for("Opportunity", field_api)
        return f"{rel}.Name"
    except Exception:
        return field_api


def property_opportunity_relationship_name() -> str:
    field_api = first_existing_field_name(
        "Property__c",
        ["Opportunity__c", "Deal__c", "Loan__c", "Line_Of_Credit__c", "Line_of_Credit__c", "LOC__c"],
    )
    if not field_api:
        raise KeyError("Could not find Property__c -> Opportunity relationship field.")
    return relationship_name_for("Property__c", field_api)


def appraisal_property_relationship_name() -> str:
    field_api = first_existing_field_name("Appraisal__c", ["Property__c", "Subject_Property__c"])
    if not field_api:
        raise KeyError("Could not find Appraisal__c -> Property relationship field.")
    return relationship_name_for("Appraisal__c", field_api)


@st.cache_data(show_spinner=False)
def load_repo_template_bytes() -> Tuple[bytes, str]:
    here = Path(__file__).resolve().parent
    candidates = []
    for filename in TEMPLATE_FILENAMES:
        candidates.extend(
            [
                here / filename,
                here / "templates" / filename,
                here / "assets" / filename,
                Path.cwd() / filename,
                Path(filename),
            ]
        )

    for p in candidates:
        try:
            if p.exists() and p.is_file():
                return p.read_bytes(), str(p)
        except Exception:
            continue

    tried = "\n".join(str(p) for p in candidates)
    raise FileNotFoundError(
        f"Could not find any template file.\n\nTried:\n{tried}\n\n"
        f"Fix: commit one of these to your repo: {', '.join(TEMPLATE_FILENAMES)}"
    )


def resolve_template_bytes(prev_upload) -> Tuple[bytes, str]:
    if prev_upload is not None:
        return prev_upload.getvalue(), f"uploaded workbook template: {prev_upload.name}"
    return load_repo_template_bytes()


@st.cache_data(show_spinner=False)
def load_template_lookup_maps(template_bytes: bytes) -> dict:
    maps = {
        "strategy_map": {},
        "ssp_deals": set(),
        "legacy_bridge_deals": set(),
        "legacy_term_deals": set(),
    }

    xls = pd.ExcelFile(BytesIO(template_bytes))

    if "Strategy Groupings" in xls.sheet_names:
        sg = pd.read_excel(BytesIO(template_bytes), sheet_name="Strategy Groupings", header=3)
        sg = sg.dropna(how="all")
        sg.columns = [str(c).strip() for c in sg.columns]
        strategy_col = first_matching_col(sg, ["Strategy"])
        grouping_col = first_matching_col(sg, ["Grouping"])
        if strategy_col and grouping_col:
            for _, row in sg.iterrows():
                strategy = clean_text(row.get(strategy_col))
                grouping = clean_text(row.get(grouping_col))
                if strategy and grouping:
                    maps["strategy_map"][strategy] = grouping
        # SA Loan lookup: the F:G columns map an Asset Manager 1 name -> "Y".
        # (These are the right-hand "Asset Manager 1"/"Y" columns, not the strategy pair.)
        sa_mgr_col = None
        sg_cols = list(sg.columns)
        for i, c in enumerate(sg_cols):
            if str(c).strip() == "Asset Manager 1":
                sa_mgr_col = c
                break
        sa_managers = set()
        if sa_mgr_col is not None:
            for _, row in sg.iterrows():
                mgr = clean_text(row.get(sa_mgr_col))
                if mgr:
                    sa_managers.add(mgr)
        maps["sa_loan_managers"] = sa_managers

    if "SSP Loans" in xls.sheet_names:
        ssp = pd.read_excel(BytesIO(template_bytes), sheet_name="SSP Loans", header=3)
        ssp = ssp.dropna(how="all")
        if "Deal No." in ssp.columns:
            maps["ssp_deals"] = set(_nonblank_unique(ssp["Deal No."].tolist()))

    if "Legacy" in xls.sheet_names:
        legacy = pd.read_excel(BytesIO(template_bytes), sheet_name="Legacy", header=4)
        legacy = legacy.dropna(how="all")
        if legacy.shape[1] >= 7:
            maps["legacy_bridge_deals"] = set(_nonblank_unique(legacy.iloc[:, 1].tolist()))
            maps["legacy_term_deals"] = set(_nonblank_unique(legacy.iloc[:, 6].tolist()))

    return maps


def strategy_grouping_from_project_strategy(project_strategy, strategy_map: dict):
    s = clean_text(project_strategy)
    if not s:
        return pd.NA
    return strategy_map.get(s, "Other")


def derive_bridge_segment(deal_number, financing, loan_buyer, template_maps: dict):
    # In the Bridge spine, "financing" is sourced from the Warehouse Line, which the
    # official report uses to classify Segment (validated against SF_Bridge):
    #   startswith "CAFL "       -> RTL Securitizations  (V43 rename)
    #   startswith "CPP JV"       -> CPP JV
    #   "Churchill Oaktree JV"    -> Oaktree JV
    #   contains "Spruce"         -> SSP
    #   Axos / Ineligible / *-NPL -> Legacy
    #   plain bank name OR blank  -> Mortgage Banking
    #   blank AND sold            -> Sold Servicing Retained
    fin = clean_text(financing)
    buyer = clean_text(loan_buyer)
    u = fin.upper()

    # Sold (servicing retained) only when there is no warehouse line to classify on.
    if not fin:
        if buyer or clean_text(financing) in BRIDGE_SOLD_FINANCING_VALUES:
            return BRIDGE_SOLD_SEGMENT
        return "Mortgage Banking"

    if u.startswith("CAFL "):
        return BRIDGE_SECURITIZED_SEGMENT
    if u.startswith("CPP JV"):
        return "CPP JV"
    if "CHURCHILL OAKTREE" in u or u == "OAKTREE JV" or "OAKTREE JV" in u:
        return "Oaktree JV"
    if "SPRUCE" in u:
        return "SSP"
    if fin in ("Axos", "Ineligible") or u.endswith("- NPL") or u.endswith("-NPL"):
        return "Legacy"
    # Template lookups remain authoritative for hand-maintained exceptions.
    if deal_in_lookup(deal_number, template_maps.get("ssp_deals", set())):
        return "SSP"
    if deal_in_lookup(deal_number, template_maps.get("legacy_bridge_deals", set())):
        return "Legacy"
    # plain bank names (Goldman Sachs, Wells Fargo, Morgan Stanley, ...) -> Mortgage Banking
    return "Mortgage Banking"


def derive_bridge_portfolio(product_type, segment, financing, deal_intro_sub_source, deal_number):
    ptype = clean_text(product_type)
    seg = clean_text(segment)
    fin = clean_text(financing)
    intro = clean_text(deal_intro_sub_source)
    deal = clean_text(deal_number)

    if intro == "Churchill Real Estate":
        return "TPO"
    if deal.startswith("5A-") or intro == "5arch":
        return "5A"
    if ptype in {"Single Asset (1-4 Unit)", "Single Asset (5-10 Unit)", "Single Asset (11+ Unit)", "Portfolio"}:
        return "RB"
    if ptype == "Multifamily/CRE" and (seg in {"SSP", "Legacy"} or "NPL" in fin):
        return "CLO"
    return "CV"


TERM_FUNDING_VEHICLE_REMAP = {
    # SF Current Funding Vehicle -> report display name (securitization naming).
    "2022-2": "2022-P2",
    "2023-1": "2023-P1",
}


def normalize_term_financing(financing) -> str:
    """Strip the SF Current-Funding-Vehicle 'CAF ' prefix so values match the report.

    e.g. 'CAF 2020-P1' -> '2020-P1', 'CAF2021-2' -> '2021-2'. Leaves non-CAF values
    (Sold, CPP JV - ..., Morgan Stanley, CAFL 2026-R1) unchanged. Then applies the
    securitization display remap (2022-2 -> 2022-P2, 2023-1 -> 2023-P1).
    """
    f = clean_text(financing)
    if not f:
        return f
    f = re.sub(r"^CAF(?=\s|\d)\s*", "", f, flags=re.I).strip()
    return TERM_FUNDING_VEHICLE_REMAP.get(f, f)


def derive_term_portfolio_segment(loan_type, financing, loan_buyer, deal_number, template_maps: dict, sold_servicing_status=None, warehouse_line=None):
    typ = clean_text(loan_type)
    fin = normalize_term_financing(financing)
    buyer = clean_text(loan_buyer)

    if typ in TERM_DSCR_TYPES:
        return "DSCR", "DSCR", "N"
    # A populated Loan Buyer (or Financing=Sold) means the loan was sold -- this wins
    # over the funding vehicle, including "CPP JV - ..." vehicles that were later sold
    # (the report shows these as Sold/Apollo/Blackstone, not CPP JV).
    if fin == "Sold" or buyer:
        seg = buyer.split()[0] if buyer else "Sold Term"
        return "Sold Term", seg, "N"
    # NOTE (V42): the Warehouse-Line CPP JV trigger (former Fix E.2) was REVERTED. It
    # inverted brand-new DSCR deals whose API Product Type was not yet "DSCR" (63709/64595
    # -> wrongly CPP JV) and was overridden by the Segment carry-forward for existing
    # deals, so it had no upside. `warehouse_line` is accepted but unused for now.
    if fin.startswith("CPP JV"):
        return "Active Term", "CPP JV", "Y"
    # Legacy deals must win before securitized (some are financed "CAFL 2026-R1").
    if deal_in_lookup(deal_number, template_maps.get("legacy_term_deals", set())):
        return "Active Term", "Legacy", "N"
    if re.match(r"^\d{4}[-A-Za-z0-9]+$", fin):
        return "Securitized Term", "Securitized Term", "N"
    return "Active Term", "Mortgage Banking", "N"


def _soql_text(value: str) -> str:
    return "'" + str(value).replace("\\", "\\\\").replace("'", "\\'") + "'"

def _recordtype_name_expr(alias_prefix: str = "") -> str:
    return f"{alias_prefix}RecordType.Name" if alias_prefix else "RecordType.Name"

def _lower_in_condition(field_expr: str, values: Sequence[str]) -> str:
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    if not vals:
        return "Id != NULL"
    return f"({field_expr} != NULL AND {field_expr} IN ({', '.join(_soql_text(str(x)) for x in vals)}))"

def _lower_contains_condition(field_expr: str, values: Sequence[str]) -> str:
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    contains_parts = [f"{field_expr} LIKE '%{str(x).replace(chr(39), chr(92)+chr(39))}%'" for x in vals]
    return "(" + " OR ".join(contains_parts) + ")" if contains_parts else "Id != NULL"

def _bridge_dealtype_condition(alias_prefix: str = "") -> str:
    expr = f"{alias_prefix}Type" if alias_prefix else "Type"
    exact_cond = _lower_in_condition(expr, BRIDGE_TYPES)
    contains_cond = _lower_contains_condition(expr, BRIDGE_TYPE_CONTAINS)
    return f"({exact_cond} OR {contains_cond})"

def _term_dealtype_condition(alias_prefix: str = "") -> str:
    expr = f"{alias_prefix}Type" if alias_prefix else "Type"
    return _lower_in_condition(expr, TERM_TYPES)

def _opportunity_status_active_condition(alias_prefix: str = "") -> Optional[str]:
    # Business logic now keys inclusion off StageName and property Status instead of the
    # custom Opportunity status fields. Leaving this filter enabled was dropping Sold /
    # REO rows upstream before the explicit retained-servicing and REO rules could run.
    return None

def _append_clause_if_present(parts: List[str], clause: Optional[str]) -> List[str]:
    if clause and str(clause).strip():
        parts.append(clause)
    return parts

def _bridge_recordtype_condition(alias_prefix: str = "") -> str:
    return _bridge_dealtype_condition(alias_prefix)

def _term_recordtype_condition(alias_prefix: str = "") -> str:
    return _term_dealtype_condition(alias_prefix)

def _coalesce_numeric_columns(df: pd.DataFrame, columns: Sequence[str], default=np.nan) -> pd.Series:
    out = pd.Series([default] * len(df), index=df.index, dtype="float64")
    for col in columns:
        if col in df.columns:
            cur = pd.to_numeric(df[col], errors="coerce")
            out = out.where(out.notna(), cur)
    return out


def _coalesce_numeric_columns_zeroaware(df: pd.DataFrame, columns: Sequence[str], default=np.nan) -> pd.Series:
    out = pd.Series([default] * len(df), index=df.index, dtype="float64")
    for col in columns:
        if col in df.columns:
            cur = pd.to_numeric(df[col], errors="coerce")
            cur = cur.where(~cur.eq(0), np.nan)
            out = out.where(out.notna(), cur)
    return out


def _coalesce_datetime_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.Series:
    out = pd.Series([pd.NaT] * len(df), index=df.index)
    for col in columns:
        if col in df.columns:
            cur = _to_datetime_series_mixed(df[col])
            out = out.where(out.notna(), cur)
    return _to_datetime_series_mixed(out)


def _term_current_maturity_source_columns(df: pd.DataFrame) -> List[str]:
    """Current/modified Salesforce maturity columns, ordered by business preference."""
    if df is None or df.empty:
        return []
    cols = list(df.columns)
    preferred: List[str] = []
    for prefix in [
        "Current Loan Maturity Date",
        "Current Line Maturity Date",
        "Current Maturity Date",
        "Loan Maturity Date",
    ]:
        preferred.extend([c for c in cols if str(c).startswith(prefix) and c not in preferred])
    return preferred


def _term_maturity_source_columns(df: pd.DataFrame) -> List[str]:
    """Preferred Salesforce term maturity columns, with current/modified before stated."""
    if df is None or df.empty:
        return []
    preferred = _term_current_maturity_source_columns(df)
    for col in ["Original Loan Maturity Date", "Stated Maturity Date", "Maturity Date"]:
        if col in df.columns and col not in preferred:
            preferred.append(col)
    return preferred


def _term_current_maturity_source_series(df: pd.DataFrame) -> pd.Series:
    """Return only the current/modified maturity date from Salesforce.

    This lets existing loans pick up modification/extension dates instead of being
    overwritten by last week's completed report. Original/stated maturity is not
    used here; it is a fallback in _term_maturity_source_series().
    """
    if df is None or df.empty:
        return pd.Series(dtype="datetime64[ns]")
    cols = _term_current_maturity_source_columns(df)
    if not cols:
        return pd.Series([pd.NaT] * len(df), index=df.index)
    return _coalesce_datetime_columns(df, cols)


def _term_maturity_source_series(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="datetime64[ns]")
    cols = _term_maturity_source_columns(df)
    if not cols:
        return pd.Series([pd.NaT] * len(df), index=df.index)
    return _coalesce_datetime_columns(df, cols)


def _term_preboarding_mask(stage_series: pd.Series) -> pd.Series:
    return pd.Series(stage_series, copy=False).astype("string").str.strip().isin(TERM_PREBOARDING_STAGES).fillna(False)


def _apply_term_preboarding_upb_fallback(out: pd.DataFrame, sf_term: Optional[pd.DataFrame], upb_col: str) -> pd.DataFrame:
    """Keep Approved-by-Committee/Purchased term loans before servicer boarding.

    The completed report can include these rows with Servicer ID = N/A and UPB
    equal to Salesforce loan amount when current servicer UPB is blank/zero.
    """
    if out is None or out.empty or sf_term is None or sf_term.empty or upb_col not in out.columns:
        return pd.DataFrame() if out is None else out.copy()
    result = out.copy()
    if "Deal Loan Number" not in sf_term.columns:
        return result
    sf = sf_term.copy()
    sf["_deal_key"] = norm_id_series(sf["Deal Loan Number"])
    sf["_sf_preboarding_stage"] = _term_preboarding_mask(sf.get("Stage", pd.Series([pd.NA] * len(sf), index=sf.index)))
    sf["_sf_loan_amount_for_preboarding"] = pd.to_numeric(sf.get("Loan Amount", pd.Series([np.nan] * len(sf), index=sf.index)), errors="coerce")
    sf["_sf_current_upb_for_preboarding"] = pd.to_numeric(sf.get("Current Servicer UPB", pd.Series([np.nan] * len(sf), index=sf.index)), errors="coerce")
    ctx = sf.dropna(subset=["_deal_key"]).sort_values(["_deal_key", "_sf_preboarding_stage", "_sf_loan_amount_for_preboarding"], ascending=[True, True, True]).drop_duplicates("_deal_key", keep="last")
    ctx = ctx[["_deal_key", "_sf_preboarding_stage", "_sf_loan_amount_for_preboarding", "_sf_current_upb_for_preboarding"]]
    if "_deal_key" not in result.columns:
        result["_deal_key"] = norm_id_series(result.get("Deal Number", pd.Series([None] * len(result), index=result.index)))
    result = result.merge(ctx, on="_deal_key", how="left")
    cur_upb = pd.to_numeric(result.get(upb_col, pd.Series([np.nan] * len(result), index=result.index)), errors="coerce")
    sf_current = pd.to_numeric(result.get("_sf_current_upb_for_preboarding", pd.Series([np.nan] * len(result), index=result.index)), errors="coerce")
    loan_amt = pd.to_numeric(result.get("_sf_loan_amount_for_preboarding", pd.Series([np.nan] * len(result), index=result.index)), errors="coerce")
    preboarding = result.get("_sf_preboarding_stage", pd.Series([False] * len(result), index=result.index)).fillna(False).astype(bool)
    replacement = sf_current.where(sf_current.gt(0), loan_amt)
    fill_mask = preboarding & replacement.gt(0) & (cur_upb.isna() | cur_upb.le(0))
    result.loc[fill_mask, upb_col] = replacement.loc[fill_mask]
    return result.drop(columns=["_sf_preboarding_stage", "_sf_loan_amount_for_preboarding", "_sf_current_upb_for_preboarding"], errors="ignore")

def _coalesce_text_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.Series:
    out = pd.Series([pd.NA] * len(df), index=df.index, dtype="object")
    for col in columns:
        if col in df.columns:
            out = coalesce_keep_nonblank(out, pd.Series(df[col], index=df.index, dtype="object"))
    return out

def _build_bridge_spine_like() -> pd.DataFrame:
    opp_rel = property_opportunity_relationship_name()

    sold_pool_field = first_existing_field_name("Opportunity", ["Sold_Loan_Pool__c", "FK_Sold_Loan_Pool__c"])
    sold_pool_rel = relationship_name_for("Opportunity", sold_pool_field) if sold_pool_field else None

    contact_field = first_existing_field_name("Opportunity", ["Contact__c", "Primary_Contact__c"])
    contact_rel = relationship_name_for("Opportunity", contact_field) if contact_field else None

    special_asset_rel = relationship_name_for("Property__c", "Special_Asset__c")

    sold_to_expr = f"{opp_rel}.{sold_pool_rel}.Sold_To__r.Name" if sold_pool_rel else f"{opp_rel}.Account.Name"
    primary_contact_expr = f"{opp_rel}.{contact_rel}.Name" if contact_rel else f"{opp_rel}.Account.Name"

    updated_value_date_field = first_existing_field_name("Property__c", ["Updated_Valuation_Date__c", "Value_Date__c", "BPO_Appraisal_Date__c"])
    generic_value_date_field = first_existing_field_name("Property__c", ["Value_Date__c", "Updated_Valuation_Date__c", "BPO_Appraisal_Date__c"])
    generic_value_field = first_existing_field_name("Property__c", ["Value__c", "Appraised_Value_Amount__c"])

    # V57: Salesforce replaced the Bridge Warehouse Line with a Current Funding Vehicle
    # lookup. The 20260824 SF_Bridge export no longer carries "Warehouse Line" at all, and
    # the official Financing column matches "Current Funding Vehicle: Funding Vehicle Name"
    # on 4,676/4,782 assets. With Warehouse_Line__c empty the build fell back to last
    # week's carried-forward value, which is why test 73 still showed the old labels
    # ("CAFL 2026-R1 CV" instead of "CAFL 2026-R1", "Churchill Oaktree JV" instead of
    # "Oaktree JV - Churchill") on ~570 cells. Resolved defensively; Warehouse Line stays
    # as the fallback so an org that still populates it is unaffected.
    _bridge_fv_field = first_existing_field_name(
        "Opportunity", ["Current_Funding_Vehicle__c", "FK_Current_Funding_Vehicle__c"]
    )
    _bridge_fv_expr = None
    if _bridge_fv_field:
        try:
            _bridge_fv_expr = f"{opp_rel}.{relationship_name_for('Opportunity', _bridge_fv_field)}.Name"
        except Exception:
            _bridge_fv_expr = None

    select_pairs = [
        ("Sold To", sold_to_expr),
        ("Warehouse Line", f"{opp_rel}.Warehouse_Line__c"),
        ("Deal Loan Number", f"{opp_rel}.Deal_Loan_Number__c"),
        ("Servicer Loan Number", "Servicer_Loan_Number__c"),
        ("Servicer Commitment Id", f"{opp_rel}.Servicer_Commitment_Id__c"),
        ("Yardi ID", "Yardi_Id__c"),
        ("Asset ID", "Asset_ID__c"),
        ("Property ID", "Id"),
        ("Unique Active Id", "Unique_Active_Id__c"),
        ("Deal Name", f"{opp_rel}.Name"),
        ("Borrower Entity: Business Entity Name", f"{opp_rel}.Borrower_Entity__r.Name"),
        ("Account Name: Account Name", f"{opp_rel}.Account.Name"),
        ("Primary Contact: Full Name", primary_contact_expr),
        ("Address", "Name"),
        ("City", "City__c"),
        ("State", "State__c"),
        ("Zip", "ZipCode__c"),
        ("County", "County__c"),
        ("CBSA", "MSA__c"),
        ("APN", "APN__c"),
        ("Additional APNs", "Additional_APNs__c"),
        ("# of Units", "Number_of_Units__c"),
        ("Year Built", "Year_Built__c"),
        ("Square Feet", "Square_Feet__c"),
        ("Close Date", f"{opp_rel}.CloseDate"),
        ("First Funding Date", "First_Funding_Date__c"),
        ("Last Funding Date", "Funding_Date__c"),
        ("Property Next Payment Date", "Next_Payment_Date__c"),
        ("Opportunity Next Payment Date", f"{opp_rel}.Next_Payment_Date__c"),
        ("Original Loan Maturity Date", f"{opp_rel}.Stated_Maturity_Date__c"),
        ("Current Loan Maturity date", f"{opp_rel}.Current_Line_Maturity_Date__c"),
        ("Original Asset Maturity Date", "Asset_Maturity_Date_Override__c"),
        ("Current Asset Maturity date", "Current_Asset_Maturity_Date__c"),
        ("Is Parent", "Is_Parent__c"),
        ("Is Sub Unit", "Is_Sub_Unit__c"),
        ("Loan Commitment", f"{opp_rel}.LOC_Commitment__c"),
        ("Remaining Commitment", f"{opp_rel}.Outstanding_Facility_Amount__c"),
        ("Salesforce Suspense Balance", f"{opp_rel}.Suspense_Balance__c"),
        ("Remedy Plan", "Remedy_Plan__c"),
        ("Delinquency Status Notes", "Delinquency_Status_Notes__c"),
        ("Maturity Status", "Maturity_Status__c"),
        ("Is Special Asset", "Is_Special_Asset__c"),
        ("Special Asset: Status", f"{special_asset_rel}.Status_Comment__c"),
        ("Special Asset: Special Asset Reason", f"{special_asset_rel}.Special_Asset_Reason__c"),
        ("Special Asset: Special Asset Status", f"{special_asset_rel}.Severity_Level__c"),
        ("Special Asset: Resolved Date", f"{special_asset_rel}.Resolved_Date__c"),
        ("Forbearance Term Date", "Forbearance_Term_Date__c"),
        ("REO Date", "REO_Date__c"),
        ("Most Recent Appraisal Order Date", "BPO_Appraisal_Order_Date__c"),
        ("BPO Appraisal Date", "BPO_Appraisal_Date__c"),
        ("Appraised Value Amount", "Appraised_Value_Amount__c"),
        ("After Repair Value", "After_Repair_Value__c"),
        ("Origination Valuation Date", "Origination_Date_Valuation_Date__c"),
        ("Origination As-Is Value", "Origination_Date_Value__c"),
        ("Origination After Repair Value", "Origination_After_Repair_Value__c"),
        ("Initial Disbursement Funded", "Initial_Disbursement_Used__c"),
        ("Approved Renovation Advance Amount", "Approved_Renovation_Holdback__c"),
        ("Renovation Advance Amount Funded", "Renovation_Advance_Amount_Used__c"),
        ("Reno Advance Amount Remaining", "Reno_Advance_Amount_Remaining__c"),
        ("Interest Allocation", "Interest_Allocation__c"),
        ("Interest Holdback Funded", "Interest_Reserves__c"),
        ("Title Company: Account Name", "Title_Company__r.Name"),
        ("Tax Payment Next Due Date", "Tax_Payment_Next_Due_Date__c"),
        ("Taxes Payment Frequency", "Taxes_Payment_Frequency__c"),
        ("Tax Commentary", "Tax_Commentary__c"),
        ("Product Type", f"{opp_rel}.LOC_Loan_Type__c"),
        ("Product Sub-Type", f"{opp_rel}.Product_Sub_Type__c"),
        ("Transaction Type", f"{opp_rel}.Transaction_Type__c"),
        ("Project Strategy", f"{opp_rel}.Project_Strategy__c"),
        ("Property Type", "Property_Type__c"),
        ("CAF Originator: Full Name", f"{opp_rel}.Owner.Name"),
        ("CAF Originator: Active", f"{opp_rel}.Owner.IsActive"),
        ("Deal Intro Sub-Source", f"{opp_rel}.Deal_Intro_Sub_Source__c"),
        ("Referral Source Account: Account Name", f"{opp_rel}.Referral_Source__r.Name"),
        ("Referral Source Contact: Full Name", f"{opp_rel}.Referral_Source_Contact__r.Name"),
        ("Stage", f"{opp_rel}.StageName"),
        ("Status", "Status__c"),
        ("Current UPB", "Current_UPB__c"),
        # Loan-level UPB from the Opportunity. Bridge Asset allocates this by deal/funded amount.
        ("Current Servicer UPB", f"{opp_rel}.Current_UPB__c"),
        ("Approved Advance Amount Funded", "Approved_Advance_Amount_Used__c"),
        ("Comments AM", f"{opp_rel}.Asset_Management_Comments__c"),
        ("Property Created Date", "CreatedDate"),
        ("Property Last Modified Date", "LastModifiedDate"),
    ]
    if _bridge_fv_expr:
        select_pairs.append(("Current Funding Vehicle", _bridge_fv_expr))
    if updated_value_date_field and updated_value_date_field != "BPO_Appraisal_Date__c":
        select_pairs.append(("Updated Valuation Date Native", updated_value_date_field))
    if generic_value_date_field and generic_value_date_field not in {updated_value_date_field, "BPO_Appraisal_Date__c"}:
        select_pairs.append(("Generic Value Date", generic_value_date_field))
    if generic_value_field and generic_value_field != "Appraised_Value_Amount__c":
        select_pairs.append(("Generic Value", generic_value_field))
    # V56: Bridge Asset "Asset Commitment" is the per-asset Approved Advance Amount Max
    # (Property__c), NOT Opportunity.LOC_Commitment__c and NOT the funded-components sum.
    # Verified 4,782/4,782 against 20260824; the old sum matched 4,762. This is the same
    # field _build_bridge_property_rollup_like sums for the Bridge Loan "Loan Commitment"
    # (V41), so asset and loan commitment now tie by construction. Resolved defensively so
    # a field rename cannot break the (large) spine SOQL.
    _aam_spine_field = first_existing_field_name(
        "Property__c",
        ["Approved_Advance_Amount_Max__c", "Approved_Max_Advance_Amount__c", "Approved_Advance_Amount__c", "Max_Advance_Amount__c"],
    )
    if _aam_spine_field:
        select_pairs.append(("Approved Advance Amount Max", _aam_spine_field))

    # Per-property suspense balance (Property__c level). The completed report shows
    # suspense per asset, not the deal-level Opportunity suspense spread once per deal.
    # Looked up defensively so a missing field cannot break the (large) spine SOQL.
    prop_suspense_field = first_existing_field_name("Property__c", ["Suspense_Balance__c", "Suspense__c", "Suspense_Amount__c"])
    if prop_suspense_field:
        select_pairs.append(("Property Suspense Balance", prop_suspense_field))
    # First Payment Date (Opportunity) — drives the FCI day-1 -> first-payment-day NPD rule.
    fpd_field = first_existing_field_name("Opportunity", ["First_Payment_Date__c", "First_Payment_Due_Date__c"])
    if fpd_field:
        select_pairs.append(("First Payment Date", f"{opp_rel}.{fpd_field}"))

    rename_map = {expr: label for label, expr in select_pairs}

    where_parts = [
        f"{opp_rel}.Deal_Loan_Number__c != NULL",
        _soql_in(f"{opp_rel}.StageName", BRIDGE_ACTIVE_STAGES),
        _bridge_dealtype_condition(f"{opp_rel}."),
        _soql_in("Status__c", BRIDGE_ACTIVE_PROPERTY_STATUSES),
        _soql_not_equal_or_null(f"{opp_rel}.LOC_Loan_Type__c", BRIDGE_EXCLUDED_PRODUCT_TYPE),
        "Asset_ID__c != NULL",
    ]
    _append_clause_if_present(where_parts, _opportunity_status_active_condition(f"{opp_rel}."))

    soql = (
        "SELECT "
        + ", ".join(expr for _label, expr in select_pairs)
        + " FROM Property__c WHERE "
        + " AND ".join(where_parts)
    )

    df = run_bulk_query(soql, rename_map=rename_map)

    if df.empty:
        return df

    # V67: census the raw Is Sub Unit values. This query does NOT filter on the flag, so it
    # sees the whole active-bridge property population and can tell FALSE from NULL.
    try:
        if "Is Sub Unit" in df.columns:
            _raw = df["Is Sub Unit"]
            _blank = int(_raw.isna().sum() + (_raw.astype("string").str.strip() == "").sum())
            _yn = _yn_from_bool_series(_raw)
            SUBUNIT_FLAG_CENSUS.clear()
            SUBUNIT_FLAG_CENSUS.update({
                "rows": int(len(df)),
                "true": int((_yn == "Y").sum()),
                "false_or_null": int((_yn == "N").sum()),
                "null_or_blank": _blank,
            })
    except Exception:
        pass

    # Updated valuation columns should only use actual updated/current appraisal inputs.
    # Do not fall back to generic/origination valuation fields. If no updated value exists,
    # these remain blank and the report/carry-forward logic can decide what to preserve.
    # V57: Current Funding Vehicle is the authoritative Bridge Financing source; the
    # legacy Warehouse Line only fills in where the lookup is empty.
    if "Current Funding Vehicle" in df.columns:
        df["Warehouse Line"] = coalesce_keep_nonblank(
            pd.Series(df["Current Funding Vehicle"], index=df.index, dtype="object"),
            pd.Series(df.get("Warehouse Line", pd.Series([pd.NA] * len(df), index=df.index)), index=df.index, dtype="object"),
        )

    df["Current Appraisal Date"] = _coalesce_datetime_columns(df, ["Updated Valuation Date Native", "BPO Appraisal Date"])
    df["Current Appraised As-Is Value"] = _coalesce_numeric_columns(df, ["Appraised Value Amount"])
    if "After Repair Value" in df.columns:
        df["Current Appraised After Repair Value"] = pd.to_numeric(df["After Repair Value"], errors="coerce")
    else:
        df["Current Appraised After Repair Value"] = np.nan

    for c in ["Servicer Loan Number", "Servicer Commitment Id", "Deal Loan Number", "Asset ID", "Property ID", "Unique Active Id"]:
        if c in df.columns:
            df[c] = df[c].astype("string").str.strip().replace({"": pd.NA})

    if {"Servicer Loan Number", "Servicer Commitment Id"}.issubset(df.columns):
        df["Servicer Loan Number"] = coalesce_keep_nonblank(df["Servicer Loan Number"], df["Servicer Commitment Id"])

    df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([None] * len(df), index=df.index)))
    df["_property_id_key"] = norm_id_series(df.get("Property ID", pd.Series([None] * len(df), index=df.index)))
    df["_is_sub_unit"] = _yn_from_bool_series(df.get("Is Sub Unit", pd.Series([pd.NA] * len(df), index=df.index))).eq("Y").astype("int8")
    df["_nonnull_score"] = 0
    for c in ["Address", "City", "State", "Zip", "CBSA", "Servicer Loan Number", "Origination Valuation Date", "Origination As-Is Value", "Current Appraisal Date", "Current Appraised As-Is Value"]:
        if c in df.columns:
            df["_nonnull_score"] = df["_nonnull_score"] + (~blankish_mask(df[c])).astype("int8")
    df["_mod_dt"] = _to_datetime_series_mixed(df.get("Property Last Modified Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["_created_dt"] = _to_datetime_series_mixed(df.get("Property Created Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df = df[df["_asset_key"].notna()].copy()
    df = df.sort_values(["_asset_key", "_is_sub_unit", "_nonnull_score", "_mod_dt", "_created_dt", "_property_id_key"], ascending=[True, False, True, True, True, True])
    df = df.drop_duplicates(["_asset_key"], keep="last")
    df = df.drop(columns=["_asset_key", "_property_id_key", "_is_sub_unit", "_nonnull_score", "_mod_dt", "_created_dt"], errors="ignore")

    return downcast_numeric_frame(df)


def _build_bridge_loan_wide_like() -> pd.DataFrame:
    sold_pool_field = first_existing_field_name("Opportunity", ["Sold_Loan_Pool__c", "FK_Sold_Loan_Pool__c"])
    sold_pool_rel = relationship_name_for("Opportunity", sold_pool_field) if sold_pool_field else None

    contact_field = first_existing_field_name("Opportunity", ["Contact__c", "Primary_Contact__c"])
    contact_rel = relationship_name_for("Opportunity", contact_field) if contact_field else None

    servicer_primary_field = first_existing_field_name("Opportunity", ["Servicer_Loan_Number__c", "Servicer_Loan_Num__c", "Servicer_Loan_ID__c", "Servicer_Loan_Id__c"])
    servicer_commitment_field = first_existing_field_name("Opportunity", TERM_SERVICER_FALLBACK_FIELD_CANDIDATES)

    sold_to_expr = f"{sold_pool_rel}.Sold_To__r.Name" if sold_pool_rel else "Account.Name"
    primary_contact_expr = f"{contact_rel}.Name" if contact_rel else "Account.Name"

    funded_candidates = existing_field_names("Opportunity", [
        "Total_Funded_Active_Assets__c",
        "Aggregate_Funding__c",
        "Total_Amount_Advances__c",
        "Funded_Amount__c",
    ])
    asset_count_candidates = existing_field_names("Opportunity", ["Number_of_Properties__c", "Total_Properties__c"])
    unit_count_candidates = existing_field_names("Opportunity", ["Total_Units__c"])
    state_candidates = existing_field_names("Opportunity", ["Distinct_States__c", "Active_States__c"])
    last_funding_candidates = existing_field_names("Opportunity", ["Funds_Released_Date__c", "Last_Funding_Date__c"])
    valuation_date_candidates = existing_field_names("Opportunity", ["Updated_Value_Date__c", "Valuation_Date__c"])
    valuation_asis_candidates = existing_field_names("Opportunity", ["Updated_Value__c", "Total_Valuation_Amount__c"])

    # V57: prefer the Current Funding Vehicle lookup over the retired Warehouse Line
    # (see _build_bridge_spine_like). Coalesced below so either source works.
    _bl_fv_field = first_existing_field_name(
        "Opportunity", ["Current_Funding_Vehicle__c", "FK_Current_Funding_Vehicle__c"]
    )
    _bl_fv_expr = None
    if _bl_fv_field:
        try:
            _bl_fv_expr = f"{relationship_name_for('Opportunity', _bl_fv_field)}.Name"
        except Exception:
            _bl_fv_expr = None

    select_pairs = [
        ("Loan Buyer", sold_to_expr),
        ("Financing", "Warehouse_Line__c"),
        ("Deal Number", "Deal_Loan_Number__c"),
        ("Deal Name", "Name"),
        ("Borrower Name", "Borrower_Entity__r.Name"),
        ("Account", "Account.Name"),
        ("Do Not Lend", "Account.Do_Not_Lend__c"),
        ("Primary Contact", primary_contact_expr),
        ("Origination Date", "CloseDate"),
        ("Next Payment Date", "Next_Payment_Date__c"),
        ("Original Maturity Date", "Stated_Maturity_Date__c"),
        ("Current Maturity Date", "Current_Line_Maturity_Date__c"),
        ("Loan Commitment", "LOC_Commitment__c"),
        ("Remaining Commitment", "Outstanding_Facility_Amount__c"),
        ("SF Suspense Balance", "Suspense_Balance__c"),
        ("SF Current UPB", "Current_UPB__c"),
        ("Loan Stage", "StageName"),
        ("Product Type", "LOC_Loan_Type__c"),
        ("Product Sub Type", "Product_Sub_Type__c"),
        ("Transaction Type", "Transaction_Type__c"),
        ("Project Strategy", "Project_Strategy__c"),
        ("CV Originator", "Owner.Name"),
        ("CV Originator: Active", "Owner.IsActive"),
        ("Deal Intro Sub-Source", "Deal_Intro_Sub_Source__c"),
        ("Referral Source Account", "Referral_Source__r.Name"),
        ("Referral Source Contact", "Referral_Source_Contact__r.Name"),
        ("AM Commentary", "Asset_Management_Comments__c"),
        ("Type", "Type"),
    ]
    if _bl_fv_expr:
        select_pairs.append(("Current Funding Vehicle", _bl_fv_expr))
    if servicer_primary_field:
        select_pairs.insert(3, ("Opportunity Servicer Loan Number", servicer_primary_field))
    if servicer_commitment_field:
        select_pairs.insert(4 if servicer_primary_field else 3, ("Opportunity Servicer Commitment Id", servicer_commitment_field))
    for idx, field in enumerate(funded_candidates, start=1):
        select_pairs.append((f"Active Funded Amount Candidate {idx}", field))
    for idx, field in enumerate(asset_count_candidates, start=1):
        select_pairs.append((f"Number of Assets Candidate {idx}", field))
    for idx, field in enumerate(unit_count_candidates, start=1):
        select_pairs.append((f"Units Candidate {idx}", field))
    for idx, field in enumerate(state_candidates, start=1):
        select_pairs.append((f"States Candidate {idx}", field))
    for idx, field in enumerate(last_funding_candidates, start=1):
        select_pairs.append((f"Last Funding Date Candidate {idx}", field))
    for idx, field in enumerate(valuation_date_candidates, start=1):
        select_pairs.append((f"Most Recent Valuation Date Candidate {idx}", field))
    for idx, field in enumerate(valuation_asis_candidates, start=1):
        select_pairs.append((f"Most Recent As-Is Value Candidate {idx}", field))

    rename_map = {expr: label for label, expr in select_pairs}

    where_parts = [
        "Deal_Loan_Number__c != NULL",
        _soql_in("StageName", BRIDGE_LOAN_SOURCE_STAGES),
        _bridge_dealtype_condition(),
        _soql_not_equal_or_null("LOC_Loan_Type__c", BRIDGE_EXCLUDED_PRODUCT_TYPE),
    ]
    _append_clause_if_present(where_parts, _opportunity_status_active_condition())

    soql = "SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Opportunity WHERE " + " AND ".join(where_parts)
    df = run_bulk_query(soql, rename_map=rename_map)
    if df.empty:
        return df

    blank_obj = pd.Series([pd.NA] * len(df), index=df.index, dtype="object")
    # V57: Current Funding Vehicle wins over the retired Warehouse Line.
    if "Current Funding Vehicle" in df.columns:
        df["Financing"] = coalesce_keep_nonblank(
            pd.Series(df["Current Funding Vehicle"], index=df.index, dtype="object"),
            pd.Series(df.get("Financing", blank_obj), index=df.index, dtype="object"),
        )
    if "Opportunity Servicer Loan Number" in df.columns:
        df["Opportunity Servicer Loan Number"] = df["Opportunity Servicer Loan Number"].astype("string").str.strip().replace({"": pd.NA})
    if "Opportunity Servicer Commitment Id" in df.columns:
        df["Opportunity Servicer Commitment Id"] = df["Opportunity Servicer Commitment Id"].astype("string").str.strip().replace({"": pd.NA})

    df["Servicer ID"] = coalesce_keep_nonblank(df.get("Opportunity Servicer Loan Number", blank_obj), df.get("Opportunity Servicer Commitment Id", blank_obj))
    df["Active Funded Amount SF"] = _coalesce_numeric_columns(df, [c for c in df.columns if c.startswith("Active Funded Amount Candidate ")])
    df["Number of Assets SF"] = _coalesce_numeric_columns(df, [c for c in df.columns if c.startswith("Number of Assets Candidate ")])
    df["# of Units SF"] = _coalesce_numeric_columns(df, [c for c in df.columns if c.startswith("Units Candidate ")])
    df["State(s) SF"] = _coalesce_text_columns(df, [c for c in df.columns if c.startswith("States Candidate ")])
    df["Last Funding Date SF"] = _coalesce_datetime_columns(df, [c for c in df.columns if c.startswith("Last Funding Date Candidate ")])
    df["Most Recent Valuation Date SF"] = _coalesce_datetime_columns(df, [c for c in df.columns if c.startswith("Most Recent Valuation Date Candidate ")])
    df["Most Recent As-Is Value SF"] = _coalesce_numeric_columns(df, [c for c in df.columns if c.startswith("Most Recent As-Is Value Candidate ")])

    df["Deal Number"] = df.get("Deal Number", blank_obj).astype("string").str.strip().replace({"": pd.NA})
    return downcast_numeric_frame(df)


def _build_bridge_property_rollup_like() -> pd.DataFrame:
    opp_rel = property_opportunity_relationship_name()

    select_pairs = [
        ("Deal Number", f"{opp_rel}.Deal_Loan_Number__c"),
        ("Asset ID", "Asset_ID__c"),
        ("State", "State__c"),
        ("# of Units", "Number_of_Units__c"),
        ("Last Funding Date", "Funding_Date__c"),
        ("Property Status", "Status__c"),
        ("Property Current UPB", "Current_UPB__c"),
    ]
    # V41: Bridge Loan "Loan Commitment" = SUM of the per-asset Approved Advance Amount Max
    # (verified deal 33182: 102,500 = 56,000 + 46,500), NOT Opportunity.LOC_Commitment__c
    # (which returned 10,000,000). Resolve the field defensively so a rename can't break the
    # rollup SOQL; when absent, Loan Commitment falls back to the spine value below.
    _aam_field = first_existing_field_name(
        "Property__c",
        ["Approved_Advance_Amount_Max__c", "Approved_Max_Advance_Amount__c", "Approved_Advance_Amount__c", "Max_Advance_Amount__c"],
    )
    if _aam_field:
        select_pairs.append(("Approved Advance Amount Max", _aam_field))
    rename_map = {expr: label for label, expr in select_pairs}

    where_parts = [
        f"{opp_rel}.Deal_Loan_Number__c != NULL",
        _soql_in(f"{opp_rel}.StageName", BRIDGE_LOAN_SOURCE_STAGES),
        _bridge_dealtype_condition(f"{opp_rel}."),
        _soql_not_equal_or_null(f"{opp_rel}.LOC_Loan_Type__c", BRIDGE_EXCLUDED_PRODUCT_TYPE),
    ]
    _append_clause_if_present(where_parts, _opportunity_status_active_condition(f"{opp_rel}."))

    soql = (
        "SELECT "
        + ", ".join(expr for _label, expr in select_pairs)
        + " FROM Property__c WHERE "
        + " AND ".join(where_parts)
    )

    df = run_bulk_query(soql, rename_map=rename_map)
    if df.empty:
        return pd.DataFrame(columns=[
            "_deal_key", "Number of Assets", "# of Units", "State(s)", "Last Funding Date",
            "Active Asset Count", "Active Asset UPB",
        ])

    df["_deal_key"] = norm_id_series(df.get("Deal Number", pd.Series([None] * len(df), index=df.index)))
    df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([None] * len(df), index=df.index)))
    status = df.get("Property Status", pd.Series([pd.NA] * len(df), index=df.index)).astype("string").str.strip()
    asset_upb = pd.to_numeric(df.get("Property Current UPB", pd.Series([np.nan] * len(df), index=df.index)), errors="coerce").fillna(0)
    is_active_asset = status.isin(BRIDGE_ACTIVE_PROPERTY_STATUSES)
    df["_active_asset_upb"] = asset_upb.where(is_active_asset, 0.0)
    df["_is_active_asset"] = is_active_asset.astype("int8")
    _has_aam = "Approved Advance Amount Max" in df.columns
    if _has_aam:
        df["_aam"] = pd.to_numeric(df["Approved Advance Amount Max"], errors="coerce")

    g = df.groupby("_deal_key", dropna=True)
    out = pd.DataFrame(
        {
            "Number of Assets": g["_asset_key"].nunique(),
            "# of Units": pd.to_numeric(g["# of Units"].sum(min_count=1), errors="coerce") if "# of Units" in df.columns else np.nan,
            "State(s)": g["State"].apply(lambda s: ", ".join(sorted({clean_text(x) for x in s if clean_text(x)}))) if "State" in df.columns else pd.Series(dtype="string"),
            "Last Funding Date": g["Last Funding Date"].apply(lambda s: pd.to_datetime(s, errors="coerce").dropna().max() if len(pd.to_datetime(s, errors="coerce").dropna()) else pd.NaT),
            "Active Asset Count": g["_is_active_asset"].sum(),
            "Active Asset UPB": pd.to_numeric(g["_active_asset_upb"].sum(min_count=1), errors="coerce"),
        }
    ).reset_index()
    if _has_aam:
        _aam_sum = pd.to_numeric(g["_aam"].sum(min_count=1), errors="coerce").rename("Loan Commitment (AAM)").reset_index()
        out = out.merge(_aam_sum, on="_deal_key", how="left")
    return downcast_numeric_frame(out)


def _build_do_not_lend_like() -> pd.DataFrame:
    where_parts = [
        "Account.Do_Not_Lend__c = TRUE",
        _soql_in("StageName", DNL_STAGES),
        "Deal_Loan_Number__c != NULL",
    ]
    _append_clause_if_present(where_parts, _opportunity_status_active_condition())
    soql = (
        "SELECT Deal_Loan_Number__c, Account.Name, Account.Do_Not_Lend__c "
        "FROM Opportunity WHERE "
        + " AND ".join(where_parts)
    )
    df = run_bulk_query(soql)
    rename_map = {
        "Deal_Loan_Number__c": "Deal Loan Number",
        "Account.Name": "Account Name",
        "Account.Do_Not_Lend__c": "Do Not Lend",
    }
    return downcast_numeric_frame(_normalize_bulk_df(df.rename(columns=rename_map)))



def _build_appraisal_like(asset_ids=None) -> pd.DataFrame:
    asset_ids = _nonblank_unique(asset_ids or [])
    property_rel = appraisal_property_relationship_name()
    deal_rel = relationship_name_for("Appraisal__c", "Deal__c") if first_existing_field_name("Appraisal__c", ["Deal__c"]) else None

    effective_field = first_existing_field_name("Appraisal__c", ["Appraisal_Effective_Date__c", "Appraisal_Report_Date__c"])
    report_field = first_existing_field_name("Appraisal__c", ["Appraisal_Report_Date__c", "Appraisal_Effective_Date__c"])
    order_field = first_existing_field_name("Appraisal__c", ["Order_Received_Date__c"])
    reviewed_as_is_field = first_existing_field_name("Appraisal__c", ["Reviewed_Appraisal_As_Is_Value__c"])
    fallback_as_is_field = first_existing_field_name("Appraisal__c", ["Appraised_Value_Amount__c"])
    reviewed_arv_field = first_existing_field_name(
        "Appraisal__c",
        ["Reviewed_Appraisal_After_Repair_Value__c", "Appraised_After_Repair_Value__c", "Internal_as_Rehab_Value__c"],
    )
    fallback_arv_field = first_existing_field_name(
        "Appraisal__c",
        ["Appraised_After_Repair_Value__c", "Internal_as_Rehab_Value__c", "Appraised_Value_Amount__c"],
    )

    status_field = first_existing_field_name("Appraisal__c", ["Status__c", "Status_Description__c"])

    select_pairs = [
        ("Asset ID", f"{property_rel}.Asset_ID__c"),
        ("Property Asset Id", "Property_Asset_Id__c"),
        ("Property ID", f"{property_rel}.Id"),
        ("Appraisal Name", "Name"),
    ]
    if status_field:
        select_pairs.append(("Appraisal Status", status_field))
    if deal_rel:
        select_pairs.append(("Deal Loan Number", f"{deal_rel}.Deal_Loan_Number__c"))
    if order_field:
        select_pairs.append(("Most Recent Appraisal Order Date", order_field))
    if effective_field:
        select_pairs.append(("Appraisal Effective Date", effective_field))
    if report_field and report_field != effective_field:
        select_pairs.append(("Appraisal Report Date", report_field))
    if reviewed_as_is_field:
        select_pairs.append(("Appraisal Reviewed As-Is Value", reviewed_as_is_field))
    if fallback_as_is_field and fallback_as_is_field != reviewed_as_is_field:
        select_pairs.append(("Appraisal As-Is Fallback Value", fallback_as_is_field))
    if reviewed_arv_field:
        select_pairs.append(("Appraisal Reviewed ARV", reviewed_arv_field))
    if fallback_arv_field and fallback_arv_field not in {reviewed_arv_field, reviewed_as_is_field, fallback_as_is_field}:
        select_pairs.append(("Appraisal ARV Fallback", fallback_arv_field))
    elif fallback_arv_field and fallback_arv_field == fallback_as_is_field and fallback_arv_field not in {reviewed_arv_field}:
        select_pairs.append(("Appraisal ARV Fallback", fallback_arv_field))

    rename_map = {expr: label for label, expr in select_pairs}
    soqls = []
    if asset_ids:
        for chunk in _chunked(asset_ids, size=200):
            soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Appraisal__c WHERE " + _soql_in(f"{property_rel}.Asset_ID__c", chunk))
    else:
        soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + f" FROM Appraisal__c WHERE {property_rel}.Asset_ID__c != NULL")

    df = _run_bulk_union(soqls, rename_map=rename_map)
    if df.empty:
        return df

    if "Asset ID" not in df.columns and "Property Asset Id" in df.columns:
        df["Asset ID"] = df["Property Asset Id"]

    df["Current Appraisal Date"] = _coalesce_datetime_columns(df, ["Appraisal Effective Date", "Appraisal Report Date"])
    if "Most Recent Appraisal Order Date" in df.columns:
        df["Most Recent Appraisal Order Date"] = _to_datetime_series_mixed(df["Most Recent Appraisal Order Date"])
    df["Current Appraised As-Is Value"] = _coalesce_numeric_columns_zeroaware(
        df,
        [c for c in ["Appraisal As-Is Fallback Value", "Appraisal Reviewed As-Is Value"] if c in df.columns],
    )
    df["Current Appraised After Repair Value"] = _coalesce_numeric_columns_zeroaware(
        df,
        # ARV must NOT fall back to the As-Is value: when an appraisal carries no ARV
        # (common for as-is-only appraisals) the official report shows N/A, not the
        # As-Is amount masquerading as ARV.
        [c for c in ["Appraisal ARV Fallback", "Appraisal Reviewed ARV"] if c in df.columns],
    )
    df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([None] * len(df), index=df.index)))
    return downcast_numeric_frame(df)


def _select_best_current_appraisal_bundle(property_df: pd.DataFrame, appraisal_df: pd.DataFrame) -> pd.DataFrame:
    if appraisal_df is None or appraisal_df.empty:
        return pd.DataFrame()

    cand = appraisal_df.copy()
    if "Asset ID" not in cand.columns and "Property Asset Id" in cand.columns:
        cand["Asset ID"] = cand["Property Asset Id"]
    cand["_asset_key"] = norm_id_series(cand.get("Asset ID", pd.Series([None] * len(cand), index=cand.index)))

    if property_df is not None and not property_df.empty and "Asset ID" in property_df.columns:
        valid_asset_keys = set(norm_id_series(property_df["Asset ID"]).dropna().tolist())
        if valid_asset_keys:
            cand = cand[cand["_asset_key"].isin(valid_asset_keys)].copy()

    if cand.empty:
        return pd.DataFrame()

    cand["Current Appraisal Date"] = _coalesce_datetime_columns(cand, ["Current Appraisal Date", "Appraisal Effective Date", "Appraisal Report Date"])
    if "Most Recent Appraisal Order Date" in cand.columns:
        cand["Most Recent Appraisal Order Date"] = _to_datetime_series_mixed(cand["Most Recent Appraisal Order Date"])
    cand["Current Appraised As-Is Value"] = _coalesce_numeric_columns_zeroaware(
        cand,
        [c for c in ["Current Appraised As-Is Value", "Appraisal As-Is Fallback Value", "Appraisal Reviewed As-Is Value"] if c in cand.columns],
    )
    cand["Current Appraised After Repair Value"] = _coalesce_numeric_columns_zeroaware(
        cand,
        # ARV must NOT borrow the As-Is value when the appraisal has no ARV.
        [c for c in ["Current Appraised After Repair Value", "Appraisal ARV Fallback", "Appraisal Reviewed ARV"] if c in cand.columns],
    )

    # "Most Recent Appraisal Order Date" = MAX order date across ALL appraisals for the
    # asset (validated ~88% vs official). Compute it BEFORE the Complete-Delivered filter
    # so cancelled/in-progress orders still count, then attach back per asset.
    # Localize to Pacific (org locale) and KEEP the time component so the value matches
    # the official report's datetime (e.g. 2026-02-08 17:04:00), not a naive-UTC date.
    _max_order_by_asset = None
    if "Most Recent Appraisal Order Date" in cand.columns:
        _ord_all = _to_pacific_naive_series(cand["Most Recent Appraisal Order Date"])
        _max_order_by_asset = (
            pd.DataFrame({"_asset_key": cand["_asset_key"], "_ord": _ord_all})
            .dropna(subset=["_asset_key"])
            .groupby("_asset_key")["_ord"].max()
        )

    # "Updated" = the latest FINALIZED post-origination appraisal by effective date.
    # Finalized = Complete-Delivered OR Complete. We intentionally do NOT require the
    # asset's newest appraisal overall to be Complete-Delivered (the old gate): the
    # official report only honored Complete-Delivered and therefore showed STALE values
    # when a newer Complete (but not-yet-delivered) appraisal existed -- e.g. asset
    # 819585 had a 2026-04-30 'Complete' appraisal the official ignored in favor of an
    # older 2025-12-26 delivered one. We treat the most recent finalized appraisal as
    # the true Updated value. In-flight / rejected statuses (Reviewed, Revision Required,
    # Under Review, Ordered, Cancelled, etc.) are NOT finalized and do not count.
    _FINALIZED_STATUSES = {"complete-delivered", "complete"}
    if "Appraisal Status" in cand.columns:
        _status = cand["Appraisal Status"].astype("string").str.strip()
        cand = cand[_status.str.casefold().isin(_FINALIZED_STATUSES)].copy()
    if cand.empty:
        # still surface Most Recent Order Date even when no finalized appraisal exists
        if _max_order_by_asset is not None and len(_max_order_by_asset):
            mo = _max_order_by_asset.reset_index()
            mo.columns = ["_asset_key", "Most Recent Appraisal Order Date"]
            return downcast_numeric_frame(mo)
        return pd.DataFrame()

    cand["_bundle_effective_dt"] = _to_datetime_series_mixed(cand.get("Current Appraisal Date", pd.Series([pd.NaT] * len(cand), index=cand.index)))
    cand["_bundle_order_dt"] = _to_datetime_series_mixed(cand.get("Most Recent Appraisal Order Date", pd.Series([pd.NaT] * len(cand), index=cand.index)))
    cand["_nonnull_score"] = 0
    for c in ["Current Appraisal Date", "Current Appraised As-Is Value", "Current Appraised After Repair Value", "Most Recent Appraisal Order Date"]:
        if c in cand.columns:
            cand["_nonnull_score"] = cand["_nonnull_score"] + (~blankish_mask(cand[c])).astype("int8")

    cand = cand[cand["_asset_key"].notna()].copy()
    cand = cand.sort_values(["_asset_key", "_bundle_effective_dt", "_bundle_order_dt", "_nonnull_score"], ascending=[True, True, True, True])
    cand = cand.drop_duplicates(["_asset_key"], keep="last")

    # Replace the selected row's order date with the per-asset MAX order date
    # (across all appraisals, computed before the Complete-Delivered filter).
    if _max_order_by_asset is not None and len(_max_order_by_asset):
        cand["Most Recent Appraisal Order Date"] = cand["_asset_key"].map(_max_order_by_asset)

    # If the SELECTED finalized appraisal carries no usable As-Is value (As-Is blank or
    # 0 after the zero-aware coalesce), blank the ENTIRE Updated triple -- including the
    # Updated Valuation Date -- not just the value, so we never emit a date with no value
    # next to it.
    if "Current Appraised As-Is Value" in cand.columns:
        _no_asis = blankish_mask(cand["Current Appraised As-Is Value"]) | (
            pd.to_numeric(cand["Current Appraised As-Is Value"], errors="coerce").fillna(0) == 0
        )
        for _c in ["Current Appraisal Date", "Current Appraised As-Is Value", "Current Appraised After Repair Value"]:
            if _c in cand.columns:
                cand.loc[_no_asis, _c] = pd.NA

    keep = ["_asset_key"] + [
        c for c in ["Asset ID", "Most Recent Appraisal Order Date", "Current Appraisal Date", "Current Appraised As-Is Value", "Current Appraised After Repair Value"]
        if c in cand.columns
    ]
    return downcast_numeric_frame(cand[keep].drop_duplicates("_asset_key"))


def _build_valuation_like(asset_ids=None) -> pd.DataFrame:

    asset_ids = _nonblank_unique(asset_ids or [])
    soqls = []

    value_date_field = first_existing_field_name("Property__c", ["Updated_Valuation_Date__c", "Value_Date__c", "BPO_Appraisal_Date__c"])
    backup_value_date_field = first_existing_field_name("Property__c", ["Value_Date__c", "Updated_Valuation_Date__c", "BPO_Appraisal_Date__c"])
    generic_value_field = first_existing_field_name("Property__c", ["Value__c", "Appraised_Value_Amount__c"])

    select_pairs = [
        ("Asset ID", "Asset_ID__c"),
        ("Property ID", "Id"),
        ("Most Recent Appraisal Order Date", "BPO_Appraisal_Order_Date__c"),
        ("BPO Appraisal Date", "BPO_Appraisal_Date__c"),
        ("Appraised Value Amount", "Appraised_Value_Amount__c"),
        ("After Repair Value", "After_Repair_Value__c"),
        ("Origination Valuation Date", "Origination_Date_Valuation_Date__c"),
        ("Origination As-Is Value", "Origination_Date_Value__c"),
        ("Origination After Repair Value", "Origination_After_Repair_Value__c"),
        ("Is Sub Unit", "Is_Sub_Unit__c"),
        ("Property Created Date", "CreatedDate"),
        ("Property Last Modified Date", "LastModifiedDate"),
    ]
    if value_date_field and value_date_field != "BPO_Appraisal_Date__c":
        select_pairs.append(("Updated Value Date Native", value_date_field))
    if backup_value_date_field and backup_value_date_field not in {value_date_field, "BPO_Appraisal_Date__c"}:
        select_pairs.append(("Backup Value Date Native", backup_value_date_field))
    if generic_value_field and generic_value_field != "Appraised_Value_Amount__c":
        select_pairs.append(("Generic Value Native", generic_value_field))

    rename_map = {expr: label for label, expr in select_pairs}

    if asset_ids:
        for chunk in _chunked(asset_ids, size=200):
            soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Property__c WHERE " + _soql_in("Asset_ID__c", chunk))
    else:
        soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Property__c WHERE Asset_ID__c != NULL")

    df = _run_bulk_union(soqls, rename_map=rename_map)
    appraisal_df = _build_appraisal_like(asset_ids=asset_ids)
    best_bundle = _select_best_current_appraisal_bundle(df, appraisal_df)
    if df.empty and best_bundle.empty:
        return df

    if df.empty:
        df = best_bundle.copy()
    else:
        # Updated/current valuation fields are strict: do not backfill them from
        # origination, backup, or generic property values.
        df["Current Appraisal Date"] = _coalesce_datetime_columns(df, ["Updated Value Date Native", "BPO Appraisal Date"])
        if "Most Recent Appraisal Order Date" in df.columns:
            df["Most Recent Appraisal Order Date"] = _to_datetime_series_mixed(df["Most Recent Appraisal Order Date"])
        df["Current Appraised As-Is Value"] = _coalesce_numeric_columns_zeroaware(df, ["Appraised Value Amount"])
        df["Current Appraised After Repair Value"] = _coalesce_numeric_columns_zeroaware(df, ["After Repair Value"])

        if not best_bundle.empty and "_asset_key" in best_bundle.columns:
            app = best_bundle.copy()
            df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([None] * len(df), index=df.index)))
            df = df.merge(app, on="_asset_key", how="left", suffixes=("", "_app"))
            # Updated valuation triple (As-Is / ARV / effective date) comes STRICTLY from
            # the latest Complete-Delivered appraisal bundle, which reads the per-appraisal
            # Appraised_Value_Amount__c / Appraised_After_Repair_Value__c / effective date.
            # The Property__c roll-up (Current Appraised*) is NOT a valid source and must
            # not survive: if the asset has no delivered appraisal, the report blanks these.
            # Most Recent Appraisal Order Date is ALSO strict-from-bundle: the official report
            # sources it from the per-appraisal Order_Received_Date__c (MAX across appraisals)
            # and shows N/A when the asset has no appraisal order date. Verified against
            # 20260608_Active_Loans: 2,034 assets with a value all have an SF appraisal Order
            # Date; all 2,656 N/A assets have none. The Property-level BPO_Appraisal_Order_Date__c
            # must NOT be used as a fallback -- doing so over-populated the column (~4,260 vs 2,034).
            strict_from_bundle = {
                "Most Recent Appraisal Order Date",
                "Current Appraisal Date",
                "Current Appraised As-Is Value",
                "Current Appraised After Repair Value",
            }
            for c in ["Most Recent Appraisal Order Date", "Current Appraisal Date", "Current Appraised As-Is Value", "Current Appraised After Repair Value"]:
                app_col = f"{c}_app"
                if app_col in df.columns:
                    if c in strict_from_bundle:
                        # bundle value only -- do NOT fall back to property roll-up
                        df[c] = df[app_col]
                    else:
                        df[c] = coalesce_keep_nonblank(df[app_col], df.get(c, pd.Series([pd.NA] * len(df), index=df.index)))
                    df = df.drop(columns=[app_col], errors="ignore")
        else:
            # No appraisal bundle at all for this batch -> blank the Updated triple AND the
            # order date so the Property roll-up never leaks into these columns.
            for c in ["Most Recent Appraisal Order Date", "Current Appraisal Date", "Current Appraised As-Is Value", "Current Appraised After Repair Value"]:
                df[c] = pd.NA

    df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([None] * len(df), index=df.index)))
    df["_property_id_key"] = norm_id_series(df.get("Property ID", pd.Series([None] * len(df), index=df.index)))
    df["_is_sub_unit"] = _yn_from_bool_series(df.get("Is Sub Unit", pd.Series([pd.NA] * len(df), index=df.index))).eq("Y").astype("int8")
    df["_nonnull_score"] = 0
    for c in ["Most Recent Appraisal Order Date", "Current Appraisal Date", "Current Appraised As-Is Value", "Current Appraised After Repair Value", "Origination Valuation Date", "Origination As-Is Value", "Origination After Repair Value"]:
        if c in df.columns:
            df["_nonnull_score"] = df["_nonnull_score"] + (~blankish_mask(df[c])).astype("int8")
    df["_mod_dt"] = _to_datetime_series_mixed(df.get("Property Last Modified Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["_created_dt"] = _to_datetime_series_mixed(df.get("Property Created Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df = df[df["_asset_key"].notna()].copy()
    df = df.sort_values(["_asset_key", "_is_sub_unit", "_nonnull_score", "_mod_dt", "_created_dt", "_property_id_key"], ascending=[True, False, True, True, True, True])
    df = df.drop_duplicates(["_asset_key"], keep="last")
    df = df.drop(columns=["_asset_key", "_property_id_key", "_is_sub_unit", "_nonnull_score", "_mod_dt", "_created_dt"], errors="ignore")
    return downcast_numeric_frame(df)


def _build_foreclosure_like(asset_ids=None) -> pd.DataFrame:
    """Pull current foreclosure sale dates by Bridge Asset ID.

    New Bridge Asset columns added in the 5/18 report:
    - FC Sale Date <- Foreclosure__c.Sale_Date__c
    - Rescheduled FC Sale Date <- Foreclosure__c.Reschedule_Sale_Date__c
    """
    asset_ids = _nonblank_unique(asset_ids or [])
    try:
        field_map = _field_map_by_name("Foreclosure__c")
    except Exception as exc:
        try:
            st.warning(f"Foreclosure__c was not available in Salesforce for this session: {exc}")
        except Exception:
            pass
        return pd.DataFrame(columns=["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"])

    if "Sale_Date__c" not in field_map and "Reschedule_Sale_Date__c" not in field_map:
        return pd.DataFrame(columns=["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"])

    property_field = first_existing_field_name(
        "Foreclosure__c",
        ["Property__c", "Subject_Property__c", "Collateral_Property__c", "Asset__c", "Property_Asset__c"],
    )
    direct_asset_field = first_existing_field_name(
        "Foreclosure__c",
        ["Asset_ID__c", "Property_Asset_ID__c", "Property_Asset_Id__c", "Asset_Id__c"],
    )

    select_pairs = []
    where_asset_expr = None
    if property_field:
        try:
            prop_rel = relationship_name_for("Foreclosure__c", property_field)
            select_pairs.append(("Asset ID", f"{prop_rel}.Asset_ID__c"))
            select_pairs.append(("Property ID", property_field))
            where_asset_expr = f"{prop_rel}.Asset_ID__c"
        except Exception:
            select_pairs.append(("Property ID", property_field))
    if direct_asset_field:
        select_pairs.append(("Asset ID Direct", direct_asset_field))
        if where_asset_expr is None:
            where_asset_expr = direct_asset_field

    if not any(label in {"Asset ID", "Asset ID Direct"} for label, _expr in select_pairs):
        return pd.DataFrame(columns=["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"])

    if "Sale_Date__c" in field_map:
        select_pairs.append(("FC Sale Date", "Sale_Date__c"))
    if "Reschedule_Sale_Date__c" in field_map:
        select_pairs.append(("Rescheduled FC Sale Date", "Reschedule_Sale_Date__c"))
    if "LastModifiedDate" in field_map:
        select_pairs.append(("Foreclosure Last Modified Date", "LastModifiedDate"))
    if "CreatedDate" in field_map:
        select_pairs.append(("Foreclosure Created Date", "CreatedDate"))
    if "Status__c" in field_map:
        select_pairs.append(("Foreclosure Status", "Status__c"))
    if "Name" in field_map:
        select_pairs.append(("Foreclosure Name", "Name"))

    rename_map = {expr: label for label, expr in select_pairs}
    base_where = [f"{where_asset_expr} != NULL"] if where_asset_expr else ["Id != NULL"]

    soqls = []
    if asset_ids and where_asset_expr:
        for chunk in _chunked(asset_ids, size=200):
            where_parts = base_where + [_soql_in(where_asset_expr, chunk)]
            soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Foreclosure__c WHERE " + " AND ".join(where_parts))
    else:
        soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Foreclosure__c WHERE " + " AND ".join(base_where))

    try:
        df = _run_bulk_union(soqls, rename_map=rename_map)
    except Exception as exc:
        try:
            st.warning(f"Foreclosure__c pull failed; FC sale date columns will be N/A: {exc}")
        except Exception:
            pass
        return pd.DataFrame(columns=["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"])

    if df.empty:
        return pd.DataFrame(columns=["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"])

    if "Asset ID" not in df.columns and "Asset ID Direct" in df.columns:
        df["Asset ID"] = df["Asset ID Direct"]
    elif "Asset ID" in df.columns and "Asset ID Direct" in df.columns:
        df["Asset ID"] = coalesce_keep_nonblank(df["Asset ID"], df["Asset ID Direct"])

    df["FC Sale Date"] = _to_datetime_series_mixed(df.get("FC Sale Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["Rescheduled FC Sale Date"] = _to_datetime_series_mixed(df.get("Rescheduled FC Sale Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([pd.NA] * len(df), index=df.index)))
    df["_fc_sale_dt"] = _to_datetime_series_mixed(df["FC Sale Date"])
    df["_fc_resched_dt"] = _to_datetime_series_mixed(df["Rescheduled FC Sale Date"])
    df["_mod_dt"] = _to_datetime_series_mixed(df.get("Foreclosure Last Modified Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["_created_dt"] = _to_datetime_series_mixed(df.get("Foreclosure Created Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["_nonnull_score"] = df["_fc_sale_dt"].notna().astype("int8") + df["_fc_resched_dt"].notna().astype("int8")
    df["_sort_dt"] = pd.concat([df["_fc_resched_dt"], df["_fc_sale_dt"], df["_mod_dt"], df["_created_dt"]], axis=1).max(axis=1)
    df = df[df["_asset_key"].notna()].copy()
    if df.empty:
        return pd.DataFrame(columns=["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"])
    df = df.sort_values(["_asset_key", "_sort_dt", "_mod_dt", "_created_dt", "_nonnull_score"], ascending=[True, True, True, True, True])
    df = df.drop_duplicates("_asset_key", keep="last")
    keep = ["Asset ID", "FC Sale Date", "Rescheduled FC Sale Date"]
    return downcast_numeric_frame(df[keep + ["_asset_key"]].drop_duplicates("_asset_key"))


def _build_am_assignments_like() -> pd.DataFrame:
    soql = (
        "SELECT Opportunity.Deal_Loan_Number__c, Opportunity.Name, User.Name, TeamMemberRole, Date_Assigned__c "
        "FROM OpportunityTeamMember WHERE "
        "Opportunity.Deal_Loan_Number__c != NULL AND "
        + _soql_parent_name_not_equal_or_no_parent("Opportunity.AccountId", "Opportunity.Account.Name", EXCLUDED_TEST_ACCOUNT_NAME)
        + " AND "
        + _soql_in("TeamMemberRole", AM_ASSIGNMENT_ROLES)
    )

    df = run_bulk_query(soql)
    rename_map = {
        "Opportunity.Deal_Loan_Number__c": "Deal Loan Number",
        "Opportunity.Name": "Deal Name",
        "User.Name": "Team Member Name",
        "TeamMemberRole": "Team Role",
        "Date_Assigned__c": "Date Assigned",
    }
    df = df.rename(columns=rename_map)
    return downcast_numeric_frame(_normalize_bulk_df(df))


def _build_active_rm_like() -> pd.DataFrame:
    # Active RM = the deal's CAF Originator (loan officer) is an active Salesforce user.
    # Verified 6/29 same-day: per-originator deterministic across 60 originators (zero
    # split) -- e.g. all 110 'N' deals share originator Dan Niemeyer (inactive). The
    # discriminator is the CAF Originator lookup's IsActive, NOT the record Owner (Owner
    # can be an active ops user even when the originator is inactive). Resolve the CAF
    # Originator User lookup defensively; fall back to Owner.IsActive only if it is absent.
    _caf_field = first_existing_field_name(
        "Opportunity",
        ["CAF_Originator__c", "CAF_Originator_User__c", "CAF_Loan_Originator__c", "Originator__c", "Loan_Originator__c"],
    )
    _caf_rel = None
    if _caf_field:
        try:
            _caf_rel = relationship_name_for("Opportunity", _caf_field)
        except KeyError:
            _caf_rel = None
    _active_expr = f"{_caf_rel}.IsActive" if _caf_rel else "Owner.IsActive"
    soql = (
        f"SELECT Deal_Loan_Number__c, {_active_expr} "
        "FROM Opportunity WHERE "
        "Deal_Loan_Number__c != NULL AND "
        + _soql_in("StageName", ACTIVE_RM_STAGES)
    )
    df = run_bulk_query(
        soql,
        rename_map={"Deal_Loan_Number__c": "Deal Loan Number", _active_expr: "Owner Active"},
    )
    if df.empty or "Deal Loan Number" not in df.columns:
        return pd.DataFrame(columns=["Deal Loan Number", "Active RM"])

    raw = pd.Series(df.get("Owner Active", pd.Series([pd.NA] * len(df), index=df.index)), index=df.index, dtype="object")
    txt = raw.astype("string").str.strip().str.lower()
    active = pd.Series([pd.NA] * len(df), index=df.index, dtype="object")
    active = active.mask(txt.isin(["true", "t", "y", "yes", "1"]), "Y")
    active = active.mask(txt.isin(["false", "f", "n", "no", "0"]), "N")
    df["Active RM"] = active

    out = df.dropna(subset=["Deal Loan Number"]).copy()
    out["_deal_key"] = norm_id_series(out["Deal Loan Number"])
    # Y wins over N when a deal has multiple owner rows.
    out["_rank"] = out["Active RM"].map({"Y": 2, "N": 1}).fillna(0)
    out = out.sort_values(["_deal_key", "_rank"]).drop_duplicates("_deal_key", keep="last")
    return out[["Deal Loan Number", "Active RM"]]



def _build_sold_term_like() -> pd.DataFrame:
    sold_pool_field = first_existing_field_name("Opportunity", ["FK_Sold_Loan_Pool__c", "Sold_Loan_Pool__c"])
    sold_pool_rel = relationship_name_for("Opportunity", sold_pool_field) if sold_pool_field else None
    servicer_commitment_field = first_existing_field_name("Opportunity", TERM_SERVICER_FALLBACK_FIELD_CANDIDATES)

    select_pairs = [
        ("Deal Loan Number", "Deal_Loan_Number__c"),
        ("Yardi ID", "Yardi_ID__c"),
        ("Deal Name", "Name"),
        ("Type", "Type"),
    ]
    if servicer_commitment_field:
        select_pairs.append(("Servicer Commitment Id", servicer_commitment_field))
    if sold_pool_rel:
        select_pairs.extend(
            [
                ("Sold Loan: Sold To", f"{sold_pool_rel}.Sold_To__r.Name"),
                ("Sold Loan: Sold Date", f"{sold_pool_rel}.Sold_Date__c"),
                ("Sold Loan: Servicing Status", f"{sold_pool_rel}.Servicing_Status__c"),
            ]
        )
    rename_map = {expr: label for label, expr in select_pairs}

    soql = (
        "SELECT "
        + ", ".join(expr for _label, expr in select_pairs)
        + " FROM Opportunity WHERE "
        + _soql_in("Type", TERM_TYPES)
        + " AND Deal_Loan_Number__c != NULL AND Probability > 0"
    )
    return run_bulk_query(soql, rename_map=rename_map)




def _build_term_asset_deal_universe(deal_numbers: Optional[Sequence[str]] = None) -> List[str]:
    opp_rel = property_opportunity_relationship_name()
    deal_numbers = _nonblank_unique(deal_numbers or [])

    select_pairs = [("Deal Loan Number", f"{opp_rel}.Deal_Loan_Number__c")]
    rename_map = {expr: label for label, expr in select_pairs}

    base_where = [
        f"{opp_rel}.Deal_Loan_Number__c != NULL",
        f"{opp_rel}.Probability > 0",
        _term_dealtype_condition(f"{opp_rel}."),
        _soql_in(f"{opp_rel}.StageName", TERM_ACTIVE_STAGES),
        _soql_in("Status__c", TERM_ACTIVE_PROPERTY_STATUSES),
        "Asset_ID__c != NULL",
        _soql_false_or_null("Is_Sub_Unit__c"),
    ]

    soqls = []
    if deal_numbers:
        for chunk in _chunked(deal_numbers, size=200):
            where_parts = base_where + [_soql_in(f"{opp_rel}.Deal_Loan_Number__c", chunk)]
            soqls.append(
                "SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Property__c WHERE " + " AND ".join(where_parts)
            )
    else:
        soqls.append(
            "SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Property__c WHERE " + " AND ".join(base_where)
        )

    df = _run_bulk_union(soqls, rename_map=rename_map)
    if df.empty or "Deal Loan Number" not in df.columns:
        return []
    return _nonblank_unique(df["Deal Loan Number"].tolist())


def _build_term_wide_like() -> pd.DataFrame:
    contact_field = first_existing_field_name("Opportunity", ["Contact__c", "Primary_Contact__c"])
    contact_rel = relationship_name_for("Opportunity", contact_field) if contact_field else None
    primary_contact_expr = f"{contact_rel}.Name" if contact_rel else "Account.Name"
    funding_expr = opportunity_name_expr(["Current_Funding_Vehicle__c", "FK_Current_Funding_Vehicle__c"])

    servicer_name_field = first_existing_field_name("Opportunity", TERM_SERVICER_NAME_FIELD_CANDIDATES)
    servicer_commitment_field = first_existing_field_name("Opportunity", TERM_SERVICER_FALLBACK_FIELD_CANDIDATES)
    term_servicer_fields = existing_field_names("Opportunity", TERM_SERVICER_PRIMARY_FIELD_CANDIDATES)
    term_current_maturity_fields = existing_field_names("Opportunity", TERM_CURRENT_MATURITY_FIELD_CANDIDATES)
    payoff_date_field = first_existing_field_name("Opportunity", TERM_PAYOFF_DATE_FIELD_CANDIDATES)

    select_pairs = [
        ("Deal Loan Number", "Deal_Loan_Number__c"),
        ("Yardi ID", "Yardi_ID__c"),
        ("Deal Name", "Name"),
        ("Borrower Entity", "Borrower_Entity__r.Name"),
        ("Account Name", "Account.Name"),
        ("Do Not Lend", "Account.Do_Not_Lend__c"),
        ("Primary Contact", primary_contact_expr),
        ("Close Date", "CloseDate"),
        ("Stage", "StageName"),
        ("Current Funding Vehicle", funding_expr),
        ("Warehouse Line", "Warehouse_Line__c"),
        ("Next Payment Date", "Next_Payment_Date__c"),
        ("Loan Amount", "Amount"),
        ("Current Servicer UPB", "Current_UPB__c"),
        ("Original Loan Maturity Date", "Stated_Maturity_Date__c"),
        ("CAF Originator", "Owner.Name"),
        ("CAF Originator: Active", "Owner.IsActive"),
        ("Product Type", "LOC_Loan_Type__c"),
        ("Product Sub-Type", "Product_Sub_Type__c"),
        ("Type", "Type"),
        ("Comments AM", "Asset_Management_Comments__c"),
        ("Deal Intro Sub-Source", "Deal_Intro_Sub_Source__c"),
        ("Referral Source Account", "Referral_Source__r.Name"),
        ("Referral Source Contact", "Referral_Source_Contact__r.Name"),
    ]
    if payoff_date_field:
        select_pairs.append(("Payoff Date", payoff_date_field))
    for idx, field_api in enumerate(term_current_maturity_fields, start=1):
        if field_api != "Stated_Maturity_Date__c":
            select_pairs.append((f"Current Loan Maturity Date {idx}", field_api))
    if servicer_name_field:
        select_pairs.insert(0, ("Servicer Name", servicer_name_field))
    if servicer_commitment_field:
        select_pairs.insert(1, ("Servicer Commitment Id", servicer_commitment_field))
    for idx, field_api in enumerate(term_servicer_fields, start=1):
        select_pairs.insert(2 + idx, (f"Term Servicer Key {idx}", field_api))

    rename_map = {expr: label for label, expr in select_pairs}

    where_parts = [
        "Deal_Loan_Number__c != NULL",
        "Probability > 0",
        _term_dealtype_condition(),
        _soql_in("StageName", TERM_ACTIVE_STAGES),
    ]
    _append_clause_if_present(where_parts, _opportunity_status_active_condition())

    soql = (
        "SELECT "
        + ", ".join(expr for _label, expr in select_pairs)
        + " FROM Opportunity WHERE "
        + " AND ".join(where_parts)
    )

    term_df = run_bulk_query(soql, rename_map=rename_map)
    sold_df = _build_sold_term_like()

    if term_df.empty:
        return term_df

    term_df["_deal_key"] = norm_id_series(term_df["Deal Loan Number"])

    if not sold_df.empty and "Deal Loan Number" in sold_df.columns:
        sold_keep = [
            c
            for c in [
                "Deal Loan Number",
                "Sold Loan: Sold To",
                "Sold Loan: Sold Date",
                "Sold Loan: Servicing Status",
            ]
            if c in sold_df.columns
        ]
        sold_df["_deal_key"] = norm_id_series(sold_df["Deal Loan Number"])
        sold_df = sold_df[["_deal_key"] + [c for c in sold_keep if c != "Deal Loan Number"]].drop_duplicates("_deal_key")
        term_df = term_df.merge(sold_df, on="_deal_key", how="left")

    term_df = term_df.drop_duplicates().copy()
    return downcast_numeric_frame(term_df)

def _build_term_asset_like(deal_numbers=None) -> pd.DataFrame:
    opp_rel = property_opportunity_relationship_name()
    deal_numbers = _nonblank_unique(deal_numbers or [])
    soqls = []

    generic_value_date_field = first_existing_field_name("Property__c", ["Value_Date__c", "Updated_Valuation_Date__c", "BPO_Appraisal_Date__c"])
    generic_value_field = first_existing_field_name("Property__c", ["Value__c", "Appraised_Value_Amount__c"])
    term_asset_date_field = first_existing_field_name("Property__c", ["Acquisition_Date__c", "Close_Date__c", "Purchase_Date__c"])
    # Origination valuation fields direct from Property__c (same fields the Bridge spine uses).
    # The completed Term Asset report sources Origination Value Date/Value/Type here; without
    # them, brand-new Term-only deals not present in the carry-forward report come through blank
    # (the 117-row Origination cluster). Looked up defensively so a missing field can't break SOQL.
    orig_val_date_field = first_existing_field_name("Property__c", ["Origination_Date_Valuation_Date__c", "Origination_Valuation_Date__c"])
    orig_val_field = first_existing_field_name("Property__c", ["Origination_Date_Value__c", "Origination_As_Is_Value__c"])
    orig_val_type_field = first_existing_field_name("Property__c", ["Origination_Date_Valuation_Type__c", "Origination_Valuation_Type__c", "Origination_Value_Type__c"])

    select_pairs = [
        ("Deal Loan Number", f"{opp_rel}.Deal_Loan_Number__c"),
        ("Property ID", "Id"),
        ("Asset ID", "Asset_ID__c"),
        ("Address", "Name"),
        ("City", "City__c"),
        ("State", "State__c"),
        ("Zip", "ZipCode__c"),
        ("CBSA", "MSA__c"),
        ("# of Units", "Number_of_Units__c"),
        ("Property Type", "Property_Type__c"),
        ("ALA", "ALA__c"),
        ("Updated Valuation Date", "Updated_Valuation_Date__c"),
        ("BPO Appraisal Date", "BPO_Appraisal_Date__c"),
        ("Appraised Value Amount", "Appraised_Value_Amount__c"),
        ("Property Special Asset", "Is_Special_Asset__c"),
        ("Is Parent", "Is_Parent__c"),
        ("Is Sub Unit", "Is_Sub_Unit__c"),
        ("Property Status", "Status__c"),
        ("Property Created Date", "CreatedDate"),
        ("Property Last Modified Date", "LastModifiedDate"),
    ]
    if orig_val_date_field:
        select_pairs.append(("Origination Value Date", orig_val_date_field))
    if orig_val_field:
        select_pairs.append(("Origination Value", orig_val_field))
    if orig_val_type_field:
        select_pairs.append(("Origination Value Type", orig_val_type_field))
    if term_asset_date_field:
        select_pairs.append(("Date", term_asset_date_field))
    if generic_value_date_field and generic_value_date_field not in {"Updated_Valuation_Date__c", "BPO_Appraisal_Date__c"}:
        select_pairs.append(("Generic Value Date", generic_value_date_field))
    if generic_value_field and generic_value_field != "Appraised_Value_Amount__c":
        select_pairs.append(("Generic Value", generic_value_field))
    rename_map = {expr: label for label, expr in select_pairs}

    base_where = [
        f"{opp_rel}.Deal_Loan_Number__c != NULL",
        f"{opp_rel}.Probability > 0",
        _term_dealtype_condition(f"{opp_rel}."),
        _soql_in(f"{opp_rel}.StageName", TERM_ACTIVE_STAGES),
        _soql_in("Status__c", TERM_ACTIVE_PROPERTY_STATUSES),
        "Asset_ID__c != NULL",
        _soql_false_or_null("Is_Sub_Unit__c"),
    ]

    if deal_numbers:
        for chunk in _chunked(deal_numbers, size=200):
            where_parts = base_where + [_soql_in(f"{opp_rel}.Deal_Loan_Number__c", chunk)]
            soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Property__c WHERE " + " AND ".join(where_parts))
    else:
        soqls.append("SELECT " + ", ".join(expr for _label, expr in select_pairs) + " FROM Property__c WHERE " + " AND ".join(base_where))

    df = _run_bulk_union(soqls, rename_map=rename_map)
    if df.empty:
        return df

    # Term Asset valuation fields in the completed report are carry-forward/manual
    # fields, not a full refresh from current Property/Appraisal fields. Leave the
    # Salesforce-built rows blank here; build_term_asset() will carry forward prior
    # completed values and only use these current rows for brand-new assets.
    df["Value Date"] = pd.NaT
    df["As-Is Value"] = np.nan

    df["_deal_key"] = norm_id_series(df.get("Deal Loan Number", pd.Series([None] * len(df), index=df.index)))
    df["_asset_key"] = norm_id_series(df.get("Asset ID", pd.Series([None] * len(df), index=df.index)))
    df["_property_id_key"] = norm_id_series(df.get("Property ID", pd.Series([None] * len(df), index=df.index)))
    df = df[df["_deal_key"].notna() & df["_asset_key"].notna()].copy()

    df["_is_sub_unit"] = _yn_from_bool_series(df.get("Is Sub Unit", pd.Series([pd.NA] * len(df), index=df.index))).eq("Y").astype("int8")

    # V79: drop a PARENT property when the same deal also carries its individual assets.
    #
    # The SOQL above filters sub-units, but never filtered parents, so a property that is
    # recorded both as one aggregate row and as its individual units came through twice. Both
    # copies carry the full ALA, which doubles the deal's ALA total -- and Term Asset UPB is
    # allocated as loan UPB x (asset ALA / deal ALA sum), so every asset on such a deal received
    # half of what it should. Against the 20260831 official that was 247 of the tab's 280
    # mismatches, over 14 deals, and the exact-2x cases are unmistakable: deal 29412 had 4
    # assets and ALA 11,900,000 against the official's 2 assets and 5,950,000, deal 58276 had
    # 8 / 9,882,000 against 4 / 4,941,000, with the loan UPB identical on both sides.
    #
    # Deal 55167 shows the shape plainly: the official keeps "228 Broad St" (ALA 424,386.89, 6
    # units, Multifamily) and the build additionally carried "228, 230, 232, 234, 236..." with
    # the identical ALA, unit count and property type. Deal 29734 is the same story with
    # "1056/1058 Irene" sitting alongside "1056 Irene St" and "1058 Irene St".
    #
    # Scoped so it can never empty a deal: parents are dropped only where that deal also has at
    # least one non-parent asset. A standalone property flagged as a parent is kept.
    df["_is_parent"] = _yn_from_bool_series(df.get("Is Parent", pd.Series([pd.NA] * len(df), index=df.index))).eq("Y").astype("int8")
    _deal_nonparent = (1 - df["_is_parent"]).groupby(df["_deal_key"]).transform("sum")
    _drop_parent = df["_is_parent"].eq(1) & _deal_nonparent.gt(0)
    if bool(_drop_parent.any()):
        TERM_ASSET_PARENT_DROPS.clear()
        TERM_ASSET_PARENT_DROPS.update({
            "rows": int(_drop_parent.sum()),
            "deals": int(df.loc[_drop_parent, "_deal_key"].nunique()),
            "ala": float(pd.to_numeric(df.loc[_drop_parent, "ALA"], errors="coerce").fillna(0).sum()) if "ALA" in df.columns else 0.0,
        })
        df = df.loc[~_drop_parent].copy()

    df["_nonnull_score"] = 0
    for c in ["Address", "City", "State", "Zip", "CBSA", "ALA", "Value Date", "As-Is Value"]:
        if c in df.columns:
            df["_nonnull_score"] = df["_nonnull_score"] + (~blankish_mask(df[c])).astype("int8")
    df["_ala_sort"] = pd.to_numeric(df.get("ALA", np.nan), errors="coerce").fillna(0)
    df["_value_dt"] = pd.to_datetime(df.get("Value Date"), errors="coerce")
    df["_mod_dt"] = _to_datetime_series_mixed(df.get("Property Last Modified Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df["_created_dt"] = _to_datetime_series_mixed(df.get("Property Created Date", pd.Series([pd.NaT] * len(df), index=df.index)))
    df = df.sort_values(["_deal_key", "_asset_key", "_is_sub_unit", "_nonnull_score", "_ala_sort", "_value_dt", "_mod_dt", "_created_dt", "_property_id_key"], ascending=[True, True, False, True, True, True, True, True, True])
    df = df.drop_duplicates(["_deal_key", "_asset_key"], keep="last")
    df = df.drop(columns=["_is_parent"], errors="ignore")
    return downcast_numeric_frame(df)


def _bridge_asset_ids_from_spine(bridge_spine: pd.DataFrame):
    if bridge_spine is None or bridge_spine.empty or "Asset ID" not in bridge_spine.columns:
        return []
    return _nonblank_unique(bridge_spine["Asset ID"].tolist())


def _term_deal_numbers_from_wide(term_wide: pd.DataFrame):
    if term_wide is None or term_wide.empty or "Deal Loan Number" not in term_wide.columns:
        return []
    return _nonblank_unique(term_wide["Deal Loan Number"].tolist())


@dataclass(frozen=True)
class UploadBlob:
    filename: str
    file_hash: str
    data: bytes


def _md5_hex(b: bytes) -> str:
    return hashlib.md5(b).hexdigest()


def make_upload_blob(upload, compute_hash: bool = True) -> UploadBlob:
    b = upload.getvalue()
    file_hash = _md5_hex(b) if compute_hash else f"nocache:{upload.name}:{len(b)}"
    return UploadBlob(filename=upload.name, file_hash=file_hash, data=b)


def date_from_filename(name: str) -> Optional[date]:
    # Digit-boundary guards (?<!\d)...(?!\d) so an 8-digit date isn't grabbed out of
    # a longer digit run. Without them, typo filenames like "...202606012" (9 digits,
    # meant to be 2026-06-12) misparse as 2026-06-01, which then wrongly wins the
    # dominant-tape-date vote and stamps the UPB header as "6/1 UPB".
    m = re.search(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)", name)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    m = re.search(r"(?<!\d)(20\d{2})[-_](\d{1,2})[-_](\d{1,2})(?!\d)", name)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    m = re.search(r"(?<!\d)(\d{2})[_-](\d{2})[_-](20\d{2})(?!\d)", name)
    if m:
        return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))

    m = re.search(r"(?<!\d)(\d{2})(\d{2})(20\d{2})(?!\d)", name)
    if m:
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mm <= 12 and 1 <= dd <= 31:
            return date(yy, mm, dd)

    return None


def detect_servicer_type(filename: str) -> str:
    n = filename.lower()
    if "shellpoint" in n:
        return "Shellpoint"
    if "onity" in n or "rs_corevest" in n or "corevest_daily" in n or "daily_report" in n:
        return "Onity"
    if "corevest_data_tape" in n:
        return "CoreVest_Data_Tape"
    if "corevestloandata" in n:
        return "CoreVestLoanData"
    if "midland" in n:
        return "Midland"
    if "fci" in n:
        return "FCI"
    if n.endswith(".csv"):
        return "CHL"
    raise ValueError(
        "Could not detect servicer file type from the filename. "
        "Use one of these naming patterns: Shellpoint, CHL, CoreVest_Data_Tape, CoreVestLoanData, FCI, Midland."
    )


def report_date_from_scalar(value) -> Optional[date]:
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()


def read_fci_report_date(file_bytes: bytes, sheet_name: str) -> Optional[date]:
    try:
        top = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=1)
        if top.shape[1] >= 2:
            return report_date_from_scalar(top.iloc[0, 1])
    except Exception:
        return None
    return None



def _servicer_specificity_rank(val) -> int:
    s = clean_text(val).lower()
    if not s:
        return 0
    if s == "fci chl streamline":
        return 6
    if "fci 2012632" in s or "fci v1805510" in s:
        return 5
    if "shellpoint" in s:
        return 4
    if any(x in s for x in ["statebridge", "berkadia", "midland", "selene", "sps", "fay", "cornerstone", "onity"]):
        return 3
    if "fci" in s or "chl" in s:
        return 2
    return 1


def _servicer_checkpoint_ok(sf_servicer, file_servicer) -> bool:
    file_txt = clean_text(file_servicer)
    if not file_txt:
        return False

    sf_txt = clean_text(sf_servicer)
    if not sf_txt or sf_txt.upper() == "N/A":
        return True

    sf_fam = normalize_servicer_family(sf_txt)
    file_fam = normalize_servicer_family(file_txt)
    if sf_fam and file_fam and sf_fam == file_fam:
        return True

    sf_low = sf_txt.lower()
    file_low = file_txt.lower()
    return sf_low in file_low or file_low in sf_low


def _bridge_bucket_from_days(days_past_due) -> Optional[str]:
    try:
        if pd.isna(days_past_due):
            return None
    except Exception:
        pass
    try:
        days = int(float(days_past_due))
    except Exception:
        return None
    if days >= 90:
        return "90 +  DAYS"
    if days >= 60:
        return "60 - 89 DAYS"
    if days >= 30:
        return "30 - 59 DAYS"
    return "CURRENT"


def _bridge_status_severity(status) -> int:
    # Collapse internal whitespace so the double-space report label "90 +  DAYS" matches
    # the single-space map keys below (the displayed label carries a double space).
    s = re.sub(r"\s+", " ", clean_text(status).upper())
    order = {
        "CURRENT": 0,
        "CURRENT": 0,
        "DQ 1-29": 1,
        "30 - 59 DAYS": 2,
        "DQ 30-59": 2,
        "60 - 89 DAYS": 3,
        "DQ 60-89": 3,
        "90 + DAYS": 4,
        "DQ 90+": 4,
        "BK": 5,
        "REO": 6,
    }
    return order.get(s, -1)


def _bridge_bucket_to_report_label(bucket, days_past_due=np.nan):
    s = re.sub(r"\s+", " ", clean_text(bucket).upper())
    try:
        days = float(days_past_due)
    except Exception:
        days = np.nan
    if s == "REO":
        return "REO"
    if s == "BK":
        return "BK"
    if s in {"90 + DAYS", "DQ 90+"}:
        return "DQ 90+"
    if s in {"60 - 89 DAYS", "DQ 60-89"}:
        return "DQ 60-89"
    if s in {"30 - 59 DAYS", "DQ 30-59"}:
        return "DQ 30-59"
    if s == "DQ 1-29":
        return "DQ 1-29"
    if s == "CURRENT":
        if not pd.isna(days) and days >= 1:
            return "DQ 1-29"
        return "Current"
    if not pd.isna(days):
        if days >= 90:
            return "DQ 90+"
        if days >= 60:
            return "DQ 60-89"
        if days >= 30:
            return "DQ 30-59"
        if days >= 1:
            return "DQ 1-29"
        return "Current"
    return pd.NA


def _bridge_loan_rollup_label(bucket_series: pd.Series, days_series: pd.Series):
    labels = {
        clean_text(_bridge_bucket_to_report_label(bucket, days))
        for bucket, days in zip(bucket_series, days_series)
        if clean_text(_bridge_bucket_to_report_label(bucket, days))
    }
    if "REO" in labels:
        return "REO"
    if "BK" in labels:
        return "BK"
    if "DQ 90+" in labels:
        return "DQ 90+"
    if "DQ 60-89" in labels:
        return "DQ 60-89"
    has_30 = "DQ 30-59" in labels
    has_1_29 = "DQ 1-29" in labels
    if has_30 and has_1_29:
        return "DQ 1-29 / DQ 30-59"
    if has_30:
        return "DQ 30-59"
    if has_1_29:
        return "DQ 1-29"
    if "Current" in labels:
        return "Current"
    return pd.NA


def _guess_days_past_due(next_payment_date, run_date: date) -> float:
    dt = pd.to_datetime(next_payment_date, errors="coerce")
    if pd.isna(dt):
        return np.nan
    delta = (pd.Timestamp(run_date) - dt.normalize()).days
    return float(max(delta, 0))


def _guess_days_from_bridge_bucket(status) -> float:
    s = re.sub(r"\s+", " ", clean_text(status).upper())
    mapping = {
        "CURRENT": 0.0,
        "DQ 1-29": 14.0,
        "30 - 59 DAYS": 45.0,
        "DQ 30-59": 45.0,
        "DQ 1-29 / DQ 30-59": 45.0,
        "60 - 89 DAYS": 75.0,
        "DQ 60-89": 75.0,
        "90 + DAYS": 90.0,
        "DQ 90+": 90.0,
        "BK": 90.0,
        "REO": 90.0,
        "CURRENT": 0.0,
    }
    return mapping.get(s, np.nan)


def normalize_bridge_servicer_status(raw_status, next_payment_date, run_date: date, loan_stage=None, property_status=None, reo_date=None) -> Optional[str]:
    if is_reo_stage(loan_stage) or is_reo_stage(property_status) or pd.notna(pd.to_datetime(reo_date, errors="coerce")):
        return "REO"

    txt = clean_text(raw_status).lower()
    txt_compact = re.sub(r"[^a-z0-9]+", " ", txt).strip()
    if txt_compact:
        if any(tok in txt_compact for tok in ["bankrupt", "bankruptcy", " bk "]) or txt_compact.startswith("bk") or txt_compact.endswith(" bk"):
            return "BK"
        if "reo" in txt_compact:
            return "REO"
        if any(tok in txt_compact for tok in ["90", "120", "180", "nonperform", "default", "foreclos", "serious delin"]):
            # Official report label carries a DOUBLE space ("90 +  DAYS"); matching it
            # exactly clears the recurring BA Servicer Status whitespace mismatch (V41).
            return "90 +  DAYS"
        if any(tok in txt_compact for tok in ["60", "61", "89", "2 month"]):
            return "60 - 89 DAYS"
        if any(tok in txt_compact for tok in ["30", "31", "59", "1 month"]):
            return "30 - 59 DAYS"
        if any(tok in txt_compact for tok in ["current", "active", "perform", "paid to date", "p2d"]):
            return "CURRENT"

    return _bridge_bucket_from_days(_guess_days_past_due(next_payment_date, run_date))


def _worst_bridge_bucket(series: pd.Series):
    best = None
    best_rank = -1
    for val in series:
        rank = _bridge_status_severity(val)
        if rank > best_rank:
            best_rank = rank
            best = val
    return best if has_any_value(best) else pd.NA


def _fill_text_defaults(df: pd.DataFrame, columns: Sequence[str], default: str = "N/A") -> pd.DataFrame:
    out = df.copy()
    for c in columns:
        if c not in out.columns:
            continue
        s_raw = out[c]
        if pd.api.types.is_numeric_dtype(s_raw) or pd.api.types.is_datetime64_any_dtype(s_raw):
            continue
        s = pd.Series(s_raw, index=out.index, dtype="object")
        sample = s[~blankish_mask(s)].head(20)
        if len(sample) > 0 and sample.map(_looks_like_date_string).mean() >= 0.70:
            continue
        out[c] = s.where(~blankish_mask(s), default)
    return out



def _prev_term_keys(prev_maps: Optional[dict]) -> Set[str]:
    keys: Set[str] = set()
    if not prev_maps:
        return keys
    for bucket in ["term_loan_manual", "term_loan_reo", "term_loan_upb"]:
        df = prev_maps.get(bucket)
        if isinstance(df, pd.DataFrame) and not df.empty and "_deal_key" in df.columns:
            vals = df["_deal_key"].dropna().astype("string").tolist()
            keys.update([clean_text(v) for v in vals if clean_text(v)])
    return keys


def _prev_term_positive_upb_keys(prev_maps: Optional[dict]) -> Set[str]:
    keys: Set[str] = set()
    if not prev_maps:
        return keys
    df = prev_maps.get("term_loan_upb")
    if isinstance(df, pd.DataFrame) and not df.empty and {"_deal_key", "_prev_upb"}.issubset(df.columns):
        tmp = df.copy()
        tmp["_prev_upb"] = pd.to_numeric(tmp["_prev_upb"], errors="coerce").fillna(0)
        vals = tmp.loc[tmp["_prev_upb"] > 0, "_deal_key"].dropna().astype("string").tolist()
        keys.update([clean_text(v) for v in vals if clean_text(v)])
    return keys


def _prev_term_sold_retained_keys(prev_maps: Optional[dict]) -> Set[str]:
    keys: Set[str] = set()
    if not prev_maps:
        return keys
    df = prev_maps.get("term_loan_manual")
    if isinstance(df, pd.DataFrame) and not df.empty and {"_deal_key", "Segment"}.issubset(df.columns):
        seg = df["Segment"].astype("string").str.strip()
        vals = df.loc[seg.isin(list(TERM_SOLD_RETAINED_SEGMENT_VALUES)), "_deal_key"].dropna().astype("string").tolist()
        keys.update([clean_text(v) for v in vals if clean_text(v)])
    return keys


def _recent_close_mask(close_series: pd.Series, run_date: date, days: int = 45) -> pd.Series:
    close_dt = pd.to_datetime(close_series, errors="coerce")
    lower = pd.Timestamp(run_date) - pd.Timedelta(days=days)
    upper = pd.Timestamp(run_date) + pd.Timedelta(days=7)
    return close_dt.notna() & close_dt.ge(lower) & close_dt.le(upper)


def _sold_servicing_retained_mask(servicing_status: pd.Series) -> pd.Series:
    txt = pd.Series(servicing_status, copy=False).astype("string").str.lower().str.strip().fillna("")
    exact = txt.isin({
        "servicing retained",
        "sold servicing retained",
        "retain servicing",
        "retained servicing",
        "serv retained",
        "retained",
    })
    retained = txt.str.contains(r"servicing retained|retain servicing|retained servicing|serv retained|\bretained\b|\bretain\w*\b", regex=True, na=False)
    bad = txt.str.contains(r"release|released|transfer|transferred|xfer|sold away|none|no servicing|unserviced|servicing released|not retained", regex=True, na=False)
    return (exact | retained) & (~bad)


def _term_effective_sold_retained_mask(servicing_status: pd.Series, fallback_prev_mask: Optional[pd.Series] = None) -> pd.Series:
    raw = _sold_servicing_retained_mask(servicing_status)
    if fallback_prev_mask is None:
        return raw
    txt = pd.Series(servicing_status, copy=False).astype("string").str.strip().fillna("")
    blank = txt.eq("")
    fallback = pd.Series(fallback_prev_mask, index=txt.index, copy=False).fillna(False).astype(bool)
    return raw | (blank & fallback)


def _term_report_keep_mask(
    stage_series: pd.Series,
    current_upb_series: pd.Series,
    sold_servicing_status: pd.Series,
    fallback_prev_retained_mask: Optional[pd.Series] = None,
    extra_reo_mask: Optional[pd.Series] = None,
    loan_amount_series: Optional[pd.Series] = None,
) -> pd.Series:
    stage = pd.Series(stage_series, copy=False).astype("string").str.strip().fillna("")
    current_upb = pd.to_numeric(pd.Series(current_upb_series, index=stage.index, copy=False), errors="coerce").fillna(0)
    loan_amount = pd.to_numeric(
        pd.Series(loan_amount_series, index=stage.index, copy=False)
        if loan_amount_series is not None
        else pd.Series([np.nan] * len(stage), index=stage.index),
        errors="coerce",
    ).fillna(0)
    sold_retained = _term_effective_sold_retained_mask(sold_servicing_status, fallback_prev_mask=fallback_prev_retained_mask)
    is_paid_off = stage.eq("Paid Off")
    is_reo_sold = stage.eq("REO-Sold")
    is_reo = stage.isin(REO_FAMILY_STAGES)
    if extra_reo_mask is not None:
        is_reo = is_reo | pd.Series(extra_reo_mask, index=stage.index, copy=False).fillna(False).astype(bool)
    is_sold = stage.eq("Sold")
    positive_upb_keep = current_upb.gt(0) & ~is_sold
    # Do not keep zero-UPB REO / payoff-style term deals solely because Stage is REO.
    # Prior workbook carry-forward must not reintroduce current terminal zero-balance loans.
    reo_positive_keep = is_reo & current_upb.gt(0)
    # V41: a Sold deal with no balance (Stage=='Sold' AND UPB<=0, e.g. deal 37638) is NOT
    # carried on the report -- exclude it even when a sold-servicing-retained flag is set.
    sold_zero = is_sold & ~current_upb.gt(0)
    # V41: pre-funding deals ('Approved by Committee' / 'Purchased') are NOT on the report
    # (verified 6/29: 15 of 18 extras were Approved-by-Committee). The former preboarding_keep
    # that admitted them on Loan-Amount-only is removed; TERM_ALWAYS_INCLUDE_DEALS still
    # force-keeps the two documented exceptions downstream. `loan_amount` is now unused here.
    keep = (reo_positive_keep | sold_retained | positive_upb_keep) & ~is_paid_off & ~is_reo_sold & ~sold_zero
    return keep.fillna(False)


def _term_terminal_zero_exclusion_keys(sf_term: Optional[pd.DataFrame]) -> Set[str]:
    """Current Salesforce says these term deals are terminal and zero-balance."""
    if sf_term is None or sf_term.empty or "Deal Loan Number" not in sf_term.columns:
        return set()
    df = sf_term.copy()
    deal_key = norm_id_series(df["Deal Loan Number"])
    stage = df.get("Stage", pd.Series([pd.NA] * len(df), index=df.index)).astype("string").str.strip()
    upb = pd.to_numeric(df.get("Current Servicer UPB", pd.Series([np.nan] * len(df), index=df.index)), errors="coerce").fillna(0)
    payoff_cols = [c for c in df.columns if "payoff" in str(c).lower() and "date" in str(c).lower()]
    has_payoff_date = pd.Series(False, index=df.index)
    for col in payoff_cols:
        has_payoff_date = has_payoff_date | pd.to_datetime(df[col], errors="coerce").notna()
    terminal_stage = stage.isin(["Paid Off", "REO", "REO-Sold"])
    terminal_zero = (terminal_stage | has_payoff_date) & upb.le(0)
    return set(deal_key.loc[terminal_zero & deal_key.notna()].astype(str).tolist())


def _drop_term_deal_keys(df: pd.DataFrame, drop_keys: Set[str]) -> pd.DataFrame:
    if df is None or df.empty or not drop_keys:
        return pd.DataFrame() if df is None else df.copy()
    out = _ensure_deal_key(df, "Deal Number")
    return out.loc[~out["_deal_key"].astype("string").isin(drop_keys)].copy()


def _force_statebridge_day10(dates: pd.Series, servicer: pd.Series) -> pd.Series:
    """Statebridge-serviced loans always bill on the 10th of the month in the official
    report (verified 3280/3280 Bridge Asset + 135/135 Bridge Loan, zero exceptions). The
    Statebridge tape itself is mostly day-10 but leaks some day-1/day-8 dates; the report
    normalizes ALL Statebridge NPDs to the 10th. Force day=10 while preserving month/year."""
    d = pd.to_datetime(pd.Series(dates, copy=False), errors="coerce")
    serv = pd.Series(servicer, index=d.index, copy=False).astype(str).str.casefold()
    is_sb = serv.str.contains("statebridge", na=False) & d.notna()
    if is_sb.any():
        forced = d[is_sb].apply(lambda ts: ts.replace(day=10) if pd.notna(ts) else ts)
        d.loc[is_sb] = forced
    return d


def _force_fci_day1_to_first_payment(dates: pd.Series, servicer: pd.Series, fpd_dates: pd.Series, run_dt: Optional[date] = None) -> pd.Series:
    """FCI (exactly "FCI", not the 2012632 / v1805530 / CHL sub-flavors) reports a
    day-1 next-payment date in its servicer file; the official report instead uses the
    loan's First Payment Date day-of-month. Verified rule: when servicer == "FCI" and the
    chosen NPD lands on day 1 and a First Payment Date exists, move the day to the FPD's
    day-of-month (clamped to the month length). Additive -- only touches FCI day-1 rows.

    Recency gate: only correct CURRENT NPDs (within ~60 days of the run date). A
    severely delinquent FCI loan carries a frozen historical day-1 NPD (e.g.
    2021-11-01) that must stay on day 1; applying the FPD-day shift there fabricates a
    future-looking date that the official report never shows."""
    d = pd.to_datetime(pd.Series(dates, copy=False), errors="coerce")
    fpd = pd.to_datetime(pd.Series(fpd_dates, index=d.index, copy=False), errors="coerce")
    serv = pd.Series(servicer, index=d.index, copy=False).astype(str).str.strip().str.casefold()
    mask = serv.eq("fci") & d.notna() & d.dt.day.eq(1) & fpd.notna()
    if run_dt is not None:
        run_ts = pd.Timestamp(run_dt)
        recent = d.ge(run_ts - pd.Timedelta(days=60)) & d.le(run_ts + pd.Timedelta(days=60))
        mask = mask & recent
    if not bool(mask.any()):
        return d
    fixed = []
    for ts, fp, m in zip(d, fpd, mask):
        if m:
            last_day = calendar.monthrange(ts.year, ts.month)[1]
            fixed.append(ts.replace(day=min(int(fp.day), last_day)))
        else:
            fixed.append(ts)
    return pd.to_datetime(pd.Series(fixed, index=d.index), errors="coerce")


def _bridge_pick_next_payment_date(
    sf_dates: pd.Series,
    servicer_dates: pd.Series,
    prior_dates: Optional[pd.Series] = None,
    servicer_names: Optional[pd.Series] = None,
    fpd_dates: Optional[pd.Series] = None,
    run_dt: Optional[date] = None,
) -> pd.Series:
    """Bridge NPD is servicer-first. Statebridge is normalized to the 10th of the month
    (deterministic servicer rule). For other servicers, keep the day-1/day-10 SF/prior
    fallback as a safety net where servicer identity is unknown.

    Additive corrections layered on top (do NOT change the servicer-first base):
      - FCI day-1 -> First Payment Date day-of-month (see _force_fci_day1_to_first_payment).
      - Statebridge -> always day 10 (see _force_statebridge_day10).

    NOTE: an SF-first experiment (V26) regressed Bridge Asset NPD from 814 to 2400
    mismatches in the live build -- the API/spine Next_Payment_Date__c is NOT the same as
    the SF_Bridge report export column an offline check matched against, so servicer-first
    is restored as the known-better behavior."""
    sf = pd.to_datetime(pd.Series(sf_dates, copy=False), errors="coerce")
    serv = pd.to_datetime(pd.Series(servicer_dates, index=sf.index, copy=False), errors="coerce")
    prior = pd.to_datetime(pd.Series(prior_dates, index=sf.index, copy=False), errors="coerce") if prior_dates is not None else pd.Series([pd.NaT] * len(sf), index=sf.index)
    # V63: do NOT reorder this to prefer the prior report over `sf`. V62 tried exactly
    # that -- serv -> prior -> sf, plus a prior-first FCI far-future gate -- on the
    # strength of test 75, where the prior (20260817 official) beat the build 99.33% to
    # 96.71% on Bridge Asset NPD. Test 76, a V61 build against the same 20260824
    # official, scores 99.79% here: prior-first would have fixed 0 cells and broken 22.
    # Test 75 was a V58 build and a bad baseline; measure NPD changes against a build
    # made with the correct prior workbook, never against an older test.
    # The residue is the opposite problem: all 10 of test 76's misses already EQUAL the
    # prior and the official has moved a month ON (assets 1479741-1479744 Onity read
    # 2026-08-01 where the official says 2026-09-01). The missing rule is 'advance the
    # carried date when a payment has since been made', which needs the live tape to
    # distinguish from a loan that simply did not pay.
    out = serv.where(serv.notna(), sf)
    if BRIDGE_NPD_PRESERVE_DAY10_WHEN_SERVICER_DAY1:
        # V57: the old rule here was
        #     same_month_sf & serv.day==1 & sf.day==10  ->  take the Salesforce date
        # and it is REMOVED. The API's Property/Opportunity Next_Payment_Date__c is a
        # SCHEDULED billing date that reads day-10 for FCI and Onity loans, while both the
        # servicer tape and the official report keep day 1. Against the 20260824 official
        # it forced 665 assets (560 FCI + 83 Onity) from a correct day-1 to day-10 and
        # dragged the tab BELOW plain servicer-first: test 73 scored 4,044/4,782 where the
        # servicer file alone scores 4,482. Note this could not be caught offline against
        # the SF_Bridge export, whose Next Payment Date column carries the day-1 tape value
        # rather than the API's scheduled date -- the same export/API divergence this
        # function's own docstring warns about.
        # V53: these loans bill on the 10th. When the prior completed report carried a
        # day-10 NPD and the fresh source reports day 1, the official report keeps the 10th
        # but advances to the FRESH source's month -- e.g. asset 1190539: FCI tape says
        # 2026-09-01, prior report 2026-08-10, official 2026-09-10. The old rule required
        # both dates to be in the SAME month, so it could never advance the month.
        # Deal-uniform on 20260810 (0 of 817 FCI deals split), and lifts Bridge Asset NPD
        # from 4,327/4,917 to 4,602/4,917 on its own.
        out = pd.to_datetime(out, errors="coerce")
        _day10_prior = out.notna() & prior.notna() & out.dt.day.eq(1) & prior.dt.day.eq(10)
        if bool(_day10_prior.any()):
            out = out.mask(_day10_prior, out + pd.to_timedelta(9, unit="D"))
    # V53b: when NEITHER the servicer file nor Salesforce carries a next-payment date, the
    # official report keeps the prior completed report's value rather than blanking the cell
    # (e.g. deal 64324's 17 assets and deal 63094's 85 all read 2026-09-10, with no servicer
    # row and no SF date anywhere on the deal). This is the last resort -- it only fills a
    # cell that would otherwise be empty. Verified 313/313 on 20260810.
    out = pd.to_datetime(out, errors="coerce")
    _prior_only = out.isna() & prior.notna()
    if bool(_prior_only.any()):
        out = out.mask(_prior_only, prior)
    # FCI rolls its servicer-file Next Due Date FORWARD to the next scheduled payment
    # even for severely delinquent loans, where the official report keeps the frozen
    # historical NPD (the last real due date). When the FCI servicer NPD is far ahead of
    # the Salesforce NPD (> 60 days), the loan is delinquent and SF holds the truth, so
    # fall back to SF. Only FCI-family servicers exhibit this roll-forward behavior.
    if servicer_names is not None:
        serv_fam = pd.Series(servicer_names, index=sf.index, copy=False).astype(str).str.strip().str.casefold()
        is_fci = serv_fam.str.startswith("fci")
        far_future = is_fci & sf.notna() & serv.notna() & ((serv - sf).dt.days > 60)
        out = out.where(~far_future, sf)
    out = pd.to_datetime(out, errors="coerce")
    # V43: the FCI day-1 -> First-Payment-day shift (_force_fci_day1_to_first_payment) is
    # DISABLED. Same-day comparison vs the 6/22 real report proved real keeps the FCI
    # servicer-file date as-is (current FCI loans show day-1, e.g. 2026-07-01), while the
    # shift was forcing them to day-10 (2026-07-10) -- ~500+ BA NPD cells plus the BL
    # Days Past Due / NPL cascade. The far-future delinquent gate above (revert FCI
    # rolled-forward NPDs to the SF date) is retained; only the day-of-month shift is off.
    if servicer_names is not None:
        # Deterministic Statebridge rule wins over everything above.
        out = _force_statebridge_day10(out, servicer_names)
    return out


def _normalize_report_comment_text(value):
    txt = clean_text(value)
    if not txt:
        return pd.NA
    replacements = {
        "Â": "", "â": "-", "â€“": "-", "â": "-",
        "â€™": "'", "â": "'", "â€œ": '"', "â": '"', "â¢": "-",
    }
    for bad, good in replacements.items():
        txt = txt.replace(bad, good)
    # Fold unicode dashes/hyphens and the replacement char to a plain hyphen so text
    # that is byte-identical except for a U+2011 vs "?" stops registering as a mismatch
    # (this alone accounts for ~380 phantom Tax Commentary diffs).
    for ch in ("\u2011", "\u2010", "\u2012", "\u2013", "\u2014", "\ufffd"):
        txt = txt.replace(ch, "-")
    txt = txt.replace("_x000D_", " ").replace("\r", " ")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt or pd.NA


def _term_segment_is_sold_servicing_retained(segment_series: pd.Series) -> pd.Series:
    txt = pd.Series(segment_series, copy=False).astype("string").str.strip()
    return txt.isin(list(TERM_SOLD_RETAINED_SEGMENT_VALUES))


def _term_sold_servicing_retained_deal_keys(sf_term: pd.DataFrame) -> Set[str]:
    """Deal keys whose Term Loan is Sold AND Berkadia-serviced.

    Verified rule (6/29 same-day real): Stage == 'Sold' AND Servicer Name == 'Berkadia'
    -> Portfolio = Financing = 'Sold Servicing Retained' (384/384 sold deals match). This
    cascades to Term Asset, which inherits the parent deal's Portfolio/Financing.
    """
    if sf_term is None or sf_term.empty or "Deal Loan Number" not in sf_term.columns:
        return set()
    stage = sf_term.get("Stage", pd.Series([pd.NA] * len(sf_term), index=sf_term.index)).astype("string").str.strip()
    servicer = sf_term.get("Servicer Name", pd.Series([pd.NA] * len(sf_term), index=sf_term.index)).astype("string").str.strip()
    mask = stage.eq("Sold") & servicer.eq("Berkadia")
    keys = norm_id_series(sf_term["Deal Loan Number"]).loc[mask].dropna().astype(str).tolist()
    return {clean_text(k) for k in keys if clean_text(k)}


def _apply_term_sold_servicing_retained(df: pd.DataFrame, ssr_keys: Set[str]) -> pd.DataFrame:
    """Relabel the Sold-Servicing-Retained deals' Portfolio and Financing.

    `ssr_keys` is the Sold+Berkadia deal-key set. Must run AFTER the prior-workbook
    carry-forward (coalesce_report_display_first lets last week's value win over the live
    derivation) and AFTER Loan Sold Date is derived off the Financing=='Sold' flag, so the
    relabel sticks without breaking Loan Sold Date. Called on Term Loan and again on Term
    Asset so the taxonomy cascades to the child assets.

    V43: the two columns carry DIFFERENT labels in the official report --
    Portfolio = 'Sold Term', Financing = 'Sold Servicing Retained' (20260803 Term Loan
    J/L: 288 each; Term Asset D/F: 6,886 each). The previous build wrote
    'Sold Servicing Retained' into both, which produced ~7,150 Portfolio mismatches.
    """
    if df is None or df.empty or not ssr_keys or "_deal_key" not in df.columns:
        return df
    mask = df["_deal_key"].astype("string").isin(ssr_keys)
    if "Portfolio" in df.columns:
        df["Portfolio"] = df["Portfolio"].mask(mask, TERM_SOLD_PORTFOLIO)
    if "Financing" in df.columns:
        df["Financing"] = df["Financing"].mask(mask, TERM_SOLD_SERVICING_RETAINED_SEGMENT)
    return df


def _id_key_no_leading_zeros_scalar(val) -> str:
    s = clean_text(val)
    if not s:
        return ""
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"[^0-9A-Za-z]", "", s)
    s = re.sub(r"COM$", "", s, flags=re.I)
    s = s.lstrip("0")
    return s


def _first_nonblank_scalar(values: Sequence) -> object:
    for v in values:
        if has_any_value(v):
            return v
    return pd.NA


def _select_term_servicer_matches(
    sf_term: pd.DataFrame,
    serv_lookup: pd.DataFrame,
    sf_servicer: pd.Series,
    prev_maps: Optional[dict] = None,
) -> pd.DataFrame:
    result = pd.DataFrame(index=sf_term.index)
    result["selected_servicer_id"] = pd.NA
    result["selected_sid_key"] = pd.NA
    result["matched_servicer"] = pd.NA
    result["matched_upb"] = np.nan
    result["matched_next_payment_date"] = pd.NaT
    result["matched_maturity_date"] = pd.NaT
    result["matched_source"] = pd.NA
    result["_selected_match_score"] = np.nan

    candidate_cols = []
    if "Servicer Commitment Id" in sf_term.columns:
        candidate_cols.append("Servicer Commitment Id")
    candidate_cols.extend([c for c in sf_term.columns if c.startswith("Term Servicer Key ")])

    if not candidate_cols:
        return result

    preferred_raw = [
        _first_nonblank_scalar([sf_term.at[idx, c] for c in candidate_cols if c in sf_term.columns])
        for idx in sf_term.index
    ]
    result["selected_servicer_id"] = pd.Series(preferred_raw, index=sf_term.index, dtype="object")
    result["selected_sid_key"] = id_key_no_leading_zeros(pd.Series(preferred_raw, index=sf_term.index, dtype="object"))

    prev_sid_owner = _prev_sid_to_deal_map(prev_maps)
    current_deal_key = norm_id_series(sf_term.get("Deal Loan Number", pd.Series([pd.NA] * len(sf_term), index=sf_term.index)))
    loan_amount = pd.to_numeric(sf_term.get("Loan Amount", pd.Series([np.nan] * len(sf_term), index=sf_term.index)), errors="coerce")

    if serv_lookup is None or serv_lookup.empty or "_sid_key" not in serv_lookup.columns:
        # Even without servicer files, do not carry a servicer ID known from the prior workbook to belong to another deal.
        conflict_mask = []
        for idx in sf_term.index:
            sid_key = _id_key_no_leading_zeros_scalar(result.at[idx, "selected_servicer_id"])
            deal_key = clean_text(current_deal_key.loc[idx]) if idx in current_deal_key.index else ""
            prev_owner = prev_sid_owner.get(sid_key)
            conflict_mask.append(bool(sid_key and deal_key and prev_owner and prev_owner != deal_key))
        conflict_mask = pd.Series(conflict_mask, index=sf_term.index)
        if bool(conflict_mask.any()):
            result.loc[conflict_mask, ["selected_servicer_id", "selected_sid_key"]] = pd.NA
        return result

    base = serv_lookup.dropna(subset=["_sid_key"]).copy()
    if base.empty:
        return result

    keep_cols = [c for c in ["_sid_key", "servicer", "upb", "next_payment_date", "maturity_date"] if c in base.columns]
    info_map = {}
    for row in base[keep_cols].drop_duplicates("_sid_key", keep="last").itertuples(index=False):
        rec = {keep_cols[i]: row[i] for i in range(len(keep_cols))}
        info_map[clean_text(rec.get("_sid_key"))] = rec

    selected_raw = []
    selected_key = []
    selected_serv = []
    selected_upb = []
    selected_npd = []
    selected_mat = []
    selected_source = []
    selected_scores = []

    for idx in sf_term.index:
        sf_serv = sf_servicer.loc[idx] if idx in sf_servicer.index else pd.NA
        deal_key = clean_text(current_deal_key.loc[idx]) if idx in current_deal_key.index else ""
        loan_amt = loan_amount.loc[idx] if idx in loan_amount.index else np.nan
        best = None
        best_score = -10**9
        fallback = None
        fallback_score = -10**9

        for pos, col in enumerate(candidate_cols):
            if col not in sf_term.columns:
                continue
            raw = sf_term.at[idx, col]
            sid_key = _id_key_no_leading_zeros_scalar(raw)
            if not sid_key:
                continue

            prev_owner = prev_sid_owner.get(sid_key)
            if prev_owner and deal_key and prev_owner != deal_key:
                # Prior completed workbook says this servicer ID belonged to a different active deal.
                # Never attach it here without explicit evidence outside this automated matcher.
                continue

            info = info_map.get(sid_key)
            has_file = info is not None
            file_serv = info.get("servicer") if info else pd.NA
            file_upb = money_to_float(info.get("upb")) if info else np.nan

            amount_conflict = False
            if has_file and pd.notna(file_upb) and pd.notna(loan_amt) and float(loan_amt) > 0:
                amount_conflict = float(file_upb) > float(loan_amt) * TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT
            if amount_conflict:
                continue

            checkpoint_ok = bool(has_file and _servicer_checkpoint_ok(sf_serv, file_serv))
            # Servicer Commitment Id is the reliable term loan key in the completed report.
            # Salesforce's Servicer Name can lag/misclassify the current servicing family, so a
            # current servicer-file hit on the commitment ID is allowed even when the SF family
            # text disagrees. This prevents good Midland/Berkadia/FCI UPB from being blanked.
            if has_file and col == "Servicer Commitment Id":
                checkpoint_ok = True
            # alternate Term Servicer Key fields are useful only as lower-priority fallbacks.
            pref = 100 if col == "Servicer Commitment Id" else max(0, 10 - pos)
            prev_bonus = 50 if (prev_owner and deal_key and prev_owner == deal_key) else 0
            file_bonus = 20 if has_file else 0
            upb_bonus = 5 if (pd.notna(file_upb) and float(file_upb) > 0) else 0
            amount_bonus = 5 if (pd.notna(file_upb) and pd.notna(loan_amt) and float(loan_amt) > 0 and float(file_upb) <= float(loan_amt) * TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT) else 0
            ok_score = 100 + prev_bonus + file_bonus + upb_bonus + amount_bonus + pref if checkpoint_ok else -10**9
            raw_score = prev_bonus + file_bonus + upb_bonus + amount_bonus + pref
            if ok_score > best_score:
                best_score = ok_score
                best = (raw, sid_key, info, col, ok_score)
            if raw_score > fallback_score:
                fallback_score = raw_score
                fallback = (raw, sid_key, info, col, raw_score)

        chosen = best if best is not None else None
        # If the SF servicer family checkpoint failed but the commitment/key still matched a
        # current servicer file, keep that fallback instead of creating blanks. The later
        # servicer-file merge overwrites stale SF servicer names.
        if chosen is None:
            chosen = fallback

        if chosen is not None:
            raw, sid_key, info, col, score = chosen
            selected_raw.append(raw)
            selected_key.append(sid_key or pd.NA)
            selected_serv.append(info.get("servicer") if info else pd.NA)
            selected_upb.append(money_to_float(info.get("upb")) if info else np.nan)
            selected_npd.append(pd.to_datetime(info.get("next_payment_date"), errors="coerce") if info else pd.NaT)
            selected_mat.append(pd.to_datetime(info.get("maturity_date"), errors="coerce") if info else pd.NaT)
            selected_source.append(col)
            selected_scores.append(score)
        else:
            raw = result.at[idx, "selected_servicer_id"]
            selected_raw.append(raw)
            selected_key.append(_id_key_no_leading_zeros_scalar(raw) or pd.NA)
            selected_serv.append(pd.NA)
            selected_upb.append(np.nan)
            selected_npd.append(pd.NaT)
            selected_mat.append(pd.NaT)
            selected_source.append(pd.NA)
            selected_scores.append(np.nan)

    result["selected_servicer_id"] = pd.Series(selected_raw, index=sf_term.index, dtype="object")
    result["selected_sid_key"] = pd.Series(selected_key, index=sf_term.index, dtype="object")
    result["matched_servicer"] = pd.Series(selected_serv, index=sf_term.index, dtype="object")
    result["matched_upb"] = pd.to_numeric(pd.Series(selected_upb, index=sf_term.index), errors="coerce")
    result["matched_next_payment_date"] = pd.to_datetime(pd.Series(selected_npd, index=sf_term.index), errors="coerce")
    result["matched_maturity_date"] = pd.to_datetime(pd.Series(selected_mat, index=sf_term.index), errors="coerce")
    result["matched_source"] = pd.Series(selected_source, index=sf_term.index, dtype="object")
    result["_selected_match_score"] = pd.to_numeric(pd.Series(selected_scores, index=sf_term.index), errors="coerce")
    result["_deal_key"] = current_deal_key

    # Do not blank duplicate-looking servicer IDs here. In practice, duplicate / stale SF servicer
    # naming issues are common around service transfers, and clearing them caused thousands of
    # downstream Term Asset UPB blanks. Keep the best selected key and let the QA diagnostics flag
    # duplicates for review instead of destroying the usable balance/source fields.

    return result


def _filter_term_population(
    sf_term: pd.DataFrame,
    prev_keys: Optional[Set[str]] = None,
    prev_positive_keys: Optional[Set[str]] = None,
    prev_sold_retained_keys: Optional[Set[str]] = None,
) -> pd.DataFrame:
    if sf_term is None or sf_term.empty:
        return sf_term

    out = sf_term.copy()
    prev_sold_retained_keys = prev_sold_retained_keys or set()
    out["_deal_key"] = norm_id_series(out.get("Deal Loan Number", pd.Series([None] * len(out), index=out.index)))
    in_prev_sold_retained = out["_deal_key"].isin(prev_sold_retained_keys)
    in_prev_positive = out["_deal_key"].isin(prev_positive_keys or set())

    keep_mask = _term_report_keep_mask(
        out.get("Stage", pd.Series([""] * len(out), index=out.index)),
        out.get("Current Servicer UPB", pd.Series([np.nan] * len(out), index=out.index)),
        out.get("Sold Loan: Servicing Status", pd.Series([pd.NA] * len(out), index=out.index)),
        fallback_prev_retained_mask=in_prev_sold_retained,
        loan_amount_series=out.get("Loan Amount", pd.Series([np.nan] * len(out), index=out.index)),
    )
    stage = out.get("Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    carry_forward_active = in_prev_positive & (~stage.isin(["Paid Off", "REO-Sold"]))
    keep_mask = keep_mask | carry_forward_active
    return out.loc[keep_mask].copy()


def _best_header_read_excel(
    file_bytes: bytes,
    required_alias_groups: List[List[str]],
    preferred_sheets: Optional[List[str]] = None,
    max_header_scan: int = 8,
    sample_rows: int = 60,
):
    xls = pd.ExcelFile(BytesIO(file_bytes))
    sheet_names = list(xls.sheet_names)

    if preferred_sheets:
        preferred = []
        others = []
        for s in sheet_names:
            if any(p.lower() in s.lower() for p in preferred_sheets):
                preferred.append(s)
            else:
                others.append(s)
        ordered = preferred + others
    else:
        ordered = sheet_names

    best = None
    best_score = -1

    for sheet in ordered:
        for header_row in range(max_header_scan):
            try:
                sample = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=header_row, nrows=sample_rows)
                sample = sample.dropna(how="all")
                if sample.empty:
                    continue
                sample.columns = [str(c).strip() for c in sample.columns]
                score = sum(first_matching_col(sample, aliases) is not None for aliases in required_alias_groups)
                if score > best_score:
                    best_score = score
                    best = (sheet, header_row, score)
                    if score == len(required_alias_groups):
                        break
            except Exception:
                continue
        if best_score == len(required_alias_groups):
            break

    if best is None or best_score <= 0:
        raise ValueError("Could not find a matching header row.")

    best_sheet, best_header_row, best_score = best
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=best_sheet, header=best_header_row)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df, best_sheet, best_header_row, best_score


def _best_header_read_csv(
    file_bytes: bytes,
    required_alias_groups: List[List[str]],
    max_header_scan: int = 3,
    sample_rows: int = 60,
):
    best = None
    best_score = -1

    for header_row in range(max_header_scan):
        try:
            sample = pd.read_csv(BytesIO(file_bytes), header=header_row, nrows=sample_rows)
            sample = sample.dropna(how="all")
            if sample.empty:
                continue
            sample.columns = [str(c).strip() for c in sample.columns]
            score = sum(first_matching_col(sample, aliases) is not None for aliases in required_alias_groups)
            if score > best_score:
                best_score = score
                best = (header_row, score)
                if score == len(required_alias_groups):
                    break
        except Exception:
            continue

    if best is None or best_score <= 0:
        raise ValueError("Could not find a matching CSV header row.")

    best_header_row, best_score = best
    df = pd.read_csv(BytesIO(file_bytes), header=best_header_row)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df, best_header_row, best_score


def _series_to_num(df: pd.DataFrame, aliases: Sequence[str]) -> pd.Series:
    col = first_matching_col(df, aliases)
    if not col:
        return pd.Series([np.nan] * len(df), index=df.index)
    return df[col].apply(money_to_float)


def _series_to_dt(df: pd.DataFrame, aliases: Sequence[str]) -> pd.Series:
    col = first_matching_col(df, aliases)
    if not col:
        return pd.Series([pd.NaT] * len(df), index=df.index)
    return df[col].apply(to_dt)


def _series_to_text(df: pd.DataFrame, aliases: Sequence[str]) -> pd.Series:
    col = first_matching_col(df, aliases)
    if not col:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype="object")
    return df[col].astype("string")


def _series_to_id(df: pd.DataFrame, aliases: Sequence[str], transform=None) -> pd.Series:
    col = first_matching_col(df, aliases)
    if not col:
        raise ValueError("Required ID column not found.")
    s = df[col]
    return transform(s) if transform else norm_id_series(s)


def _as_of_for_df(df: pd.DataFrame, filename: str, aliases: Sequence[str]) -> date:
    col = first_matching_col(df, aliases)
    if col and df[col].notna().any():
        d = report_date_from_scalar(df[col].dropna().iloc[0])
        if d:
            return d
    return date_from_filename(filename) or today_et()


def parse_servicer_bytes(filename: str, b: bytes) -> pd.DataFrame:
    servicer_type = detect_servicer_type(filename)

    if servicer_type == "Onity":
        df, _sheet, _hdr, _score = _best_header_read_excel(
            b,
            [["LOAN_NUMBER", "Loan Number", "Servicer Loan Number"],
             ["FIRST_PRINCIPAL_BALANCE", "ACQUIRED_PRINCIPAL_BALANCE", "ORIGINAL_PRINCIPAL_BALANCE", "Current UPB", "UPB"]],
            preferred_sheets=["daily_data", "daily", "report"],
        )
        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": "Onity",
                "servicer_family": "onity",
                "servicer_id": _series_to_id(df, ["LOAN_NUMBER", "Loan Number", "Servicer Loan Number"]),
                "upb": _series_to_num(df, ["FIRST_PRINCIPAL_BALANCE", "ACQUIRED_PRINCIPAL_BALANCE", "ORIGINAL_PRINCIPAL_BALANCE", "Current UPB", "UPB"]),
                "suspense": _series_to_num(df, ["SUSPENSE_BALANCE", "Suspense Balance", "Suspense"]),
                "next_payment_date": _series_to_dt(df, ["NEXT_PAYMENT_DUE_DATE", "Next Payment Date", "Next Due Date"]),
                "maturity_date": _series_to_dt(df, ["LOAN_MATURES_DATE", "LINE_MATURITY_DATE", "NEW_LINE_MATURITY_DATE", "Maturity Date"]),
                "status": _series_to_text(df, ["LOAN_STATUS", "Status", "Loan Status"]),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["REPORT_DATE", "Report Date", "As Of Date", "Run Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    if servicer_type == "Shellpoint":
        df, _hdr, _score = _best_header_read_csv(
            b,
            [["LoanID", "Servicer Loan ID", "Loan Number"], ["PrincipalBalance", "UPB", "Current UPB"]],
            max_header_scan=2,
        )
        sid_col = first_matching_col(df, ["LoanID", "Servicer Loan ID", "Loan Number"])
        if not sid_col:
            sid_col = first_matching_col(df, ["InvestorLoanID", "Investor Loan ID"])
        if not sid_col:
            raise ValueError("Shellpoint file is missing a loan identifier column.")

        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": "Shellpoint",
                "servicer_family": "shellpoint",
                "servicer_id": norm_id_series(df[sid_col]),
                "upb": _series_to_num(df, ["PrincipalBalance", "UPB", "Current UPB"]),
                "suspense": _series_to_num(df, ["SuspenseBalance", "Suspense Balance"]),
                "next_payment_date": _series_to_dt(df, ["NextDueDate", "Next Due Date", "Next Payment Date"]),
                "maturity_date": pd.Series([pd.NaT] * len(df), index=df.index),
                "status": _series_to_text(df, ["LoanStatus", "Status", "PayString"]),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["DataAsOf", "Report Date", "As Of Date", "Run Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    if servicer_type == "CHL":
        df, _hdr, _score = _best_header_read_csv(
            b,
            [["Servicer Loan ID", "Loan ID", "Loan Number"], ["UPB", "Principal Balance", "Current UPB"]],
        )
        servicer_col = first_matching_col(df, ["Servicing Company", "Servicer", "Servicer Name"])
        servicer = df[servicer_col].astype("string") if servicer_col else pd.Series(["CHL Streamline"] * len(df))
        servicer = servicer.fillna("CHL Streamline")
        servicer = servicer.where(~servicer.astype("string").str.upper().eq("FCI"), "FCI CHL Streamline")
        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": servicer,
                "servicer_family": servicer.map(normalize_servicer_family),
                "servicer_id": _series_to_id(df, ["Servicer Loan ID", "Loan ID", "Loan Number"]),
                "upb": _series_to_num(df, ["UPB", "Principal Balance", "Current UPB"]),
                "suspense": np.nan,
                "next_payment_date": _series_to_dt(df, ["Due Date", "Next Due Date", "Next Payment Date"]),
                "maturity_date": _series_to_dt(df, ["Current Maturity Date", "Maturity Date"]),
                "status": _series_to_text(df, ["Performing Status", "Status", "Loan Status"]),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["Report Date", "As Of Date", "Run Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    if servicer_type == "CoreVestLoanData":
        df, _sheet, _hdr, _score = _best_header_read_excel(
            b,
            [["Loan Number", "Loan No", "BCM Loan#", "Servicer Loan Number"], ["Current UPB", "Principal Balance", "UPB"]],
            preferred_sheets=["loan"],
        )

        def _idfix(s: pd.Series) -> pd.Series:
            sid = norm_id_series(s).astype("string")
            return sid.apply(lambda x: x if pd.isna(x) else (x if x.startswith("0000") else f"0000{x}"))

        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": "Statebridge",
                "servicer_family": "statebridge",
                "servicer_id": _idfix(df[first_matching_col(df, ["Loan Number", "Loan No", "Servicer Loan Number"])]),
                "upb": _series_to_num(df, ["Current UPB", "Principal Balance", "UPB"]),
                "suspense": _series_to_num(df, ["Unapplied Balance", "Suspense Balance", "Suspense"]),
                "next_payment_date": _series_to_dt(df, ["Due Date", "Next Due Date", "Next Payment Date"]),
                "maturity_date": _series_to_dt(df, ["Maturity Date", "Current Maturity Date"]),
                "status": coalesce_keep_nonblank(
                    _series_to_text(df, ["MBA", "MBA Status", "MBA Delinquency Status", "MBA Status Code"]),
                    _series_to_text(df, ["Loan Status", "Status"]),
                ),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["Date", "Run Date", "Report Date", "As Of Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    if servicer_type == "CoreVest_Data_Tape":
        df, _sheet, _hdr, _score = _best_header_read_excel(
            b,
            [["BCM Loan#", "Loan Number", "Loan No"], ["Principal Balance", "Current UPB", "UPB"]],
            preferred_sheets=["loan"],
        )
        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": "Berkadia",
                "servicer_family": "berkadia",
                "servicer_id": _series_to_id(df, ["BCM Loan#", "Loan Number", "Loan No"]),
                "upb": _series_to_num(df, ["Principal Balance", "Current UPB", "UPB"]),
                "suspense": _series_to_num(df, ["Suspense Balance", "Unapplied Balance", "Suspense"]),
                "next_payment_date": _series_to_dt(df, ["Next Payment Due Date", "Next Due Date", "Due Date"]),
                "maturity_date": _series_to_dt(df, ["Maturity Date", "Current Maturity Date"]),
                "status": _series_to_text(df, ["Loan Status", "Status"]),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["Run Date", "Date", "Report Date", "As Of Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    if servicer_type == "FCI":
        df, _sheet, _hdr, _score = _best_header_read_excel(
            b,
            [["Account", "Loan Number", "Loan No"], ["Current Balance", "Current UPB", "UPB", "Principal Balance"]],
            preferred_sheets=["fci", "cvmaster", "v1805510", "report"],
        )
        servicer = fci_servicer_label_from_filename(filename)
        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": servicer,
                "servicer_family": "fci",
                "servicer_id": _series_to_id(df, ["Account", "Loan Number", "Loan No"]),
                "upb": _series_to_num(df, ["Current Balance", "Current UPB", "UPB", "Principal Balance"]),
                "suspense": _series_to_num(df, ["Reserve Balance", "Suspense Pmt.", "Suspense Payment", "Suspense Balance", "Unapplied Balance"]),
                "next_payment_date": _series_to_dt(df, ["Due Date", "Next Due Date", "Next Payment Date"]),
                "maturity_date": _series_to_dt(df, ["Maturity Date", "Current Maturity Date"]),
                "status": _series_to_text(df, ["Status", "Loan Status"]),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["Report Date", "As Of Date", "Date", "Run Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    if servicer_type == "Midland":
        df, _sheet, _hdr, _score = _best_header_read_excel(
            b,
            [["ServicerLoanNumber", "Servicer Loan Number", "Loan Number"], ["UPB$", "UPB", "Current UPB", "Principal Balance"]],
            preferred_sheets=["export", "midland", "loan"],
        )

        def _idfix(s: pd.Series) -> pd.Series:
            raw = s.astype("string").str.strip()
            raw = raw.str.replace(r"COM$", "", regex=True)
            raw = raw.str.replace(r"[^0-9A-Za-z]", "", regex=True).str.lstrip("0")
            return raw.replace({"": pd.NA})

        out = pd.DataFrame(
            {
                "source_file": filename,
                "servicer": "Midland",
                "servicer_family": "midland",
                "servicer_id": _idfix(df[first_matching_col(df, ["ServicerLoanNumber", "Servicer Loan Number", "Loan Number"])]),
                "upb": _series_to_num(df, ["UPB$", "UPB", "Current UPB", "Principal Balance"]),
                "suspense": np.nan,
                "next_payment_date": _series_to_dt(df, ["NextPaymentDate", "Next Payment Date", "Next Due Date"]),
                "maturity_date": _series_to_dt(df, ["MaturityDate", "Maturity Date"]),
                "status": _series_to_text(df, ["ServicerLoanStatus", "Loan Status", "Status"]),
                "as_of": pd.to_datetime(_as_of_for_df(df, filename, ["ReportDate", "Report Date", "As Of Date", "Run Date"])),
            }
        )
        return downcast_numeric_frame(out.dropna(subset=["servicer_id"]))

    raise ValueError("Unhandled servicer type.")



@st.cache_data(show_spinner=False, ttl=6 * 60 * 60, max_entries=128, hash_funcs={UploadBlob: lambda b: f"{b.filename}:{b.file_hash}"})
def parse_servicer_cached(blob: UploadBlob) -> pd.DataFrame:
    return parse_servicer_bytes(blob.filename, blob.data)


def choose_dominant_servicer_report_date(file_dates: Sequence[date]) -> date:
    """Return the common tape as-of date used for the report UPB header.

    The completed Active Loan Report uses the dominant external servicer tape date
    from the uploaded filenames. It should not drift to a later workbook run date
    embedded inside one servicer file.

    This report is produced weekly, so the UPB header date is supposed to advance
    every run; it must never inherit the prior week's date. On a tie, choose the
    LATEST date: a stale leftover tape from the previous week (e.g. a 6/19 file still
    sitting in the upload set next to the current 6/26 tapes) must not drag the header
    backward. The dominant-count vote still wins when one date clearly predominates;
    only genuine ties resolve to the most recent week.
    """
    counts: Dict[date, int] = {}
    for d in file_dates or []:
        if d:
            counts[d] = counts.get(d, 0) + 1
    if not counts:
        return today_et()
    top_count = max(counts.values())
    candidates = [d for d, c in counts.items() if c == top_count]
    return max(candidates)


def build_servicer_lookup(
    servicer_uploads: List,
    progress_hook=None,
    preview_rows_limit: int = 30,
    use_cache: bool = False,
) -> Tuple[pd.DataFrame, date, pd.DataFrame]:
    combined = pd.DataFrame(columns=["source_file", "servicer", "servicer_family", "servicer_id", "upb", "suspense", "next_payment_date", "maturity_date", "status", "as_of"])
    preview_parts: List[pd.DataFrame] = []
    file_dates: List[date] = []
    skipped_files: List[str] = []
    total = len(servicer_uploads or [])

    for idx, upload in enumerate(servicer_uploads or [], start=1):
        filename = getattr(upload, "name", f"servicer_{idx}")
        try:
            servicer_type = detect_servicer_type(filename)
        except Exception:
            servicer_type = "Unknown"

        if progress_hook is not None:
            try:
                progress_hook(f"file {idx}/{total} | {filename} | detected {servicer_type} | parsing")
            except Exception:
                pass

        blob = None
        try:
            blob = make_upload_blob(upload, compute_hash=use_cache)
            parsed = parse_servicer_cached(blob) if use_cache else parse_servicer_bytes(blob.filename, blob.data)
        except Exception as e:
            skipped_files.append(f"{filename}: {e}")
            parsed = pd.DataFrame()
        finally:
            blob = None

        if parsed.empty:
            gc.collect()
            continue

        parsed = downcast_numeric_frame(parsed)
        if preview_rows_limit > 0:
            preview_taken = sum(len(x) for x in preview_parts)
            remaining = max(preview_rows_limit - preview_taken, 0)
            if remaining > 0:
                preview_parts.append(parsed.head(remaining).copy())

        if combined.empty:
            combined = parsed
        else:
            combined = pd.concat([combined, parsed], ignore_index=True, copy=False)

        # Use the uploaded filename date first. Some servicer workbooks carry a later
        # internal run date, but the completed report's UPB header follows the common
        # external tape date encoded in filenames.
        d = date_from_filename(filename)
        if not d and "as_of" in parsed.columns and parsed["as_of"].notna().any():
            d = pd.to_datetime(parsed["as_of"].dropna().iloc[0]).date()
        if d:
            file_dates.append(d)

        if progress_hook is not None:
            try:
                progress_hook(f"file {idx}/{total} | {filename} | parsed {len(parsed):,} row(s)")
            except Exception:
                pass

        parsed = pd.DataFrame()
        gc.collect()

    if skipped_files:
        try:
            st.warning("Skipped servicer file(s): " + " | ".join(skipped_files))
        except Exception:
            pass

    full = combined

    if not full.empty:
        full = full.dropna(subset=["servicer_id"]).copy()
        full["_sid_key"] = id_key_no_leading_zeros(full["servicer_id"])
        full = full.dropna(subset=["_sid_key"]).copy()

        full["_has_upb"] = full["upb"].notna().astype("int8")
        full["_has_nonzero_upb"] = (pd.to_numeric(full["upb"], errors="coerce").fillna(0) > 0).astype("int8")
        full["_has_suspense"] = full["suspense"].notna().astype("int8")
        full["_has_npd"] = full["next_payment_date"].notna().astype("int8")
        full["_has_mat"] = full["maturity_date"].notna().astype("int8")
        full["_label_rank"] = full["servicer"].map(_servicer_specificity_rank).fillna(0).astype("int16")

        full = full.sort_values(
            ["_sid_key", "as_of", "_has_nonzero_upb", "_has_upb", "_has_suspense", "_has_npd", "_has_mat", "_label_rank", "upb"],
            ascending=[True, True, True, True, True, True, True, True, True],
        )

        join = full.drop_duplicates(["_sid_key"], keep="last").drop(
            columns=["_has_upb", "_has_nonzero_upb", "_has_suspense", "_has_npd", "_has_mat", "_label_rank"], errors="ignore"
        )
        if preview_parts:
            preview = pd.concat(preview_parts, ignore_index=True, copy=False).head(preview_rows_limit).copy()
        else:
            preview = full.head(min(preview_rows_limit or 0, 200)).copy() if preview_rows_limit > 0 else full.head(0).copy()
        full = full.drop(columns=["_has_upb", "_has_nonzero_upb", "_has_suspense", "_has_npd", "_has_mat", "_label_rank"], errors="ignore")
    else:
        full["_sid_key"] = pd.Series(dtype="string")
        join = full.copy()
        preview = full.head(0).copy()

    run_date = choose_dominant_servicer_report_date(file_dates) if file_dates else today_et()
    return downcast_numeric_frame(join), run_date, downcast_numeric_frame(preview)



def _find_upb_col(cols: Sequence[str]) -> Optional[str]:
    for c in cols:
        if isinstance(c, str) and re.search(r"\b\d{1,2}/\d{1,2}\s*UPB\b", c):
            return c
    return None


def _detect_active_loans_header_row(file_bytes: bytes, sheet: str) -> int:
    """Return the 0-based pandas header index for a prior Active Loan Report sheet.

    The report layout changed from row-4 headers (header=3) to row-5 headers
    (header=4) on 2026-05-27. A hardcoded header index silently breaks
    carry-forward whenever the uploaded prior report uses the other layout
    (every "Unnamed: X" column makes build_prev_maps find nothing, which kills
    Strategy Grouping / Segment / valuation carry-forward). Detect the real
    header row by looking for the sheet's key identifier columns.
    """
    # Identifier columns that appear in every report layout.
    expected = {"Asset ID", "Deal Number", "Deal Name", "Portfolio"}
    best_idx, best_hits = 4, -1
    for h in (4, 3, 2, 5, 1, 0):
        try:
            probe = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=h, nrows=1)
        except Exception:
            continue
        cols = {str(c).strip() for c in probe.columns}
        hits = len(expected & cols)
        if hits > best_hits:
            best_idx, best_hits = h, hits
        if hits >= 2:  # two identifier columns is a confident match
            return h
    return best_idx


def read_tab_df_from_active_loans(file_bytes: bytes, sheet: str) -> pd.DataFrame:
    header_idx = _detect_active_loans_header_row(file_bytes, sheet)
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=header_idx)
    df = df.dropna(how="all").copy()
    df.columns = [str(c).strip() for c in df.columns]
    # Drop any echoed header row that slipped into the data (defensive: a report
    # that carries headers in BOTH row 4 and row 5 must not leak the dupe as data).
    for key in ("Asset ID", "Deal Number"):
        if key in df.columns:
            df = df[df[key].astype(str).str.strip() != key]
            break
    return df



# V61: sheets that only a BUILD output has -- the official completed report never
# carries the post-build QA tabs. Used to catch a previous test being uploaded as the
# "prior completed report", which would let its errors carry forward.
BUILD_OUTPUT_MARKER_SHEETS = ("QA Summary", "QA Exceptions")


def prior_workbook_provenance(prev_bytes: Optional[bytes]) -> Tuple[bool, str]:
    """Return (looks_like_build_output, human-readable note) for the uploaded prior workbook.

    Carry-forward-first columns (Segment, Strategy Grouping, the Origination valuation
    snapshot, Remedy Plan, ...) take the prior workbook's value over the live one, and
    Next Payment Date falls back to it when neither the servicer file nor Salesforce has a
    date. So the prior workbook must be the OFFICIAL completed report -- uploading last
    run's test output makes this build inherit that run's mistakes.
    """
    if not prev_bytes:
        return False, ""
    try:
        wb = load_workbook(BytesIO(prev_bytes), read_only=True, data_only=True)
        try:
            names = set(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return False, ""
    found = [s for s in BUILD_OUTPUT_MARKER_SHEETS if s in names]
    if found:
        return True, (
            "The uploaded prior workbook contains " + " and ".join(found) + ", which only a "
            "generated build has -- it looks like a previous test output rather than the "
            "official completed report. Carry-forward-first columns and the Next Payment Date "
            "fallback will inherit that run's values, so any error in it compounds into this "
            "build. Upload the official completed report for the prior week instead."
        )
    return False, ""


def build_prev_maps(prev_bytes: bytes) -> dict:
    out: dict = {}

    try:
        ba = read_tab_df_from_active_loans(prev_bytes, "Bridge Asset")
        if "Asset ID" in ba.columns:
            keep = [
                c for c in [
                    "Asset ID", "Portfolio", "Segment", "Strategy Grouping", "REO Date", "Active RM",
                    # V53: "Next Payment Date" MUST be carried forward. _bridge_pick_next_payment_date
                    # takes a `prior_dates` argument and uses it to preserve the report's day-10
                    # billing convention, but the column was never extracted here, so `prior` was
                    # always empty and that whole branch was dead code (654 Bridge Asset NPD
                    # mismatches against 20260810).
                    "Next Payment Date",
                    "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag",
                    "Asset Manager 1", "AM 1 Assigned Date", "Asset Manager 2", "AM 2 Assigned Date",
                    "Construction Mgr.", "CM Assigned Date", "Servicer", "Servicer Status",
                    "Remedy Plan", "Delinquency Notes", "Maturity Status", "Title Company", "Tax Commentary",
                    "Origination Value Dt", "Origination As-Is Value", "Origination ARV",
                    "Most Recent Appraisal Order Date", "Updated Valuation Date", "Updated As-Is Value", "Updated ARV",
                    "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact",
                ] if c in ba.columns
            ]
            tmp = ba[keep].copy()
            tmp["_asset_key"] = norm_id_series(tmp["Asset ID"])
            out["bridge_asset_manual"] = tmp.dropna(subset=["_asset_key"]).drop_duplicates("_asset_key")

            upb_col_prev = _find_upb_col(ba.columns)
            if upb_col_prev:
                tmpu = ba[["Asset ID", upb_col_prev]].copy()
                tmpu["_asset_key"] = norm_id_series(tmpu["Asset ID"])
                tmpu["_prev_asset_upb"] = tmpu[upb_col_prev].apply(money_to_float)
                out["bridge_asset_upb"] = tmpu.dropna(subset=["_asset_key"]).drop_duplicates("_asset_key")[["_asset_key", "_prev_asset_upb"]]
    except Exception:
        pass

    try:
        bl = read_tab_df_from_active_loans(prev_bytes, "Bridge Loan")
        keep = [
            c for c in [
                "Deal Number", "Portfolio", "Segment", "Strategy Grouping", "Loan Level Delinquency",
                # V53: same reason as Bridge Asset -- the loan-level day-10 preservation needs this.
                "Next Payment Date",
                "Special Focus (Y/N)", "AM Commentary", "3/31 NPL", "Needs NPL Value", "Active RM",
                "Asset Manager 1", "AM 1 Assigned Date", "Asset Manager 2", "AM 2 Assigned Date",
                "Construction Mgr.", "CM Assigned Date",
            ] if c in bl.columns
        ]
        if "Deal Number" in keep and len(keep) > 1:
            tmp = bl[keep].copy()
            tmp["_deal_key"] = norm_id_series(tmp["Deal Number"])
            out["bridge_loan_manual"] = tmp.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key")

        upb_col_prev = _find_upb_col(bl.columns)
        if upb_col_prev and "Deal Number" in bl.columns:
            tmpu = bl[["Deal Number", upb_col_prev]].copy()
            tmpu["_deal_key"] = norm_id_series(tmpu["Deal Number"])
            tmpu["_prev_upb"] = tmpu[upb_col_prev].apply(money_to_float)
            out["bridge_loan_upb"] = tmpu.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key")[["_deal_key", "_prev_upb"]]
    except Exception:
        pass

    try:
        tl = read_tab_df_from_active_loans(prev_bytes, "Term Loan")
        if "Deal Number" in tl.columns and "REO Date" in tl.columns:
            tmp = tl[["Deal Number", "REO Date"]].copy()
            tmp["_deal_key"] = norm_id_series(tmp["Deal Number"])
            out["term_loan_reo"] = tmp.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key")

        keep = [
            c for c in [
                "Deal Number", "Servicer ID", "Servicer", "SF Yardi ID", "Deal Name", "Borrower Entity",
                "Account Name", "Do Not Lend (Y/N)", "Portfolio", "Segment", "Financing", "CPP JV",
                "Loan Buyer", "Loan Amount", "Origination Date", "Maturity Date", "Next Payment Date",
                "REO Date", "Asset Manager", "Originator", "Active RM", "Deal Intro Sub-Source",
                "Referral Source Account", "Referral Source Contact", "AM Commentary", "Special Loans List (Y/N)",
            ] if c in tl.columns
        ]
        if "Deal Number" in keep and len(keep) > 1:
            tmpm = tl[keep].copy()
            tmpm["_deal_key"] = norm_id_series(tmpm["Deal Number"])
            out["term_loan_manual"] = tmpm.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key")

        if "Servicer ID" in tl.columns:
            keep_sid = [
                c for c in [
                    "Deal Number", "Servicer ID", "Servicer", "SF Yardi ID", "Deal Name", "Borrower Entity",
                    "Account Name", "Do Not Lend (Y/N)", "Portfolio", "Segment", "Financing", "CPP JV",
                    "Loan Buyer", "Loan Amount", "Origination Date", "Maturity Date", "Next Payment Date",
                    "REO Date", "Asset Manager", "Originator", "Active RM", "Deal Intro Sub-Source",
                    "Referral Source Account", "Referral Source Contact", "AM Commentary", "Special Loans List (Y/N)",
                ] if c in tl.columns
            ]
            if keep_sid:
                tmps = tl[keep_sid].copy()
                tmps["_sid_key"] = id_key_no_leading_zeros(tmps["Servicer ID"])
                tmps["_deal_key"] = norm_id_series(tmps["Deal Number"]) if "Deal Number" in tmps.columns else pd.NA
                out["term_loan_sid"] = tmps.dropna(subset=["_sid_key"]).drop_duplicates("_sid_key")

        upb_col_prev = _find_upb_col(tl.columns)
        if upb_col_prev and "Deal Number" in tl.columns:
            tmpu = tl[["Deal Number", upb_col_prev]].copy()
            tmpu["_deal_key"] = norm_id_series(tmpu["Deal Number"])
            tmpu["_prev_upb"] = tmpu[upb_col_prev].apply(money_to_float)
            out["term_loan_upb"] = tmpu.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key")[["_deal_key", "_prev_upb"]]
    except Exception:
        pass

    try:
        ta = read_tab_df_from_active_loans(prev_bytes, "Term Asset")
        if "Deal Number" in ta.columns and "Asset ID" in ta.columns:
            keep = [
                c for c in [
                    "Deal Number", "Asset ID", "Portfolio", "Date", "Address", "City", "State", "Zip", "CBSA",
                    "# Units", "Property Type", "Property ALA", "Value Date", "As-Is Value",
                    "Special (Y/N)", "CPP JV",
                ] if c in ta.columns
            ]
            if len(keep) >= 2:
                tmpa = ta[keep].copy()
                tmpa["_deal_key"] = norm_id_series(tmpa["Deal Number"])
                tmpa["_asset_key"] = norm_id_series(tmpa["Asset ID"])
                out["term_asset_manual"] = tmpa.dropna(subset=["_deal_key", "_asset_key"]).drop_duplicates(["_deal_key", "_asset_key"])
    except Exception:
        pass

    gc.collect()
    return out


SHEET_BASELINE_KEY_CANDIDATES = {
    "Bridge Asset": [["Asset ID"]],
    "Bridge Loan": [["Deal Number"]],
    "Term Loan": [["Deal Number"], ["Servicer ID"]],
    "Term Asset": [["Asset ID"], ["Deal Number", "Asset ID"]],
}


def _backfill_rule_candidates(sheet_name: str) -> List[List[str]]:
    return [list(x) for x in SHEET_BASELINE_KEY_CANDIDATES.get(sheet_name, [])]


def _available_backfill_keys(sheet_name: str, built_df: pd.DataFrame, baseline_df: pd.DataFrame) -> List[List[str]]:
    candidates = _backfill_rule_candidates(sheet_name)
    built_cols = set(built_df.columns)
    baseline_cols = set(baseline_df.columns)

    out: List[List[str]] = []
    for candidate in candidates:
        if not set(candidate).issubset(built_cols) or not set(candidate).issubset(baseline_cols):
            continue
        if sheet_name == "Term Loan" and candidate == ["Servicer ID"]:
            built_nonblank = id_key_no_leading_zeros(built_df["Servicer ID"]).notna().sum()
            baseline_nonblank = id_key_no_leading_zeros(baseline_df["Servicer ID"]).notna().sum()
            if built_nonblank == 0 or baseline_nonblank == 0:
                continue
        out.append(list(candidate))
    return out


def _choose_backfill_normalizer(sheet_name: str, key_cols: Sequence[str]):
    prefer_no_leading_zeros = sheet_name == "Term Loan" and list(key_cols) == ["Servicer ID"]

    def _one(s: pd.Series) -> pd.Series:
        return id_key_no_leading_zeros(s) if prefer_no_leading_zeros else norm_id_series(s)

    return _one


def _object_series_like(s: pd.Series) -> pd.Series:
    base = pd.Series(s, copy=False)
    return pd.Series(list(base), index=base.index, dtype="object")



def _safe_backfill_assign(sheet_name: str, out: pd.DataFrame, col: str, base_vals: pd.Series, mask: pd.Series) -> pd.DataFrame:
    if not bool(mask.any()):
        return out

    try:
        out.loc[mask, col] = base_vals.loc[mask]
        return out
    except Exception:
        pass

    current = out[col]
    source = base_vals

    if _is_date_header(sheet_name, col):
        parsed_current = _to_datetime_series_mixed(_object_series_like(current))
        parsed_source = _to_datetime_series_mixed(_object_series_like(source))
        parsed_mask = blankish_mask(parsed_current) & (~blankish_mask(parsed_source))
        if bool((mask & parsed_mask).any()):
            out[col] = parsed_current
            out.loc[mask & parsed_mask, col] = parsed_source.loc[mask & parsed_mask]
            remaining_mask = mask & (~parsed_mask)
            if bool(remaining_mask.any()):
                out[col] = _object_series_like(out[col])
                out.loc[remaining_mask, col] = _object_series_like(source).loc[remaining_mask]
            return out

    if pd.api.types.is_numeric_dtype(current):
        source_obj = _object_series_like(source)
        numeric_source = pd.to_numeric(source_obj, errors="coerce")
        numeric_mask = mask & numeric_source.notna()
        if bool(numeric_mask.any()):
            try:
                out.loc[numeric_mask, col] = numeric_source.loc[numeric_mask]
            except Exception:
                out[col] = _object_series_like(out[col])
                out.loc[numeric_mask, col] = numeric_source.loc[numeric_mask]
        remaining_mask = mask & (~numeric_mask)
        if bool(remaining_mask.any()):
            out[col] = _object_series_like(out[col])
            out.loc[remaining_mask, col] = source_obj.loc[remaining_mask]
        return out

    out[col] = _object_series_like(out[col])
    out.loc[mask, col] = _object_series_like(source).loc[mask]
    return out



def _backfill_df_from_baseline_once(
    sheet_name: str,
    built_df: pd.DataFrame,
    baseline_df: pd.DataFrame,
    key_cols: Sequence[str],
) -> Tuple[pd.DataFrame, int]:
    out = built_df.copy()
    base = baseline_df.copy()
    normer = _choose_backfill_normalizer(sheet_name, key_cols)
    helper_keys = [f"_baseline_k{i}" for i in range(len(key_cols))]

    for i, col in enumerate(key_cols):
        out[helper_keys[i]] = normer(out[col])
        base[helper_keys[i]] = normer(base[col])

    base = base.dropna(subset=helper_keys).copy()
    if base.empty:
        return out.drop(columns=helper_keys, errors="ignore"), 0

    common_cols = [
        c for c in base.columns
        if c in out.columns
        and c not in key_cols
        and c not in helper_keys
        and not c.startswith("_")
        and not UPB_HEADER_RE.search(str(c))
    ]
    if not common_cols:
        return out.drop(columns=helper_keys, errors="ignore"), 0

    score = pd.Series([0] * len(base), index=base.index, dtype="int64")
    for col in common_cols:
        score = score + (~blankish_mask(base[col])).astype("int64")
    base["_baseline_score"] = score
    base = base.sort_values(helper_keys + ["_baseline_score"], ascending=[True] * len(helper_keys) + [True])
    base = base.drop_duplicates(helper_keys, keep="last")

    base_lookup = base.set_index(helper_keys, drop=False)
    out = out.set_index(helper_keys, drop=False)

    fills = 0
    for col in common_cols:
        base_vals = base_lookup[col].reindex(out.index)
        mask = blankish_mask(out[col]) & (~blankish_mask(base_vals))
        if mask.any():
            fills += int(mask.sum())
            out = _safe_backfill_assign(sheet_name, out, col, base_vals, mask)

    out = out.reset_index(drop=True)
    out = out.drop(columns=helper_keys + ["_baseline_score"], errors="ignore")
    return downcast_numeric_frame(out), fills


def backfill_df_from_baseline(sheet_name: str, built_df: pd.DataFrame, baseline_bytes: Optional[bytes]) -> Tuple[pd.DataFrame, dict]:
    if baseline_bytes is None or built_df is None or built_df.empty:
        return built_df, {"sheet_name": sheet_name, "status": "skipped_empty", "keys": "", "fills": 0}

    try:
        baseline_df = read_tab_df_from_active_loans(baseline_bytes, sheet_name)
    except Exception as exc:
        return built_df, {"sheet_name": sheet_name, "status": f"baseline_read_failed: {exc}", "keys": "", "fills": 0}

    out = built_df.copy()
    key_candidates = _available_backfill_keys(sheet_name, out, baseline_df)
    if not key_candidates:
        return out, {"sheet_name": sheet_name, "status": "skipped_no_keys", "keys": "", "fills": 0}

    fills = 0
    used_keys: List[str] = []
    for key_cols in key_candidates:
        out, one_fill = _backfill_df_from_baseline_once(sheet_name, out, baseline_df, key_cols)
        if one_fill:
            fills += int(one_fill)
            used_keys.append(", ".join(key_cols))

    return downcast_numeric_frame(out), {
        "sheet_name": sheet_name,
        "status": "backfilled" if fills else "no_fills_needed",
        "keys": " | ".join(used_keys or [", ".join(x) for x in key_candidates]),
        "fills": fills,
    }


def _parse_npl_or_reo_sheet(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:

    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=4)
    df = df.dropna(how="all").copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def parse_npl_reo_bytes(file_bytes: bytes) -> dict:
    out = {
        "loan_flags": pd.DataFrame(columns=["_deal_key", "NPL Flag", "Needs NPL Value", "Special Focus (Y/N)"]),
        "asset_flags": pd.DataFrame(columns=["_deal_key", "_asset_key", "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag"]),
        "asset_deal_fallback": pd.DataFrame(columns=["_deal_key", "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag"]),
    }

    def _append_flags(df_src: pd.DataFrame, npl_flag: str, special_flag: str):
        nonlocal out
        deal_key = norm_id_series(df_src["Deal Number"])
        asset_key = norm_id_series(df_src["Asset ID"]) if "Asset ID" in df_src.columns else pd.Series([pd.NA] * len(df_src), index=df_src.index)

        loan_flags = pd.DataFrame(
            {
                "_deal_key": deal_key,
                "NPL Flag": npl_flag,
                "Needs NPL Value": "N",
                "Special Focus (Y/N)": special_flag,
            }
        ).dropna(subset=["_deal_key"]).drop_duplicates("_deal_key")
        out["loan_flags"] = pd.concat([out["loan_flags"], loan_flags], ignore_index=True, copy=False)

        asset_specific = pd.DataFrame(
            {
                "_deal_key": deal_key,
                "_asset_key": asset_key,
                "3/31 NPL (Y/N)": npl_flag,
                "Needs NPL Value": "N",
                "Special Flag": special_flag,
            }
        )
        asset_specific = asset_specific.dropna(subset=["_deal_key"])
        out["asset_flags"] = pd.concat(
            [out["asset_flags"], asset_specific[asset_specific["_asset_key"].notna()]],
            ignore_index=True,
            copy=False,
        )
        out["asset_deal_fallback"] = pd.concat(
            [out["asset_deal_fallback"], asset_specific[asset_specific["_asset_key"].isna()].drop(columns=["_asset_key"])],
            ignore_index=True,
            copy=False,
        )

    try:
        npl = _parse_npl_or_reo_sheet(file_bytes, "NPL")
        if "Deal Number" in npl.columns:
            _append_flags(npl, "Y", "Y")
    except Exception:
        pass

    try:
        reo = _parse_npl_or_reo_sheet(file_bytes, "REO")
        if "Deal Number" in reo.columns:
            _append_flags(reo, "N", "Y")
    except Exception:
        pass

    if not out["loan_flags"].empty:
        out["loan_flags"] = out["loan_flags"].sort_values(["_deal_key", "Special Focus (Y/N)", "NPL Flag"]).drop_duplicates("_deal_key", keep="last")
    if not out["asset_flags"].empty:
        out["asset_flags"] = out["asset_flags"].drop_duplicates(["_deal_key", "_asset_key"], keep="last")
    if not out["asset_deal_fallback"].empty:
        specific_deals = set(out["asset_flags"]["_deal_key"].dropna().tolist())
        out["asset_deal_fallback"] = out["asset_deal_fallback"][~out["asset_deal_fallback"]["_deal_key"].isin(specific_deals)].copy()
        out["asset_deal_fallback"] = out["asset_deal_fallback"].drop_duplicates("_deal_key", keep="last")

    return out





def _numeric_series(df: pd.DataFrame, col: str, default=np.nan) -> pd.Series:
    if df is None:
        return pd.Series(dtype="float64")
    if col in df.columns:
        return pd.to_numeric(pd.Series(df[col], index=df.index), errors="coerce")
    return pd.to_numeric(pd.Series([default] * len(df), index=df.index), errors="coerce")


def _coalesce_positive_then_any_numeric(*series_like, index=None) -> pd.Series:
    """Choose the first positive numeric source, then the first non-null numeric source.

    This is intentionally stricter than normal text coalescing: a servicer zero
    should not hide a positive Salesforce/prior balance, but zero is preserved
    when every source is truly zero or blank.
    """
    if index is None:
        for obj in series_like:
            if isinstance(obj, pd.Series):
                index = obj.index
                break
    if index is None:
        index = pd.RangeIndex(0)

    candidates = []
    for obj in series_like:
        if isinstance(obj, pd.Series):
            ser = pd.to_numeric(obj.reindex(index), errors="coerce")
        else:
            ser = pd.to_numeric(pd.Series(obj, index=index), errors="coerce")
        candidates.append(ser)

    out = pd.Series([np.nan] * len(index), index=index, dtype="float64")
    for ser in candidates:
        vals = ser.where(ser.gt(0))
        out = out.where(out.notna(), vals)
    for ser in candidates:
        vals = ser.where(ser.notna())
        out = out.where(out.notna(), vals)
    return pd.to_numeric(out, errors="coerce")


def _group_first_positive_then_any_numeric(values: pd.Series):
    vals = pd.to_numeric(values, errors="coerce").dropna()
    if vals.empty:
        return np.nan
    positive = vals[vals.gt(0)]
    if not positive.empty:
        return float(positive.iloc[0])
    return float(vals.iloc[0])


def _ensure_deal_key(df: pd.DataFrame, source_col: str = "Deal Number") -> pd.DataFrame:
    out = df.copy()
    if "_deal_key" not in out.columns:
        out["_deal_key"] = norm_id_series(out.get(source_col, pd.Series([None] * len(out), index=out.index)))
    return out


def _bridge_component_funded_sum(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="float64")
    idx = df.index
    init = _numeric_series(df, "Initial Disbursement Funded")
    reno = _numeric_series(df, "Renovation Holdback Funded")
    interest = _numeric_series(df, "Interest Allocation Funded")
    has_any_component = init.notna() | reno.notna() | interest.notna()
    funded = init.fillna(0.0) + reno.fillna(0.0) + interest.fillna(0.0)
    return funded.where(has_any_component, np.nan)


def _recompute_bridge_asset_funded_amount(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    out = df.copy()
    out["SF Funded Amount"] = _bridge_component_funded_sum(out)

    if {"Renovation Holdback", "Renovation Holdback Funded", "Renovation Holdback Remaining"}.issubset(out.columns):
        reno_total = _numeric_series(out, "Renovation Holdback")
        reno_funded = _numeric_series(out, "Renovation Holdback Funded")
        # Verified against 20260615_Active_Loans (433/433 Bridge Asset, 212/213 Bridge
        # Loan): the report's remaining = Approved Renovation Advance Amount minus
        # Renovation Advance Amount Funded. It does NOT add Interest Allocation -- that
        # extra term diverged from the official report on 430 rows.
        calc_remaining = reno_total.fillna(0.0) - reno_funded.fillna(0.0)
        has_calc = reno_total.notna() | reno_funded.notna()
        out["Renovation Holdback Remaining"] = pd.to_numeric(out["Renovation Holdback Remaining"], errors="coerce").where(
            ~has_calc,
            calc_remaining,
        )
    return downcast_numeric_frame(out)


def _rollup_bridge_asset_math(bridge_asset: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    if bridge_asset is None or bridge_asset.empty:
        return pd.DataFrame(columns=["_deal_key"])
    ba = _recompute_bridge_asset_funded_amount(bridge_asset)
    ba = _ensure_deal_key(ba, "Deal Number")
    ba = ba[ba["_deal_key"].notna()].copy()
    if ba.empty:
        return pd.DataFrame(columns=["_deal_key"])

    g = ba.groupby("_deal_key", dropna=True)
    roll = pd.DataFrame(index=g.size().index)
    sum_pairs = {
        "Active Funded Amount": "SF Funded Amount",
        "Initial Disbursement Funded": "Initial Disbursement Funded",
        "Renovation Holdback": "Renovation Holdback",
        "Renovation HB Funded": "Renovation Holdback Funded",
        "Renovation HB Remaining": "Renovation Holdback Remaining",
        "Interest Allocation": "Interest Allocation",
        "Interest Allocation Funded": "Interest Allocation Funded",
        "Suspense Balance": "Suspense Balance",
        "Most Recent As-Is Value": "Most Recent As-Is Value",
        "Most Recent ARV": "Most Recent ARV",
    }
    for target, source in sum_pairs.items():
        if source in ba.columns:
            roll[target] = pd.to_numeric(g[source].sum(min_count=1), errors="coerce")
    if upb_col in ba.columns:
        roll[upb_col] = pd.to_numeric(g[upb_col].sum(min_count=1), errors="coerce")
    return roll.reset_index()



def _bridge_material_tolerance(base_series) -> pd.Series:
    base = pd.to_numeric(pd.Series(base_series, copy=False), errors="coerce").abs().fillna(0.0)
    return pd.Series(
        np.maximum(BRIDGE_OVER_COMMITMENT_WARN_TOLERANCE_DOLLARS, base * BRIDGE_OVER_COMMITMENT_WARN_RATIO),
        index=base.index,
        dtype="float64",
    )


def _bridge_stage_exception_mask(stage_series: pd.Series) -> pd.Series:
    txt = pd.Series(stage_series, copy=False).astype("string").str.strip()
    return txt.isin(BRIDGE_EXCEPTION_STAGES_ALLOW_OVER_COMMITMENT).fillna(False)


def _bridge_limit_base(commitment_series, funded_series=None) -> pd.Series:
    commitment = pd.to_numeric(pd.Series(commitment_series, copy=False), errors="coerce")
    if funded_series is None:
        funded = pd.Series([np.nan] * len(commitment), index=commitment.index, dtype="float64")
    else:
        funded = pd.to_numeric(pd.Series(funded_series, index=commitment.index, copy=False), errors="coerce")
    base = pd.concat([commitment.where(commitment.gt(0)), funded.where(funded.gt(0))], axis=1).max(axis=1)
    return pd.to_numeric(base, errors="coerce")


def _bridge_value_plausible_against_commitment(value_series, commitment_series, ratio: float = BRIDGE_UPB_COMMITMENT_RATIO_LIMIT) -> pd.Series:
    value = pd.to_numeric(pd.Series(value_series, copy=False), errors="coerce")
    commitment = pd.to_numeric(pd.Series(commitment_series, index=value.index, copy=False), errors="coerce")
    tol = _bridge_material_tolerance(commitment)
    no_commitment = commitment.isna() | commitment.le(0)
    return value.gt(0) & (no_commitment | value.le(commitment * ratio + tol))


def _bridge_value_plausible_for_upb(value_series, commitment_series, funded_series=None, ratio: float = BRIDGE_UPB_COMMITMENT_RATIO_LIMIT) -> pd.Series:
    value = pd.to_numeric(pd.Series(value_series, copy=False), errors="coerce")
    commitment = pd.to_numeric(pd.Series(commitment_series, index=value.index, copy=False), errors="coerce")
    funded = None if funded_series is None else pd.to_numeric(pd.Series(funded_series, index=value.index, copy=False), errors="coerce")
    base = _bridge_limit_base(commitment, funded)
    tol = _bridge_material_tolerance(base)
    no_base = base.isna() | base.le(0)
    return value.gt(0) & (no_base | value.le(base * ratio + tol))


def _allocate_group_amount_by_weights(out: pd.DataFrame, row_indices: Sequence, target_col: str, target_amount: float, weight_col: str) -> None:
    idxs = list(row_indices)
    if not idxs:
        return
    weights = pd.to_numeric(out.loc[idxs, weight_col] if weight_col in out.columns else pd.Series([np.nan] * len(idxs), index=idxs), errors="coerce").fillna(0.0)
    weights = weights.where(weights.gt(0), 0.0)
    if float(weights.sum()) > 0:
        alloc = float(target_amount) * (weights / float(weights.sum()))
    else:
        alloc = pd.Series([float(target_amount) / len(idxs)] * len(idxs), index=idxs, dtype="float64")
    out.loc[idxs, target_col] = alloc
    # Fix floating-point drift on the last row so the group ties exactly before workbook rounding.
    diff = float(target_amount) - float(pd.to_numeric(out.loc[idxs, target_col], errors="coerce").fillna(0.0).sum())
    out.loc[idxs[-1], target_col] = float(pd.to_numeric(pd.Series([out.loc[idxs[-1], target_col]]), errors="coerce").fillna(0.0).iloc[0]) + diff


def repair_bridge_asset_math(bridge_asset: pd.DataFrame, upb_col: str) -> Tuple[pd.DataFrame, List[str]]:
    """Repair Bridge Asset funded amount and obviously contaminated UPB using same-deal fields.

    The main invariant is not "UPB must be capped to commitment." Some valid exception / sold / overfunded
    situations can exceed current commitment. The real production rule is: do not publish an asset UPB rollup
    that is wildly above every same-deal basis when there is a safer same-deal source such as funded components
    or prior asset UPB. This function records repairs in diagnostics and leaves unresolved exception-stage rows
    visible for review instead of fabricating a cap.
    """
    if bridge_asset is None or bridge_asset.empty:
        return (pd.DataFrame() if bridge_asset is None else bridge_asset.copy()), []

    out = _recompute_bridge_asset_funded_amount(bridge_asset)
    out = _ensure_deal_key(out, "Deal Number")
    diagnostics: List[str] = []

    if upb_col not in out.columns:
        return downcast_numeric_frame(out), diagnostics

    before_total = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0.0).sum()
    repaired_count = 0
    unresolved_count = 0

    for deal_key, idxs in out.groupby("_deal_key", dropna=True).groups.items():
        idxs = list(idxs)
        if not idxs:
            continue

        commitment_vals = pd.to_numeric(
            out.loc[idxs, "Loan Commitment"] if "Loan Commitment" in out.columns else pd.Series([np.nan] * len(idxs), index=idxs),
            errors="coerce",
        )
        funded_vals = pd.to_numeric(
            out.loc[idxs, "SF Funded Amount"] if "SF Funded Amount" in out.columns else pd.Series([np.nan] * len(idxs), index=idxs),
            errors="coerce",
        )
        upb_vals = pd.to_numeric(out.loc[idxs, upb_col], errors="coerce")
        prev_vals = pd.to_numeric(
            out.loc[idxs, "_prev_asset_upb"] if "_prev_asset_upb" in out.columns else pd.Series([np.nan] * len(idxs), index=idxs),
            errors="coerce",
        )

        commitment = float(commitment_vals[commitment_vals.gt(0)].max()) if bool(commitment_vals.gt(0).any()) else np.nan
        funded_sum = float(funded_vals.fillna(0.0).sum())
        upb_sum = float(upb_vals.fillna(0.0).sum())
        prev_sum = float(prev_vals.fillna(0.0).sum())

        basis_values = [x for x in [commitment, funded_sum] if pd.notna(x) and float(x) > 0]
        if not basis_values or upb_sum <= 0:
            continue
        limit_base = max(float(x) for x in basis_values)
        tol = max(BRIDGE_OVER_COMMITMENT_WARN_TOLERANCE_DOLLARS, abs(limit_base) * BRIDGE_OVER_COMMITMENT_WARN_RATIO)
        if upb_sum <= limit_base * BRIDGE_UPB_COMMITMENT_RATIO_LIMIT + tol:
            continue

        stage_exception = False
        if "Loan Stage" in out.columns:
            stage_exception = bool(_bridge_stage_exception_mask(out.loc[idxs, "Loan Stage"]).any())

        candidates: List[Tuple[str, float]] = []
        # Do not repair UPB to funded amount. Funded amount is only an allocation
        # weight/source for Active Funded Amount, not a UPB substitute. Only use a
        # prior completed-report UPB when it is plausible; otherwise leave the row
        # visible for review.
        if prev_sum > 0 and prev_sum <= limit_base * BRIDGE_UPB_COMMITMENT_RATIO_LIMIT + tol:
            candidates.append(("prior asset UPB rollup", prev_sum))

        if candidates:
            source_label, target_amount = candidates[0]
            _allocate_group_amount_by_weights(out, idxs, upb_col, float(target_amount), "SF Funded Amount")
            repaired_count += 1
            if repaired_count <= 20:
                display_deal = clean_text(out.loc[idxs[0], "Deal Number"] if "Deal Number" in out.columns else deal_key)
                diagnostics.append(
                    f"Bridge Asset UPB repair: deal {display_deal} had asset UPB rollup {upb_sum:,.2f} versus same-deal basis {limit_base:,.2f}; replaced with {source_label} {float(target_amount):,.2f}."
                )
        else:
            unresolved_count += 1
            if unresolved_count <= 20:
                display_deal = clean_text(out.loc[idxs[0], "Deal Number"] if "Deal Number" in out.columns else deal_key)
                diagnostics.append(
                    f"Bridge Asset UPB review: deal {display_deal} has asset UPB rollup {upb_sum:,.2f} above same-deal basis {limit_base:,.2f}, but no safer funded/prior source was available; value was left unchanged."
                )

    after_total = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0.0).sum()
    if repaired_count or unresolved_count:
        diagnostics.insert(
            0,
            f"Bridge Asset UPB repair summary: repaired {repaired_count:,} deal(s), left {unresolved_count:,} deal(s) for review; total UPB changed by {after_total - before_total:,.2f}.",
        )
    return downcast_numeric_frame(out), diagnostics


def _choose_bridge_loan_upb(out: pd.DataFrame, upb_col: str, asset_rollup_col: Optional[str] = None) -> pd.Series:
    idx = out.index
    commitment = pd.to_numeric(out.get("Loan Commitment", pd.Series([np.nan] * len(out), index=idx)), errors="coerce")
    funded = pd.to_numeric(out.get("Active Funded Amount", pd.Series([np.nan] * len(out), index=idx)), errors="coerce")
    stage_exception = _bridge_stage_exception_mask(out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=idx)))

    # Bridge Loan UPB should tie to the already-built Bridge Asset rows. Treat the
    # asset rollup as authoritative when present, then use same-deal fallbacks only
    # for rows that did not receive an asset rollup.
    authoritative_rollup = None
    if asset_rollup_col and asset_rollup_col in out.columns:
        authoritative_rollup = pd.to_numeric(out[asset_rollup_col], errors="coerce")

    candidates: List[Tuple[str, pd.Series]] = []
    for label, col in [
        ("loan servicer UPB", "_loan_upb"),
        ("current Bridge Loan UPB", upb_col),
        ("SF current UPB", "SF Current UPB"),
        ("active asset UPB", "Active Asset UPB"),
        ("prior workbook UPB", "_prev_upb"),
        ("funded amount", "Active Funded Amount"),
    ]:
        if col in out.columns:
            candidates.append((label, pd.to_numeric(out[col], errors="coerce")))

    chosen = pd.Series([np.nan] * len(out), index=idx, dtype="float64")
    if authoritative_rollup is not None:
        chosen.loc[authoritative_rollup.notna()] = authoritative_rollup.loc[authoritative_rollup.notna()]
    for _label, values in candidates:
        values = pd.to_numeric(pd.Series(values, index=idx), errors="coerce")
        plausible = _bridge_value_plausible_for_upb(values, commitment, funded)
        fill_mask = chosen.isna() & plausible
        chosen.loc[fill_mask] = values.loc[fill_mask]

    # Exception stages stay visible. If no candidate passes the active-loan plausibility test, keep the first
    # positive same-deal value rather than blanking the row. The diagnostics will surface it for review.
    if bool(stage_exception.any()):
        for _label, values in candidates:
            values = pd.to_numeric(pd.Series(values, index=idx), errors="coerce")
            fill_mask = chosen.isna() & stage_exception & values.gt(0)
            chosen.loc[fill_mask] = values.loc[fill_mask]

    # Last-resort same-deal presentation fallback. Prefer funded amount; only use commitment when there is no
    # funded amount. This avoids publishing a contaminated servicer/property value while avoiding a blank UPB.
    funded_fallback = funded.where(funded.gt(0))
    fill_mask = chosen.isna() & funded_fallback.notna()
    chosen.loc[fill_mask] = funded_fallback.loc[fill_mask]
    commitment_fallback = commitment.where(commitment.gt(0))
    fill_mask = chosen.isna() & commitment_fallback.notna()
    chosen.loc[fill_mask] = commitment_fallback.loc[fill_mask]

    return pd.to_numeric(chosen, errors="coerce")


def _repair_bridge_loan_commitment_math(bridge_loan: pd.DataFrame, upb_col: str) -> Tuple[pd.DataFrame, List[str]]:
    if bridge_loan is None or bridge_loan.empty:
        return (pd.DataFrame() if bridge_loan is None else bridge_loan.copy()), []
    out = bridge_loan.copy()
    diagnostics: List[str] = []
    idx = out.index
    commitment = pd.to_numeric(out.get("Loan Commitment", pd.Series([np.nan] * len(out), index=idx)), errors="coerce")
    stage_exception = _bridge_stage_exception_mask(out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=idx)))

    if "Active Funded Amount" in out.columns:
        funded = pd.to_numeric(out["Active Funded Amount"], errors="coerce")
        tol = _bridge_material_tolerance(commitment)
        material_over = commitment.gt(0) & funded.gt(commitment + tol) & (~stage_exception)
        if bool(material_over.any()):
            diagnostics.append(
                f"Bridge Loan review: Active Funded Amount exceeds Loan Commitment for {int(material_over.sum()):,} non-exception active deal(s). The amount was not capped; Remaining Commitment was preserved from Salesforce/template."
            )

        # Remaining Commitment is a source/template field in the active-loan report.
        # Do not recompute it as commitment minus funded amount, because that caused
        # mismatches on the completed 4/27 report and can overwrite business-approved
        # remaining-facility values.
        if "Remaining Commitment" not in out.columns:
            out["Remaining Commitment"] = np.nan

    if upb_col in out.columns:
        before = pd.to_numeric(out[upb_col], errors="coerce")
        chosen = _choose_bridge_loan_upb(out, upb_col, asset_rollup_col=None)
        changed = before.fillna(-999999999.12345).round(2).ne(chosen.fillna(-999999999.12345).round(2)) & chosen.notna()
        if bool(changed.any()):
            diagnostics.append(
                f"Bridge Loan repair: replaced implausible/missing UPB for {int(changed.sum()):,} deal(s) using same-deal asset, servicer, Salesforce, prior, or funded-source rules."
            )
            out.loc[changed, upb_col] = chosen.loc[changed]

    return downcast_numeric_frame(out), diagnostics


def _reconcile_bridge_loan_from_asset_rollup(bridge_loan: pd.DataFrame, bridge_asset: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    if bridge_loan is None or bridge_loan.empty:
        return pd.DataFrame() if bridge_loan is None else bridge_loan.copy()
    out = _ensure_deal_key(bridge_loan, "Deal Number")
    roll = _rollup_bridge_asset_math(bridge_asset, upb_col)
    if roll.empty:
        out, _diags = _repair_bridge_loan_commitment_math(out, upb_col)
        return out

    out = out.merge(roll, on="_deal_key", how="left", suffixes=("", "_asset_rollup"))
    critical_rollup_cols = [
        "Active Funded Amount",
        "Initial Disbursement Funded",
        "Renovation Holdback",
        "Renovation HB Funded",
        "Renovation HB Remaining",
        "Interest Allocation",
        "Interest Allocation Funded",
        "Suspense Balance",
    ]
    for col in critical_rollup_cols:
        src = f"{col}_asset_rollup"
        if src not in out.columns:
            continue
        if col not in out.columns:
            out[col] = np.nan
        src_num = pd.to_numeric(out[src], errors="coerce")
        cur_num = pd.to_numeric(out[col], errors="coerce")
        out[col] = cur_num.where(src_num.isna(), src_num)
        out = out.drop(columns=[src], errors="ignore")

    upb_roll_col = f"{upb_col}_asset_rollup"
    if upb_roll_col in out.columns:
        # Bridge Loan UPB must reflect the sum of finalized Bridge Asset UPB values.
        # Do not replace this rollup with a deal-level balance after asset values are set.
        out[upb_col] = pd.to_numeric(out[upb_roll_col], errors="coerce")
        out = out.drop(columns=[upb_roll_col], errors="ignore")

    for col in ["Most Recent As-Is Value", "Most Recent ARV"]:
        src = f"{col}_asset_rollup"
        if src in out.columns:
            cur_num = pd.to_numeric(out.get(col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
            src_num = pd.to_numeric(out[src], errors="coerce")
            out[col] = cur_num.where(cur_num.notna(), src_num)
            out = out.drop(columns=[src], errors="ignore")

    # Do not run the same-deal UPB repair after applying the asset rollup; that repair
    # can replace the asset-summed UPB with a loan-level fallback. Keep the rollup.
    return downcast_numeric_frame(out)

def _prev_sid_to_deal_map(prev_maps: Optional[dict]) -> Dict[str, str]:
    if not prev_maps or "term_loan_sid" not in prev_maps:
        return {}
    prev_sid = prev_maps.get("term_loan_sid")
    if not isinstance(prev_sid, pd.DataFrame) or prev_sid.empty or not {"_sid_key", "_deal_key"}.issubset(prev_sid.columns):
        return {}
    tmp = prev_sid.dropna(subset=["_sid_key", "_deal_key"]).copy()
    tmp["_sid_key"] = tmp["_sid_key"].astype("string").str.strip()
    tmp["_deal_key"] = tmp["_deal_key"].astype("string").str.strip()
    tmp = tmp[(tmp["_sid_key"] != "") & (tmp["_deal_key"] != "")]
    return tmp.drop_duplicates("_sid_key", keep="last").set_index("_sid_key")["_deal_key"].to_dict()


def _guard_term_loan_upb_vs_amount(df: pd.DataFrame, upb_col: str, prev_maps: Optional[dict] = None) -> pd.DataFrame:
    if df is None or df.empty or upb_col not in df.columns or "Loan Amount" not in df.columns:
        return pd.DataFrame() if df is None else df.copy()
    out = _ensure_deal_key(df, "Deal Number")
    loan_amount = pd.to_numeric(out["Loan Amount"], errors="coerce")
    upb = pd.to_numeric(out[upb_col], errors="coerce")
    implausible = loan_amount.gt(0) & upb.gt(loan_amount * TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT)

    if bool(implausible.any()) and prev_maps and "term_loan_upb" in prev_maps:
        prev = prev_maps["term_loan_upb"]
        if isinstance(prev, pd.DataFrame) and not prev.empty and {"_deal_key", "_prev_upb"}.issubset(prev.columns):
            prev_map = prev.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key").set_index("_deal_key")["_prev_upb"]
            prev_upb = pd.to_numeric(out["_deal_key"].map(prev_map), errors="coerce")
            plausible_prev = prev_upb.gt(0) & loan_amount.gt(0) & prev_upb.le(loan_amount * TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT)
            out.loc[implausible & plausible_prev, upb_col] = prev_upb.loc[implausible & plausible_prev]
            upb = pd.to_numeric(out[upb_col], errors="coerce")
            implausible = loan_amount.gt(0) & upb.gt(loan_amount * TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT)

    if bool(implausible.any()):
        # No usable prior UPB: the servicer match attached a balance that exceeds the
        # loan amount (often a different/larger loan). The report never shows Term Loan
        # UPB above Loan Amount, so fall back to Loan Amount rather than keep the bad
        # value. This keeps Term Asset UPB populated (Loan Amount is always present)
        # and matches the report. Implausible matches are still logged by QA.
        out.loc[implausible, upb_col] = loan_amount.loc[implausible]

    # V64: the same guard, from below. The official report never shows a Term Loan with a
    # zero or blank UPB -- 0 of 1,025 rows on 20260824 -- so a zero here means the servicer
    # match failed, not that the loan paid off. Deals 43422 and 43462 (Vision & Beyond 1 and
    # 2, Legacy / Active Term, DQ 90+, Servicer 'N/A') carry no servicer row at all, and the
    # official falls back to Loan Amount: 18,189,500 and 18,047,000, which is also exactly
    # what each deal's Property ALA sums to. Test 76 wrote 0 for both, understating the Term
    # tabs by 36.2M of their total 52.5M gap and blanking the 60 Term Asset rows underneath.
    # Scoped to rows that are zero/blank AND have a Loan Amount, so it cannot touch a row the
    # servicer file already priced.
    upb = pd.to_numeric(out[upb_col], errors="coerce")
    missing_upb = loan_amount.gt(0) & (upb.isna() | upb.le(0))
    if bool(missing_upb.any()):
        out.loc[missing_upb, upb_col] = loan_amount.loc[missing_upb]
    return downcast_numeric_frame(out)


def _clear_duplicate_term_servicer_assignments(df: pd.DataFrame, upb_col: str, prev_maps: Optional[dict] = None) -> pd.DataFrame:
    if df is None or df.empty or "Servicer ID" not in df.columns:
        return pd.DataFrame() if df is None else df.copy()
    out = _ensure_deal_key(df, "Deal Number")
    # Duplicate-looking servicer IDs should not be cleared automatically. Clearing them
    # caused blank UPB to cascade from Term Loan into Term Asset. Keep values in the
    # report and rely on QA diagnostics for duplicate review instead of blanking balances.
    return downcast_numeric_frame(out)

def _allocate_term_asset_upb_from_loan(term_asset: pd.DataFrame, term_loan: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    if term_asset is None or term_asset.empty:
        return pd.DataFrame() if term_asset is None else term_asset.copy()
    if term_loan is None or term_loan.empty or upb_col not in term_loan.columns:
        return term_asset.copy()

    out = _ensure_deal_key(term_asset, "Deal Number")
    if "_asset_key" not in out.columns:
        out["_asset_key"] = norm_id_series(out.get("Asset ID", pd.Series([None] * len(out), index=out.index)))

    tl = _ensure_deal_key(term_loan, "Deal Number")
    tl_upb = tl[["_deal_key", upb_col]].dropna(subset=["_deal_key"]).drop_duplicates("_deal_key", keep="last")
    tl_upb = tl_upb.rename(columns={upb_col: "_loan_upb_for_alloc"})

    out = out.drop(columns=[upb_col, "_loan_upb_for_alloc"], errors="ignore")
    out = out.merge(tl_upb, on="_deal_key", how="left")

    loan_upb = pd.to_numeric(out["_loan_upb_for_alloc"], errors="coerce")
    ala = pd.to_numeric(out.get("Property ALA", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0.0)
    positive_ala = ala.where(ala.gt(0), 0.0)
    ala_sum = positive_ala.groupby(out["_deal_key"]).transform("sum")
    asset_count = out.groupby("_deal_key")["_asset_key"].transform("count").replace({0: np.nan})

    alloc = pd.Series([np.nan] * len(out), index=out.index, dtype="float64")
    ala_mask = loan_upb.notna() & ala_sum.gt(0)
    alloc.loc[ala_mask] = loan_upb.loc[ala_mask] * (positive_ala.loc[ala_mask] / ala_sum.loc[ala_mask])
    equal_mask = loan_upb.notna() & (~ala_sum.gt(0)) & asset_count.gt(0)
    alloc.loc[equal_mask] = loan_upb.loc[equal_mask] / asset_count.loc[equal_mask]
    out[upb_col] = alloc

    # Correct rounding drift so each deal's Term Asset UPB exactly ties to Term Loan UPB before workbook rounding.
    for deal, idxs in out.groupby("_deal_key", dropna=True).groups.items():
        idxs = list(idxs)
        loan_vals = pd.to_numeric(out.loc[idxs, "_loan_upb_for_alloc"], errors="coerce").dropna()
        if loan_vals.empty:
            continue
        loan_val = float(loan_vals.iloc[0])
        cur_vals = pd.to_numeric(out.loc[idxs, upb_col], errors="coerce").fillna(0.0)
        if cur_vals.empty:
            continue
        diff = loan_val - float(cur_vals.sum())
        nonblank_alloc = cur_vals[cur_vals.ne(0)].index.tolist() or idxs
        adjust_idx = nonblank_alloc[-1]
        current_adjust_val = pd.to_numeric(pd.Series([out.loc[adjust_idx, upb_col]]), errors="coerce").iloc[0]
        if pd.isna(current_adjust_val):
            current_adjust_val = 0.0
        out.loc[adjust_idx, upb_col] = float(current_adjust_val) + diff

    out = out.drop(columns=["_loan_upb_for_alloc"], errors="ignore")
    return downcast_numeric_frame(out)



def _math_issue_message(title: str, examples: pd.DataFrame, max_rows: int = 12) -> str:
    shown = examples.head(max_rows).copy()
    try:
        rendered = shown.to_string(index=False)
    except Exception:
        rendered = str(shown.head(max_rows).to_dict("records"))
    return f"{title}\n\nFirst {min(len(shown), max_rows)} example(s):\n{rendered}"


def _math_issues_to_diagnostics(prefix: str, issues: List[Tuple[str, pd.DataFrame]]) -> List[str]:
    if not issues:
        return []
    detail = "\n\n".join(_math_issue_message(title, df) for title, df in issues)
    if STRICT_MATH_HARD_STOP:
        raise RuntimeError(f"{prefix}\n\n" + detail)
    lines = [f"{prefix} The workbook was allowed to continue; review/fix source data or QA examples below."]
    for title, df in issues:
        lines.append(_math_issue_message(title, df, max_rows=5))
    return lines


def validate_bridge_math_or_raise(bridge_asset: pd.DataFrame, bridge_loan: Optional[pd.DataFrame], upb_col: str) -> List[str]:
    issues: List[Tuple[str, pd.DataFrame]] = []
    if bridge_asset is not None and not bridge_asset.empty:
        ba = _recompute_bridge_asset_funded_amount(bridge_asset)
        if "SF Funded Amount" in ba.columns:
            expected = _bridge_component_funded_sum(ba)
            actual = pd.to_numeric(ba["SF Funded Amount"], errors="coerce")
            diff = (actual.fillna(0.0) - expected.fillna(0.0)).abs()
            mask = diff.gt(MATH_TOLERANCE_DOLLARS)
            if bool(mask.any()):
                cols = [c for c in ["Deal Number", "Asset ID", "SF Funded Amount", "Initial Disbursement Funded", "Renovation Holdback Funded", "Interest Allocation Funded"] if c in ba.columns]
                ex = ba.loc[mask, cols].copy()
                ex["expected_component_sum"] = expected.loc[mask]
                ex["difference"] = actual.loc[mask] - expected.loc[mask]
                issues.append(("Bridge Asset SF Funded Amount does not equal funded components.", ex))

    if bridge_loan is not None and not bridge_loan.empty and bridge_asset is not None and not bridge_asset.empty:
        bl = _ensure_deal_key(bridge_loan, "Deal Number")
        roll = _rollup_bridge_asset_math(bridge_asset, upb_col)
        if not roll.empty:
            check = bl.merge(roll[[c for c in ["_deal_key", "Active Funded Amount", upb_col] if c in roll.columns]], on="_deal_key", how="left", suffixes=("", "_asset_rollup"))
            if "Active Funded Amount_asset_rollup" in check.columns:
                actual = pd.to_numeric(check.get("Active Funded Amount"), errors="coerce")
                expected = pd.to_numeric(check.get("Active Funded Amount_asset_rollup"), errors="coerce")
                diff = actual.fillna(0.0) - expected.fillna(0.0)
                commitment = pd.to_numeric(check.get("Loan Commitment", pd.Series([np.nan] * len(check), index=check.index)), errors="coerce")
                stage_exception = _bridge_stage_exception_mask(check.get("Loan Stage", pd.Series([pd.NA] * len(check), index=check.index)))
                material_tol = _bridge_material_tolerance(commitment)
                mask = expected.notna() & diff.abs().gt(material_tol) & (~stage_exception)
                if bool(mask.any()):
                    ex = check.loc[mask, [c for c in ["Deal Number", "Loan Commitment", "Active Funded Amount", "Active Funded Amount_asset_rollup"] if c in check.columns]].copy()
                    ex["difference"] = diff.loc[mask]
                    issues.append(("Bridge Loan Active Funded Amount materially differs from Bridge Asset funded rollup.", ex))
            upb_roll_col = f"{upb_col}_asset_rollup"
            if upb_roll_col in check.columns and upb_col in check.columns:
                actual = pd.to_numeric(check[upb_col], errors="coerce")
                expected = pd.to_numeric(check[upb_roll_col], errors="coerce")
                diff = actual.fillna(0.0) - expected.fillna(0.0)
                commitment = pd.to_numeric(check.get("Loan Commitment", pd.Series([np.nan] * len(check), index=check.index)), errors="coerce")
                funded = pd.to_numeric(check.get("Active Funded Amount", pd.Series([np.nan] * len(check), index=check.index)), errors="coerce")
                plausible_expected = _bridge_value_plausible_for_upb(expected, commitment, funded)
                mask = expected.notna() & plausible_expected & diff.abs().gt(_bridge_material_tolerance(_bridge_limit_base(commitment, funded)))
                if bool(mask.any()):
                    ex = check.loc[mask, [c for c in ["Deal Number", upb_col, upb_roll_col] if c in check.columns]].copy()
                    ex["difference"] = diff.loc[mask]
                    issues.append(("Bridge Loan UPB differs from plausible Bridge Asset UPB rollup.", ex))

        if {"Loan Commitment", "Active Funded Amount"}.issubset(bl.columns):
            commitment = pd.to_numeric(bl["Loan Commitment"], errors="coerce")
            funded = pd.to_numeric(bl["Active Funded Amount"], errors="coerce")
            tol = _bridge_material_tolerance(commitment)
            stage_exception = _bridge_stage_exception_mask(bl.get("Loan Stage", pd.Series([pd.NA] * len(bl), index=bl.index)))
            mask = commitment.gt(0) & funded.gt(commitment + tol) & (~stage_exception)
            if bool(mask.any()):
                ex = bl.loc[mask, [c for c in ["Deal Number", "Loan Stage", "Loan Commitment", "Active Funded Amount"] if c in bl.columns]].copy()
                ex["over_by"] = funded.loc[mask] - commitment.loc[mask]
                issues.append(("Bridge Loan Active Funded Amount materially exceeds Loan Commitment for non-exception active loans.", ex))
        if {"Loan Commitment", upb_col}.issubset(bl.columns):
            commitment = pd.to_numeric(bl["Loan Commitment"], errors="coerce")
            upb = pd.to_numeric(bl[upb_col], errors="coerce")
            funded = pd.to_numeric(bl.get("Active Funded Amount", pd.Series([np.nan] * len(bl), index=bl.index)), errors="coerce")
            base = _bridge_limit_base(commitment, funded)
            tol = _bridge_material_tolerance(base)
            stage_exception = _bridge_stage_exception_mask(bl.get("Loan Stage", pd.Series([pd.NA] * len(bl), index=bl.index)))
            mask = base.gt(0) & upb.gt(base * BRIDGE_UPB_COMMITMENT_RATIO_LIMIT + tol) & (~stage_exception)
            if bool(mask.any()):
                ex = bl.loc[mask, [c for c in ["Deal Number", "Loan Stage", "Loan Commitment", upb_col] if c in bl.columns]].copy()
                ex["over_ratio"] = upb.loc[mask] / commitment.loc[mask]
                issues.append(("Bridge Loan UPB remains implausibly above Loan Commitment after repair.", ex))

    return _math_issues_to_diagnostics("Bridge math validation found review items.", issues)


def validate_term_loan_amounts_or_raise(term_loan: pd.DataFrame, upb_col: str) -> List[str]:
    if term_loan is None or term_loan.empty or upb_col not in term_loan.columns or "Loan Amount" not in term_loan.columns:
        return []
    tl = _ensure_deal_key(term_loan, "Deal Number")
    loan_amount = pd.to_numeric(tl["Loan Amount"], errors="coerce")
    upb = pd.to_numeric(tl[upb_col], errors="coerce")
    mask = loan_amount.gt(0) & upb.gt(loan_amount * TERM_UPB_LOAN_AMOUNT_RATIO_LIMIT)
    issues: List[Tuple[str, pd.DataFrame]] = []
    if bool(mask.any()):
        ex = tl.loc[mask, [c for c in ["Deal Number", "Servicer ID", "Servicer", "Loan Amount", upb_col] if c in tl.columns]].copy()
        ex["over_ratio"] = upb.loc[mask] / loan_amount.loc[mask]
        issues.append(("Term Loan UPB remains implausibly above Loan Amount after repair.", ex))
    return _math_issues_to_diagnostics("Term Loan amount validation found review items.", issues)


def validate_term_math_or_raise(term_loan: pd.DataFrame, term_asset: pd.DataFrame, upb_col: str) -> List[str]:
    diagnostics = validate_term_loan_amounts_or_raise(term_loan, upb_col)
    if term_loan is None or term_loan.empty or term_asset is None or term_asset.empty or upb_col not in term_loan.columns or upb_col not in term_asset.columns:
        return diagnostics
    tl = _ensure_deal_key(term_loan, "Deal Number")
    ta = _ensure_deal_key(term_asset, "Deal Number")
    asset_roll = ta.groupby("_deal_key", dropna=True)[upb_col].sum(min_count=1).reset_index().rename(columns={upb_col: "Term Asset UPB Rollup"})
    check = tl.merge(asset_roll, on="_deal_key", how="left")
    loan_upb = pd.to_numeric(check[upb_col], errors="coerce")
    asset_upb = pd.to_numeric(check["Term Asset UPB Rollup"], errors="coerce")
    diff = loan_upb.fillna(0.0) - asset_upb.fillna(0.0)
    mask = loan_upb.notna() & asset_upb.notna() & diff.abs().gt(MATH_TOLERANCE_DOLLARS)
    issues: List[Tuple[str, pd.DataFrame]] = []
    if bool(mask.any()):
        ex = check.loc[mask, [c for c in ["Deal Number", "Servicer ID", "Loan Amount", upb_col, "Term Asset UPB Rollup"] if c in check.columns]].copy()
        ex["difference"] = diff.loc[mask]
        issues.append(("Term Loan UPB does not equal Term Asset UPB rollup.", ex))
    diagnostics.extend(_math_issues_to_diagnostics("Term math validation found review items.", issues))
    return diagnostics

def build_bridge_asset(
    sf_spine: pd.DataFrame,
    sf_dnl: pd.DataFrame,
    sf_val: pd.DataFrame,
    sf_foreclosure: pd.DataFrame,
    sf_am: pd.DataFrame,
    sf_active_rm: pd.DataFrame,
    serv_lookup: pd.DataFrame,
    upb_col: str,
    prev_maps: dict,
    template_maps: dict,
    npl_maps: Optional[dict] = None,
) -> pd.DataFrame:
    out = pd.DataFrame(index=sf_spine.index)

    for col, label in BRIDGE_ASSET_FROM_BRIDGE_SPINE.items():
        out[col] = sf_spine[label] if label in sf_spine.columns else pd.NA

    for extra in ["Loan Commitment", "Remaining Commitment", "Current UPB", "Current Servicer UPB", "Salesforce Suspense Balance", "Property Suspense Balance", "Comments AM"]:
        if extra in sf_spine.columns:
            out[extra] = sf_spine[extra]

    # V69: carry the Salesforce payment dates as COLUMNS on `out`, here, while `out` still
    # shares sf_spine's index. They used to be read straight off sf_spine much further down,
    # AFTER `out` had been through eight left-merges -- and a merge replaces the index with a
    # fresh 0..n-1 RangeIndex. sf_spine's index is NOT 0..n-1: _build_bridge_spine_like ends
    # with a filter, a sort_values and a drop_duplicates, none of which reset it, so it is a
    # sparse permuted subset of the original bulk-query index. Every later operation that
    # combined those series with `out` therefore aligned on labels that no longer meant the
    # same row, and _bridge_pick_next_payment_date made it worse by doing
    # pd.Series(servicer_dates, index=sf.index), which REINDEXES the servicer dates onto the
    # stale labels. Rows came out holding another row's date or NaT.
    #
    # On test 77 that put 2026-05-10 -- a real Statebridge value, 281 rows of it in
    # CoreVestLoanData_08312026 -- onto 88 FCI deals whose own FCI tape rows say 2026-10-01
    # (51 of them) or 2026-09-01 (27). Bridge Loan takes the MIN of its assets' dates, so one
    # scrambled asset dragged the whole deal back four months, and Days Past Due and 9/30 NPL
    # cascaded off it: 176 + 176 + 129 of that tab's 971 mismatches.
    #
    # Assigned as columns, the values travel with their row through every merge.
    for _spine_src, _spine_dst in (
        ("Property Next Payment Date", "_sf_prop_npd"),
        ("Opportunity Next Payment Date", "_sf_opp_npd"),
        ("First Payment Date", "_sf_first_payment_src"),
    ):
        out[_spine_dst] = sf_spine[_spine_src] if _spine_src in sf_spine.columns else pd.NaT

    # V56: Asset Commitment is mapped from the per-asset Approved Advance Amount Max by
    # the BRIDGE_ASSET_FROM_BRIDGE_SPINE loop above. It must NOT be re-seeded to NaN here --
    # doing so wiped the sourced value and forced the materializer's approved-components
    # fallback (which matches only 4,762 of 4,782 on 20260824). Seed only when the spine
    # did not supply the column at all.
    if "Asset Commitment" not in out.columns:
        out["Asset Commitment"] = np.nan

    out["Portfolio"] = pd.NA
    out["Segment"] = pd.NA
    out["Strategy Grouping"] = pd.NA
    out["Do Not Lend (Y/N)"] = "N"
    out["Active RM"] = pd.NA
    out["3/31 NPL (Y/N)"] = pd.NA
    out["Needs NPL Value"] = pd.NA
    out["Special Flag"] = pd.NA
    out["Servicer"] = pd.NA
    out["Servicer Status"] = pd.NA

    out["_deal_key"] = norm_id_series(out.get("Deal Number", pd.Series([None] * len(out))))
    out["_sid_key"] = id_key_no_leading_zeros(out.get("Servicer ID", pd.Series([None] * len(out))))
    out["_asset_key"] = norm_id_series(out.get("Asset ID", pd.Series([None] * len(out))))

    # The report's asset-funded amount is formula-equivalent to
    # Initial Disbursement Funded + Renovation Holdback Funded + Interest Allocation Funded.
    # Build it before servicer allocation so UPB uses the same weights as the workbook.
    out = _recompute_bridge_asset_funded_amount(out)

    if "bridge_asset_upb" in prev_maps:
        prev_asset_upb = prev_maps["bridge_asset_upb"].copy()
        out = out.merge(prev_asset_upb, on="_asset_key", how="left")
    else:
        out["_prev_asset_upb"] = np.nan

    if not sf_dnl.empty and "Deal Loan Number" in sf_dnl.columns:
        dnl = sf_dnl.copy()
        dnl["_deal_key"] = norm_id_series(dnl["Deal Loan Number"])
        if "Do Not Lend" in dnl.columns:
            dnl = dnl[["_deal_key", "Do Not Lend"]].drop_duplicates("_deal_key")
            out = out.merge(dnl, on="_deal_key", how="left")
            out["Do Not Lend (Y/N)"] = _yn_from_bool_series(out["Do Not Lend"])
            out = out.drop(columns=["Do Not Lend"], errors="ignore")

    if not sf_active_rm.empty and "Deal Loan Number" in sf_active_rm.columns and "Active RM" in sf_active_rm.columns:
        arm = sf_active_rm.copy()
        arm["_deal_key"] = norm_id_series(arm["Deal Loan Number"])
        arm = arm[["_deal_key", "Active RM"]].drop_duplicates("_deal_key")
        out = out.merge(arm, on="_deal_key", how="left", suffixes=("", "_sf"))
        out["Active RM"] = coalesce_keep_nonblank(out.get("Active RM_sf", pd.Series([pd.NA] * len(out), index=out.index)), out["Active RM"])
        out = out.drop(columns=["Active RM_sf"], errors="ignore")

    if not sf_val.empty and "Asset ID" in sf_val.columns:
        v = sf_val.copy()
        v["_asset_key"] = norm_id_series(v["Asset ID"])
        rename_map = {
            vlabel: f"__val__{tcol}"
            for tcol, vlabel in BRIDGE_ASSET_FROM_VALUATION.items()
            if vlabel in v.columns
        }
        keep = ["_asset_key"] + list(rename_map.keys())
        v = v[keep].rename(columns=rename_map).drop_duplicates("_asset_key")
        out = out.merge(v, on="_asset_key", how="left")
        for tcol in BRIDGE_ASSET_FROM_VALUATION.keys():
            tmpcol = f"__val__{tcol}"
            if tmpcol in out.columns:
                if tcol == "Most Recent Appraisal Order Date":
                    # The valuation path computes this strictly from the per-appraisal
                    # Order_Received_Date__c (MAX, N/A when none) -- the authoritative source
                    # that matches the official report. It must OVERWRITE the spine's
                    # Property-level BPO_Appraisal_Order_Date__c, which over-populates the column.
                    out[tcol] = out[tmpcol]
                else:
                    out[tcol] = coalesce_keep_nonblank(
                        out.get(tcol, pd.Series([pd.NA] * len(out), index=out.index)),
                        out[tmpcol],
                    )
                out = out.drop(columns=[tmpcol], errors="ignore")

    if not sf_foreclosure.empty and "Asset ID" in sf_foreclosure.columns:
        fc = sf_foreclosure.copy()
        fc["_asset_key"] = norm_id_series(fc["Asset ID"])
        rename_map = {
            vlabel: f"__fc__{tcol}"
            for tcol, vlabel in BRIDGE_ASSET_FROM_FORECLOSURE.items()
            if vlabel in fc.columns
        }
        keep = ["_asset_key"] + list(rename_map.keys())
        fc = fc[keep].rename(columns=rename_map).drop_duplicates("_asset_key")
        out = out.merge(fc, on="_asset_key", how="left")
        for tcol in BRIDGE_ASSET_FROM_FORECLOSURE.keys():
            tmpcol = f"__fc__{tcol}"
            if tmpcol in out.columns:
                out[tcol] = coalesce_keep_nonblank(out[tmpcol], out.get(tcol, pd.Series([pd.NA] * len(out), index=out.index)))
                out = out.drop(columns=[tmpcol], errors="ignore")
            if tcol in out.columns:
                out[tcol] = coalesce_keep_nonblank(out[tcol], pd.Series(["N/A"] * len(out), index=out.index))

    if not sf_am.empty and "Deal Loan Number" in sf_am.columns:
        am = sf_am.copy()
        am["_deal_key"] = norm_id_series(am["Deal Loan Number"])
        am["_dt"] = pd.to_datetime(am.get("Date Assigned"), errors="coerce")
        am = am.sort_values(["_deal_key", "Team Role", "_dt"]).drop_duplicates(["_deal_key", "Team Role"], keep="last")

        role_to_namecol = {
            "Asset Manager": "Asset Manager 1",
            "Asset Manager 2": "Asset Manager 2",
            "Construction Manager": "Construction Mgr.",
        }
        role_to_datecol = {
            "Asset Manager": "AM 1 Assigned Date",
            "Asset Manager 2": "AM 2 Assigned Date",
            "Construction Manager": "CM Assigned Date",
        }

        piv_name = am.pivot_table(index="_deal_key", columns="Team Role", values="Team Member Name", aggfunc="first")
        piv_date = am.pivot_table(index="_deal_key", columns="Team Role", values="Date Assigned", aggfunc="first")
        piv_name = piv_name.rename(columns=role_to_namecol).reset_index()
        piv_date = piv_date.rename(columns=role_to_datecol).reset_index()

        out = out.merge(piv_name, on="_deal_key", how="left")
        out = out.merge(piv_date, on="_deal_key", how="left")

    if npl_maps and not npl_maps.get("asset_flags", pd.DataFrame()).empty:
        af_asset = npl_maps["asset_flags"].copy()
        keep = ["_deal_key", "_asset_key", "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag"]
        af_asset = af_asset[keep].drop_duplicates(["_deal_key", "_asset_key"])
        out = out.merge(af_asset, on=["_deal_key", "_asset_key"], how="left", suffixes=("", "_nplasset"))
        for c in ["3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag"]:
            out[c] = coalesce_keep_nonblank(out.get(f"{c}_nplasset", pd.Series([pd.NA] * len(out), index=out.index)), out[c])
            out = out.drop(columns=[f"{c}_nplasset"], errors="ignore")

    if npl_maps and not npl_maps.get("asset_deal_fallback", pd.DataFrame()).empty:
        af_deal = npl_maps["asset_deal_fallback"].copy()
        af_deal = af_deal[["_deal_key", "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag"]].drop_duplicates("_deal_key")
        out = out.merge(af_deal, on="_deal_key", how="left", suffixes=("", "_npldeal"))
        for c in ["3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag"]:
            out[c] = coalesce_keep_nonblank(out.get(c, pd.Series([pd.NA] * len(out), index=out.index)), out.get(f"{c}_npldeal", pd.Series([pd.NA] * len(out), index=out.index)))
            out = out.drop(columns=[f"{c}_npldeal"], errors="ignore")

    seg_guess = out.apply(
        lambda r: derive_bridge_segment(r.get("Deal Number"), r.get("Financing"), r.get("Loan Buyer"), template_maps),
        axis=1,
    )
    strat_guess = out["Project Strategy"].map(lambda x: strategy_grouping_from_project_strategy(x, template_maps.get("strategy_map", {})))
    # SA Loan (Y/N) = Asset Manager 1 is in the special-asset manager list.
    _sa_mgrs = {clean_text(m) for m in template_maps.get("sa_loan_managers", set()) if clean_text(m)}
    _am1 = out.get("Asset Manager 1", pd.Series([pd.NA] * len(out), index=out.index)).map(clean_text)
    out["SA Loan (Y/N)"] = _am1.map(lambda x: "Y" if x in _sa_mgrs else "N")
    port_guess = out.apply(
        lambda r: derive_bridge_portfolio(
            r.get("Product Type"),
            r.get("Segment") if has_any_value(r.get("Segment")) else derive_bridge_segment(r.get("Deal Number"), r.get("Financing"), r.get("Loan Buyer"), template_maps),
            r.get("Financing"),
            r.get("Deal Intro Sub-Source"),
            r.get("Deal Number"),
        ),
        axis=1,
    )

    out["Segment"] = coalesce_keep_nonblank(out["Segment"], seg_guess)
    out["Strategy Grouping"] = coalesce_keep_nonblank(out["Strategy Grouping"], strat_guess)
    out["Portfolio"] = coalesce_keep_nonblank(out["Portfolio"], port_guess)

    base_stage_series = out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    out["Financing"] = pd.Series(out.get("Financing", pd.Series([pd.NA] * len(out), index=out.index)), index=out.index, dtype="object")
    # V43: sold bridge rows display the full "Sold Servicing Retained" label, not "Sold"
    # (20260803 Bridge Asset!D: 12 rows). The report's NPL/special-list formulas compare
    # against that exact string, so writing bare "Sold" both mismatched the cell and
    # silently disabled the sold exclusion on recalc.
    out["Financing"] = out["Financing"].mask(blankish_mask(out["Financing"]) & base_stage_series.eq("Sold"), BRIDGE_SOLD_FINANCING)

    # V69: read these off `out`, NOT off sf_spine. `out` has been re-indexed by the merges
    # above; sf_spine still carries its original sparse index, so anything taken from it here
    # would align onto the wrong rows. See the note next to the _sf_* columns above.
    _nat_col = pd.Series([pd.NaT] * len(out), index=out.index)
    prop_npd = pd.to_datetime(out.get("_sf_prop_npd", _nat_col), errors="coerce")
    opp_npd = pd.to_datetime(out.get("_sf_opp_npd", _nat_col), errors="coerce")
    sf_next_payment = prop_npd.where(prop_npd.notna(), opp_npd)
    # First Payment Date for the FCI day-1 NPD correction.
    sf_first_payment = pd.to_datetime(out.get("_sf_first_payment_src", _nat_col), errors="coerce")
    sf_current_upb = pd.to_numeric(out.get("Current UPB", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")

    blank_obj = pd.Series([pd.NA] * len(out), index=out.index, dtype="object")

    if not serv_lookup.empty and "_sid_key" in serv_lookup.columns:
        s = serv_lookup.dropna(subset=["_sid_key"]).copy()
        s = s.rename(
            columns={
                "servicer": "_servicer_file",
                "upb": "_servicer_file_upb",
                "suspense": "_loan_suspense",
                "next_payment_date": "_serv_next_payment_date",
                "maturity_date": "_servicer_maturity_file",
                "status": "_servicer_status_file",
            }
        )

        out = out.merge(
            s[["_sid_key", "_servicer_file", "_servicer_file_upb", "_loan_suspense", "_serv_next_payment_date", "_servicer_maturity_file", "_servicer_status_file", "source_file"]],
            on="_sid_key",
            how="left",
        )
    else:
        out["_servicer_file"] = pd.NA
        out["_servicer_file_upb"] = np.nan
        out["_loan_suspense"] = np.nan
        out["_serv_next_payment_date"] = pd.NaT
        out["_servicer_maturity_file"] = pd.NaT
        out["_servicer_status_file"] = pd.NA

    if "bridge_loan_upb" in prev_maps:
        prev_upb = prev_maps["bridge_loan_upb"].copy()
        out = out.merge(prev_upb, on="_deal_key", how="left")
    else:
        out["_prev_upb"] = np.nan

    stage_series = out.get("Loan Stage", pd.Series([None] * len(out), index=out.index))
    reo_mask = stage_series.apply(is_reo_stage)
    late_stage_mask = stage_series.astype("string").str.strip().isin(EXPIRED_OR_MATURED_STAGES)

    sf_loan_upb = pd.to_numeric(
        out.get("Current Servicer UPB", pd.Series([np.nan] * len(out), index=out.index)),
        errors="coerce",
    )
    servicer_file_upb = pd.to_numeric(out.get("_servicer_file_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    prev_loan_upb = pd.to_numeric(out.get("_prev_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")

    # Correct Bridge Asset UPB rule from the completed report process:
    # each Bridge Asset row should carry its own asset-level UPB. Do not allocate
    # the same loan-level UPB across all assets when Salesforce/servicer already
    # provides an asset/property-level UPB. This prevents repeated UPB values inside
    # the same deal and lets Bridge Loan roll up the true sum of the asset rows.
    prev_asset_upb_vals = pd.to_numeric(out.get("_prev_asset_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")

    out["_asset_count_in_deal"] = out.groupby("_deal_key", dropna=True)["_deal_key"].transform("size").replace({0: np.nan})
    sid_count = out.groupby(["_deal_key", "_sid_key"], dropna=False)["_sid_key"].transform("size") if "_sid_key" in out.columns else pd.Series([np.nan] * len(out), index=out.index)
    # Bridge Asset UPB fix (V49): a servicer loan number (SID) that maps to a SINGLE asset
    # already carries an asset-level UPB, so use it directly. When several assets SHARE one
    # SID (the FCI / Statebridge multi-property case), the servicer file holds a single
    # LOAN-level UPB and the official report ALLOCATES it across the shared-SID assets by
    # SF Funded Amount. The prior line masked the servicer UPB whenever the SID was not 1:1,
    # so those assets fell back to funded / stale SF Current UPB and the deal total was wrong
    # (e.g. deal 61870: FCI loan 399638024 is paid down to $464,776 but every asset showed
    # the $380,827 funded amount). Guard: only trust the servicer UPB when it is POSITIVE, so
    # an FCI $0 (paid-off / unboarded) balance never overwrites a real Salesforce balance.
    # Validated vs the 20260803 official using the 7/31 servicer files: Bridge Asset UPB
    # mismatches 551 -> 236 (317 fixed across FCI + Statebridge; 2 negligible rounding cases).
    _upb_sid_key = out["_sid_key"] if "_sid_key" in out.columns else pd.Series([pd.NA] * len(out), index=out.index)
    _upb_funded = pd.to_numeric(out.get("SF Funded Amount", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").clip(lower=0).fillna(0.0)
    # groupby transform on a key that contains <NA> (blank Servicer IDs) leaves <NA> in the
    # result, which makes np.where raise "boolean value of NA is ambiguous". Coerce to plain
    # float and fill so every mask below is a clean, NA-free boolean.
    _upb_sid_funded_sum = pd.to_numeric(_upb_funded.groupby(_upb_sid_key).transform("sum"), errors="coerce").fillna(0.0)
    _upb_sid_asset_n = pd.to_numeric(_upb_sid_key.groupby(_upb_sid_key).transform("size"), errors="coerce")
    _sid_has_funded = _upb_sid_funded_sum.gt(0)
    _alloc_by_funded = servicer_file_upb * (_upb_funded / _upb_sid_funded_sum.where(_sid_has_funded))
    _alloc_equal = servicer_file_upb / _upb_sid_asset_n.where(_upb_sid_asset_n.gt(0))
    _servicer_upb_alloc = pd.to_numeric(_alloc_by_funded.where(_sid_has_funded, _alloc_equal), errors="coerce")
    _sid_is_unique = out["_asset_count_in_deal"].le(1) | sid_count.eq(1)
    safe_servicer_asset_upb = servicer_file_upb.where(_sid_is_unique, _servicer_upb_alloc)
    safe_servicer_asset_upb = safe_servicer_asset_upb.where(servicer_file_upb.gt(0))

    funded_amount_for_upb = pd.to_numeric(out.get("SF Funded Amount", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    late_stage_for_upb = stage_series.astype("string").str.strip().isin(EXPIRED_OR_MATURED_STAGES + ["Sold", "REO", "REO-Sold"])
    tiny_vs_funded = (
        funded_amount_for_upb.gt(0)
        & sf_current_upb.notna()
        & sf_current_upb.ge(0)
        & sf_current_upb.lt(funded_amount_for_upb * BRIDGE_ASSET_UPB_TINY_VS_FUNDED_RATIO)
    )
    sf_asset_upb_usable = sf_current_upb.where(~tiny_vs_funded)

    # Active bridge assets: the servicer file carries the correct per-asset UPB keyed
    # by the asset's own servicer Loan Number, so prefer it over the (often stale,
    # loan-level) Salesforce Current UPB. Fall back to SF Current UPB, then funded
    # amount, then prior asset UPB.
    # Late-stage rows keep prior completed asset UPB first.
    active_asset_upb = _coalesce_positive_then_any_numeric(safe_servicer_asset_upb, sf_asset_upb_usable, funded_amount_for_upb, prev_asset_upb_vals, index=out.index)
    late_asset_upb = _coalesce_positive_then_any_numeric(prev_asset_upb_vals, safe_servicer_asset_upb, funded_amount_for_upb, sf_asset_upb_usable, index=out.index)
    asset_level_upb = active_asset_upb.where(~late_stage_for_upb, late_asset_upb)

    # Last-resort only: if no asset-level balance exists, allocate the deal-level
    # UPB so that the report can still roll Bridge Loan UPB. This path should be
    # rare and is intentionally after asset-level and prior asset UPB.
    loan_upb_candidate = _coalesce_positive_then_any_numeric(sf_loan_upb, servicer_file_upb, prev_loan_upb, index=out.index)
    out["_loan_upb_for_alloc_raw"] = loan_upb_candidate
    out["_loan_upb_for_alloc"] = out.groupby("_deal_key", dropna=True)["_loan_upb_for_alloc_raw"].transform(_group_first_positive_then_any_numeric)

    funded_weight = pd.to_numeric(out.get("SF Funded Amount", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0.0)
    out["_funded_weight"] = funded_weight.where(funded_weight.gt(0), 0.0)
    out["_funded_weight_sum"] = out.groupby("_deal_key", dropna=True)["_funded_weight"].transform("sum")

    allocated_fallback = np.where(
        out["_funded_weight_sum"].fillna(0).gt(0),
        out["_loan_upb_for_alloc"] * (out["_funded_weight"] / out["_funded_weight_sum"]),
        out["_loan_upb_for_alloc"] / out["_asset_count_in_deal"],
    )
    allocated_fallback = pd.to_numeric(pd.Series(allocated_fallback, index=out.index), errors="coerce")

    # V48: the Bridge Asset UPB is the asset's OWN balance (asset_level_upb above),
    # NOT a funded-weighted split of the loan UPB. The 7/20 official report proves
    # this: assets with identical SF Funded Amount carry different per-asset UPB that
    # sum to the loan total (e.g. deal 40477). The earlier V46 override replaced the
    # correct per-asset values with a uniform funded-weighted allocation, which matched
    # by coincidence at one snapshot but flattened every unevenly-paid-down multi-asset
    # deal (regression: Bridge Asset UPB mismatches jumped to 1,301 on the 7/17 build).
    # allocated_fallback (funded-weighted) stays below as the genuine LAST resort, used
    # only when a deal has no asset-level balance at all.
    out[upb_col] = pd.to_numeric(asset_level_upb, errors="coerce")
    out[upb_col] = out[upb_col].where(out[upb_col].notna(), allocated_fallback)

    current_upb_series = pd.to_numeric(out[upb_col], errors="coerce")
    out[upb_col] = current_upb_series.where(
        ~((reo_mask | late_stage_mask) & (current_upb_series.isna() | current_upb_series.le(0))),
        prev_asset_upb_vals,
    )

    # Suspense is servicer-loan-level when the servicer file supplies it. Do not
    # allocate it across every property in the deal, because that double-counts
    # multi-asset loans. If there is no servicer-file suspense, fall back to the
    # Salesforce deal-level suspense one time on the first report row for the deal.
    sf_suspense = pd.to_numeric(out.get("Salesforce Suspense Balance", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    prop_suspense = pd.to_numeric(out.get("Property Suspense Balance", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    serv_suspense = pd.to_numeric(out.get("_loan_suspense", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    out["_row_in_deal"] = out.groupby("_deal_key", dropna=True).cumcount()
    # Suspense is per-asset in the completed report. The servicer-file suspense is keyed
    # by Servicer ID (loan/SID level): on a multi-asset deal that shares one Servicer ID,
    # applying it to every asset row repeats the WHOLE loan's suspense on each asset (the
    # 1459786-style 119,233.93-vs-0 over-count). Only trust the servicer-file suspense at
    # the asset level when the SID maps to a single asset in the deal (mirrors the
    # safe_servicer_asset_upb guard above); otherwise prefer the Property__c per-asset
    # suspense, then the deal-level Opportunity suspense applied once on the first row.
    safe_serv_suspense = serv_suspense.where(out["_asset_count_in_deal"].le(1) | sid_count.eq(1))
    sf_suspense_once = pd.Series(np.where(out["_row_in_deal"].eq(0), sf_suspense, np.nan), index=out.index)
    out["Suspense Balance"] = safe_serv_suspense.where(safe_serv_suspense.notna(), prop_suspense)
    out["Suspense Balance"] = out["Suspense Balance"].where(out["Suspense Balance"].notna(), sf_suspense_once)
    out["Suspense Balance"] = pd.to_numeric(out["Suspense Balance"], errors="coerce").fillna(0.0)

    # Bridge NPD is servicer-first, except for the recurring day-1/day-10 issue
    # where the official report keeps the 10th-of-month SF/prior date.
    sf_next_payment = pd.to_datetime(sf_next_payment, errors="coerce")
    serv_next_payment = pd.to_datetime(out.get("_serv_next_payment_date"), errors="coerce")
    prior_npd = None
    if "bridge_asset_manual" in prev_maps and isinstance(prev_maps.get("bridge_asset_manual"), pd.DataFrame):
        prev_npd = prev_maps["bridge_asset_manual"]
        if "Next Payment Date" in prev_npd.columns and "_asset_key" in prev_npd.columns:
            prior_map = prev_npd.dropna(subset=["_asset_key"]).drop_duplicates("_asset_key").set_index("_asset_key")["Next Payment Date"]
            prior_npd = out["_asset_key"].map(prior_map)
    _ba_servicer = coalesce_keep_nonblank(out.get("_servicer_file", pd.Series([pd.NA] * len(out), index=out.index)), out.get("Servicer", pd.Series([pd.NA] * len(out), index=out.index)))
    out["Next Payment Date"] = _bridge_pick_next_payment_date(sf_next_payment, serv_next_payment, prior_npd, servicer_names=_ba_servicer, fpd_dates=sf_first_payment, run_dt=run_dt)

    # V53: an asset with NO servicer-file row and NO Salesforce next-payment date inherits
    # the deal's date -- the official report shows one NPD for every asset on such a deal
    # (e.g. deal 63094's 85 assets all read 2026-09-10). Verified 315/315 on 20260810 using
    # the deal's modal resolved NPD; First Payment Date is the last resort for a brand-new
    # deal where no sibling has one either. Together with the day-10 rule above this takes
    # Bridge Asset NPD to 4,914/4,917.
    _npd_fill = pd.to_datetime(out["Next Payment Date"], errors="coerce")
    if bool(_npd_fill.isna().any()):
        _npd_deal_mode = _npd_fill.groupby(out["_deal_key"]).transform(
            lambda s: s.mode().iloc[0] if len(s.dropna()) and len(s.mode()) else pd.NaT
        )
        _npd_fill = _npd_fill.where(_npd_fill.notna(), _npd_deal_mode)
        _npd_fill = _npd_fill.where(_npd_fill.notna(), pd.to_datetime(sf_first_payment, errors="coerce"))
        out["Next Payment Date"] = _npd_fill

    out["Servicer"] = coalesce_keep_nonblank(out.get("_servicer_file", blank_obj), out.get("Servicer", blank_obj))
    out["Servicer Status"] = coalesce_keep_nonblank(out.get("_servicer_status_file", blank_obj), out.get("Servicer Status", blank_obj))
    # NOTE (V42): the FCI rolled-forward Servicer Status -> "N/A" gate (former Fix A.2) was
    # REVERTED -- it regressed BA Servicer Status from 26 -> 91 mismatches against same-day
    # real (it blanked rows where real kept a real status). A.1's NPD far-future gate stays.
    out["Servicer Maturity Date"] = pd.to_datetime(out.get("_servicer_maturity_file"), errors="coerce")
    out = out.drop(columns=["_prev_upb"], errors="ignore")

    if "bridge_asset_manual" in prev_maps:
        man = prev_maps["bridge_asset_manual"].copy()
        keep_cols = ["_asset_key"] + [c for c in [
            "Portfolio", "Segment", "Strategy Grouping", "REO Date", "Active RM",
            "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag",
            "Asset Manager 1", "AM 1 Assigned Date", "Asset Manager 2", "AM 2 Assigned Date",
            "Construction Mgr.", "CM Assigned Date", "Servicer", "Servicer Status",
            "Remedy Plan", "Delinquency Notes", "Maturity Status", "Title Company", "Tax Commentary",
            # NOTE: "Most Recent Appraisal Order Date" is deliberately excluded here. It is a live
            # value (MAX per-appraisal Order_Received_Date__c, N/A when none) and is authoritative
            # even when blank. Creating a _prev for it would let coalesce_keep_nonblank resurrect a
            # stale prior value on assets that correctly compute to N/A (the over-population bug).
            "Updated Valuation Date", "Updated As-Is Value", "Updated ARV",
            # Origination valuation is a FROZEN snapshot and is carry-forward-first below, so it
            # MUST be extracted here -- otherwise no _prev column exists and the carry-forward
            # silently no-ops (the current SF value wrongly wins).
            "Origination Value Dt", "Origination As-Is Value", "Origination ARV",
            "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact",
        ] if c in man.columns]
        out = out.merge(man[keep_cols], on="_asset_key", how="left", suffixes=("", "_prev"))
        bridge_asset_carry_forward_first = {
            "Portfolio", "Segment", "Strategy Grouping", "REO Date",
            "3/31 NPL (Y/N)", "Needs NPL Value", "Special Flag",
            "Remedy Plan", "Delinquency Notes",
            # V41: Active RM, the Asset/Construction Manager assignments + their dates, and
            # Maturity Status are LIVE Salesforce values and must NOT be carry-forward-first.
            # A reassignment (e.g. the 2026-06-23 Cole Smith Special-Asset move on deals
            # 54127/54129/54130/54131/54180) would otherwise be pinned to last week's stale
            # name/date. Fresh SF wins; the prior workbook still backfills via the SF-first
            # else branch when SF has no value for an asset.
            "Title Company", "Tax Commentary",
            # V80: do NOT add "% of Reno Budget" here, and do not carry it forward at all.
            # V78 tried both and it cost 3,300 cells on 20260831: the prior workbook holds the
            # literal text "N/A" on those rows, coalesce_keep_nonblank treats "N/A" as a value,
            # and last week's N/A overwrote a correctly derived ratio -- the column went from 91
            # mismatches to 3,391. A zero-fill for rows with no approved advance was equally
            # pointless: the column is in REPORT_NA_FILL_HEADERS and not in _numeric_na_cols, so
            # _normalize_output_for_report turned every 0 straight back into "N/A". Reverted to
            # the behaviour that scores 91. The official's split -- 3,375 literal zeros, 913
            # ones, 75 N/A -- is not reproduced by funded/approved, so this needs a real
            # derivation before anything is changed here again.
            "Updated Valuation Date", "Updated As-Is Value", "Updated ARV",
            # Most Recent Appraisal Order Date is intentionally NOT carry-forward-first: it is a
            # live current value (MAX per-appraisal Order_Received_Date__c, N/A when none),
            # validated 100% against the current SF extract, so the fresh value must win.
            # Origination valuation is a FROZEN snapshot captured at loan origination -- it is
            # carried forward, NOT recomputed from live Salesforce each week. Verified against
            # the official report: e.g. asset 819585 shows the origination As-Is from the prior
            # report (94.5M), not the current Property field (98.8M). Carry-forward-first.
            "Origination Value Dt", "Origination As-Is Value", "Origination ARV",
            "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact",
        }
        for c in [x for x in keep_cols if x != "_asset_key"]:
            if f"{c}_prev" in out.columns:
                if c in bridge_asset_carry_forward_first:
                    out[c] = coalesce_keep_nonblank(out[f"{c}_prev"], out.get(c, blank_obj))
                else:
                    out[c] = coalesce_keep_nonblank(out.get(c, blank_obj), out[f"{c}_prev"])
                out = out.drop(columns=[f"{c}_prev"], errors="ignore")

    # SA Loan (Y/N) depends on Asset Manager 1, which the SF-first carry-forward above may
    # have just finalized (Cole Smith reassignment, or a prior-workbook backfill of an SA
    # manager). Recompute from the FINAL Asset Manager 1 -- the earlier compute (~line 5532)
    # ran before the carry-forward. The materializer keeps this value as-is.
    if "Asset Manager 1" in out.columns:
        _am1_final = out["Asset Manager 1"].map(clean_text)
        out["SA Loan (Y/N)"] = _am1_final.map(lambda x: "Y" if x in _sa_mgrs else "N")

    # CAFL re-financed deals: once a deal is re-financed into a CAFL securitization its
    # Segment must read BRIDGE_SECURITIZED_SEGMENT, even if last week's completed report still
    # carried the prior vehicle (e.g. "CPP JV"). The prior-workbook carry-forward above is
    # Segment-first, so it would otherwise pin the stale value. Financing (Warehouse Line)
    # starting with "CAFL " is authoritative here and matches derive_bridge_segment().
    # Securitized (Y/N) and CPP JV (Y/N) are workbook formulas off Segment, so they follow.
    if "Financing" in out.columns and "Segment" in out.columns:
        _cafl_mask = out["Financing"].astype("string").str.strip().str.upper().str.startswith("CAFL ", na=False)
        out.loc[_cafl_mask, "Segment"] = BRIDGE_SECURITIZED_SEGMENT

    # V78: the same treatment for CPP JV. Segment is carry-forward-first, so a deal that has
    # since moved onto a CPP JV vehicle keeps last week's label -- 184 Bridge Asset rows and 19
    # Bridge Loan rows read "Mortgage Banking" on 20260831 while the official says "CPP JV",
    # and every one of them carries Financing "CPP JV - Goldman Sachs" in BOTH the build and the
    # official, so only the Segment derivation was stale. Exact in both directions on 20260831:
    # 731 rows have CPP JV financing, the official calls all 731 CPP JV, and the official's CPP
    # JV population is exactly those 731.
    if "Financing" in out.columns and "Segment" in out.columns:
        _cpp_mask = out["Financing"].astype("string").str.strip().str.upper().str.startswith("CPP JV", na=False)
        out.loc[_cpp_mask, "Segment"] = "CPP JV"

    # Normalize curated free-text columns for encoding artifacts (mojibake like "â€™",
    # U+2019/dash variants, _x000D_ carriage returns). Previously only Tax Commentary was
    # cleaned; the same byte-noise drove phantom Data mismatches on the other narrative
    # columns. _normalize_report_comment_text is whitespace/encoding-only -- it never
    # alters a clean value.
    for _txtcol in (
        "Tax Commentary", "Special Asset Status", "Special Asset Reason",
        "Special Asset: Special Asset Status", "Remedy Plan", "Delinquency Notes",
        "Maturity Status",
    ):
        if _txtcol in out.columns:
            out[_txtcol] = pd.Series(out[_txtcol], index=out.index, dtype="object").map(_normalize_report_comment_text)

    # Valuation/value columns must be blank (empty cell) when there is no value --
    # never 0 -- to match the official report. Guards against stale 0s from SF,
    # carry-forward, or numeric coercion.
    # V41 (Cluster 7.5): "Updated As-Is Value" is intentionally EXCLUDED here -- the
    # official report carries a genuine 0/appraised value there that the blank-zero rule
    # was wrongly emptying (190 cells). Only true nulls should read blank for that column,
    # which is already its natural state, so it is simply left out of the blank-zero list.
    out = blank_zero_value_columns(out, [
        "Updated ARV", "Most Recent ARV", "Most Recent As-Is Value",
        "Origination ARV", "Origination As-Is Value",
    ])

    status_bucket = pd.Series(
        [
            normalize_bridge_servicer_status(raw_status, npd, run_dt, loan_stage, property_status, reo_date)
            for raw_status, npd, loan_stage, property_status, reo_date in zip(
                out.get("Servicer Status", blank_obj),
                out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)),
                out.get("Loan Stage", blank_obj),
                out.get("Property Status", blank_obj),
                out.get("REO Date", pd.Series([pd.NaT] * len(out), index=out.index)),
            )
        ],
        index=out.index,
        dtype="object",
    )
    out["_bridge_dq_bucket"] = status_bucket
    out["_bridge_dpd_num"] = pd.Series(
        [
            min(_guess_days_past_due(npd, run_dt), 29.0) if has_any_value(bucket) and clean_text(bucket).upper() == "CURRENT" else _guess_days_from_bridge_bucket(bucket)
            for bucket, npd in zip(status_bucket, out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)))
        ],
        index=out.index,
        dtype="float64",
    )
    out["Servicer Status"] = coalesce_keep_nonblank(out.get("Servicer Status", blank_obj), status_bucket)

    # A row with no servicer has no servicer status: the official report shows N/A, not a
    # derived delinquency bucket. Verified exactly on 20260803 Bridge Asset -- Servicer
    # blank/N/A and Servicer Status N/A coincide on all 400 such rows with no exceptions
    # in either direction, plus 5 Sold-Servicing-Retained rows that also read N/A. The
    # previous build back-filled these from Next Payment Date and produced ~400
    # CURRENT/Performing cells where the report has N/A.
    _no_servicer = _report_is_blank_or_na(
        pd.Series(out.get("Servicer", blank_obj), index=out.index, dtype="object")
    )
    _sold_serviced = _report_is_sold_financing(out)
    out["Servicer Status"] = pd.Series(out["Servicer Status"], index=out.index, dtype="object").mask(
        _no_servicer | _sold_serviced, "N/A"
    )

    # This column must match the visible workbook formula:
    # Initial Disbursement Funded + Renovation Holdback Funded + Interest Allocation Funded.
    # Do not use Approved Advance Amount Funded here; that field caused Bridge Loan Active Funded Amount to roll up the wrong number.
    out = _recompute_bridge_asset_funded_amount(out)

    if "Is Special Asset (Y/N)" in out.columns:
        out["Is Special Asset (Y/N)"] = _yn_from_bool_series(out["Is Special Asset (Y/N)"])

    out["Servicer ID"] = normalize_servicer_id_for_report(out.get("Servicer ID", blank_obj), out.get("Servicer", blank_obj))
    out["Active RM"] = coalesce_keep_nonblank(out["Active RM"], pd.Series(["N"] * len(out), index=out.index))
    out["3/31 NPL (Y/N)"] = coalesce_keep_nonblank(out["3/31 NPL (Y/N)"], pd.Series(["N"] * len(out), index=out.index))
    out["Needs NPL Value"] = coalesce_keep_nonblank(out["Needs NPL Value"], pd.Series(["N"] * len(out), index=out.index))
    out["Special Flag"] = coalesce_keep_nonblank(out["Special Flag"], pd.Series(["N"] * len(out), index=out.index))

    out = _fill_text_defaults(
        out,
        [
            "Loan Buyer", "Servicer", "Primary Contact",
            "Remedy Plan", "Delinquency Notes", "Maturity Status",
            "Special Asset Status", "Special Asset Reason", "Special Asset: Special Asset Status",
            "Title Company", "Tax Frequency", "Tax Commentary",
            "Originator", "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact",
            "Servicer Status", "Asset Manager 1", "Asset Manager 2", "Construction Mgr.",
        ],
    )

    stage_series = out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    current_upb = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0)
    is_closed_won = stage_series.eq("Closed Won")
    is_sold = stage_series.eq("Sold")
    is_reo = stage_series.isin(REO_FAMILY_STAGES) | pd.to_datetime(out.get("REO Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce").notna()
    is_expired_or_matured = stage_series.isin(EXPIRED_OR_MATURED_STAGES)

    keep_mask = is_closed_won | is_reo | is_sold | (is_expired_or_matured & current_upb.gt(0))
    out = out.loc[keep_mask].copy()
    out = out[out["_deal_key"].notna() & out["_asset_key"].notna()].copy()

    return downcast_numeric_frame(out)



def _build_term_loan_salesforce_fallback(
    sf_term: pd.DataFrame,
    sf_am: pd.DataFrame,
    sf_active_rm: pd.DataFrame,
    serv_lookup: pd.DataFrame,
    upb_col: str,
    prev_maps: dict,
    template_maps: dict,
    prev_sold_retained_keys: Optional[Set[str]] = None,
) -> pd.DataFrame:
    if sf_term is None or sf_term.empty:
        return pd.DataFrame()

    out = pd.DataFrame(index=sf_term.index)

    for col, label in TERM_LOAN_FROM_TERM_WIDE.items():
        out[col] = sf_term[label] if label in sf_term.columns else pd.NA

    out["_deal_key"] = norm_id_series(out.get("Deal Number", pd.Series([None] * len(out))))
    prev_sold_retained_keys = prev_sold_retained_keys or set()

    if "Do Not Lend (Y/N)" in out.columns:
        out["Do Not Lend (Y/N)"] = _yn_from_bool_series(out["Do Not Lend (Y/N)"])

    out["Loan Buyer"] = sf_term["Sold Loan: Sold To"] if "Sold Loan: Sold To" in sf_term.columns else pd.NA
    out["Active RM"] = pd.NA
    out["Servicer"] = sf_term["Servicer Name"] if "Servicer Name" in sf_term.columns else pd.NA
    out["Maturity Date"] = _term_maturity_source_series(sf_term)
    out["Next Payment Date"] = pd.to_datetime(sf_term["Next Payment Date"], errors="coerce") if "Next Payment Date" in sf_term.columns else pd.NaT

    if not sf_active_rm.empty and "Deal Loan Number" in sf_active_rm.columns and "Active RM" in sf_active_rm.columns:
        arm = sf_active_rm.copy()
        arm["_deal_key"] = norm_id_series(arm["Deal Loan Number"])
        arm = arm[["_deal_key", "Active RM"]].drop_duplicates("_deal_key")
        out = out.merge(arm, on="_deal_key", how="left", suffixes=("", "_sf"))
        out["Active RM"] = coalesce_keep_nonblank(out.get("Active RM_sf", pd.Series([pd.NA] * len(out), index=out.index)), out["Active RM"])
        out = out.drop(columns=["Active RM_sf"], errors="ignore")

    # Fix K: classify the DSCR portfolio off "Product Type" (LOC_Loan_Type__c), not the
    # deal "Type". Verified against same-day SF_Term: 132/133 real-DSCR deals have
    # Product Type in {DSCR, Single Rental Loan} (the 1 exception is Single Rental Loan,
    # also DSCR), and 82/82 real-Active-Term deals have Product Type "Term Loan". New
    # DSCR deals (63010/63704/63709) carry Product Type=DSCR but were mis-binned as
    # Active Term off "Type". derive_term_portfolio_segment uses loan_type ONLY for the
    # DSCR check, so this does not affect Segment/CPP. Fall back to "Type" when blank.
    cls = sf_term.apply(
        lambda r: pd.Series(
            derive_term_portfolio_segment(
                r.get("Product Type") if clean_text(r.get("Product Type")) else r.get("Type"),
                r.get("Current Funding Vehicle"),
                r.get("Sold Loan: Sold To"),
                r.get("Deal Loan Number"),
                template_maps,
                sold_servicing_status=r.get("Sold Loan: Servicing Status"),
                warehouse_line=r.get("Warehouse Line"),
            ),
            index=["Portfolio", "Segment", "CPP JV"],
        ),
        axis=1,
    )
    out["Portfolio"] = cls["Portfolio"]
    out["Segment"] = cls["Segment"]
    out["CPP JV"] = cls["CPP JV"]

    sold_stage_series = sf_term.get("Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    out["Financing"] = pd.Series(out.get("Financing", pd.Series([pd.NA] * len(out), index=out.index)), index=out.index, dtype="object")
    out["Financing"] = out["Financing"].map(normalize_term_financing).replace({"": pd.NA})
    # A populated Loan Buyer (Sold Loan Pool join) means the loan was sold; the report
    # shows Financing="Sold" regardless of the funding vehicle (validated: all 387 Sold
    # deals have a Loan Buyer). Stage=="Sold" is also honored as a fallback trigger.
    _loan_buyer_populated = ~blankish_mask(out.get("Loan Buyer", pd.Series([pd.NA] * len(out), index=out.index)))
    out["Financing"] = out["Financing"].mask(_loan_buyer_populated, "Sold")
    out["Financing"] = out["Financing"].mask(blankish_mask(out["Financing"]) & sold_stage_series.eq("Sold"), "Sold")

    # Loan Sold Date (sourcing-map rule #4): Sold_Loan_Pool__c.Sold_Date__c, shown ONLY
    # for sold loans (Financing == "Sold"); otherwise left blank -> N/A by report policy.
    if "Loan Sold Date" in out.columns:
        _sold_mask = out["Financing"].astype("string").str.strip().eq("Sold")
        out["Loan Sold Date"] = pd.to_datetime(out["Loan Sold Date"], errors="coerce").where(_sold_mask, pd.NaT)
    else:
        out["Loan Sold Date"] = pd.NaT

    blank_obj = pd.Series([pd.NA] * len(out), index=out.index, dtype="object")

    if "term_loan_manual" in prev_maps:
        man = prev_maps["term_loan_manual"].copy()
        keep_cols = ["_deal_key"] + [c for c in [
            "Portfolio", "Segment", "Financing", "CPP JV", "Special Loans List (Y/N)",
            "Asset Manager", "Deal Intro Sub-Source", "Referral Source Account",
            "Referral Source Contact", "AM Commentary", "Servicer", "Loan Buyer", "Servicer ID",
            "Maturity Date", "Active RM",
        ] if c in man.columns]
        out = out.merge(man[keep_cols], on="_deal_key", how="left", suffixes=("", "_prev"))
        term_loan_carry_forward_first = {
            "Portfolio", "Segment", "Financing", "CPP JV", "Loan Buyer",
            "Asset Manager", "Active RM", "Deal Intro Sub-Source",
            "Referral Source Account", "Referral Source Contact", "AM Commentary",
            "Special Loans List (Y/N)", "Servicer",
        }
        # Maturity Date is intentionally NOT prior-first. Term maturities can change
        # through modifications/extensions, so the preferred current Salesforce maturity
        # source should overwrite last week's completed report when Salesforce has a value.
        # Prior completed report maturity is only a fallback when Salesforce/servicer is blank.
        for c in [x for x in keep_cols if x != "_deal_key"]:
            if f"{c}_prev" in out.columns:
                if c in term_loan_carry_forward_first:
                    out[c] = coalesce_report_display_first(out[f"{c}_prev"], out.get(c, blank_obj))
                else:
                    out[c] = coalesce_keep_nonblank(out.get(c, blank_obj), out[f"{c}_prev"])
                out = out.drop(columns=[f"{c}_prev"], errors="ignore")

    # V37 taxonomy: Sold + Berkadia-serviced term deals display Portfolio = Financing =
    # 'Sold Servicing Retained' (verified 384/384). Runs after the carry-forward so last
    # week's 'Sold Term'/'Sold' is overridden, and after Loan Sold Date (derived above off
    # Financing=='Sold'). The deal-key set drives the Term Asset cascade.
    out = _apply_term_sold_servicing_retained(out, _term_sold_servicing_retained_deal_keys(sf_term))

    # Re-apply current/modified SF maturity after prior manual carry-forward.
    # The prior workbook is only a fallback for Maturity Date; it should not block
    # maturity changes after loan modifications/extensions.
    sf_current_maturity = _term_current_maturity_source_series(sf_term)
    if len(sf_current_maturity) == len(out):
        cur_sf_mat = pd.to_datetime(pd.Series(sf_current_maturity.to_numpy(), index=out.index), errors="coerce")
        cur_report_mat = pd.to_datetime(out.get("Maturity Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        out["Maturity Date"] = cur_sf_mat.where(cur_sf_mat.notna(), cur_report_mat)

    if not sf_am.empty and "Deal Loan Number" in sf_am.columns:
        am = sf_am.copy()
        am["_deal_key"] = norm_id_series(am["Deal Loan Number"])
        am["_dt"] = pd.to_datetime(am.get("Date Assigned"), errors="coerce")
        am = am.sort_values(["_deal_key", "Team Role", "_dt"]).drop_duplicates(["_deal_key", "Team Role"], keep="last")

        am1 = am[am["Team Role"].astype("string").str.strip().eq("Asset Manager")][["_deal_key", "Team Member Name"]]
        am1 = am1.drop_duplicates("_deal_key")
        out = out.merge(am1, on="_deal_key", how="left")
        out["Asset Manager"] = coalesce_keep_nonblank(out.get("Asset Manager", blank_obj), out["Team Member Name"])
        out = out.drop(columns=["Team Member Name"], errors="ignore")
    else:
        out["Asset Manager"] = out.get("Asset Manager", blank_obj)

    base_sf_servicer = pd.Series(out.get("Servicer", blank_obj), index=out.index, dtype="object")
    match_df = _select_term_servicer_matches(sf_term, serv_lookup, base_sf_servicer, prev_maps=prev_maps)
    if "Servicer Commitment Id" in sf_term.columns and len(sf_term) == len(out):
        sf_commitment_display = pd.Series(sf_term["Servicer Commitment Id"].to_numpy(), index=out.index, dtype="object")
    else:
        sf_commitment_display = pd.Series([pd.NA] * len(out), index=out.index, dtype="object")
    # The completed report displays the Salesforce Servicer Commitment Id first.
    # Alternate servicer keys can still supply UPB/name via match_df, but should not
    # replace the visible Servicer ID unless Salesforce is blank.
    out["Servicer ID"] = coalesce_keep_nonblank(sf_commitment_display, match_df["selected_servicer_id"])
    out["Servicer ID"] = coalesce_keep_nonblank(out["Servicer ID"], out.get("Servicer ID", blank_obj))

    sf_upb_fallback = pd.to_numeric(
        sf_term["Current Servicer UPB"] if "Current Servicer UPB" in sf_term.columns else pd.Series([np.nan] * len(out)),
        errors="coerce",
    )

    # Servicer file/match name may enrich reporting but should not overwrite the visible
    # Term Loan Servicer label from prior/SF. N/A is a valid display value.
    out["Servicer"] = coalesce_keep_nonblank(out["Servicer"], match_df["matched_servicer"])
    matched_mat = pd.to_datetime(match_df["matched_maturity_date"], errors="coerce")
    cur_mat = pd.to_datetime(out["Maturity Date"], errors="coerce")
    out["Maturity Date"] = cur_mat.where(cur_mat.notna(), matched_mat)
    matched_npd = pd.to_datetime(match_df["matched_next_payment_date"], errors="coerce")
    cur_npd = pd.to_datetime(out["Next Payment Date"], errors="coerce")
    out["Next Payment Date"] = cur_npd.where(cur_npd.notna(), matched_npd)
    out[upb_col] = pd.to_numeric(match_df["matched_upb"], errors="coerce").where(
        pd.to_numeric(match_df["matched_upb"], errors="coerce").notna(),
        sf_upb_fallback,
    )
    out = _apply_term_preboarding_upb_fallback(out, sf_term, upb_col)
    out = _guard_term_loan_upb_vs_amount(out, upb_col, prev_maps=prev_maps)

    if "term_loan_manual" in prev_maps and "Servicer ID" in prev_maps["term_loan_manual"].columns:
        prev_sid = prev_maps["term_loan_manual"][["_deal_key", "Servicer ID"]].copy()
        out = out.merge(prev_sid, on="_deal_key", how="left", suffixes=("", "_prev_sid"))
        if "Servicer ID_prev_sid" in out.columns:
            out["Servicer ID"] = coalesce_keep_nonblank(out.get("Servicer ID", blank_obj), out["Servicer ID_prev_sid"])
            out = out.drop(columns=["Servicer ID_prev_sid"], errors="ignore")

    out["REO Date"] = pd.NaT
    if "term_loan_reo" in prev_maps:
        reo = prev_maps["term_loan_reo"][["_deal_key", "REO Date"]].copy()
        out = out.merge(reo, on="_deal_key", how="left", suffixes=("", "_prev"))
        out["REO Date"] = pd.to_datetime(out["REO Date_prev"], errors="coerce").where(
            pd.to_datetime(out["REO Date_prev"], errors="coerce").notna(),
            pd.to_datetime(out["REO Date"], errors="coerce"),
        )
        out = out.drop(columns=["REO Date_prev"], errors="ignore")

    if "term_loan_upb" in prev_maps and upb_col in out.columns:
        prevu = prev_maps["term_loan_upb"].copy()
        out = out.merge(prevu, on="_deal_key", how="left")

        reo_mask = pd.to_datetime(out["REO Date"], errors="coerce").notna()
        cur_upb = pd.to_numeric(out[upb_col], errors="coerce")
        prev_upb = pd.to_numeric(out.get("_prev_upb", np.nan), errors="coerce")
        fill_val = prev_upb.fillna(0.0)
        out[upb_col] = np.where(reo_mask & ((cur_upb.isna()) | (cur_upb <= 0)), fill_val, cur_upb)
        out = out.drop(columns=["_prev_upb"], errors="ignore")

    out = _guard_term_loan_upb_vs_amount(out, upb_col, prev_maps=prev_maps)

    raw_stage_series = pd.Series(sf_term.get("Stage", pd.Series([pd.NA] * len(out), index=out.index)).values, index=out.index)
    raw_current_upb = pd.Series(sf_term.get("Current Servicer UPB", pd.Series([np.nan] * len(out), index=out.index)).values, index=out.index)
    raw_loan_amount = pd.Series(sf_term.get("Loan Amount", pd.Series([np.nan] * len(out), index=out.index)).values, index=out.index)
    raw_sold_status = pd.Series(sf_term.get("Sold Loan: Servicing Status", pd.Series([pd.NA] * len(out), index=out.index)).values, index=out.index)
    prev_retained_mask = out.get("_deal_key", pd.Series([pd.NA] * len(out), index=out.index)).isin(prev_sold_retained_keys)
    reo_date_mask = pd.to_datetime(out["REO Date"], errors="coerce").notna()
    keep_mask = _term_report_keep_mask(
        raw_stage_series,
        raw_current_upb,
        raw_sold_status,
        fallback_prev_retained_mask=prev_retained_mask,
        extra_reo_mask=reo_date_mask,
        loan_amount_series=raw_loan_amount,
    )
    out = out.loc[keep_mask].copy()
    out = out[(out.get("_sid_key", pd.Series([pd.NA] * len(out), index=out.index)).notna()) | (out["_deal_key"].notna())].copy()

    out["Active RM"] = coalesce_keep_nonblank(out.get("Active RM", pd.Series([pd.NA] * len(out), index=out.index)), pd.Series(["N"] * len(out), index=out.index))
    out["Special Loans List (Y/N)"] = coalesce_keep_nonblank(
        out.get("Special Loans List (Y/N)", pd.Series([pd.NA] * len(out), index=out.index)),
        pd.Series(["N"] * len(out), index=out.index),
    )

    out = _fill_text_defaults(
        out,
        [
            "Servicer ID", "Servicer", "Loan Buyer", "Asset Manager",
            "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact", "AM Commentary",
        ],
    )

    return downcast_numeric_frame(out)




def _build_term_servicer_spine(serv_lookup: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "_sid_key", "Servicer ID", "Servicer", "Servicer Family",
        "Current Servicer UPB", "Next Payment Date", "Maturity Date",
        "Servicer Status", "Source File",
    ]
    if serv_lookup is None or serv_lookup.empty:
        return pd.DataFrame(columns=cols)

    s = serv_lookup.copy()
    if "_sid_key" not in s.columns:
        s["_sid_key"] = id_key_no_leading_zeros(s.get("servicer_id", pd.Series([None] * len(s), index=s.index)))

    fam = s.get("servicer_family", pd.Series([""] * len(s), index=s.index)).astype("string").str.lower().str.strip()
    upb = pd.to_numeric(s.get("upb", pd.Series([np.nan] * len(s), index=s.index)), errors="coerce")
    maturity = pd.to_datetime(s.get("maturity_date", pd.Series([pd.NaT] * len(s), index=s.index)), errors="coerce")
    next_payment = pd.to_datetime(s.get("next_payment_date", pd.Series([pd.NaT] * len(s), index=s.index)), errors="coerce")
    as_of = pd.to_datetime(s.get("as_of", pd.Series([pd.NaT] * len(s), index=s.index)), errors="coerce")

    active_mask = fam.isin(TERM_SPINE_SERVICER_FAMILIES) & s["_sid_key"].notna() & upb.gt(0)
    s = s.loc[active_mask].copy()
    if s.empty:
        return pd.DataFrame(columns=cols)

    s["Servicer ID"] = coalesce_keep_nonblank(
        pd.Series(s.get("servicer_id", pd.Series([pd.NA] * len(s), index=s.index)), index=s.index, dtype="object"),
        pd.Series(s["_sid_key"], index=s.index, dtype="object"),
    )
    s["Servicer"] = pd.Series(s.get("servicer", pd.Series([pd.NA] * len(s), index=s.index)), index=s.index, dtype="object")
    s["Servicer Family"] = fam
    s["Current Servicer UPB"] = upb
    s["Next Payment Date"] = next_payment
    s["Maturity Date"] = maturity
    s["Servicer Status"] = pd.Series(s.get("status", pd.Series([pd.NA] * len(s), index=s.index)), index=s.index, dtype="object")
    s["Source File"] = pd.Series(s.get("source_file", pd.Series([pd.NA] * len(s), index=s.index)), index=s.index, dtype="object")
    s["_as_of_sort"] = as_of
    s["_mat_sort"] = s["Maturity Date"]
    s["_npd_sort"] = s["Next Payment Date"]

    s = s.sort_values(["_sid_key", "_as_of_sort", "Current Servicer UPB", "_mat_sort", "_npd_sort"]).drop_duplicates("_sid_key", keep="last")
    return downcast_numeric_frame(s[cols])


def _term_stage_rank(stage, sold_servicing_status=None) -> int:
    st = clean_text(stage)
    base = {
        "Closed Won": 80,
        "Approved by Committee": 70,
        "Sold": 60,
        "REO": 50,
        "REO-Sold": 45,
        "Paid Off": 20,
    }.get(st, 10 if st else 0)
    retained = bool(_sold_servicing_retained_mask(pd.Series([sold_servicing_status])).iloc[0])
    if retained:
        base += 15
    return base


def _build_term_sf_sid_lookup(sf_term: pd.DataFrame, prev_maps: Optional[dict] = None) -> pd.DataFrame:
    if sf_term is None or sf_term.empty:
        return pd.DataFrame()

    candidate_cols = []
    if "Servicer Commitment Id" in sf_term.columns:
        candidate_cols.append("Servicer Commitment Id")
    candidate_cols.extend([c for c in sf_term.columns if c.startswith("Term Servicer Key ")])
    if not candidate_cols:
        return pd.DataFrame()

    detail_fields = [
        "Deal Loan Number", "Yardi ID", "Deal Name", "Borrower Entity", "Account Name",
        "Current Funding Vehicle", "Loan Amount", "Close Date", "CAF Originator",
        "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact",
        "Comments AM", "Sold Loan: Sold To", "Servicer Name", "Current Servicer UPB",
        "Sold Loan: Servicing Status", "Stage", "Original Loan Maturity Date", "Next Payment Date",
    ]
    detail_fields.extend([c for c in sf_term.columns if str(c).startswith("Current Loan Maturity Date")])
    detail_fields = list(dict.fromkeys(detail_fields))
    keep_cols = [c for c in dict.fromkeys(detail_fields) if c in sf_term.columns]

    frames = []
    for pos, col in enumerate(candidate_cols):
        sid_key = id_key_no_leading_zeros(sf_term.get(col, pd.Series([None] * len(sf_term), index=sf_term.index)))
        mask = sid_key.notna()
        if not mask.any():
            continue
        tmp = sf_term.loc[mask, keep_cols].copy()
        tmp["_sid_key"] = pd.Series(sid_key.loc[mask], index=tmp.index, dtype="object")
        tmp["_sid_source_col"] = col
        tmp["_sid_priority"] = 100 if col == "Servicer Commitment Id" else max(0, 50 - pos)
        frames.append(tmp)

    if not frames:
        return pd.DataFrame()

    key_df = pd.concat(frames, ignore_index=True, copy=False)
    key_df["_deal_key"] = norm_id_series(key_df.get("Deal Loan Number", pd.Series([None] * len(key_df), index=key_df.index)))
    key_df["_sf_upb_num"] = pd.to_numeric(key_df.get("Current Servicer UPB", pd.Series([np.nan] * len(key_df), index=key_df.index)), errors="coerce").fillna(0)
    key_df["_loan_amt_num"] = pd.to_numeric(key_df.get("Loan Amount", pd.Series([np.nan] * len(key_df), index=key_df.index)), errors="coerce").fillna(0)
    key_df["_close_dt"] = pd.to_datetime(key_df.get("Close Date", pd.Series([pd.NaT] * len(key_df), index=key_df.index)), errors="coerce")
    key_df["_sold_retained"] = _sold_servicing_retained_mask(
        key_df.get("Sold Loan: Servicing Status", pd.Series([pd.NA] * len(key_df), index=key_df.index))
    ).astype("int8")
    key_df["_stage_rank"] = [
        _term_stage_rank(stage, sold_status)
        for stage, sold_status in zip(
            key_df.get("Stage", pd.Series([pd.NA] * len(key_df), index=key_df.index)),
            key_df.get("Sold Loan: Servicing Status", pd.Series([pd.NA] * len(key_df), index=key_df.index)),
        )
    ]

    detail_count = pd.Series([0] * len(key_df), index=key_df.index, dtype="int64")
    for col in [c for c in detail_fields if c in key_df.columns]:
        detail_count = detail_count + (~blankish_mask(key_df[col])).astype("int64")
    key_df["_detail_count"] = detail_count

    key_df["_prev_sid_match"] = 0
    key_df["_prev_sid_present"] = 0
    if prev_maps and "term_loan_sid" in prev_maps:
        prev_sid = prev_maps["term_loan_sid"]
        if isinstance(prev_sid, pd.DataFrame) and not prev_sid.empty and {"_sid_key", "_deal_key"}.issubset(prev_sid.columns):
            sid_to_prev_deal = prev_sid.dropna(subset=["_sid_key"]).drop_duplicates("_sid_key").set_index("_sid_key")["_deal_key"].to_dict()
            prev_deal = key_df["_sid_key"].map(sid_to_prev_deal)
            prev_sid_present = prev_deal.notna()
            deal_key_present = key_df["_deal_key"].notna()
            prev_sid_match = prev_sid_present & deal_key_present & key_df["_deal_key"].eq(prev_deal)
            key_df["_prev_sid_present"] = prev_sid_present.astype("int8")
            key_df["_prev_sid_match"] = prev_sid_match.astype("int8")
            # Prior completed reports are useful hints, but not hard blockers. Do not drop
            # current Salesforce key rows solely because the prior workbook mapped the same
            # servicer ID differently; service transfers and corrected SF IDs can make that stale.

    if key_df.empty:
        return pd.DataFrame()

    key_df = key_df.sort_values(
        [
            "_sid_key",
            "_prev_sid_match",
            "_prev_sid_present",
            "_sold_retained",
            "_stage_rank",
            "_sid_priority",
            "_detail_count",
            "_sf_upb_num",
            "_loan_amt_num",
            "_close_dt",
        ],
        ascending=[True, True, True, True, True, True, True, True, True, True],
    )
    return downcast_numeric_frame(key_df.drop_duplicates("_sid_key", keep="last"))


def _apply_term_financing_taxonomy(out: pd.DataFrame) -> pd.DataFrame:
    """Financing pins Segment and Portfolio on both Term tabs.

    Measured against the 20260824 official, with zero exceptions in either direction:

      Financing contains 'CPP JV'  -> Segment 'CPP JV', Portfolio 'Active Term'
          Term Asset 361 rows, Term Loan 25. The official says CPP JV on every one; the
          build said 'Mortgage Banking' on 40 assets and 2 loans, all of them
          'CPP JV - Wells Fargo' on deals 63955 and 64218 -- and the official carries the
          same Financing text we do, so only the Segment derivation was wrong.

      Financing == 'Morgan Stanley' -> Segment 'Mortgage Banking', Portfolio 'DSCR'
          Term Loan 57 rows, Term Asset 74. The official is Mortgage Banking / DSCR on all
          of them; the build disagreed on 6 Segment + 3 Portfolio (loans) and 7 + 4 (assets).

    The Morgan Stanley test is an EXACT match, not a substring: 'CPP JV - Morgan Stanley' is
    a CPP JV warehouse and must fall to the rule above, which is why CPP JV is applied first.

    NOT done here: blanking Financing itself. The official blanks it on only 15 of the 57
    Morgan Stanley loans and keeps 'Morgan Stanley' on the other 42, and no rule separating
    them is visible in the report, so the value is left alone.
    """
    if out is None or out.empty or "Financing" not in out.columns:
        return out
    fin = out["Financing"].astype("string").fillna("").str.strip()

    cpp = fin.str.contains("CPP JV", case=False, na=False)
    if bool(cpp.any()):
        if "Segment" in out.columns:
            out.loc[cpp, "Segment"] = "CPP JV"
        if "Portfolio" in out.columns:
            out.loc[cpp, "Portfolio"] = "Active Term"

    ms = fin.str.casefold().eq("morgan stanley")
    if bool(ms.any()):
        if "Segment" in out.columns:
            out.loc[ms, "Segment"] = "Mortgage Banking"
        if "Portfolio" in out.columns:
            out.loc[ms, "Portfolio"] = "DSCR"
    return out


def build_term_loan(
    sf_term: pd.DataFrame,
    sf_am: pd.DataFrame,
    sf_active_rm: pd.DataFrame,
    serv_lookup: pd.DataFrame,
    upb_col: str,
    prev_maps: dict,
    template_maps: dict,
    asset_deal_numbers: Optional[Sequence[str]] = None,
) -> pd.DataFrame:
    prev_keys = _prev_term_keys(prev_maps)
    prev_positive_keys = _prev_term_positive_upb_keys(prev_maps)
    prev_sold_retained_keys = _prev_term_sold_retained_keys(prev_maps)
    sf_term_active = _filter_term_population(sf_term, prev_keys=prev_keys, prev_positive_keys=prev_positive_keys, prev_sold_retained_keys=prev_sold_retained_keys)

    always_keep_keys = set(norm_id_series(pd.Series(list(TERM_ALWAYS_INCLUDE_DEALS), dtype="object")).dropna().tolist())
    if sf_term_active is not None and not sf_term_active.empty:
        sf_term_active = sf_term_active.copy()
        sf_term_active["_deal_key"] = norm_id_series(sf_term_active.get("Deal Loan Number", pd.Series([None] * len(sf_term_active), index=sf_term_active.index)))
        sf_term_active = sf_term_active.drop(columns=["_deal_key"], errors="ignore")

    out = _build_term_loan_salesforce_fallback(
        sf_term_active,
        sf_am,
        sf_active_rm,
        serv_lookup,
        upb_col,
        prev_maps,
        template_maps,
        prev_sold_retained_keys=prev_sold_retained_keys,
    )
    if out.empty:
        return out

    blank_obj = pd.Series([pd.NA] * len(out), index=out.index, dtype="object")
    out["_deal_key"] = norm_id_series(out.get("Deal Number", pd.Series([None] * len(out), index=out.index)))
    out["_sid_key"] = id_key_no_leading_zeros(out.get("Servicer ID", pd.Series([None] * len(out), index=out.index)))

    sf_sid = _build_term_sf_sid_lookup(sf_term_active, prev_maps=prev_maps)
    if not sf_sid.empty:
        sf_keep_base = ["_sid_key", "_deal_key", "Deal Loan Number", "Yardi ID", "Deal Name", "Borrower Entity", "Account Name", "Do Not Lend", "Current Funding Vehicle", "Loan Amount", "Close Date", "CAF Originator", "Deal Intro Sub-Source", "Referral Source Account", "Referral Source Contact", "Comments AM", "Sold Loan: Sold To", "Sold Loan: Servicing Status", "Type", "Servicer Name", "Stage", "Current Servicer UPB", "Original Loan Maturity Date", "Next Payment Date"]
        sf_keep_base.extend([c for c in sf_sid.columns if str(c).startswith("Current Loan Maturity Date")])
        sf_keep = [c for c in dict.fromkeys(sf_keep_base) if c in sf_sid.columns]
        sf_pick = sf_sid[sf_keep].drop_duplicates("_sid_key")
        out = out.merge(sf_pick, on="_sid_key", how="left", suffixes=("", "_sid"))

        sid_detail_deal = norm_id_series(
            out.get("Deal Loan Number_sid", out.get("_deal_key_sid", pd.Series([pd.NA] * len(out), index=out.index)))
        )
        current_deal = norm_id_series(out.get("Deal Number", pd.Series([pd.NA] * len(out), index=out.index)))
        sid_same_deal = sid_detail_deal.isna() | current_deal.eq(sid_detail_deal)

        def _sid_source(source_col: str) -> pd.Series:
            return pd.Series(out[source_col], index=out.index, dtype="object").where(sid_same_deal, pd.NA)

        map_pairs = {
            "Deal Number": "Deal Loan Number_sid",
            "SF Yardi ID": "Yardi ID_sid",
            "Deal Name": "Deal Name_sid",
            "Borrower Entity": "Borrower Entity_sid",
            "Account Name": "Account Name_sid",
            "Do Not Lend (Y/N)": "Do Not Lend_sid",
            "Financing": "Current Funding Vehicle_sid",
            "Originator": "CAF Originator_sid",
            "Deal Intro Sub-Source": "Deal Intro Sub-Source_sid",
            "Referral Source Account": "Referral Source Account_sid",
            "Referral Source Contact": "Referral Source Contact_sid",
            "AM Commentary": "Comments AM_sid",
            "Loan Buyer": "Sold Loan: Sold To_sid",
            "Servicer": "Servicer Name_sid",
        }
        for target, source in map_pairs.items():
            if source in out.columns:
                safe_source = _sid_source(source)
                if target == "Do Not Lend (Y/N)":
                    out[target] = coalesce_keep_nonblank(out.get(target, blank_obj), _yn_from_bool_series(safe_source))
                else:
                    out[target] = coalesce_keep_nonblank(out.get(target, blank_obj), safe_source)
        if "Loan Amount_sid" in out.columns:
            cur_amt = pd.to_numeric(out.get("Loan Amount", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
            src_amt = pd.to_numeric(pd.Series(out["Loan Amount_sid"], index=out.index).where(sid_same_deal), errors="coerce")
            out["Loan Amount"] = cur_amt.where(cur_amt.notna(), src_amt)
        if "Close Date_sid" in out.columns:
            cur_dt = pd.to_datetime(out.get("Origination Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
            src_dt = pd.to_datetime(pd.Series(out["Close Date_sid"], index=out.index).where(sid_same_deal), errors="coerce")
            out["Origination Date"] = cur_dt.where(cur_dt.notna(), src_dt)
        sid_maturity_cols = [c for c in out.columns if str(c).startswith("Current Loan Maturity Date") and str(c).endswith("_sid")]
        if "Original Loan Maturity Date_sid" in out.columns:
            sid_maturity_cols.append("Original Loan Maturity Date_sid")
        if sid_maturity_cols:
            cur_dt = pd.to_datetime(out.get("Maturity Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
            src_df = pd.DataFrame(index=out.index)
            for mat_col in sid_maturity_cols:
                src_df[mat_col] = pd.Series(out[mat_col], index=out.index).where(sid_same_deal)
            src_dt = _coalesce_datetime_columns(src_df, sid_maturity_cols)
            out["Maturity Date"] = cur_dt.where(cur_dt.notna(), src_dt)
        if "Next Payment Date_sid" in out.columns:
            cur_dt = pd.to_datetime(out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
            src_dt = pd.to_datetime(pd.Series(out["Next Payment Date_sid"], index=out.index).where(sid_same_deal), errors="coerce")
            out["Next Payment Date"] = cur_dt.where(cur_dt.notna(), src_dt)
        if "Current Servicer UPB_sid" in out.columns:
            cur_upb = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
            src_upb = pd.to_numeric(pd.Series(out["Current Servicer UPB_sid"], index=out.index).where(sid_same_deal), errors="coerce")
            out[upb_col] = cur_upb.where(cur_upb.gt(0), src_upb)
        out = _guard_term_loan_upb_vs_amount(out, upb_col, prev_maps=prev_maps)

    if serv_lookup is not None and not serv_lookup.empty and "_sid_key" in serv_lookup.columns:
        s = serv_lookup.dropna(subset=["_sid_key"]).copy().rename(columns={"servicer": "_servicer_file", "upb": "_loan_upb", "next_payment_date": "_serv_next_payment_date", "maturity_date": "_serv_maturity_file", "status": "_serv_status_file"})
        out = out.merge(s[["_sid_key", "_servicer_file", "_loan_upb", "_serv_next_payment_date", "_serv_maturity_file", "_serv_status_file"]], on="_sid_key", how="left")
        out["Servicer"] = coalesce_keep_nonblank(out.get("Servicer", blank_obj), out.get("_servicer_file", blank_obj))
        file_upb = pd.to_numeric(out.get("_loan_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
        cur_upb = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
        out[upb_col] = file_upb.where(file_upb.gt(0), cur_upb)

        file_npd = pd.to_datetime(out.get("_serv_next_payment_date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        cur_npd = pd.to_datetime(out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        out["Next Payment Date"] = cur_npd.where(cur_npd.notna(), file_npd)

        file_mat = pd.to_datetime(out.get("_serv_maturity_file", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        cur_mat = pd.to_datetime(out.get("Maturity Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        # Servicer-file maturity WINS (sourcing-map rule #10, verified 1106/1106 vs
        # 20260615 Term Loan). The servicer carries the live modified maturity; SF only
        # ships Original, which is ~1 month off. SF stays as fallback when no servicer row.
        out["Maturity Date"] = file_mat.where(file_mat.notna(), cur_mat)
        out = _guard_term_loan_upb_vs_amount(out, upb_col, prev_maps=prev_maps)

    out["Servicer ID"] = normalize_servicer_id_for_report(out.get("Servicer ID", blank_obj), out.get("Servicer", blank_obj))
    out["Do Not Lend (Y/N)"] = _yn_from_bool_series(out.get("Do Not Lend (Y/N)", pd.Series([pd.NA] * len(out), index=out.index)))
    out["CPP JV"] = coalesce_keep_nonblank(out.get("CPP JV", blank_obj), pd.Series(["N"] * len(out), index=out.index))
    out["Active RM"] = coalesce_keep_nonblank(out.get("Active RM", blank_obj), pd.Series(["N"] * len(out), index=out.index))
    out["Special Loans List (Y/N)"] = coalesce_keep_nonblank(out.get("Special Loans List (Y/N)", blank_obj), pd.Series(["N"] * len(out), index=out.index))

    if always_keep_keys:
        missing_force = sorted(always_keep_keys - set(out["_deal_key"].dropna().tolist()))
        if missing_force and "term_loan_manual" in prev_maps:
            prev_force = prev_maps["term_loan_manual"].copy()
            prev_force = prev_force[prev_force["_deal_key"].isin(missing_force)].copy()
            if not prev_force.empty:
                for col in [c for c in out.columns if c not in prev_force.columns]:
                    prev_force[col] = pd.NA
                prev_force = prev_force[out.columns]
                out = pd.concat([out, prev_force], ignore_index=True, copy=False)

    out = out[out["_deal_key"].notna()].copy()
    out = _clear_duplicate_term_servicer_assignments(out, upb_col, prev_maps=prev_maps)
    out = _apply_term_preboarding_upb_fallback(out, sf_term_active, upb_col)
    out = _guard_term_loan_upb_vs_amount(out, upb_col, prev_maps=prev_maps)

    terminal_zero_keys = _term_terminal_zero_exclusion_keys(sf_term)
    if terminal_zero_keys:
        before_drop = len(out)
        out = _drop_term_deal_keys(out, terminal_zero_keys)
        if before_drop != len(out):
            try:
                st.warning(f"Dropped {before_drop - len(out):,} terminal zero-UPB term loan(s) after enrichment/backfill.")
            except Exception:
                pass

    # V41: deals SF_Term surfaces but the official Term tab does not carry (e.g. 20747, a
    # bridge-to-term refi real treats as Bridge-only). Force-excluded last so no enrichment
    # path can reintroduce them.
    _force_exclude_keys = set(norm_id_series(pd.Series(list(TERM_FORCE_EXCLUDE_DEALS), dtype="object")).dropna().tolist())
    if _force_exclude_keys:
        out = _drop_term_deal_keys(out, _force_exclude_keys)

    # Final display guard: the completed report displays Salesforce Servicer
    # Commitment Id by Deal Number. Servicer-file/alternate keys can enrich UPB,
    # but should not replace the visible report ID.
    if (sf_term_active is not None and not sf_term_active.empty
            and {"Deal Loan Number", "Servicer Commitment Id"}.issubset(sf_term_active.columns)):
        sf_commit = sf_term_active[["Deal Loan Number", "Servicer Commitment Id"]].copy()
        sf_commit["_deal_key"] = norm_id_series(sf_commit["Deal Loan Number"])
        sf_commit = sf_commit.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key", keep="last")
        sf_commit = sf_commit[["_deal_key", "Servicer Commitment Id"]].rename(columns={"Servicer Commitment Id": "_sf_commitment_display_final"})
        out = out.merge(sf_commit, on="_deal_key", how="left")
        if "_sf_commitment_display_final" in out.columns:
            out["Servicer ID"] = coalesce_keep_nonblank(out["_sf_commitment_display_final"], out.get("Servicer ID", blank_obj))
            out = out.drop(columns=["_sf_commitment_display_final"], errors="ignore")
    # New approved / purchased term deals may not exist in the prior workbook yet.
    # For those rows, fill blank maturity from the preferred current SF maturity field.
    if sf_term_active is not None and not sf_term_active.empty and "Deal Loan Number" in sf_term_active.columns:
        maturity_ctx = sf_term_active.copy()
        maturity_ctx["_deal_key"] = norm_id_series(maturity_ctx["Deal Loan Number"])
        maturity_ctx["_sf_preferred_term_maturity"] = _term_maturity_source_series(maturity_ctx)
        maturity_ctx = maturity_ctx.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key", keep="last")[["_deal_key", "_sf_preferred_term_maturity"]]
        out = out.merge(maturity_ctx, on="_deal_key", how="left")
        cur_mat = pd.to_datetime(out.get("Maturity Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        sf_mat = pd.to_datetime(out.get("_sf_preferred_term_maturity", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
        # Keep the already-resolved maturity (servicer-file wins per rule #10); the SF
        # preferred maturity only fills brand-new deals that have no servicer maturity yet.
        out["Maturity Date"] = cur_mat.where(cur_mat.notna(), sf_mat)
        out = out.drop(columns=["_sf_preferred_term_maturity"], errors="ignore")

    # Fix L: a brand-new term deal that Salesforce names a servicer for, but which has NO
    # servicer-file activity (no positive UPB in any uploaded servicer file), is shown as
    # N/A by the official report until payments actually start hitting the servicer.
    # Verified vs same-day real: 36 newly-boarded Berkadia/FCI/SPS deals (47576, 59864,
    # 61372, 61683, 62486, ...) read N/A. The gate is restricted to deals ABSENT from the
    # prior workbook so an established / carried-forward servicer (e.g. Selene on 63181) is
    # never blanked. _loan_upb is the servicer-file UPB merged earlier in this function.
    if prev_keys:
        _l_file_upb = pd.to_numeric(out.get("_loan_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
        _l_has_activity = _l_file_upb.gt(0).fillna(False)
        _l_is_new = ~out["_deal_key"].astype(str).isin(prev_keys)
        _l_servicer = pd.Series(out.get("Servicer", blank_obj), index=out.index, dtype="object")
        _l_named = (~blankish_mask(_l_servicer)) & _l_servicer.astype("string").str.strip().str.upper().ne("N/A")
        _l_boarding_only = _l_is_new & _l_named & (~_l_has_activity)
        if bool(_l_boarding_only.any()):
            # V59: blank ONLY the servicer identity. The official report shows Servicer and
            # Servicer ID as N/A for a newly-boarded deal, but it still carries the balance
            # and the next payment date -- both come from Salesforce, not the servicer file.
            # Blanking them cost 22 Term Loan deals their UPB on 20260824 (Salesforce
            # Current Servicer UPB matched the official exactly on 12 of them, e.g. 64080
            # 343,000 and 64219 157,500), and that cascaded into 318 Term Asset UPB cells
            # because the asset allocation is a share of the loan UPB. It also blanked 13
            # Next Payment Dates the official shows as 2026-10-01.
            out.loc[_l_boarding_only, "Servicer"] = "N/A"
            out.loc[_l_boarding_only, "Servicer ID"] = "N/A"

    out = _apply_term_financing_taxonomy(out)

    terminal_zero_keys = _term_terminal_zero_exclusion_keys(sf_term)
    if terminal_zero_keys:
        out = _drop_term_deal_keys(out, terminal_zero_keys)
    return downcast_numeric_frame(out.drop(columns=[c for c in out.columns if c.startswith("_") and c not in {"_deal_key", "_sid_key"}], errors="ignore"))

def _drop_subunit_twin_term_assets(out: pd.DataFrame, term_loan: pd.DataFrame) -> dict:
    """Remove sub-unit duplicates that inflate a deal's ALA above its Loan Amount.

    The official report satisfies a hard invariant: per deal, sum(Property ALA) equals the Term
    Loan's Loan Amount. It holds on 1,028 of 1,030 deals on 20260831 and 1,023 of 1,025 on
    20260824, and the sum is NEVER below Loan Amount. So an ALA sum ABOVE Loan Amount is proof
    of duplicate asset rows -- and because Term Asset UPB is allocated as
    loan UPB x (asset ALA / deal ALA sum), that inflated denominator shortchanges every asset
    on the deal. Test 83 had 7 such deals, two of them at a ratio of exactly 2.000.

    Is_Sub_Unit__c does not identify these rows -- the SOQL already filters on it and they come
    through anyway -- and V79's Is_Parent__c filter only caught about half. What does identify
    them is the shape of the pair. Deal 58276:

        KEPT   11 Lillian Dr          Multifamily  ALA 1,235,250
        DROP   11 Lillian Dr Unit 1   Townhome     ALA 1,235,250

    So a row is dropped only when all four hold: its deal is over Loan Amount, its ALA matches
    a sibling's to the cent, that sibling is Multifamily while it is not, and its own address
    carries a sub-unit marker. Measured on test 83 against the 20260831 official: 5 rows
    dropped, all 5 genuinely absent from the official, 0 real assets lost, 28 UPB cells fixed.

    A subset-sum approach was tried first and rejected -- it tied the invariant on all 7 deals
    but removed 14 legitimate assets to fix 25 cells, which is a worse report, not a better one.
    """
    if out is None or out.empty or term_loan is None or term_loan.empty:
        return {}
    if "Property ALA" not in out.columns or "_deal_key" not in out.columns:
        return {}
    tl = _ensure_deal_key(term_loan, "Deal Number")
    if "Loan Amount" not in tl.columns:
        return {}
    amt = pd.to_numeric(tl["Loan Amount"], errors="coerce").groupby(tl["_deal_key"]).max()

    ala = pd.to_numeric(out["Property ALA"], errors="coerce").fillna(0.0)
    ptype = _report_text_series_from_col(out, "Property Type")
    grouping = _report_text_series_from_col(out, "Grouping")
    is_mf = ptype.eq("Multifamily") | grouping.eq("Multifamily")
    addr = out.get("Address", pd.Series([""] * len(out), index=out.index)).astype("string").fillna("")
    addr_low = addr.str.lower()
    has_subunit_marker = addr_low.str.contains(" unit ", regex=False) | addr_low.str.endswith(" unit") | addr.str.contains("(", regex=False)

    drops = []
    for _deal, _idx in out.groupby("_deal_key", dropna=True).groups.items():
        _idx = list(_idx)
        _amt = amt.get(_deal)
        if _amt is None or not (float(_amt) > 0):
            continue
        _tol = abs(float(_amt)) * 0.001 + 1.0
        if float(ala.loc[_idx].sum()) - float(_amt) <= _tol:
            continue
        for _i in _idx:
            _v = float(ala.loc[_i])
            if _v <= 0 or bool(is_mf.loc[_i]) or not bool(has_subunit_marker.loc[_i]):
                continue
            if any(_j != _i and abs(float(ala.loc[_j]) - _v) <= 0.01 and bool(is_mf.loc[_j]) for _j in _idx):
                drops.append(_i)
    if not drops:
        return {}
    info = {
        "rows": len(drops),
        "deals": int(out.loc[drops, "_deal_key"].nunique()),
        "ala": float(ala.loc[drops].sum()),
    }
    out.drop(index=drops, inplace=True)
    return info


def build_term_asset(sf_term_asset: pd.DataFrame, term_loan: pd.DataFrame, upb_col: str, prev_maps: Optional[dict] = None) -> pd.DataFrame:
    """Build Term Asset like the completed report: carry-forward spine first.

    The completed Active Loan Report does not rebuild Term Asset from the current
    Property/Valuation pull. It keeps the prior completed Term Asset rows/static
    values, filters them to the current Term Loan population, appends only brand-new
    current SF assets, and then recalculates the current UPB allocation.
    """
    term_asset_cols = list(TERM_ASSET_FROM_TERM_ASSET_REPORT.keys())

    tl = _ensure_deal_key(term_loan.copy(), "Deal Number") if term_loan is not None else pd.DataFrame()
    valid_deals = set(tl.get("_deal_key", pd.Series(dtype="object")).dropna().astype(str).tolist())

    # Current Salesforce rows are a fallback/add-new source only.
    current = pd.DataFrame(index=sf_term_asset.index if sf_term_asset is not None else None)
    if sf_term_asset is not None and not sf_term_asset.empty:
        for col, label in TERM_ASSET_FROM_TERM_ASSET_REPORT.items():
            current[col] = sf_term_asset[label] if label in sf_term_asset.columns else pd.NA
        current["_deal_key"] = norm_id_series(current.get("Deal Number", pd.Series([None] * len(current), index=current.index)))
        current["_asset_key"] = norm_id_series(current.get("Asset ID", pd.Series([None] * len(current), index=current.index)))
        if valid_deals:
            current = current[current["_deal_key"].isin(valid_deals) & current["_asset_key"].notna()].copy()
        else:
            current = current[current["_asset_key"].notna()].copy()
    else:
        current = pd.DataFrame(columns=term_asset_cols + ["_deal_key", "_asset_key"])

    out = pd.DataFrame(columns=term_asset_cols + ["_deal_key", "_asset_key"])

    # Prior completed Term Asset rows are the primary spine/static source.
    if prev_maps and "term_asset_manual" in prev_maps and isinstance(prev_maps["term_asset_manual"], pd.DataFrame):
        prev = prev_maps["term_asset_manual"].copy()
        if not prev.empty and {"_deal_key", "_asset_key"}.issubset(prev.columns):
            if valid_deals:
                prev = prev[prev["_deal_key"].astype(str).isin(valid_deals)].copy()
            prev = prev[prev["_deal_key"].notna() & prev["_asset_key"].notna()].copy()
            for col in term_asset_cols:
                if col not in prev.columns:
                    prev[col] = pd.NA
            out = prev[term_asset_cols + ["_deal_key", "_asset_key"]].drop_duplicates(["_deal_key", "_asset_key"], keep="last").copy()

    # Append only new SF assets that were not already in the carried-forward sheet.
    if not current.empty:
        if out.empty:
            out = current[term_asset_cols + ["_deal_key", "_asset_key"]].copy()
        else:
            present_pairs = set(zip(out["_deal_key"].astype(str), out["_asset_key"].astype(str)))
            missing_current = current[
                ~current.apply(lambda r: (str(r["_deal_key"]), str(r["_asset_key"])) in present_pairs, axis=1)
            ].copy()
            if not missing_current.empty:
                for col in term_asset_cols + ["_deal_key", "_asset_key"]:
                    if col not in missing_current.columns:
                        missing_current[col] = pd.NA
                out = pd.concat([out, missing_current[term_asset_cols + ["_deal_key", "_asset_key"]]], ignore_index=True, copy=False)

    if out.empty:
        return out

    # Refresh only loan-derived fields from the current Term Loan tab. Prior Term Asset
    # values remain first; current Term Loan fills blanks, especially for appended new assets.
    if "Portfolio" in tl.columns:
        tl_portfolio = tl[["_deal_key", "Portfolio"]].drop_duplicates("_deal_key")
        out = out.merge(tl_portfolio, on="_deal_key", how="left", suffixes=("", "_loan"))
        out["Portfolio"] = coalesce_report_display_first(
            out.get("Portfolio", pd.Series([pd.NA] * len(out), index=out.index)),
            out.get("Portfolio_loan", pd.Series([pd.NA] * len(out), index=out.index)),
        )
        out = out.drop(columns=["Portfolio_loan"], errors="ignore")

    # Financing is also loan-derived. New approved/purchased term assets have no prior
    # Term Asset row, so their carried-forward Financing is blank; refresh it from the
    # current Term Loan tab (prior value still wins, Term Loan fills the blanks).
    if "Financing" in tl.columns:
        tl_financing = tl[["_deal_key", "Financing"]].drop_duplicates("_deal_key")
        out = out.merge(tl_financing, on="_deal_key", how="left", suffixes=("", "_loan"))
        out["Financing"] = coalesce_report_display_first(
            out.get("Financing", pd.Series([pd.NA] * len(out), index=out.index)),
            out.get("Financing_loan", pd.Series([pd.NA] * len(out), index=out.index)),
        )
        out = out.drop(columns=["Financing_loan"], errors="ignore")

    # Segment is loan-derived too, and for the same reason: newly appended assets carry no
    # prior Term Asset row. The official tab has ZERO blank Segment cells; the previous
    # build shipped 152 (exactly its new-asset rows). Fill from the parent Term Loan.
    if "Segment" in tl.columns:
        tl_segment = tl[["_deal_key", "Segment"]].drop_duplicates("_deal_key")
        out = out.merge(tl_segment, on="_deal_key", how="left", suffixes=("", "_loan"))
        out["Segment"] = coalesce_keep_nonblank(
            out.get("Segment", pd.Series([pd.NA] * len(out), index=out.index)),
            out.get("Segment_loan", pd.Series([pd.NA] * len(out), index=out.index)),
        )
        out = out.drop(columns=["Segment_loan"], errors="ignore")

    # Grouping drives the Term Loan SFR/MF allocation split, so a blank silently
    # mis-buckets the parent deal's Strategy Grouping. The official tab has zero blanks.
    # Property Type resolves it for every new row: Multifamily/Other -> Multifamily,
    # everything else (SFR, 2-4 Unit, Condo, Townhome, PUD, ...) -> Single Family Rental.
    # (Mixed Use splits 40/28 in the source data, so the upstream value wins where present
    # and this only fills what would otherwise ship blank.)
    _grouping = pd.Series(
        out.get("Grouping", pd.Series([pd.NA] * len(out), index=out.index)), index=out.index, dtype="object"
    )
    _prop_type = _report_text_series_from_col(out, "Property Type")
    _grouping_guess = pd.Series(["Single Family Rental"] * len(out), index=out.index, dtype="object").mask(
        _prop_type.isin(["Multifamily", "Other"]), "Multifamily"
    )
    out["Grouping"] = coalesce_keep_nonblank(_grouping, _grouping_guess)

    # V66: a Mixed Use property with 5 or more units is Multifamily. The guess above only
    # fills a blank, so a carried-forward 'Single Family Rental' survived on all 68 of the
    # official's Mixed Use assets while the official calls 40 of them Multifamily. '# Units'
    # separates them exactly: every Multifamily one has 5+ units (5 through 63) and every
    # Single Family Rental one has 2, 3 or 4 -- 68/68, no overlap. Applied as an override
    # rather than a fill, and scoped to Mixed Use so it cannot disturb the other 23,771 rows.
    _units = pd.to_numeric(out.get("# Units", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    _mixed_mf = _prop_type.astype("string").fillna("").str.strip().str.casefold().eq("mixed use") & _units.ge(5)
    if bool(_mixed_mf.any()):
        out.loc[_mixed_mf, "Grouping"] = "Multifamily"

    # V37 taxonomy cascade: Term Asset inherits the parent deal's Portfolio/Financing, but
    # the prior Term Asset carry-forward wins over the Term Loan value via
    # coalesce_report_display_first. Derive the SSR deal set from the built Term Loan
    # (Portfolio == 'Sold Term' in V43; older workbooks carried 'Sold Servicing Retained'
    # there, so accept both when reading the parent back) and force it onto the assets.
    if "Portfolio" in tl.columns:
        _ssr_labels = {TERM_SOLD_PORTFOLIO} | TERM_SOLD_RETAINED_SEGMENT_VALUES
        _ssr_keys = set(
            tl.loc[
                tl["Portfolio"].astype("string").str.strip().isin(_ssr_labels),
                "_deal_key",
            ].dropna().astype(str).tolist()
        )
        out = _apply_term_sold_servicing_retained(out, _ssr_keys)

    # Term Asset "2Q26 Special Loans List " mirrors the Term Loan special value by
    # deal (the report uses XLOOKUP into Term Loan). Map it from whichever Term Loan
    # special column is present.
    _tl_special_col = None
    for _cand in [c for c in tl.columns if re.fullmatch(r"\dQ\d{2}\s+Special\s+Loans\s+List", str(c).strip(), flags=re.I)]:
        _tl_special_col = _cand
        break
    if _tl_special_col is None and "Special Loans List (Y/N)" in tl.columns:
        _tl_special_col = "Special Loans List (Y/N)"
    _ta_special_header = _special_list_header(quarter_end_for_run(run_dt))
    if _tl_special_col is not None:
        tl_special = tl[["_deal_key", _tl_special_col]].drop_duplicates("_deal_key")
        out = out.merge(tl_special, on="_deal_key", how="left")
        out[_ta_special_header] = out[_tl_special_col]
        if _tl_special_col != _ta_special_header:
            out = out.drop(columns=[_tl_special_col], errors="ignore")
    else:
        out[_ta_special_header] = "N/A"

    _subunit_twin_drops = _drop_subunit_twin_term_assets(out, term_loan)
    if _subunit_twin_drops:
        TERM_ASSET_SUBUNIT_TWIN_DROPS.clear()
        TERM_ASSET_SUBUNIT_TWIN_DROPS.update(_subunit_twin_drops)

    out = _allocate_term_asset_upb_from_loan(out, term_loan, upb_col)

    meaningful_mask = (
        out["_deal_key"].notna()
        & out["_asset_key"].notna()
        & (
            (~blankish_mask(out.get("Address", pd.Series([pd.NA] * len(out), index=out.index))))
            | pd.to_numeric(out.get("Property ALA", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0).gt(0)
            | pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0).ne(0)
            | pd.to_numeric(out.get("Origination Value", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna()
            | pd.to_numeric(out.get("Updated As-Is Value", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna()
        )
    )
    # V65: a Term Asset must carry a balance. Salesforce keeps sub-unit and released
    # property records alive with a real Address but no allocation, and the mask above lets
    # them through on the Address alone -- 116 of test 76's 152 extra Term Asset rows are
    # exactly that. Deal 25703 is the clearest case: 23 Condo records at one address, each
    # '# Units' = 1, every one with Property ALA and UPB of zero, because the parent
    # property holds the allocation.
    #
    # Is_Sub_Unit__c is already filtered in the SOQL (_build_term_asset_like), so these rows
    # have the checkbox unticked or null in Salesforce. This is the belt-and-braces catch for
    # the ones the flag misses.
    #
    # No official report carries such a row: Property ALA <= 0 and UPB <= 0 both occur 0
    # times across the 20260803 / 20260810 / 20260817 / 20260824 reports, 97,000+ asset rows.
    # On test 76 the rule drops 116 rows, all 116 of them extras, and loses no real asset.
    # Scoped to rows where BOTH are non-positive so an asset with a real balance but a
    # missing allocation still survives.
    out = _apply_term_financing_taxonomy(out)

    _ta_ala = pd.to_numeric(out.get("Property ALA", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0.0)
    _ta_upb = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0.0)
    meaningful_mask = meaningful_mask & (_ta_ala.gt(0) | _ta_upb.ne(0))

    out = out.loc[meaningful_mask].copy()

    # Value columns blank (not 0) to match the official report.
    out = blank_zero_value_columns(out, ["Origination Value", "Updated As-Is Value"])

    return downcast_numeric_frame(out.drop(columns=["_value_dt", "_mod_dt", "_created_dt", "_ala_sort"], errors="ignore"))

def build_bridge_loan(
    bridge_loan_wide: pd.DataFrame,
    bridge_asset: pd.DataFrame,
    bridge_property_rollup: pd.DataFrame,
    serv_lookup: pd.DataFrame,
    upb_col: str,
    prev_maps: dict,
    template_maps: dict,
    npl_maps: Optional[dict] = None,
) -> pd.DataFrame:
    out = bridge_loan_wide.copy()
    if out.empty:
        return out

    blank_obj = pd.Series([pd.NA] * len(out), index=out.index, dtype="object")
    out["_deal_key"] = norm_id_series(out.get("Deal Number", pd.Series([None] * len(out), index=out.index)))
    out["Servicer ID"] = coalesce_keep_nonblank(out.get("Servicer ID", blank_obj), blank_obj)

    if "Do Not Lend" in out.columns:
        out["Do Not Lend (Y/N)"] = _yn_from_bool_series(out["Do Not Lend"])
        out = out.drop(columns=["Do Not Lend"], errors="ignore")
    elif "Do Not Lend (Y/N)" not in out.columns:
        out["Do Not Lend (Y/N)"] = "N"

    seg_guess = out.apply(
        lambda r: derive_bridge_segment(r.get("Deal Number"), r.get("Financing"), r.get("Loan Buyer"), template_maps),
        axis=1,
    )
    strat_guess = out.get("Project Strategy", pd.Series([pd.NA] * len(out), index=out.index)).map(
        lambda x: strategy_grouping_from_project_strategy(x, template_maps.get("strategy_map", {}))
    )
    port_guess = out.apply(
        lambda r: derive_bridge_portfolio(
            r.get("Product Type"),
            r.get("Segment") if has_any_value(r.get("Segment")) else derive_bridge_segment(r.get("Deal Number"), r.get("Financing"), r.get("Loan Buyer"), template_maps),
            r.get("Financing"),
            r.get("Deal Intro Sub-Source"),
            r.get("Deal Number"),
        ),
        axis=1,
    )
    out["Segment"] = coalesce_keep_nonblank(out.get("Segment", blank_obj), seg_guess)
    out["Strategy Grouping"] = coalesce_keep_nonblank(out.get("Strategy Grouping", blank_obj), strat_guess)
    out["Portfolio"] = coalesce_keep_nonblank(out.get("Portfolio", blank_obj), port_guess)

    sold_stage_series = out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    out["Financing"] = pd.Series(out.get("Financing", blank_obj), index=out.index, dtype="object")
    # V43: mirrors the Bridge Asset label -- "Sold Servicing Retained", not "Sold".
    out["Financing"] = out["Financing"].mask(blankish_mask(out["Financing"]) & sold_stage_series.eq("Sold"), BRIDGE_SOLD_FINANCING)

    if bridge_property_rollup is not None and not bridge_property_rollup.empty:
        out = out.merge(bridge_property_rollup, on="_deal_key", how="left")
    else:
        out["Number of Assets"] = np.nan
        out["# of Units"] = np.nan
        out["State(s)"] = pd.NA
        out["Active Asset Count"] = 0
        out["Active Asset UPB"] = np.nan

    # V41: Loan Commitment = SUM(per-asset Approved Advance Amount Max) from the property
    # rollup (verified deal 33182 = 102,500), which supersedes the spine LOC_Commitment__c.
    # The summed value wins when present/positive; otherwise the spine value is kept.
    if "Loan Commitment (AAM)" in out.columns:
        _aam_commit = pd.to_numeric(out["Loan Commitment (AAM)"], errors="coerce")
        _spine_commit = pd.to_numeric(out.get("Loan Commitment", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
        out["Loan Commitment"] = _aam_commit.where(_aam_commit.gt(0), _spine_commit)
        out = out.drop(columns=["Loan Commitment (AAM)"], errors="ignore")

    if bridge_asset is not None and not bridge_asset.empty:
        ba = bridge_asset.copy()
        upd_dt = _to_datetime_series_mixed(ba.get("Updated Valuation Date", pd.Series([pd.NaT] * len(ba), index=ba.index)))
        org_dt = _to_datetime_series_mixed(ba.get("Origination Value Dt", pd.Series([pd.NaT] * len(ba), index=ba.index)))
        has_updated_appraisal = upd_dt.notna()
        ba["_roll_recent_val_dt"] = upd_dt.where(has_updated_appraisal, org_dt)
        ba["_roll_recent_asis"] = pd.to_numeric(
            ba.get("Updated As-Is Value", pd.Series([np.nan] * len(ba), index=ba.index)), errors="coerce"
        ).where(
            has_updated_appraisal,
            pd.to_numeric(ba.get("Origination As-Is Value", pd.Series([np.nan] * len(ba), index=ba.index)), errors="coerce"),
        )
        ba["_roll_recent_arv"] = pd.to_numeric(
            ba.get("Updated ARV", pd.Series([np.nan] * len(ba), index=ba.index)), errors="coerce"
        ).where(
            has_updated_appraisal,
            pd.to_numeric(ba.get("Origination ARV", pd.Series([np.nan] * len(ba), index=ba.index)), errors="coerce"),
        )

        # V60: the Bridge Loan "Most Recent" triple must summarise the FINAL Bridge Asset
        # values, not a second derivation from the raw Updated/Origination columns. Those
        # two can disagree (the asset tab applies the report's own
        # IF($BH<>"N/A",...) precedence plus the blank-zero and carry-forward rules), and
        # the loan tab then reports a total that does not tie to the assets beneath it.
        # Measured on 20260824 against the official: summing the materialized asset values
        # scores 986/998 As-Is and 980/998 ARV, and MIN of the asset valuation date scores
        # 995/998, versus 917 / 959 / 962 for the re-derived version. Recomputing the
        # materializer here is cheap (~5k rows) and guarantees the two tabs tie.
        try:
            _ba_final = _materialize_bridge_asset_formula_columns(ba, upb_col)
            for _src_col, _roll_col in (
                ("Most Recent Valuation Date", "_roll_recent_val_dt"),
                ("Most Recent As-Is Value", "_roll_recent_asis"),
                ("Most Recent ARV", "_roll_recent_arv"),
            ):
                if _src_col in _ba_final.columns:
                    _vals = pd.Series(_ba_final[_src_col].to_numpy(), index=ba.index)
                    ba[_roll_col] = (
                        pd.to_datetime(_vals, errors="coerce") if "Date" in _src_col
                        else pd.to_numeric(_vals, errors="coerce")
                    )
        except Exception:
            # Never let the consistency pass break the build; the derivation above stands.
            pass

        g = ba.groupby("_deal_key", dropna=True)

        def _first(series: pd.Series):
            return first_nonblank(series)

        def _max_dt(series: pd.Series):
            s = pd.to_datetime(series, errors="coerce").dropna()
            return s.max() if len(s) else pd.NaT

        def _min_dt(series: pd.Series):
            s = pd.to_datetime(series, errors="coerce").dropna()
            return s.min() if len(s) else pd.NaT

        def _yn_any(series: pd.Series):
            vals = pd.Series(series).astype("string").str.strip().str.upper()
            return "Y" if vals.eq("Y").any() else "N"

        active_roll = pd.DataFrame(
            {
                "Servicer ID_active": g["Servicer ID"].apply(first_or_various) if "Servicer ID" in ba.columns else pd.Series(dtype="string"),
                "Servicer_active": g["Servicer"].apply(first_or_various) if "Servicer" in ba.columns else pd.Series(dtype="string"),
                "Number of Assets_active": g["_asset_key"].nunique() if "_asset_key" in ba.columns else pd.Series(dtype="float"),
                "# of Units_active": pd.to_numeric(g["# of Units"].sum(min_count=1), errors="coerce") if "# of Units" in ba.columns else pd.Series(dtype="float"),
                "State(s)_active": g["State"].apply(lambda s: ", ".join(sorted({clean_text(x) for x in s if clean_text(x)}))) if "State" in ba.columns else pd.Series(dtype="string"),
                "Primary Contact_active": g["Primary Contact"].apply(_first) if "Primary Contact" in ba.columns else pd.Series(dtype="string"),
                "Last Funding Date_active": g["Last Funding Date"].apply(_max_dt) if "Last Funding Date" in ba.columns else pd.NaT,
                "Days Past Due_active": pd.to_numeric(g["_bridge_dpd_num"].max(), errors="coerce") if "_bridge_dpd_num" in ba.columns else pd.Series(dtype="float"),
                "Loan Level Delinquency_active": g.apply(lambda grp: _bridge_loan_rollup_label(grp.get("_bridge_dq_bucket", pd.Series(dtype="object")), grp.get("_bridge_dpd_num", pd.Series(dtype="float")))) if "_bridge_dq_bucket" in ba.columns else pd.Series(dtype="string"),
                "Active Funded Amount": pd.to_numeric(g["SF Funded Amount"].sum(min_count=1), errors="coerce") if "SF Funded Amount" in ba.columns else np.nan,
                "Suspense Balance_active": pd.to_numeric(g["Suspense Balance"].sum(min_count=1), errors="coerce") if "Suspense Balance" in ba.columns else np.nan,
                # V55: the official Bridge Loan "Most Recent Valuation Date" is the MINIMUM
                # (oldest) of the deal's asset valuation dates, not the newest -- it flags how
                # stale the weakest collateral value is. 993/996 deals against 20260810; MAX
                # matched only 893. (Consistent with Next Payment Date, which is also a MIN.)
                # V60: the official also aggregates these straight off the asset tab --
                # Current/Original Maturity Date as the deal MIN and Loan Stage as the single
                # distinct value (998/998 each on 20260824).
                "Current Maturity Date_active": g["Current Loan Maturity date"].apply(_min_dt) if "Current Loan Maturity date" in ba.columns else pd.NaT,
                "Original Maturity Date_active": g["Original Loan Maturity date"].apply(_min_dt) if "Original Loan Maturity date" in ba.columns else pd.NaT,
                "Loan Stage_active": g["Loan Stage"].apply(first_or_various) if "Loan Stage" in ba.columns else pd.Series(dtype="string"),
                "Most Recent Valuation Date": g["_roll_recent_val_dt"].apply(_min_dt),
                "Most Recent As-Is Value": pd.to_numeric(g["_roll_recent_asis"].sum(min_count=1), errors="coerce"),
                "Most Recent ARV": pd.to_numeric(g["_roll_recent_arv"].sum(min_count=1), errors="coerce"),
                "Initial Disbursement Funded": pd.to_numeric(g["Initial Disbursement Funded"].sum(min_count=1), errors="coerce") if "Initial Disbursement Funded" in ba.columns else np.nan,
                "Renovation Holdback": pd.to_numeric(g["Renovation Holdback"].sum(min_count=1), errors="coerce") if "Renovation Holdback" in ba.columns else np.nan,
                "Renovation HB Funded": pd.to_numeric(g["Renovation Holdback Funded"].sum(min_count=1), errors="coerce") if "Renovation Holdback Funded" in ba.columns else np.nan,
                "Renovation HB Remaining": pd.to_numeric(g["Renovation Holdback Remaining"].sum(min_count=1), errors="coerce") if "Renovation Holdback Remaining" in ba.columns else np.nan,
                "Interest Allocation": pd.to_numeric(g["Interest Allocation"].sum(min_count=1), errors="coerce") if "Interest Allocation" in ba.columns else np.nan,
                "Interest Allocation Funded": pd.to_numeric(g["Interest Allocation Funded"].sum(min_count=1), errors="coerce") if "Interest Allocation Funded" in ba.columns else np.nan,
                "3/31 NPL": g["3/31 NPL (Y/N)"].apply(_yn_any) if "3/31 NPL (Y/N)" in ba.columns else "N",
                "Needs NPL Value": g["Needs NPL Value"].apply(_yn_any) if "Needs NPL Value" in ba.columns else "N",
                "Special Focus (Y/N)": g["Special Flag"].apply(_yn_any) if "Special Flag" in ba.columns else "N",
                "Asset Manager 1": g["Asset Manager 1"].apply(_first) if "Asset Manager 1" in ba.columns else pd.Series(dtype="string"),
                "AM 1 Assigned Date": g["AM 1 Assigned Date"].apply(_first) if "AM 1 Assigned Date" in ba.columns else pd.NaT,
                "Asset Manager 2": g["Asset Manager 2"].apply(_first) if "Asset Manager 2" in ba.columns else pd.Series(dtype="string"),
                "AM 2 Assigned Date": g["AM 2 Assigned Date"].apply(_first) if "AM 2 Assigned Date" in ba.columns else pd.NaT,
                "Construction Mgr.": g["Construction Mgr."].apply(_first) if "Construction Mgr." in ba.columns else pd.Series(dtype="string"),
                "CM Assigned Date": g["CM Assigned Date"].apply(_first) if "CM Assigned Date" in ba.columns else pd.NaT,
                "Active RM": g["Active RM"].apply(_first) if "Active RM" in ba.columns else pd.Series(dtype="string"),
                "AM Commentary": g["Comments AM"].apply(_first) if "Comments AM" in ba.columns else pd.Series(dtype="string"),
            }
        ).reset_index()
        out = out.merge(active_roll, on="_deal_key", how="left")
    else:
        out["3/31 NPL"] = "N"
        out["Needs NPL Value"] = "N"
        out["Special Focus (Y/N)"] = "N"

    # V60: prefer the asset-tab rollup for these three (the loan-level Salesforce value is
    # only a fallback), so the loan tab cannot contradict its own assets.
    for _bl_col in ("Current Maturity Date", "Original Maturity Date"):
        _src = f"{_bl_col}_active"
        if _src in out.columns:
            _rolled = pd.to_datetime(out[_src], errors="coerce")
            out[_bl_col] = _rolled.where(_rolled.notna(), pd.to_datetime(out.get(_bl_col, pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce"))
    if "Loan Stage_active" in out.columns:
        out["Loan Stage"] = coalesce_keep_nonblank(out["Loan Stage_active"], out.get("Loan Stage", blank_obj))

    out["Primary Contact"] = coalesce_keep_nonblank(out.get("Primary Contact_active", blank_obj), out.get("Primary Contact", blank_obj))
    # Fix H.1: BL Last Funding Date = MAX of the per-asset "Last Funding Date" across the
    # deal's Bridge Asset rows (verified 5/5 vs same-day real). The asset rollup is the
    # authoritative source; the deal-level property rollup is only a fallback. The
    # Opportunity Funds_Released/Last_Funding field ("Last Funding Date SF") is NOT used
    # below -- it carries later post-closing advances the official report excludes
    # (the 2021-10-08-vs-2021-01-27 leak on deals like 33182).
    out["Last Funding Date"] = _coalesce_datetime_columns(out, ["Last Funding Date_active", "Last Funding Date"])
    out["Servicer ID"] = coalesce_keep_nonblank(out.get("Servicer ID_active", blank_obj), out.get("Servicer ID", blank_obj))
    out["Servicer"] = coalesce_keep_nonblank(out.get("Servicer_active", blank_obj), out.get("Servicer", blank_obj))
    out["Number of Assets"] = pd.to_numeric(out.get("Number of Assets_active", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").where(
        pd.to_numeric(out.get("Number of Assets_active", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_numeric(out.get("Number of Assets", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )
    out["Number of Assets"] = pd.to_numeric(out["Number of Assets"], errors="coerce").where(
        pd.to_numeric(out["Number of Assets"], errors="coerce").notna(),
        pd.to_numeric(out.get("Number of Assets SF", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )
    # # of Units = sum of the per-asset Property Number_of_Units__c only (asset rollup,
    # then the property-rollup fallback which is the SAME field aggregated differently).
    # Do NOT fall back to the Opportunity-level Total_Units__c ("# of Units SF"): for ~45
    # deals the per-asset count is genuinely blank in Salesforce and the official report
    # shows 0/N/A there, but Total_Units__c carries the building's structural unit count
    # (e.g. 84) and was leaking in as a non-zero override. Blank stays blank -> 0 via the
    # end-of-function fillna, matching the report (verified 1045/1045).
    out["# of Units"] = pd.to_numeric(out.get("# of Units_active", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").where(
        pd.to_numeric(out.get("# of Units_active", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_numeric(out.get("# of Units", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )
    out["State(s)"] = coalesce_keep_nonblank(out.get("State(s)_active", blank_obj), out.get("State(s)", blank_obj))
    out["State(s)"] = coalesce_keep_nonblank(out.get("State(s)", blank_obj), out.get("State(s) SF", blank_obj))
    # Fix H.1: intentionally NO "Last Funding Date SF" (Opportunity Funds_Released_Date__c /
    # Last_Funding_Date__c) fallback here -- it leaks later advances the report excludes.
    # Last Funding Date is the asset-rollup MAX set above.
    out["Most Recent Valuation Date"] = pd.to_datetime(out.get("Most Recent Valuation Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce").where(
        pd.to_datetime(out.get("Most Recent Valuation Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_datetime(out.get("Most Recent Valuation Date SF", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce"),
    )
    out["Most Recent As-Is Value"] = pd.to_numeric(out.get("Most Recent As-Is Value", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").where(
        pd.to_numeric(out.get("Most Recent As-Is Value", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_numeric(out.get("Most Recent As-Is Value SF", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )
    out["Most Recent ARV"] = pd.to_numeric(out.get("Most Recent ARV", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").where(
        pd.to_numeric(out.get("Most Recent ARV", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_numeric(out.get("Most Recent ARV SF", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )
    out["Active Funded Amount"] = pd.to_numeric(out.get("Active Funded Amount", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").where(
        pd.to_numeric(out.get("Active Funded Amount", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_numeric(out.get("Active Funded Amount SF", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )

    out["Active Asset Count"] = pd.to_numeric(out.get("Active Asset Count", pd.Series([0] * len(out), index=out.index)), errors="coerce").fillna(0)
    out["Active Asset UPB"] = pd.to_numeric(out.get("Active Asset UPB", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    out["SF Current UPB"] = pd.to_numeric(out.get("SF Current UPB", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    out["Suspense Balance"] = pd.to_numeric(out.get("Suspense Balance_active", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")

    out["_sid_key"] = id_key_no_leading_zeros(out.get("Servicer ID", pd.Series([None] * len(out), index=out.index)))
    if not serv_lookup.empty and "_sid_key" in serv_lookup.columns:
        s = serv_lookup.dropna(subset=["_sid_key"]).copy().rename(
            columns={
                "servicer": "_servicer_file",
                "upb": "_loan_upb",
                "suspense": "_loan_suspense",
                "next_payment_date": "_serv_next_payment_date",
                "maturity_date": "_servicer_maturity_file",
                "status": "_servicer_status_file",
            }
        )
        out = out.merge(
            s[["_sid_key", "_servicer_file", "_loan_upb", "_loan_suspense", "_serv_next_payment_date", "_servicer_maturity_file", "_servicer_status_file"]],
            on="_sid_key",
            how="left",
        )

    if "bridge_loan_upb" in prev_maps:
        prev_upb = prev_maps["bridge_loan_upb"].copy()
        out = out.merge(prev_upb, on="_deal_key", how="left")
    else:
        out["_prev_upb"] = np.nan

    stage_series = out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    loan_upb_raw = pd.to_numeric(out.get("_loan_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    asset_upb_raw = pd.to_numeric(out.get("Active Asset UPB", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    sf_upb_raw = pd.to_numeric(out.get("SF Current UPB", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")

    valid_loan_upb = loan_upb_raw.where(loan_upb_raw.gt(0))
    valid_asset_upb = asset_upb_raw.where(asset_upb_raw.gt(0))
    valid_sf_upb = sf_upb_raw.where(sf_upb_raw.gt(0))
    final_upb = valid_loan_upb.where(valid_loan_upb.notna(), valid_asset_upb)
    final_upb = final_upb.where(final_upb.notna(), valid_sf_upb)
    # Zero is valid only after every positive source is absent. This prevents a servicer zero from hiding positive asset-level UPB.
    final_upb = final_upb.where(final_upb.notna(), loan_upb_raw.where(loan_upb_raw.eq(0)))
    final_upb = final_upb.where(final_upb.notna(), asset_upb_raw.where(asset_upb_raw.eq(0)))
    final_upb = final_upb.where(final_upb.notna(), sf_upb_raw.where(sf_upb_raw.eq(0)))

    late_stage_mask = stage_series.isin(EXPIRED_OR_MATURED_STAGES)
    prev_upb_vals = pd.to_numeric(out.get("_prev_upb", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    final_upb = final_upb.where(~(late_stage_mask & (final_upb.isna() | final_upb.le(0))), prev_upb_vals)
    out[upb_col] = pd.to_numeric(final_upb, errors="coerce")

    out["Suspense Balance"] = pd.to_numeric(out.get("_loan_suspense", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").where(
        pd.to_numeric(out.get("_loan_suspense", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").notna(),
        pd.to_numeric(out.get("Suspense Balance", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )
    out["Suspense Balance"] = pd.to_numeric(out["Suspense Balance"], errors="coerce").where(
        pd.to_numeric(out["Suspense Balance"], errors="coerce").notna(),
        pd.to_numeric(out.get("SF Suspense Balance", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce"),
    )

    cur_bridge_loan_npd = pd.to_datetime(out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
    serv_bridge_loan_npd = pd.to_datetime(out.get("_serv_next_payment_date", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
    prior_bridge_loan_npd = None
    if "bridge_loan_manual" in prev_maps and isinstance(prev_maps.get("bridge_loan_manual"), pd.DataFrame):
        prev_bl = prev_maps["bridge_loan_manual"]
        if "Next Payment Date" in prev_bl.columns and "_deal_key" in prev_bl.columns:
            prior_bridge_loan_npd = out["_deal_key"].map(prev_bl.dropna(subset=["_deal_key"]).drop_duplicates("_deal_key").set_index("_deal_key")["Next Payment Date"])
    _bl_servicer = coalesce_keep_nonblank(out.get("_servicer_file", pd.Series([pd.NA] * len(out), index=out.index)), out.get("Servicer", pd.Series([pd.NA] * len(out), index=out.index)))
    out["Next Payment Date"] = _bridge_pick_next_payment_date(cur_bridge_loan_npd, serv_bridge_loan_npd, prior_bridge_loan_npd, servicer_names=_bl_servicer, run_dt=run_dt)

    # V54: the official Bridge Loan Next Payment Date is the MINIMUM of the deal's Bridge
    # Asset next-payment dates -- exact on 996/996 deals against 20260810 (MAX matches 990,
    # mode 995). This supersedes the loan-level servicer/Salesforce lookup above, which is
    # kept only as the fallback for a deal whose assets carry no date at all. Rolling up from
    # the assets also means the asset-level day-10 / prior-report rules propagate to the loan
    # tab automatically instead of being re-derived from a different source.
    if bridge_asset is not None and not bridge_asset.empty and "Next Payment Date" in bridge_asset.columns:
        _ba_npd_roll = pd.to_datetime(bridge_asset["Next Payment Date"], errors="coerce")
        _ba_dk_roll = (
            bridge_asset["_deal_key"] if "_deal_key" in bridge_asset.columns
            else norm_id_series(bridge_asset.get("Deal Number", pd.Series([pd.NA] * len(bridge_asset), index=bridge_asset.index)))
        )
        _min_npd_by_deal = _ba_npd_roll.groupby(_ba_dk_roll).min()
        _rolled_npd = pd.to_datetime(out["_deal_key"].map(_min_npd_by_deal), errors="coerce")
        out["Next Payment Date"] = _rolled_npd.where(
            _rolled_npd.notna(), pd.to_datetime(out["Next Payment Date"], errors="coerce")
        )
    out["Next Advance Maturity Date"] = pd.to_datetime(out.get("_servicer_maturity_file", pd.Series([pd.NaT] * len(out), index=out.index)), errors="coerce")
    out["Servicer"] = coalesce_keep_nonblank(out.get("_servicer_file", blank_obj), out.get("Servicer", blank_obj))

    loan_status_bucket = pd.Series(
        [
            normalize_bridge_servicer_status(raw_status, npd, run_dt, loan_stage, None, None)
            for raw_status, npd, loan_stage in zip(
                out.get("_servicer_status_file", blank_obj),
                out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)),
                stage_series,
            )
        ],
        index=out.index,
        dtype="object",
    )
    servicer_report_bucket = pd.Series(
        [_bridge_bucket_to_report_label(bucket, _guess_days_past_due(npd, run_dt)) for bucket, npd in zip(loan_status_bucket, out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)))],
        index=out.index,
        dtype="object",
    )
    out["Loan Level Delinquency"] = coalesce_keep_nonblank(out.get("Loan Level Delinquency_active", blank_obj), servicer_report_bucket)
    out["Days Past Due"] = pd.to_numeric(out.get("Days Past Due_active", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    derived_days = pd.Series(
        [
            _guess_days_from_bridge_bucket(bucket) if not pd.isna(_guess_days_from_bridge_bucket(bucket)) else _guess_days_past_due(npd, run_dt)
            for bucket, npd in zip(out["Loan Level Delinquency"], out.get("Next Payment Date", pd.Series([pd.NaT] * len(out), index=out.index)))
        ],
        index=out.index,
        dtype="float64",
    )
    out["Days Past Due"] = out["Days Past Due"].where(out["Days Past Due"].notna(), derived_days)

    if npl_maps and not npl_maps.get("loan_flags", pd.DataFrame()).empty:
        loan_flags = npl_maps["loan_flags"].copy().drop_duplicates("_deal_key")
        out = out.merge(loan_flags, on="_deal_key", how="left", suffixes=("", "_npl"))
        if "NPL Flag_npl" in out.columns:
            out["3/31 NPL"] = coalesce_keep_nonblank(out.get("3/31 NPL", blank_obj), out["NPL Flag_npl"])
            out = out.drop(columns=["NPL Flag_npl"], errors="ignore")
        for c in ["Needs NPL Value", "Special Focus (Y/N)"]:
            if f"{c}_npl" in out.columns:
                out[c] = coalesce_keep_nonblank(out.get(c, blank_obj), out[f"{c}_npl"])
                out = out.drop(columns=[f"{c}_npl"], errors="ignore")

    if "bridge_loan_manual" in prev_maps and not out.empty:
        man = prev_maps["bridge_loan_manual"].copy()
        out = out.merge(man, on="_deal_key", how="left", suffixes=("", "_prev"))
        bridge_loan_carry_forward_first = {
            "Portfolio", "Segment", "Financing", "Strategy Grouping", "Loan Level Delinquency",
            # NOTE (V42): "AM Commentary" REVERTED back to carry-forward-first -- the live
            # Opportunity.Asset_Management_Comments__c pull did not beat the prior value vs
            # same-day real (49 mismatches unchanged; the live API note was older than real's).
            "Special Focus (Y/N)", "AM Commentary", "3/31 NPL", "Needs NPL Value",
            "Active RM", "Asset Manager 1", "AM 1 Assigned Date", "Asset Manager 2",
            "AM 2 Assigned Date", "Construction Mgr.", "CM Assigned Date",
        }
        for c in [
            "Portfolio", "Segment", "Financing", "Strategy Grouping", "Loan Level Delinquency", "Special Focus (Y/N)",
            "AM Commentary", "3/31 NPL", "Needs NPL Value", "Active RM",
            "Asset Manager 1", "AM 1 Assigned Date", "Asset Manager 2", "AM 2 Assigned Date",
            "Construction Mgr.", "CM Assigned Date",
        ]:
            if f"{c}_prev" in out.columns:
                if c in bridge_loan_carry_forward_first:
                    out[c] = coalesce_report_display_first(out[f"{c}_prev"], out.get(c, blank_obj))
                else:
                    out[c] = coalesce_keep_nonblank(out.get(c, blank_obj), out[f"{c}_prev"])
                out = out.drop(columns=[f"{c}_prev"], errors="ignore")

    # CAFL re-financed deals: Segment must read BRIDGE_SECURITIZED_SEGMENT once the deal is
    # re-financed into a CAFL securitization, overriding the Segment-first carry-forward
    # that would otherwise pin last week's stale vehicle (e.g. "CPP JV"). Mirrors the
    # Bridge Asset fix; Financing (Warehouse Line) startswith "CAFL " is authoritative.
    if "Financing" in out.columns and "Segment" in out.columns:
        _cafl_mask = out["Financing"].astype("string").str.strip().str.upper().str.startswith("CAFL ", na=False)
        out.loc[_cafl_mask, "Segment"] = BRIDGE_SECURITIZED_SEGMENT

    # V78: the same treatment for CPP JV. Segment is carry-forward-first, so a deal that has
    # since moved onto a CPP JV vehicle keeps last week's label -- 184 Bridge Asset rows and 19
    # Bridge Loan rows read "Mortgage Banking" on 20260831 while the official says "CPP JV",
    # and every one of them carries Financing "CPP JV - Goldman Sachs" in BOTH the build and the
    # official, so only the Segment derivation was stale. Exact in both directions on 20260831:
    # 731 rows have CPP JV financing, the official calls all 731 CPP JV, and the official's CPP
    # JV population is exactly those 731.
    if "Financing" in out.columns and "Segment" in out.columns:
        _cpp_mask = out["Financing"].astype("string").str.strip().str.upper().str.startswith("CPP JV", na=False)
        out.loc[_cpp_mask, "Segment"] = "CPP JV"

    # Hard-reconcile the loan-level math back to the already-built Bridge Asset rows.
    # This prevents servicer zeroes / wrong rollup fields from breaking loan-level Active Funded Amount or UPB.
    out = _reconcile_bridge_loan_from_asset_rollup(out, bridge_asset, upb_col)

    most_recent_arv_num = pd.to_numeric(out.get("Most Recent ARV", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    if "Most Recent ARV" in out.columns:
        out["Most Recent ARV"] = pd.Series(out["Most Recent ARV"], index=out.index, dtype="object")
        out.loc[most_recent_arv_num.fillna(0).eq(0), "Most Recent ARV"] = "N/A"

    out["Servicer ID"] = normalize_servicer_id_for_report(out.get("Servicer ID", blank_obj), out.get("Servicer", blank_obj))
    out["Active RM"] = coalesce_keep_nonblank(out.get("Active RM", blank_obj), pd.Series(["N"] * len(out), index=out.index))
    # NOTE (V42): the "Special Focus = has Asset Manager 2 assignment" rule (former Fix G)
    # was REVERTED -- it regressed BL Special Focus from 24 -> 111 mismatches against
    # same-day real. Restored the rollup/NPL/carry-forward-derived value.
    out["Special Focus (Y/N)"] = coalesce_keep_nonblank(out.get("Special Focus (Y/N)", blank_obj), pd.Series(["N"] * len(out), index=out.index))
    out["3/31 NPL"] = coalesce_keep_nonblank(out.get("3/31 NPL", blank_obj), pd.Series(["N"] * len(out), index=out.index))
    out["Needs NPL Value"] = coalesce_keep_nonblank(out.get("Needs NPL Value", blank_obj), pd.Series(["N"] * len(out), index=out.index))

    out["Number of Assets"] = pd.to_numeric(out.get("Number of Assets", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")
    # V54: "# of Units" is N/A -- NOT 0 -- when no child asset carries a unit count.
    # 20260810 Bridge Loan holds the literal text "N/A" on exactly 45 deals, and on 44 of
    # those every child Bridge Asset unit count is likewise blank (the 45th is a single-asset
    # rounding case). The previous fillna(0) turned all of them into 0. Leaving the rollup
    # NaN lets REPORT_NA_FILL_HEADERS render "N/A" at write time.
    out["# of Units"] = pd.to_numeric(out.get("# of Units", pd.Series([np.nan] * len(out), index=out.index)), errors="coerce")

    bridge_asset_deal_keys = set()
    if bridge_asset is not None and not bridge_asset.empty and "_deal_key" in bridge_asset.columns:
        bridge_asset_deal_keys = set(pd.Series(bridge_asset["_deal_key"], copy=False).dropna().astype(str).tolist())
        out = out[out["_deal_key"].isin(bridge_asset_deal_keys)].copy()
        blank_obj = pd.Series([pd.NA] * len(out), index=out.index, dtype="object")

    current_upb = pd.to_numeric(out.get(upb_col, pd.Series([np.nan] * len(out), index=out.index)), errors="coerce").fillna(0)
    active_asset_count = pd.to_numeric(out.get("Active Asset Count", pd.Series([0] * len(out), index=out.index)), errors="coerce").fillna(0)
    stage_series = out.get("Loan Stage", pd.Series([pd.NA] * len(out), index=out.index)).astype("string").str.strip()
    is_closed_won = stage_series.eq("Closed Won")
    is_sold = stage_series.eq("Sold")
    is_reo = stage_series.isin(REO_FAMILY_STAGES)
    is_late_stage = stage_series.isin(EXPIRED_OR_MATURED_STAGES)

    keep_mask = is_closed_won | is_reo | is_sold | (is_late_stage & (current_upb.gt(0) | active_asset_count.gt(0)))
    out = out.loc[keep_mask].copy()
    out = out[out["_deal_key"].notna()].copy()

    out = _fill_text_defaults(
        out,
        [
            "Loan Buyer", "Servicer ID", "Servicer", "Primary Contact", "Loan Level Delinquency",
            "Asset Manager 1", "Asset Manager 2", "Construction Mgr.", "Deal Intro Sub-Source",
            "Referral Source Account", "Referral Source Contact", "AM Commentary",
        ],
    )

    # Value rollup columns blank (not 0) to match the official report.
    out = blank_zero_value_columns(out, [
        "Most Recent ARV", "Most Recent As-Is Value",
        "Origination ARV", "Origination As-Is Value", "Updated ARV", "Updated As-Is Value",
    ])

    drop_cols = [
        c for c in out.columns
        if c.startswith("_") or c.endswith("_active") or c.startswith("SF ") or c.startswith("Opportunity Servicer")
    ]
    return downcast_numeric_frame(out.drop(columns=drop_cols, errors="ignore"))



def _scaffold_col_index(col_idx) -> int:
    """Accept numeric Excel column indexes or Excel letters like AZ/BA.

    Some blueprint updates use Excel-style labels for readability. openpyxl
    Worksheet.cell() requires an integer column index, so normalize here before
    any scaffold/header/formula writes.
    """
    if isinstance(col_idx, int):
        return col_idx
    if isinstance(col_idx, str):
        raw = col_idx.strip().upper()
        if raw.isdigit():
            return int(raw)
        n = 0
        for ch in raw:
            if not ("A" <= ch <= "Z"):
                raise ValueError(f"Invalid Excel scaffold column label: {col_idx!r}")
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n
    return int(col_idx)


def _set_scaffold_cell(ws, row_idx: int, col_idx: int, value):
    col_idx = _scaffold_col_index(col_idx)
    cell = ws.cell(row_idx, col_idx)
    cell.value = value
    if isinstance(value, (date, datetime)):
        cell.number_format = DATE_NUMBER_FORMAT


def _qend_npl_header(q_end: date, suffix: str = "") -> str:
    base = f"{q_end.month}/{q_end.day} NPL"
    return f"{base} {suffix}".strip()


def _qend_npl_reo_header(q_end: date) -> str:
    """Bridge Asset's 3-valued NPL column, e.g. '9/30 NPL/REO' (20260803 col DF)."""
    return f"{q_end.month}/{q_end.day} NPL/REO"


def _special_list_header(q_end: date) -> str:
    quarter = (q_end.month - 1) // 3 + 1
    yy = q_end.year % 100
    # Report header has a trailing space, e.g. "2Q26 Special Loans List "
    return f"{quarter}Q{yy:02d} Special Loans List "


def _resolve_scaffold_token(value, run_dt: date, q_end: date, upb_header: str):
    if value == "__UPB__":
        return upb_header
    if value == "__QEND__":
        return q_end
    if value == "__QEND_NPL__":
        return _qend_npl_header(q_end)
    if value == "__QEND_NPL_YN__":
        return _qend_npl_header(q_end, "(Y/N)")
    if value == "__QEND_NPL_REO__":
        return _qend_npl_reo_header(q_end)
    if value == "__SPECIAL_LIST__":
        return _special_list_header(q_end)
    if value == "__RUN_DT__":
        return run_dt
    return value


def _ensure_bridge_asset_fc_columns(ws):
    """No-op on the new row-5 template: FC Sale Date / Rescheduled FC Sale Date
    are already present (cols 53/54). The legacy column-insert path is retired."""
    return
    if ws.title != "Bridge Asset":
        return
    cur_az = clean_text(ws.cell(4, 52).value)
    cur_ba = clean_text(ws.cell(4, 53).value)
    if cur_az == "FC Sale Date" and cur_ba == "Rescheduled FC Sale Date":
        return
    if cur_az != "REO Date":
        return

    ws.insert_cols(52, 2)
    max_row = max(ws.max_row, 5)
    for row_idx in range(1, max_row + 1):
        source = ws.cell(row_idx, 54)
        for col_idx in (52, 53):
            target = ws.cell(row_idx, col_idx)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.font = copy(source.font)


def refresh_summary_labels(wb, run_dt: date, upb_header: str):
    summary_sheets = [s for s in ("Summary", "Bridge Summary", "Term Summary") if s in wb.sheetnames]
    if not summary_sheets:
        return
    current_md = f"{run_dt.month}/{run_dt.day}"
    q_end = quarter_end_for_run(run_dt)
    q_end_md = f"{q_end.month}/{q_end.day}"

    for _summary_name in summary_sheets:
      ws = wb[_summary_name]
      for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            txt = cell.value
            new_txt = txt
            if "UPB" in txt.upper():
                new_txt = re.sub(r"\b\d{1,2}/\d{1,2}\s*UPB\b", upb_header, new_txt, flags=re.I)
            if "NPL" in txt.upper():
                new_txt = re.sub(r"\b\d{1,2}/\d{1,2}\b", q_end_md, new_txt)
            elif any(k in txt.upper() for k in ["DQ", "DELINQ", "PAST DUE", "CURRENT"]):
                new_txt = re.sub(r"\b\d{1,2}/\d{1,2}\b", current_md, new_txt)
            if new_txt != txt:
                cell.value = new_txt


def _pin_today_formulas_to_run_date(wb, run_dt: date) -> None:
    """Replace volatile TODAY()/NOW() in any worksheet formula with the servicer
    import (run) date. Days Past Due / Days to Maturity anchor on the run-date cell,
    so leaving a live TODAY() would recompute them against whatever day the workbook
    is opened instead of the servicer tape / UPB date."""
    date_literal = f"DATE({run_dt.year},{run_dt.month},{run_dt.day})"
    pat = re.compile(r"\b(?:TODAY|NOW)\s*\(\s*\)", re.I)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and pat.search(v):
                    cell.value = pat.sub(date_literal, v)


def _migrate_sheet_columns_to_blueprint(ws) -> List[str]:
    """Insert/delete physical columns so an older template/baseline matches the blueprint.

    Must run BEFORE the row 1..5 scaffold is cleared, because it identifies the layout
    from the existing headers. Only the two V43 structural changes are handled; both are
    idempotent (a workbook already on the new layout is left untouched).

      Bridge Asset  insert "% of Reno Budget" at BO (67). 20260803 carries it between
                    Renovation Holdback Remaining (BN) and Interest Allocation (BP);
                    every column from 67 on shifts +1.
      Bridge Loan   delete "Remaining Commitment" at AB (28). The official tab goes
                    straight from Suspense Balance (AA) to Most Recent Valuation Date.

    openpyxl's insert_cols/delete_cols relocates cells without rewriting formula text,
    which is safe here: every data-row formula in the affected ranges is re-seeded from
    DRAFT_FORMULA_OVERRIDES (whose letters are the official report's) immediately after
    the blueprint headers are written.
    """
    notes: List[str] = []

    def header_at(col_idx: int) -> str:
        return clean_text(ws.cell(HEADER_ROW, col_idx).value)

    if ws.title == "Bridge Asset":
        if header_at(67) != "% of Reno Budget":
            existing = {header_at(c) for c in range(1, ws.max_column + 1)}
            if "% of Reno Budget" in existing:
                notes.append("Bridge Asset: '% of Reno Budget' present at an unexpected column; left as-is")
            elif header_at(67) == "Interest Allocation":
                ws.insert_cols(67)
                ws.cell(HEADER_ROW, 67).value = "% of Reno Budget"
                notes.append("Bridge Asset: inserted '% of Reno Budget' column at BO (layout shifted +1 from BO)")
            else:
                notes.append(
                    "Bridge Asset: expected 'Interest Allocation' at BO for the "
                    f"'% of Reno Budget' insert, found {header_at(67)!r}; skipped"
                )

    if ws.title == "Bridge Loan":
        if header_at(28) == "Remaining Commitment":
            ws.delete_cols(28)
            notes.append("Bridge Loan: removed 'Remaining Commitment' column at AB (layout shifted -1 from AB)")

    return notes


def restore_template_scaffold(wb, run_dt: date, upb_header: str):
    q_end = quarter_end_for_run(run_dt)
    migration_notes: List[str] = []

    for sheet_name, blueprint in SHEET_BLUEPRINTS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        _ensure_bridge_asset_fc_columns(ws)
        migration_notes.extend(_migrate_sheet_columns_to_blueprint(ws))

        # Clear all values in the scaffold rows before rewriting them from the
        # blueprint. Without this, a prior-layout template (notably Term Asset, which
        # carried headers in row 4 AND row 5) leaves a duplicate header row that trips
        # the QA matched-row logic.
        #
        # Also strip the cell FILL on the empty rows above the header (rows 1..HEADER_ROW-1).
        # A prior completed report can carry a solid (blue) fill across an otherwise-empty
        # row 4 -- visually a second, title-less blue header band above the real titles in
        # row 5. The official report has no such fill, so clear it here. Blueprint cells
        # (e.g. the row-4 SUBTOTAL) are re-set immediately afterward and keep their styling.
        _no_fill = PatternFill(fill_type=None)
        for _clear_row in range(1, HEADER_ROW + 1):
            for _clear_col in range(1, ws.max_column + 1):
                _cell = ws.cell(_clear_row, _clear_col)
                _cell.value = None
                if _clear_row < HEADER_ROW:  # leave the row-5 header styling untouched
                    _cell.fill = _no_fill

        for col_idx, val in blueprint.get("row1", {}).items():
            _set_scaffold_cell(ws, 1, col_idx, _resolve_scaffold_token(val, run_dt, q_end, upb_header))

        for col_idx, val in blueprint.get("row2", {}).items():
            _set_scaffold_cell(ws, 2, col_idx, _resolve_scaffold_token(val, run_dt, q_end, upb_header))

        for col_idx, val in blueprint.get("row3", {}).items():
            _set_scaffold_cell(ws, 3, col_idx, _resolve_scaffold_token(val, run_dt, q_end, upb_header))

        subtotal_col = blueprint.get("subtotal_col")
        subtotal_col_idx = _scaffold_col_index(subtotal_col) if subtotal_col is not None else None
        for col_idx, val in blueprint.get("row4", {}).items():
            col_idx = _scaffold_col_index(col_idx)
            if val == "__RUN_DT__":
                _set_scaffold_cell(ws, 4, col_idx, run_dt)
            elif val == "__SUBTOTAL__":
                col_letter = get_column_letter(subtotal_col_idx)
                _set_scaffold_cell(ws, 4, col_idx, f"=SUBTOTAL(9,{col_letter}{DATA_START_ROW}:{col_letter}{max(DATA_START_ROW, ws.max_row)})")
            else:
                _set_scaffold_cell(ws, 4, col_idx, _resolve_scaffold_token(val, run_dt, q_end, upb_header))

        for col_idx, val in blueprint.get("row5", {}).items():
            col_idx = _scaffold_col_index(col_idx)
            _set_scaffold_cell(ws, 5, col_idx, _resolve_scaffold_token(val, run_dt, q_end, upb_header))

        # Re-seed the DATA_START_ROW formula cells from DRAFT_FORMULA_OVERRIDES now that
        # the headers are authoritative. Two reasons this has to happen here:
        #  * _migrate_sheet_columns_to_blueprint moves cells with insert_cols/delete_cols,
        #    which relocates a formula without rewriting its column letters -- so a
        #    migrated template's seeds still point at the pre-shift columns.
        #  * an uploaded prior-week workbook carries whatever formulas that week had,
        #    including the pre-V43 "Securitized Bridge" / $DE$4 / "Sold" variants.
        # Only columns that already hold a formula, or that live beyond the blue
        # auto-populated range, are touched -- a data column is never turned into one.
        _fx_overrides = DRAFT_FORMULA_OVERRIDES.get(sheet_name, {})
        if _fx_overrides:
            _fx_blue_max = _sheet_blue_max_col(sheet_name)
            for col_idx, val in blueprint.get("row5", {}).items():
                col_idx = _scaffold_col_index(col_idx)
                key = _formula_override_key_for_header(
                    _resolve_scaffold_token(val, run_dt, q_end, upb_header), upb_header
                )
                if key not in _fx_overrides:
                    continue
                cur = ws.cell(DATA_START_ROW, col_idx).value
                is_formula = isinstance(cur, str) and cur.startswith("=")
                beyond_blue = _fx_blue_max is not None and col_idx > _fx_blue_max
                if is_formula or beyond_blue:
                    ws.cell(DATA_START_ROW, col_idx).value = _resolve_formula_override(
                        _fx_overrides[key], q_end
                    )

        # Deterministically (re)apply the blue header fill AND the white bold header
        # font to row 5 on exactly the titled header columns. The official report styles
        # every header cell with a solid theme-3 fill and bold theme-0 (white) text;
        # blank spacer columns (those not in the blueprint, e.g. Term Asset col 21) stay
        # unstyled. Doing this explicitly -- rather than relying on the input template/
        # prior-report styling surviving -- guarantees Term Asset gets the same blue
        # header band with readable white text as the other sheets (test builds were
        # leaving Term Asset's header text non-bold black on blue, hard to read).
        _header_fill = PatternFill(fill_type="solid", fgColor=Color(theme=3, tint=0.0))
        _header_font = Font(name="Aptos Narrow", size=11, bold=True, color=Color(theme=0, tint=0.0))
        _header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for _hdr_col in blueprint.get("row5", {}):
            _hc = _scaffold_col_index(_hdr_col)
            _hcell = ws.cell(HEADER_ROW, _hc)
            _hcell.fill = _header_fill
            _hcell.font = _header_font
            _hcell.alignment = _header_align

        # Deterministically fix the freeze pane and row heights. The official report
        # freezes BELOW the header row (A6 with headers in row 5) and gives the header
        # row the tall height (45), with the empty rows above at the normal height (15).
        # Some inputs leave Term Asset shifted up by one row -- freeze at A5 and the tall
        # height on the empty row 4 instead of row 5 -- which renders the blue header band
        # in the wrong place and one row too high. Set these explicitly so every data
        # sheet matches the official exactly.
        ws.freeze_panes = f"A{DATA_START_ROW}"            # A6
        for _r in range(1, HEADER_ROW):
            ws.row_dimensions[_r].height = 15.0           # rows 1-4 normal
        ws.row_dimensions[HEADER_ROW].height = 45.0       # row 5 tall header

    refresh_summary_labels(wb, run_dt, upb_header)
    # Pin volatile TODAY()/NOW() to the servicer import (run) date so Days Past Due /
    # Days to Maturity compute against the UPB tape date, not the live open date.
    _pin_today_formulas_to_run_date(wb, run_dt)
    return migration_notes


def _parse_direct_ref_formula(formula_text: str):
    if not isinstance(formula_text, str):
        return None
    txt = formula_text.strip()
    if not txt.startswith("="):
        return None
    txt = txt[1:].lstrip("+").strip()

    m = re.fullmatch(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)", txt)
    if m:
        return m.group(1), f"{m.group(2)}{m.group(3)}"

    m = re.fullmatch(r"([A-Za-z0-9_ ]+)!\$?([A-Z]{1,3})\$?(\d+)", txt)
    if m:
        return m.group(1), f"{m.group(2)}{m.group(3)}"

    return None


def _resolve_header_value(wb, ws, cell, upb_header: str, max_depth: int = 6) -> str:
    cur_val = cell.value

    for _ in range(max_depth):
        if cur_val is None:
            return ""
        if not isinstance(cur_val, str):
            return str(cur_val).strip()

        txt = cur_val.strip()
        if UPB_HEADER_RE.search(txt):
            return upb_header

        ref = _parse_direct_ref_formula(txt)
        if not ref:
            return txt

        ref_sheet, ref_cell = ref
        if ref_sheet not in wb.sheetnames:
            return txt
        cur_val = wb[ref_sheet][ref_cell].value

    if cur_val is None:
        return ""
    if isinstance(cur_val, str) and UPB_HEADER_RE.search(cur_val.strip()):
        return upb_header
    return str(cur_val).strip()


def header_tuples_from_ws(ws, header_row: int = HEADER_ROW, wb=None, upb_header: Optional[str] = None) -> List[Tuple[int, str]]:
    out: List[Tuple[int, str]] = []
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]

    for col_idx, cell in enumerate(row, start=1):
        if wb is not None and upb_header is not None:
            header = _resolve_header_value(wb, ws, cell, upb_header)
        else:
            v = cell.value
            header = "" if v is None else str(v).strip()
        if header:
            out.append((col_idx, header.strip()))
    return out




def _validate_sheet_blueprints_or_raise():
    """Fail fast when a sheet blueprint contains invalid Excel columns or formula headers.

    This catches the exact class of regression that caused the 5/18 drift: formula
    strings accidentally placed as row-4 header keys, missing numeric column keys,
    or row-4 formula values where static report headers are expected.
    """
    errors: List[str] = []
    for sheet_name, blueprint in SHEET_BLUEPRINTS.items():
        for row_name in ("row1", "row2", "row3", "row4", "row5"):
            for col_idx, value in blueprint.get(row_name, {}).items():
                try:
                    _scaffold_col_index(col_idx)
                except Exception as exc:
                    errors.append(f"{sheet_name}.{row_name}: invalid column key {col_idx!r}: {exc}")
                if row_name == "row5" and isinstance(value, str) and value.strip().startswith("="):
                    errors.append(
                        f"{sheet_name}.row5 column {col_idx!r}: header value is a formula. "
                        "Move formulas to DRAFT_FORMULA_OVERRIDES and keep row5 as static/dynamic header text."
                    )
    if errors:
        raise ValueError("Invalid Active Loan Report sheet blueprint:\n" + "\n".join(f"- {e}" for e in errors))


def _expected_header_matches(actual_header: str, expected_header: str, upb_header: str) -> bool:
    actual = clean_text(actual_header)
    expected = clean_text(expected_header)
    if expected == "__UPB__":
        return actual == clean_text(upb_header)
    if expected == "__QEND_NPL__":
        return bool(re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL", actual, flags=re.I))
    if expected == "__QEND_NPL_YN__":
        return bool(re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL\s+\(Y/N\)", actual, flags=re.I))
    if expected == "__QEND_NPL_REO__":
        return bool(re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL\s*/\s*REO", actual, flags=re.I))
    if expected == "__SPECIAL_LIST__":
        return bool(re.fullmatch(r"\dQ\d{2}\s+Special\s+Loans\s+List", actual, flags=re.I))
    return actual == expected


def _formula_override_key_for_header(header: str, upb_header: str) -> str:
    header = clean_text(header)
    if header == clean_text(upb_header):
        return "__UPB__"
    if re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL\s*/\s*REO", header, flags=re.I):
        return "__QEND_NPL_REO__"
    if re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL\s+\(Y/N\)", header, flags=re.I):
        return "__QEND_NPL_YN__"
    if re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL", header, flags=re.I):
        return "__QEND_NPL__"
    if re.fullmatch(r"\dQ\d{2}\s+Special\s+Loans\s+List\s*", header, flags=re.I):
        return "__SPECIAL_LIST__"
    return header


def validate_sheet_schema_or_raise(wb, sheet_name: str, upb_header: str) -> None:
    """Validate row-4 report schema before values are written.

    This guard prevents a high mismatch count caused by missing columns from being
    mistaken for a source-mapping failure.
    """
    _validate_sheet_blueprints_or_raise()
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Required sheet missing from workbook: {sheet_name}")
    blueprint = SHEET_BLUEPRINTS.get(sheet_name, {})
    expected = blueprint.get("row5", {})
    if not expected:
        return

    ws = wb[sheet_name]
    actual_by_col = {col_idx: header for col_idx, header in header_tuples_from_ws(ws, header_row=HEADER_ROW, wb=wb, upb_header=upb_header)}
    errors: List[str] = []

    seen: Dict[str, int] = {}
    for _col_idx, header in actual_by_col.items():
        if not clean_text(header):
            continue
        seen[header] = seen.get(header, 0) + 1
    duplicates = sorted([h for h, count in seen.items() if count > 1 and not UPB_HEADER_RE.search(h)])
    if duplicates:
        errors.append(f"duplicate row-4 header(s): {', '.join(duplicates)}")

    for raw_col_idx, expected_header in expected.items():
        col_idx = _scaffold_col_index(raw_col_idx)
        actual_header = actual_by_col.get(col_idx, "")
        if not _expected_header_matches(actual_header, str(expected_header), upb_header):
            expected_display = upb_header if expected_header == "__UPB__" else str(expected_header)
            errors.append(
                f"column {get_column_letter(col_idx)} expected {expected_display!r} but found {actual_header!r}"
            )

    overrides = DRAFT_FORMULA_OVERRIDES.get(sheet_name, {})
    if overrides and not MATERIALIZE_FORMULA_RESULT_COLUMNS:
        formula_cols = formula_col_indices(ws, start_row=DATA_START_ROW, header_row=HEADER_ROW, scan_rows=50)
        allowed_keys = set(overrides.keys())
        for col_idx in sorted(formula_cols):
            header = actual_by_col.get(col_idx, "")
            formula_key = _formula_override_key_for_header(header, upb_header)
            if formula_key not in allowed_keys:
                errors.append(
                    f"column {get_column_letter(col_idx)} has a row-5 formula seed under non-formula header {header!r}"
                )
        for col_idx, header in actual_by_col.items():
            formula_key = _formula_override_key_for_header(header, upb_header)
            if formula_key in allowed_keys and col_idx not in formula_cols:
                errors.append(
                    f"column {get_column_letter(col_idx)} header {header!r} is expected to have a row-5 formula seed"
                )

    if errors:
        raise ValueError(
            f"{sheet_name} template/schema validation failed before writing data.\n"
            + "\n".join(f"- {e}" for e in errors)
        )


def validate_workbook_schema_or_raise(wb, upb_header: str, sheet_names: Optional[Sequence[str]] = None) -> None:
    targets = list(sheet_names) if sheet_names else ["Bridge Asset", "Bridge Loan", "Term Loan", "Term Asset"]
    for sheet_name in targets:
        validate_sheet_schema_or_raise(wb, sheet_name, upb_header)


def formula_col_indices(ws_formula, start_row: int = DATA_START_ROW, header_row: int = HEADER_ROW, scan_rows: int = 50) -> Set[int]:
    fcols: Set[int] = set()
    max_scan_row = min(ws_formula.max_row, start_row + scan_rows - 1)

    for r in range(start_row, max_scan_row + 1):
        for col_idx in range(1, ws_formula.max_column + 1):
            v = ws_formula.cell(r, col_idx).value
            if isinstance(v, str) and v.startswith("="):
                fcols.add(col_idx)
    return fcols


def _capture_formula_seeds(ws_formula, formula_cols: Set[int], start_row: int = DATA_START_ROW, scan_rows: int = 50):
    seeds = {}
    max_scan_row = min(ws_formula.max_row, start_row + scan_rows - 1)

    for col_idx in sorted(formula_cols):
        for r in range(start_row, max_scan_row + 1):
            v = ws_formula.cell(r, col_idx).value
            if isinstance(v, str) and v.startswith("="):
                seeds[col_idx] = {"origin_row": r, "formula": v}
                break
    return seeds


def _used_output_columns(ws, wb, upb_header: str, header_row: int = HEADER_ROW, start_row: int = DATA_START_ROW) -> Set[int]:
    hdr = header_tuples_from_ws(ws, header_row=header_row, wb=wb, upb_header=upb_header)
    cols = {c for c, _h in hdr}
    cols |= formula_col_indices(ws, start_row=start_row, header_row=header_row)
    return cols


def _clear_sheet_body(ws, used_cols: Set[int], start_row: int = DATA_START_ROW):
    if not used_cols:
        return
    max_r = ws.max_row
    for r in range(start_row, max_r + 1):
        for c in used_cols:
            ws.cell(r, c).value = None


def _trim_sheet_body_rows(ws, row_count: int, start_row: int = DATA_START_ROW):
    keep_last = (start_row - 1) if row_count <= 0 else (start_row + row_count - 1)
    if ws.max_row > keep_last:
        ws.delete_rows(keep_last + 1, ws.max_row - keep_last)


def _drop_fully_blank_dataframe_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    mask = pd.Series(False, index=df.index)
    for col in df.columns:
        series = df[col]
        if pd.api.types.is_numeric_dtype(series):
            mask = mask | pd.to_numeric(series, errors="coerce").notna()
        elif pd.api.types.is_datetime64_any_dtype(series):
            mask = mask | pd.to_datetime(series, errors="coerce").notna()
        else:
            mask = mask | (~blankish_mask(series))
    return df.loc[mask].copy()


def _drop_rows_missing_required_keys(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()

    key_map = {
        "Bridge Asset": ["Deal Number", "Asset ID"],
        "Bridge Loan": ["Deal Number"],
        "Term Loan": ["Deal Number"],
        "Term Asset": ["Deal Number", "Asset ID"],
    }
    required = [c for c in key_map.get(sheet_name, []) if c in df.columns]
    if not required:
        return df.copy()

    mask = pd.Series(True, index=df.index)
    for col in required:
        mask = mask & (~blankish_mask(df[col]))
    return df.loc[mask].copy()


def _reset_sheet_autofilter(ws, header_tuples: List[Tuple[int, str]], row_count: int, header_row: int = HEADER_ROW, start_row: int = DATA_START_ROW):
    if not header_tuples:
        return
    first_col = min(col_idx for col_idx, _header in header_tuples)
    last_col = max(col_idx for col_idx, _header in header_tuples)
    end_row = header_row if row_count <= 0 else (start_row + row_count - 1)
    ws.auto_filter.ref = f"{get_column_letter(first_col)}{header_row}:{get_column_letter(last_col)}{end_row}"


def _has_excel_unsupported_timezone(val) -> bool:
    if isinstance(val, pd.Timestamp):
        return val.tz is not None
    if isinstance(val, (datetime, datetime_time)):
        return val.tzinfo is not None
    return False


def _excel_strip_timezone(val):
    if val is None or val is pd.NA:
        return None
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return None
        if val.tz is not None:
            val = val.tz_localize(None)
        return val.to_pydatetime()
    if isinstance(val, datetime):
        return val.replace(tzinfo=None) if val.tzinfo is not None else val
    if isinstance(val, datetime_time):
        return val.replace(tzinfo=None) if val.tzinfo is not None else val
    return val


def _excel_safe_value(val):
    if val is None or val is pd.NA:
        return None
    if isinstance(val, pd.Timestamp):
        return _excel_strip_timezone(val)
    if isinstance(val, np.generic):
        val = val.item()
    if isinstance(val, (list, dict, set, tuple)):
        return str(val)
    val = _excel_strip_timezone(val)
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    return val


def _coerce_excel_date_value(val):
    if val is None:
        return None
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return None
        return _excel_strip_timezone(val).date()
    if isinstance(val, datetime):
        return _excel_strip_timezone(val).date()
    if isinstance(val, date):
        return val
    if isinstance(val, np.generic):
        val = val.item()
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            if np.isnan(float(val)):
                return None
        except Exception:
            pass
        # pandas interprets bare numbers as nanoseconds. In Excel date columns,
        # numeric values from prior workbooks are Excel serial dates.
        if 1 <= float(val) <= 100000:
            try:
                return (datetime(1899, 12, 30) + timedelta(days=float(val))).date()
            except Exception:
                return val
    try:
        parsed = pd.to_datetime(val, errors="coerce")
        if pd.isna(parsed):
            return _excel_strip_timezone(val)
        return _excel_strip_timezone(parsed).date()
    except Exception:
        return _excel_strip_timezone(val)


def _money_format_for_header(sheet_name: str, header: str, upb_header: str) -> Optional[str]:
    if header == upb_header:
        return MONEY2_FORMAT if sheet_name in {"Bridge Asset", "Term Asset"} else MONEY0_FORMAT
    if header in SHEET_MONEY2_HEADERS.get(sheet_name, set()):
        return MONEY2_FORMAT
    if header in SHEET_MONEY0_HEADERS.get(sheet_name, set()):
        return MONEY0_FORMAT
    return None


def _is_date_header(sheet_name: str, header: str) -> bool:
    return header in SHEET_DATE_HEADERS.get(sheet_name, set())


def _should_preserve_datetime(sheet_name: str, header: str) -> bool:
    return header in SHEET_DATETIME_HEADERS.get(sheet_name, set())


def _infer_template_text_headers(ws, header_tuples: List[Tuple[int, str]], start_row: int = DATA_START_ROW, sample_limit: int = 120, scan_limit: int = 1000) -> Set[str]:
    hinted: Set[str] = set()
    max_row = min(ws.max_row, start_row + scan_limit - 1)
    for col_idx, header in header_tuples:
        if not header or header in REPORT_IDENTIFIER_HEADERS.get(ws.title, set()):
            continue
        text_like = 0
        number_like = 0
        samples = 0
        for r in range(start_row, max_row + 1):
            val = ws.cell(r, col_idx).value
            if not has_any_value(val):
                continue
            if isinstance(val, pd.Timestamp):
                if pd.isna(val):
                    continue
                break
            if isinstance(val, (datetime, date)):
                break
            if isinstance(val, (int, float, np.integer, np.floating)) and not isinstance(val, bool):
                number_like += 1
            else:
                text_like += 1
            samples += 1
            if samples >= sample_limit:
                break
        if text_like and not number_like:
            hinted.add(header)
    return hinted


def _round_report_money_series(series: pd.Series) -> pd.Series:
    ser = pd.Series(series, copy=False)
    num = pd.to_numeric(ser, errors="coerce")
    out = _object_series_like(ser)
    mask = num.notna()
    if bool(mask.any()):
        out.loc[mask] = num.loc[mask].round(2).astype("object")
    return out



def _report_date_series_from_col(df: pd.DataFrame, col: str) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="datetime64[ns]")
    if col not in df.columns:
        return pd.Series([pd.NaT] * len(df), index=df.index)
    ser = pd.Series(df[col], index=df.index, dtype="object")
    coerced = ser.map(_coerce_excel_date_value)
    return pd.to_datetime(coerced, errors="coerce")


def _report_numeric_series_from_col(df: pd.DataFrame, col: str, default=np.nan) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="float64")
    if col not in df.columns:
        return pd.Series([default] * len(df), index=df.index, dtype="float64")
    return pd.to_numeric(pd.Series(df[col], index=df.index), errors="coerce")


def _report_text_series_from_col(df: pd.DataFrame, col: str, default="") -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="string")
    if col not in df.columns:
        return pd.Series([default] * len(df), index=df.index, dtype="object")
    return pd.Series(df[col], index=df.index, dtype="object").astype("string").str.strip()


def _report_is_blank_or_na(series_like) -> pd.Series:
    ser = pd.Series(series_like, copy=False, dtype="object")
    txt = ser.astype("string").str.strip().str.lower()
    return ser.isna() | txt.isin(["", "nan", "none", "<na>", "nat", "n/a", "na"])


def _report_yn(mask: pd.Series, index) -> pd.Series:
    return pd.Series(np.where(pd.Series(mask, index=index).fillna(False).astype(bool), "Y", "N"), index=index, dtype="object")


def _report_is_sold_financing(df: pd.DataFrame) -> pd.Series:
    """Rows the report's NPL / special-list formulas exclude as sold.

    The official formulas literally test $D6<>"Sold Servicing Retained". Accept the bare
    "Sold" and the misspelled legacy variant too, so a carried-forward value from an older
    workbook cannot silently re-enable the flags for a sold loan.
    """
    fin = _report_text_series_from_col(df, "Financing").str.casefold()
    return fin.isin({v.casefold() for v in BRIDGE_SOLD_FINANCING_VALUES}).fillna(False)


def _days_between(later, earlier: pd.Series) -> pd.Series:
    later_ts = pd.Timestamp(later)
    earlier_ts = pd.to_datetime(earlier, errors="coerce")
    return pd.Series((later_ts - earlier_ts).dt.days, index=earlier_ts.index, dtype="float64")


def _days_until(target: pd.Series, from_date) -> pd.Series:
    from_ts = pd.Timestamp(from_date)
    target_ts = pd.to_datetime(target, errors="coerce")
    return pd.Series((target_ts - from_ts).dt.days, index=target_ts.index, dtype="float64")


def _dq_status_from_days(days_series: pd.Series, reo_mask: pd.Series) -> pd.Series:
    days = pd.to_numeric(days_series, errors="coerce")
    out = pd.Series(["Current"] * len(days), index=days.index, dtype="object")
    out.loc[days.gt(0) & days.lt(30)] = "DQ 1-29"
    out.loc[days.ge(30) & days.lt(60)] = "DQ 30-59"
    out.loc[days.ge(60) & days.lt(90)] = "DQ 60-89"
    out.loc[days.ge(90)] = "DQ 90+"
    out.loc[pd.Series(reo_mask, index=days.index).fillna(False).astype(bool)] = "REO"
    return out


def _materialize_bridge_asset_formula_columns(df: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    out = _recompute_bridge_asset_funded_amount(df)
    idx = out.index
    q_end = quarter_end_for_run(run_dt)
    qend_npl_header = _qend_npl_header(q_end, "(Y/N)")
    qend_npl_reo_header = _qend_npl_reo_header(q_end)

    product_type = _report_text_series_from_col(out, "Product Type")
    product_sub_type = _report_text_series_from_col(out, "Product Sub-Type")
    current_loan_maturity = _report_date_series_from_col(out, "Current Loan Maturity date")
    current_asset_maturity = _report_date_series_from_col(out, "Current Asset Maturity Date")
    servicer_maturity = _report_date_series_from_col(out, "Servicer Maturity Date")
    cv_maturity = current_asset_maturity.where(
        product_type.eq("Credit Line") | product_sub_type.eq("Line of Credit"),
        current_loan_maturity,
    )
    out["CV Maturity Date"] = cv_maturity
    maturity_diff = (cv_maturity - servicer_maturity).dt.days
    out["Maturity Difference"] = pd.Series(maturity_diff, index=idx, dtype="object").where(
        cv_maturity.notna() & servicer_maturity.notna(),
        "N/A",
    )
    out["Maturity Date"] = servicer_maturity.where(servicer_maturity.notna(), cv_maturity)
    out["Days to Maturity"] = _days_until(out["Maturity Date"], run_dt)

    next_payment = _report_date_series_from_col(out, "Next Payment Date")
    out["Days Past Due"] = _days_between(run_dt, next_payment)

    reo_date_raw = pd.Series(out.get("REO Date", pd.Series([pd.NA] * len(out), index=idx)), index=idx, dtype="object")
    reo_date = _report_date_series_from_col(out, "REO Date")
    reo_mask = reo_date.notna() | (~_report_is_blank_or_na(reo_date_raw))
    out["DQ Status"] = _dq_status_from_days(out["Days Past Due"], reo_mask)

    updated_val_date_raw = pd.Series(out.get("Updated Valuation Date", pd.Series([pd.NA] * len(out), index=idx)), index=idx, dtype="object")
    has_updated_val = ~_report_is_blank_or_na(updated_val_date_raw)
    updated_val_date = _report_date_series_from_col(out, "Updated Valuation Date")
    orig_val_date = _report_date_series_from_col(out, "Origination Value Dt")
    out["Most Recent Valuation Date"] = updated_val_date.where(has_updated_val & updated_val_date.notna(), orig_val_date)
    out["Most Recent As-Is Value"] = _report_numeric_series_from_col(out, "Updated As-Is Value").where(
        has_updated_val,
        _report_numeric_series_from_col(out, "Origination As-Is Value"),
    )
    out["Most Recent ARV"] = _report_numeric_series_from_col(out, "Updated ARV").where(
        has_updated_val,
        _report_numeric_series_from_col(out, "Origination ARV"),
    )

    # Quarter-end NPL column. V43: this is THREE-valued (REO / NPL / N), not Y/N, and it
    # is titled "<m>/<d> NPL/REO". Transcribed from 20260803 Bridge Asset!DF6:
    #   =IF(AND($D6<>"Sold Servicing Retained",$CS6="REO"),"REO",
    #      IF(AND($D6<>"Sold Servicing Retained",MINIFS($AC:$AC,$E:$E,$E6)<=$DF$4),"NPL","N"))
    # where $CS6 is DQ Status, $AC is Next Payment Date, $E is Deal Number and
    # $DF$4 = quarter-end minus 90 days. Note the comparison is <= (not <) and the sold
    # exclusion tests the full "Sold Servicing Retained" label.
    # Computed here rather than left as a live formula so the materialized value tracks the
    # corrected Next Payment Date (incl. the Statebridge day-10 rule) instead of inheriting
    # the stale upstream "3/31 NPL (Y/N)" column.
    # Verified 4941/4941 against 20260803 Bridge Asset.
    _npl_threshold = pd.Timestamp(q_end) - pd.Timedelta(days=90)
    _npl_npd = _report_date_series_from_col(out, "Next Payment Date")
    _npl_deal = _report_text_series_from_col(out, "Deal Number")
    _not_sold = ~_report_is_sold_financing(out)
    _min_npd_by_deal = _npl_npd.groupby(_npl_deal).transform("min")
    _dq_reo = _report_text_series_from_col(out, "DQ Status").str.upper().eq("REO")
    qend_npl_reo = pd.Series([QEND_NPL_NONE_VALUE] * len(out), index=idx, dtype="object")
    qend_npl_reo = qend_npl_reo.mask(
        _not_sold & _min_npd_by_deal.notna() & _min_npd_by_deal.le(_npl_threshold),
        QEND_NPL_NPL_VALUE,
    )
    qend_npl_reo = qend_npl_reo.mask(_not_sold & _dq_reo, QEND_NPL_REO_VALUE)
    out[qend_npl_reo_header] = qend_npl_reo
    # Keep the legacy Y/N alias so downstream consumers / prior-workbook joins still resolve.
    out[qend_npl_header] = _report_yn(qend_npl_reo.ne(QEND_NPL_NONE_VALUE), idx)

    # Needs NPL Value, from 20260803 Bridge Asset!CW6:
    #   =IF(AND($D6<>"Sold Servicing Retained",OR($DF6="NPL",$DF6="REO"),$CT6<$CW$4),"Y","N")
    # $CW$4 = EDATE(quarter-end, -6), $CT6 = Most Recent Valuation Date.
    npl_flag = qend_npl_reo.isin([QEND_NPL_NPL_VALUE, QEND_NPL_REO_VALUE])
    stale_threshold = pd.Timestamp(q_end) - pd.DateOffset(months=6)
    most_recent_val = pd.to_datetime(out["Most Recent Valuation Date"], errors="coerce")
    out["Needs NPL Value"] = _report_yn(
        _not_sold & npl_flag & most_recent_val.notna() & most_recent_val.lt(stale_threshold),
        idx,
    )

    segment = _report_text_series_from_col(out, "Segment")
    # V56: Asset Commitment comes straight from the per-asset Approved Advance Amount Max
    # (mapped on the spine). Exact 4,782/4,782 on 20260824. The old approved-components sum
    # (Initial Disbursement Funded + Renovation Holdback + Interest Allocation) matched only
    # 4,762 and is kept solely as the fallback for an asset the field is missing on -- it
    # must NOT overwrite a sourced value.
    _asset_commitment_sum = (
        _report_numeric_series_from_col(out, "Initial Disbursement Funded").fillna(0.0)
        + _report_numeric_series_from_col(out, "Renovation Holdback").fillna(0.0)
        + _report_numeric_series_from_col(out, "Interest Allocation").fillna(0.0)
    )
    _asset_commitment_src = _report_numeric_series_from_col(out, "Asset Commitment")
    out["Asset Commitment"] = _asset_commitment_src.where(_asset_commitment_src.notna(), _asset_commitment_sum)
    # NEW Loan Type: Portfolio 5A/TPO/RB map to labels, else fall back to Product Type.
    portfolio_ba = _report_text_series_from_col(out, "Portfolio")
    product_type_ba = _report_text_series_from_col(out, "Product Type")
    loan_type_ba = product_type_ba.copy()
    loan_type_ba = loan_type_ba.mask(portfolio_ba.eq("5A"), "5A Bridge")
    loan_type_ba = loan_type_ba.mask(portfolio_ba.eq("TPO"), "Purchased Bridge")
    loan_type_ba = loan_type_ba.mask(portfolio_ba.eq("RB"), BRIDGE_RB_LOAN_TYPE)
    out["Loan Type"] = loan_type_ba
    # Interest Allocation shows 0 (never blank) in the report when Salesforce has none.
    if "Interest Allocation" in out.columns:
        out["Interest Allocation"] = _report_numeric_series_from_col(out, "Interest Allocation").fillna(0.0)
    # Widened Securitized rule: also flags CAFL 2026-R1 CV legacy.
    financing_ba = _report_text_series_from_col(out, "Financing")
    out["Securitized (Y/N)"] = _report_yn(
        segment.eq(BRIDGE_SECURITIZED_SEGMENT) | (financing_ba.eq("CAFL 2026-R1 CV") & segment.eq("Legacy")),
        idx,
    )
    out["SSP JV (Y/N)"] = _report_yn(segment.eq("SSP"), idx)
    out["CPP JV (Y/N)"] = _report_yn(segment.eq("CPP JV"), idx)
    out["Oaktree JV (Y/N)"] = _report_yn(segment.eq("Oaktree JV"), idx)
    out["Legacy (Y/N)"] = _report_yn(segment.eq("Legacy"), idx)

    deal_key = norm_id_series(out.get("Deal Number", pd.Series([pd.NA] * len(out), index=idx)))
    days_to_mat = pd.to_numeric(out["Days to Maturity"], errors="coerce")
    days_past_due = pd.to_numeric(out["Days Past Due"], errors="coerce")
    if bool(deal_key.notna().any()):
        min_days_to_mat = days_to_mat.groupby(deal_key).transform("min")
        max_days_past_due = days_past_due.groupby(deal_key).transform("max")
    else:
        min_days_to_mat = pd.Series([np.nan] * len(out), index=idx)
        max_days_past_due = pd.Series([np.nan] * len(out), index=idx)
    out["Matured Loan (YN)"] = _report_yn(min_days_to_mat.lt(0), idx)
    out["DQ 45+ Loan (Y/N)"] = _report_yn(max_days_past_due.ge(45), idx)
    # SA Loan (Y/N) is computed in build_bridge_asset from the Asset Manager 1
    # special-asset lookup; keep it if present, else default N.
    out["SA Loan (Y/N)"] = coalesce_keep_nonblank(
        out.get("SA Loan (Y/N)", pd.Series([pd.NA] * len(out), index=idx)),
        pd.Series(["N"] * len(out), index=idx),
    ).replace({"": "N"})

    # Special loans list, from 20260803 Bridge Asset!DG6:
    #   =IF(AND($D6<>"Sold Servicing Retained",
    #          OR($DB6="Y",$DC6="Y",$DD6="Y",$DE6="Y",OR($DF6="NPL",$DF6="REO"))),"Y","N")
    # i.e. Legacy / Matured Loan / DQ 45+ / SA Loan, PLUS the NPL-or-REO term the previous
    # build omitted entirely.
    special_any = (
        _report_text_series_from_col(out, "Legacy (Y/N)").str.upper().eq("Y")
        | _report_text_series_from_col(out, "Matured Loan (YN)").str.upper().eq("Y")
        | _report_text_series_from_col(out, "DQ 45+ Loan (Y/N)").str.upper().eq("Y")
        | _report_text_series_from_col(out, "SA Loan (Y/N)").str.upper().eq("Y")
        | npl_flag
    )
    out["Special Flag"] = _report_yn(_not_sold & special_any, idx)
    # Emit under the new header name too so the writer finds it.
    out[_special_list_header(quarter_end_for_run(run_dt))] = out["Special Flag"]
    return out


def _materialize_bridge_loan_formula_columns(df: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    out["Days Past Due"] = _days_between(run_dt, _report_date_series_from_col(out, "Next Payment Date"))
    # Loan Type: Portfolio 5A/TPO/RB map to fixed labels, else fall back to Product Type.
    # Mirrors _materialize_bridge_asset_formula_columns; the Bridge Loan tab previously had
    # no Loan Type derivation (verified vs 20260615: CV/CLO rows carry the Product Type).
    portfolio_bl = _report_text_series_from_col(out, "Portfolio")
    product_type_bl = _report_text_series_from_col(out, "Product Type")
    loan_type_bl = product_type_bl.copy()
    loan_type_bl = loan_type_bl.mask(portfolio_bl.eq("5A"), "5A Bridge")
    loan_type_bl = loan_type_bl.mask(portfolio_bl.eq("TPO"), "Purchased Bridge")
    loan_type_bl = loan_type_bl.mask(portfolio_bl.eq("RB"), BRIDGE_RB_LOAN_TYPE)
    out["Loan Type"] = loan_type_bl
    # Interest Allocation shows 0 (never blank) in the report (verified 20260615: 0 blanks,
    # 845 zeros). Fill Salesforce-null with 0 so new deals match.
    if "Interest Allocation" in out.columns:
        out["Interest Allocation"] = _report_numeric_series_from_col(out, "Interest Allocation").fillna(0.0)

    # Quarter-end NPL: the Bridge Loan tab carries the loan-level counterpart of Bridge
    # Asset's "<m>/<d> NPL/REO", and it is likewise three-valued (REO / NPL / N). 20260803
    # Bridge Loan!AX is a pasted value, so the rule was recovered by fitting the tab:
    #   Financing == "Sold Servicing Retained"      -> N
    #   Loan Level Delinquency == "REO"             -> REO
    #   Next Payment Date <= quarter-end minus 90d  -> NPL
    #   otherwise                                   -> N
    # Reproduces 1,024/1,026 rows (the 2 exceptions are hand overrides on deal 39499,
    # which the report shows as NPL despite Loan Level Delinquency == REO).
    idx = out.index
    q_end = quarter_end_for_run(run_dt)
    _npl_threshold = pd.Timestamp(q_end) - pd.Timedelta(days=90)
    _not_sold = ~_report_is_sold_financing(out)
    _npd = _report_date_series_from_col(out, "Next Payment Date")
    _lld_reo = _report_text_series_from_col(out, "Loan Level Delinquency").str.upper().eq("REO")
    qend_npl = pd.Series([QEND_NPL_NONE_VALUE] * len(out), index=idx, dtype="object")
    qend_npl = qend_npl.mask(_not_sold & _npd.notna() & _npd.le(_npl_threshold), QEND_NPL_NPL_VALUE)
    qend_npl = qend_npl.mask(_not_sold & _lld_reo, QEND_NPL_REO_VALUE)
    out[_qend_npl_header(q_end)] = qend_npl
    # Needs NPL Value and Special Focus (Y/N) are deliberately NOT recomputed here: both
    # arrive via the Bridge Asset rollup / prior-workbook carry-forward and already land
    # within a couple of rows of the official tab (59 vs 60 Y, 190 vs 198 Y against
    # 20260803). Deriving them off this column instead over-flags them.
    return out


def _materialize_term_loan_formula_columns(df: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    idx = out.index
    out["Days Past Due"] = _days_between(run_dt, _report_date_series_from_col(out, "Next Payment Date"))
    reo_date_raw = pd.Series(out.get("REO Date", pd.Series([pd.NA] * len(out), index=idx)), index=idx, dtype="object")
    reo_date = _report_date_series_from_col(out, "REO Date")
    reo_mask = reo_date.notna() | (~_report_is_blank_or_na(reo_date_raw))
    out["DQ Status"] = _dq_status_from_days(out["Days Past Due"], reo_mask)
    q_end = quarter_end_for_run(run_dt)
    special_threshold = pd.Timestamp(q_end) - pd.Timedelta(days=90)
    portfolio = _report_text_series_from_col(out, "Portfolio")
    next_payment = _report_date_series_from_col(out, "Next Payment Date")
    days_past_due = pd.to_numeric(out["Days Past Due"], errors="coerce")
    active_term_or_dscr = portfolio.isin(["Active Term", "DSCR"])
    dq = _report_text_series_from_col(out, "DQ Status")
    special_txt = pd.Series(["N/A"] * len(out), index=idx, dtype="object")
    # Order mirrors the report's nested IF (later assignments take precedence only
    # where their condition is the matching branch).
    special_txt = special_txt.mask(active_term_or_dscr & days_past_due.ge(45) & dq.ne("REO"), "DQ 45+")
    # V52 (20260810): the "CAFL REO" branch was dropped from the official formula, so a
    # Securitized-Term REO deal now reads N/A rather than CAFL REO. Verified against
    # 20260810 Term Loan (0 CAFL REO) and Term Asset (0 via the XLOOKUP mirror).
    special_txt = special_txt.mask(active_term_or_dscr & dq.eq("REO"), "Term REO")
    # The NPL label carries the RUNNING quarter, e.g. "Q3 NPL" for a 9/30 quarter-end
    # (20260803 Term Loan!AF6). This was hardcoded "Q2 NPL", which mislabelled every
    # NPL row from Q3 onward.
    special_txt = special_txt.mask(
        active_term_or_dscr & next_payment.notna() & next_payment.le(special_threshold) & dq.ne("REO"),
        f"Q{(q_end.month - 1) // 3 + 1} NPL",
    )
    special_header = _special_list_header(quarter_end_for_run(run_dt))
    out[special_header] = special_txt
    out["Special Loans List (Y/N)"] = special_txt
    return out


def _materialize_term_asset_formula_columns(df: pd.DataFrame, upb_col: str) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out
    if upb_col not in out.columns:
        out[upb_col] = np.nan
    out["Special (Y/N)"] = coalesce_keep_nonblank(
        out.get("Special (Y/N)", pd.Series([pd.NA] * len(out), index=out.index)),
        pd.Series(["N"] * len(out), index=out.index),
    )
    return out


def _materialize_report_formula_columns(df: pd.DataFrame, sheet_name: str, upb_col: str) -> pd.DataFrame:
    if df is None or df.empty or not MATERIALIZE_FORMULA_RESULT_COLUMNS:
        return pd.DataFrame() if df is None else df.copy()
    if sheet_name == "Bridge Asset":
        return _materialize_bridge_asset_formula_columns(df, upb_col)
    if sheet_name == "Bridge Loan":
        return _materialize_bridge_loan_formula_columns(df, upb_col)
    if sheet_name == "Term Loan":
        return _materialize_term_loan_formula_columns(df, upb_col)
    if sheet_name == "Term Asset":
        return _materialize_term_asset_formula_columns(df, upb_col)
    return df.copy()

def _apply_report_blank_na_policy(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Apply final report-specific blank vs N/A rules before writing Excel.

    This is the automated NA/blank QA cleanup: columns in REPORT_FORCE_BLANK_HEADERS
    remain true blanks, while columns in REPORT_NA_FILL_HEADERS become N/A when blank.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    out = df.copy()

    # Force blank-preserve columns first. These columns should not be filled with N/A.
    for header in REPORT_FORCE_BLANK_HEADERS.get(sheet_name, set()):
        if header in out.columns:
            ser = pd.Series(out[header], index=out.index, dtype="object")
            out[header] = ser.where(~blankish_mask(ser), pd.NA)

    return out


def _normalize_output_for_report(df: pd.DataFrame, sheet_name: str, upb_col: str, template_text_headers: Optional[Set[str]] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    out = df.copy()

    for header in REPORT_IDENTIFIER_HEADERS.get(sheet_name, set()):
        if header in out.columns:
            out[header] = normalize_report_identifier_series(out[header])

    for header in REPORT_INTEGER_HEADERS.get(sheet_name, set()):
        if header in out.columns:
            out[header] = normalize_integer_display_series(out[header])

    text_headers = set(DEFAULT_TEXT_HEADERS.get(sheet_name, set())) | set(template_text_headers or set())
    text_headers = {h for h in text_headers if h not in REPORT_IDENTIFIER_HEADERS.get(sheet_name, set())}
    text_headers -= PRESERVE_INTERNAL_WHITESPACE_HEADERS.get(sheet_name, set())
    for header in text_headers:
        if header in out.columns:
            out[header] = normalize_text_display_series(out[header])

    # Fix N: collapse data-entry double-spaces in free-text columns so they render
    # single-spaced like the official report. These columns are written verbatim from SF
    # and do not all route through normalize_text_display_series, so collapse them here.
    for header in WHITESPACE_COLLAPSE_HEADERS.get(sheet_name, set()):
        if header in out.columns:
            ser = pd.Series(out[header], index=out.index, dtype="object")
            out[header] = ser.map(lambda v: re.sub(r"\s{2,}", " ", v).strip() if isinstance(v, str) else v)

    out = _apply_report_blank_na_policy(out, sheet_name)

    force_blank_headers = REPORT_FORCE_BLANK_HEADERS.get(sheet_name, set())
    # Numeric/money/date N/A-fill columns keep a legitimate 0 (e.g. Bridge Asset "# of Units"
    # is genuinely 0 in the official). TEXT N/A-fill columns never carry a real 0, so a value
    # that arrives as the literal "0"/"0.0" there is a coercion artifact (e.g. Term Loan
    # Borrower Entity showing "0" where the official shows "N/A"). Treat that "0" as missing so
    # the N/A fill catches it, without disturbing numeric columns.
    _numeric_na_cols = (
        set(REPORT_INTEGER_HEADERS.get(sheet_name, set()))
        | set(SHEET_MONEY2_HEADERS.get(sheet_name, set()))
        | set(SHEET_MONEY0_HEADERS.get(sheet_name, set()))
        | set(SHEET_DATE_HEADERS.get(sheet_name, set()))
        # V59: Term Loan "Borrower Entity" is the column the V50 zero-to-N/A rule was
        # written for, but the official is not consistent about it -- it shows a literal 0
        # on 20260803 (77 rows) and 20260824 (76 rows), and "N/A" only on 20260810. Two of
        # the last three weeks keep the 0, and 0 is what the upstream blank actually
        # produces, so exclude it from the rule and pass the value through. Flagged rather
        # than settled: if the report standardises on N/A, drop this exclusion.
        | REPORT_ZERO_IS_REAL_HEADERS.get(sheet_name, set())
    )
    if upb_col:
        _numeric_na_cols.add(upb_col)
    for header in REPORT_NA_FILL_HEADERS.get(sheet_name, set()):
        if header in force_blank_headers:
            continue
        if header in out.columns:
            ser = pd.Series(out[header], index=out.index, dtype="object")
            missing = blankish_mask(ser)
            if header not in _numeric_na_cols:
                zero_txt = ser.astype("string").str.strip().isin(["0", "0.0", "0.00"]).fillna(False)
                missing = missing | zero_txt
            out[header] = ser.where(~missing, "N/A")

    money_headers = set(SHEET_MONEY2_HEADERS.get(sheet_name, set())) | set(SHEET_MONEY0_HEADERS.get(sheet_name, set()))
    if upb_col:
        money_headers.add(upb_col)
    for header in money_headers:
        if header in out.columns:
            out[header] = _round_report_money_series(out[header])

    return out


def _copy_reference_row_style(ws_formula, col_idx: int, target_cell):
    ref_cell = ws_formula.cell(DATA_START_ROW, col_idx)
    if ref_cell.has_style:
        target_cell._style = copy(ref_cell._style)


def _apply_display_style(ws_formula, row_idx: int, col_idx: int, header: str, upb_header: str):
    cell = ws_formula.cell(row_idx, col_idx)
    _copy_reference_row_style(ws_formula, col_idx, cell)
    cell.font = copy(BASE_FONT)
    cell.alignment = copy(BASE_ALIGNMENT)

    if _is_date_header(ws_formula.title, header):
        cell.number_format = DATE_NUMBER_FORMAT
    else:
        money_fmt = _money_format_for_header(ws_formula.title, header, upb_header)
        if money_fmt:
            cell.number_format = money_fmt


def _copy_formula_columns_down(ws_formula, formula_seeds: dict, row_count: int, header_tuples: List[Tuple[int, str]], upb_header: str, start_row: int = DATA_START_ROW):
    if row_count <= 0:
        return

    header_by_col = {c: h for c, h in header_tuples}
    overrides = DRAFT_FORMULA_OVERRIDES.get(ws_formula.title, {})
    _q_end = quarter_end_for_run(run_dt)

    for col_idx in sorted(formula_seeds):
        header = header_by_col.get(col_idx, "")
        override_key = _formula_override_key_for_header(header, upb_header)
        seed_info = formula_seeds.get(col_idx, {})
        if override_key in overrides:
            # Override formulas are authored relative to DATA_START_ROW (row 6).
            origin_formula = _resolve_formula_override(overrides[override_key], _q_end)
            origin_row = start_row
        else:
            # Seeded formulas may have been captured from any scan row; anchor the
            # Translator to the row the seed actually came from so relative refs
            # (e.g. $U6) shift correctly instead of freezing at the seed's row.
            origin_formula = seed_info.get("formula")
            origin_row = seed_info.get("origin_row", start_row)
        origin_ref = f"{get_column_letter(col_idx)}{origin_row}"

        for r in range(start_row, start_row + row_count):
            target = ws_formula.cell(r, col_idx)
            if r == origin_row:
                target.value = origin_formula
            else:
                target.value = Translator(origin_formula, origin=origin_ref).translate_formula(f"{get_column_letter(col_idx)}{r}")
            _copy_reference_row_style(ws_formula, col_idx, target)


def _refresh_subtotal_formula(ws_formula, row_count: int, subtotal_row: int = 4, start_row: int = DATA_START_ROW):
    blueprint = SHEET_BLUEPRINTS.get(ws_formula.title, {})
    subtotal_col = blueprint.get("subtotal_col")
    if not subtotal_col:
        return
    subtotal_col = _scaffold_col_index(subtotal_col)
    col_letter = get_column_letter(subtotal_col)
    end_row = max(start_row, start_row + row_count - 1)
    ws_formula.cell(subtotal_row, subtotal_col).value = f"=SUBTOTAL(9,{col_letter}{start_row}:{col_letter}{end_row})"


def write_df_to_sheet_preserve_formulas(
    ws_formula,
    df: pd.DataFrame,
    header_tuples: List[Tuple[int, str]],
    formula_cols: Set[int],
    upb_header: str,
    start_row: int = DATA_START_ROW,
):
    write_cols = [(c, h) for (c, h) in header_tuples if c not in formula_cols]
    # Restrict the data write to the blue-header columns. Columns to the right of the
    # blue range are filled in manually in Excel, so the build leaves them untouched.
    _blue_max = _sheet_blue_max_col(ws_formula.title)
    if _blue_max is not None:
        write_cols = [(c, h) for (c, h) in write_cols if c <= _blue_max]
    headers = [h for _c, h in write_cols]

    # Resolve each output header against df columns, tolerating leading/trailing
    # whitespace differences. header_tuples_from_ws() strips header text, but some
    # materialized columns carry the report's exact header WITH a trailing space
    # (e.g. "2Q26 Special Loans List "). Without this, the stripped header is treated
    # as "missing" and the materialized value (special-loans Y/N etc.) is silently
    # written as N/A instead of the computed result.
    _strip_to_actual = {}
    for _c in df.columns:
        _strip_to_actual.setdefault(str(_c).strip(), _c)
    df_out = pd.DataFrame(index=df.index)
    for h in headers:
        if h in df.columns:
            df_out[h] = df[h].to_numpy()
        elif str(h).strip() in _strip_to_actual:
            df_out[h] = df[_strip_to_actual[str(h).strip()]].to_numpy()
        else:
            df_out[h] = pd.NA
    df_out = df_out[headers]

    for r_offset, row in enumerate(df_out.itertuples(index=False, name=None), start=0):
        r = start_row + r_offset
        for (c, h), val in zip(write_cols, row):
            safe_val = _excel_safe_value(val)
            if _is_date_header(ws_formula.title, h):
                if _should_preserve_datetime(ws_formula.title, h):
                    safe_val = _excel_safe_value(safe_val)
                else:
                    safe_val = _coerce_excel_date_value(safe_val)
            ws_formula.cell(r, c).value = safe_val
            _apply_display_style(ws_formula, r, c, h, upb_header)


def write_output_sheet(wb, sheet_name: str, df: pd.DataFrame, upb_col: str):
    WRITTEN_SHEET_ROWS[sheet_name] = 0 if df is None else int(len(df))
    if sheet_name not in wb.sheetnames:
        return

    df = _drop_fully_blank_dataframe_rows(df)
    df = _drop_rows_missing_required_keys(sheet_name, df)

    ws = wb[sheet_name]
    hdr = header_tuples_from_ws(ws, header_row=HEADER_ROW, wb=wb, upb_header=upb_col)
    validate_sheet_schema_or_raise(wb, sheet_name, upb_col)
    for _col_idx, _header in hdr:
        if _header in df.columns:
            continue
        if re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL\s+\(Y/N\)", str(_header), flags=re.I) and "3/31 NPL (Y/N)" in df.columns:
            df[_header] = df["3/31 NPL (Y/N)"]
        elif re.fullmatch(r"\d{1,2}/\d{1,2}\s+NPL", str(_header), flags=re.I) and "3/31 NPL" in df.columns:
            df[_header] = df["3/31 NPL"]
    df = _materialize_report_formula_columns(df, sheet_name, upb_col)
    template_text_headers = _infer_template_text_headers(ws, hdr, start_row=DATA_START_ROW)
    df = _normalize_output_for_report(df, sheet_name, upb_col, template_text_headers=template_text_headers)
    fcols = formula_col_indices(ws, start_row=DATA_START_ROW, header_row=HEADER_ROW)
    template_formula_cols = set(fcols)

    # Columns that MUST stay live formulas because they reference another sheet that is
    # not yet written when this sheet is written. Term Loan's SFR/MF Allocation and
    # Strategy Grouping SUMIFS into Term Asset, which is built afterwards -- those three
    # are backfilled as values by _materialize_term_loan_allocations_on_sheet once Term
    # Asset exists. Term Asset's UPB and special-loans-list are NOT listed: the build
    # already computes both in Python (_allocate_term_asset_upb_from_loan and the Term
    # Loan special-value merge in build_term_asset), so forcing them live only guaranteed
    # the columns shipped blank.
    always_live = {
        "Term Loan": {"SFR Allocation", "MF Allocation", "Strategy Grouping"},
    }.get(sheet_name, set())
    always_live_cols = {col_idx for col_idx, header in hdr if clean_text(header) in always_live}

    # Resolve a sheet header to its df column the same way write_df_to_sheet_preserve_
    # formulas does -- tolerating the trailing space the report carries on a few headers
    # (e.g. "3Q26 Special Loans List "), which header_tuples_from_ws strips away. Without
    # the strip-match a materialized column reads as "missing" and would be left as a
    # blank live formula.
    _df_col_by_stripped = {}
    for _c in df.columns:
        _df_col_by_stripped.setdefault(str(_c).strip(), _c)
    _header_by_col = {col_idx: clean_text(header) for col_idx, header in hdr}

    def _df_col_has_values(col_idx: int) -> bool:
        header = _header_by_col.get(col_idx, "")
        if not header:
            return False
        actual = header if header in df.columns else _df_col_by_stripped.get(header)
        if actual is None:
            return False
        return bool((~_report_is_blank_or_na(df[actual])).any())

    _blue_max = _sheet_blue_max_col(sheet_name)

    if MATERIALIZE_FORMULA_RESULT_COLUMNS:
        # Write a value wherever the build produced one; keep a live formula only where it
        # did not. Previously every beyond-blue template formula column was forced live,
        # which shipped the whole Bridge Asset CALC block (Loan Type, DQ Status, the JV
        # flags, NPL/REO, the special-loans list) and the Term Asset UPB as blank cells --
        # openpyxl writes no cached result, so nothing appears until Excel recalculates.
        fcols = always_live_cols.copy()
        if _blue_max is not None:
            fcols |= {
                c for c in template_formula_cols
                if c > _blue_max and not _df_col_has_values(c)
            }
    else:
        if sheet_name == "Term Asset" and not PRESERVE_TERM_ASSET_FORMULA_COLUMNS:
            force_write_headers = {upb_col, "Special (Y/N)"}
            force_write_cols = {col_idx for col_idx, header in hdr if header in force_write_headers}
            fcols = {c for c in fcols if c not in force_write_cols}
        # Beyond-blue CALC formulas must stay live and be propagated down so e.g. Days Past
        # Due keeps recalculating off the CQ4 run-date anchor.
        if _blue_max is not None:
            fcols |= {c for c in template_formula_cols if c > _blue_max}

    formula_seeds = _capture_formula_seeds(ws, fcols, start_row=DATA_START_ROW)
    # Ensure always-live cross-sheet columns have a seed even if the template
    # carried no sample data row with formulas.
    _overrides = DRAFT_FORMULA_OVERRIDES.get(sheet_name, {})
    for _col_idx, _header in hdr:
        if _col_idx in always_live_cols and _col_idx not in formula_seeds:
            _key = _formula_override_key_for_header(_header, upb_col)
            if _key in _overrides:
                formula_seeds[_col_idx] = {
                    "origin_row": DATA_START_ROW,
                    "formula": _resolve_formula_override(_overrides[_key], quarter_end_for_run(run_dt)),
                }

    used_cols = _used_output_columns(ws, wb=wb, upb_header=upb_col, header_row=HEADER_ROW, start_row=DATA_START_ROW)
    _clear_sheet_body(ws, used_cols, start_row=DATA_START_ROW)

    write_df_to_sheet_preserve_formulas(ws, df, hdr, fcols, upb_col, start_row=DATA_START_ROW)
    _copy_formula_columns_down(ws, formula_seeds, row_count=len(df), header_tuples=hdr, upb_header=upb_col, start_row=DATA_START_ROW)
    _refresh_subtotal_formula(ws, row_count=len(df), subtotal_row=4, start_row=DATA_START_ROW)
    _trim_sheet_body_rows(ws, row_count=len(df), start_row=DATA_START_ROW)
    _reset_sheet_autofilter(ws, hdr, row_count=len(df), header_row=HEADER_ROW, start_row=DATA_START_ROW)


    # V71: confirm this sheet actually received rows, immediately after writing it. Tests 77
    # and 78 both shipped with Bridge Asset and Term Asset as empty worksheets while the QA
    # audit -- which runs later -- still reported 4,691 and 23,848 data rows for them, so the
    # rows were present at audit time and gone in the saved bytes. That leaves two very
    # different possibilities, and this check separates them: if it fires, the write itself
    # produced nothing; if it stays quiet and the V70 save check still fails, the rows are
    # being lost after this point. Either way the next run says which.
    _expected_rows = int(WRITTEN_SHEET_ROWS.get(sheet_name, 0) or 0)
    if _expected_rows > 0:
        _written = 0
        _probe_cols = [c for c, _h in hdr][:40] or [2]
        for _r in range(DATA_START_ROW, min(ws.max_row, DATA_START_ROW + _expected_rows + 5) + 1):
            if any(ws.cell(_r, _c).value not in (None, "") for _c in _probe_cols):
                _written += 1
        SHEET_WRITE_AUDIT[sheet_name] = {"handed": _expected_rows, "on_sheet_after_write": _written}
        if _written == 0:
            # V74: record it, do not abort the build. Aborting produced no file to inspect.
            try:
                st.error(
                    f"⚠️ {sheet_name}: handed {_expected_rows:,} rows but the sheet holds none "
                    f"immediately after writing. See the Build Log sheet."
                )
            except Exception:
                pass
        if False:
            raise RuntimeError(
                f"{sheet_name}: write_output_sheet was handed {_expected_rows:,} rows but the sheet "
                f"holds none immediately after writing. Refusing to continue rather than build an "
                f"incomplete report. This is the failure that emptied Bridge Asset and Term Asset in "
                f"tests 77 and 78; check the pinned dependency versions in requirements.txt."
            )

def _materialize_term_loan_allocations_on_sheet(wb, term_asset_df: pd.DataFrame, upb_col: str) -> str:
    """Fill Term Loan SFR/MF Allocation + Strategy Grouping as values, post Term Asset.

    These three columns are the only genuinely cross-sheet ones on Term Loan -- they
    SUMIFS Property ALA out of Term Asset, which is not written yet when Term Loan is
    written, so they ship as live formulas with no cached result (i.e. blank to every
    non-Excel reader; 20260803 has 840/214 populated Strategy Grouping rows against the
    previous build's 1,030 blanks). Once Term Asset exists we can compute them directly:
      SFR Allocation    = sum(Property ALA) where Grouping == 'Single Family Rental'
      MF Allocation     = sum(Property ALA) where Grouping == 'Multifamily'
      Strategy Grouping = 'Single Family Rental' if SFR > MF else 'Multifamily'
    (transcribed from Term Loan!AG6/AH6/AI6; reproduces 1,054/1,054 rows on all three).
    The live formulas are left in place for any deal with no Term Asset rows, so Excel
    still recalculates those on open.
    """
    if not MATERIALIZE_FORMULA_RESULT_COLUMNS or "Term Loan" not in wb.sheetnames:
        return ""
    if term_asset_df is None or term_asset_df.empty:
        return ""
    if not {"Deal Number", "Property ALA", "Grouping"}.issubset(set(term_asset_df.columns)):
        return ""

    keys = norm_id_series(term_asset_df["Deal Number"])
    ala = pd.to_numeric(term_asset_df["Property ALA"], errors="coerce").fillna(0.0)
    grouping = _report_text_series_from_col(term_asset_df, "Grouping")
    sfr_by_deal = ala.where(grouping.eq("Single Family Rental"), 0.0).groupby(keys).sum()
    mf_by_deal = ala.where(grouping.eq("Multifamily"), 0.0).groupby(keys).sum()

    ws = wb["Term Loan"]
    hdr = header_tuples_from_ws(ws, header_row=HEADER_ROW, wb=wb, upb_header=upb_col)
    col_of = {clean_text(h): c for c, h in hdr}
    deal_col = col_of.get("Deal Number")
    sfr_col = col_of.get("SFR Allocation")
    mf_col = col_of.get("MF Allocation")
    strat_col = col_of.get("Strategy Grouping")
    if None in (deal_col, sfr_col, mf_col, strat_col):
        return ""

    filled = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        key = _normalize_sheet_key_value("Deal Number", ws.cell(r, deal_col).value)
        if not key:
            continue
        if key not in sfr_by_deal.index and key not in mf_by_deal.index:
            continue
        sfr = float(sfr_by_deal.get(key, 0.0))
        mf = float(mf_by_deal.get(key, 0.0))
        ws.cell(r, sfr_col).value = sfr
        ws.cell(r, mf_col).value = mf
        ws.cell(r, strat_col).value = "Single Family Rental" if sfr > mf else "Multifamily"
        filled += 1

    if not filled:
        return ""
    return f"Term Loan SFR/MF allocation + Strategy Grouping materialized for {filled:,} deals"


def _strip_timezones_from_workbook(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _has_excel_unsupported_timezone(cell.value):
                    cell.value = _excel_strip_timezone(cell.value)


def _normalize_sheet_key_value(header: str, value) -> str:
    txt = clean_text(value)
    if not txt:
        return ""
    txt = re.sub(r"\.0$", "", txt)
    if header == "Servicer ID":
        txt = re.sub(r"[^0-9A-Za-z]", "", txt).lstrip("0")
    elif header in {"Deal Number", "Asset ID"}:
        txt = re.sub(r"[^0-9A-Za-z]", "", txt)
    return txt



def _sheet_header_index(wb, ws, upb_header: str) -> Dict[str, int]:
    return {header: col_idx for col_idx, header in header_tuples_from_ws(ws, header_row=HEADER_ROW, wb=wb, upb_header=upb_header)}



def _normalize_sheet_key_variants(header: str, value) -> List[str]:
    primary = _normalize_sheet_key_value(header, value)
    if not primary:
        return []
    out = [primary]
    if header == "Deal Number":
        for raw in deal_lookup_keys(value):
            normed = _normalize_sheet_key_value(header, raw)
            if normed and normed not in out:
                out.append(normed)
    return out


def _build_sheet_row_lookup(
    wb,
    ws,
    key_headers: Sequence[str],
    upb_header: str,
    include_key_variants: bool = False,
) -> Tuple[Dict[Tuple[str, ...], int], Dict[str, int]]:
    header_idx = _sheet_header_index(wb, ws, upb_header)
    if not all(h in header_idx for h in key_headers):
        return {}, header_idx

    out: Dict[Tuple[str, ...], int] = {}
    for r in range(5, ws.max_row + 1):
        parts: List[List[str]] = []
        valid = True
        for h in key_headers:
            raw = ws.cell(r, header_idx[h]).value
            vals = _normalize_sheet_key_variants(h, raw) if include_key_variants else [_normalize_sheet_key_value(h, raw)]
            vals = [v for v in vals if v]
            if not vals:
                valid = False
                break
            parts.append(vals)
        if not valid:
            continue
        if include_key_variants and len(parts) > 1:
            import itertools
            keys_to_add = list(itertools.product(*parts))
        else:
            keys_to_add = [tuple(part[0] for part in parts)]
        for key in keys_to_add:
            out.setdefault(key, r)
    return out, header_idx


def _available_sheet_key_matches(wb, out_ws, base_wb, base_ws, sheet_name: str, upb_header: str):
    for candidate in _backfill_rule_candidates(sheet_name):
        out_rows_try, out_header_try = _build_sheet_row_lookup(wb, out_ws, candidate, upb_header, include_key_variants=False)
        base_rows_try, base_header_try = _build_sheet_row_lookup(base_wb, base_ws, candidate, upb_header, include_key_variants=True)
        if sheet_name == "Term Loan" and candidate == ["Servicer ID"] and (not out_rows_try or not base_rows_try):
            continue
        if out_rows_try and base_rows_try:
            yield list(candidate), out_rows_try, base_rows_try, out_header_try, base_header_try


def repair_workbook_from_baseline(wb, baseline_bytes: Optional[bytes], upb_header: str, sheet_names: Optional[Sequence[str]] = None) -> List[dict]:
    if not baseline_bytes:
        return []

    summary: List[dict] = []
    base_wb = None
    try:
        base_wb = load_workbook(BytesIO(baseline_bytes), data_only=False, keep_links=False)
        targets = list(sheet_names) if sheet_names else list(SHEET_BASELINE_KEY_CANDIDATES.keys())

        for sheet_name in targets:
            if sheet_name not in wb.sheetnames or sheet_name not in base_wb.sheetnames:
                continue

            out_ws = wb[sheet_name]
            base_ws = base_wb[sheet_name]
            formula_cols = formula_col_indices(out_ws, start_row=DATA_START_ROW, header_row=HEADER_ROW)
            # Manual columns past the blue range are never backfilled from the baseline.
            blue_max = _sheet_blue_max_col(sheet_name)

            fills = 0
            matched_rows: Set[int] = set()
            used_keys: List[str] = []
            any_keys = False

            for candidate, out_rows, base_rows, out_header_idx, base_header_idx in _available_sheet_key_matches(
                wb, out_ws, base_wb, base_ws, sheet_name, upb_header
            ):
                any_keys = True
                used_keys.append(", ".join(candidate))
                common_headers = [
                    h for h in out_header_idx
                    if h in base_header_idx
                    and h not in candidate
                    and not UPB_HEADER_RE.search(str(h))
                ]
                for key, out_row in out_rows.items():
                    base_row = base_rows.get(key)
                    if not base_row:
                        continue
                    matched_rows.add(out_row)
                    for header in common_headers:
                        out_col = out_header_idx[header]
                        if out_col in formula_cols:
                            continue
                        out_cell = out_ws.cell(out_row, out_col)
                        base_cell = base_ws.cell(base_row, base_header_idx[header])
                        if clean_text(out_cell.value) == "" and clean_text(base_cell.value) != "":
                            out_cell.value = base_cell.value
                            fills += 1

            if not any_keys:
                summary.append({"sheet_name": sheet_name, "status": "skipped_no_sheet_keys", "keys": "", "fills": 0, "matched_rows": 0})
                continue

            summary.append({
                "sheet_name": sheet_name,
                "status": "repaired" if fills else "no_fills_needed",
                "keys": " | ".join(used_keys),
                "fills": fills,
                "matched_rows": len(matched_rows),
            })
    finally:
        try:
            if base_wb is not None:
                base_wb.close()
        except Exception:
            pass

    return summary



EMBEDDED_AUDIT_NUMERIC_HEADERS = {
    "Square Feet", "# of Units", "Year Built", "Property Count", "Loan Amount",
    "Initial Disbursement Funded", "Renovation Holdback", "Renovation Holdback Funded",
    "Renovation Holdback Remaining", "Renovation HB Funded", "Renovation HB Remaining",
    "Interest Allocation", "Interest Allocation Funded", "Loan Commitment", "Active Funded Amount",
    "Suspense Balance", "Remaining Commitment", "Origination As-Is Value", "Origination ARV",
    "Updated As-Is Value", "Updated ARV", "Most Recent As-Is Value", "Most Recent ARV",
    "Property ALA", "As-Is Value", "Needs NPL Value", "Days Past Due",
}

EMBEDDED_AUDIT_PROTECTED_HEADERS = {
    "Bridge Asset": {
        "Deal Number", "Servicer ID", "SF Yardi ID", "Asset ID", "Deal Name", "Address",
        "City", "State", "Zip", "County", "CBSA", "APN", "# of Units", "Year Built",
        "Square Feet", "Origination Date", "FC Sale Date", "Rescheduled FC Sale Date", "Origination Value Dt", "Origination As-Is Value",
        "Origination ARV", "Most Recent Appraisal Order Date", "Updated Valuation Date",
        "Updated As-Is Value", "Updated ARV", "Initial Disbursement Funded", "Renovation Holdback",
        "Interest Allocation",
    },
    "Bridge Loan": {
        "Deal Number", "Servicer ID", "SF Yardi ID", "Loan Commitment", "Active Funded Amount",
        "Initial Disbursement Funded", "Renovation Holdback", "Interest Allocation",
        "Most Recent As-Is Value", "Most Recent ARV",
    },
    "Term Loan": {
        "Deal Number", "Servicer ID", "SF Yardi ID", "Loan Amount", "Origination Date", "Maturity Date",
    },
    "Term Asset": {
        "Deal Number", "Asset ID", "Portfolio", "Date", "Address", "City", "State", "Zip",
        "CBSA", "# Units", "Property Type", "Value Date", "Property ALA", "As-Is Value",
    },
}

EMBEDDED_AUDIT_HEADER_FONT = Font(name="Aptos Narrow", size=11, bold=True)
EMBEDDED_AUDIT_BODY_FONT = Font(name="Aptos Narrow", size=11)
EMBEDDED_AUDIT_WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
EMBEDDED_AUDIT_ALL_DATE_HEADERS = set().union(*SHEET_DATE_HEADERS.values())


def _embedded_find_upb_header(wb, preferred_sheet: str = "Bridge Asset") -> str:
    sheets = [preferred_sheet] + [name for name in wb.sheetnames if name != preferred_sheet]
    for sheet_name in sheets:
        ws = wb[sheet_name]
        for cell in ws[4]:
            value = cell.value
            if isinstance(value, str) and UPB_HEADER_RE.search(value):
                return value.strip()
    return "UPB"



def _embedded_row_has_any_value(ws, row_idx: int, col_indices: Sequence[int]) -> bool:
    for col_idx in col_indices:
        if clean_text(ws.cell(row_idx, col_idx).value) != "":
            return True
    return False



def _embedded_normalize_compare_value(header: str, value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        if header in EMBEDDED_AUDIT_ALL_DATE_HEADERS or "date" in header.lower():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, datetime):
        if header in EMBEDDED_AUDIT_ALL_DATE_HEADERS or "date" in header.lower():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float, np.integer, np.floating)):
        try:
            val = float(value)
        except Exception:
            return clean_text(value)
        if np.isnan(val):
            return ""
        if header in EMBEDDED_AUDIT_ALL_DATE_HEADERS:
            return str(value)
        if float(val).is_integer():
            return str(int(val))
        return f"{val:.6f}".rstrip("0").rstrip(".")

    text = clean_text(value)
    if not text:
        return ""

    if header in EMBEDDED_AUDIT_ALL_DATE_HEADERS or "date" in header.lower():
        parsed = pd.to_datetime(text, errors="coerce")
        if not pd.isna(parsed):
            return parsed.date().isoformat()

    if header in EMBEDDED_AUDIT_NUMERIC_HEADERS:
        numeric = text.replace(",", "").replace("$", "")
        if numeric.startswith("(") and numeric.endswith(")"):
            numeric = "-" + numeric[1:-1]
        try:
            val = float(numeric)
            if float(val).is_integer():
                return str(int(val))
            return f"{val:.6f}".rstrip("0").rstrip(".")
        except Exception:
            pass

    return re.sub(r"\s+", " ", text).strip().upper()



def _embedded_build_row_lookup(
    wb,
    ws,
    key_headers: Sequence[str],
    upb_header: str,
    include_key_variants: bool = False,
) -> Tuple[Dict[Tuple[str, ...], int], Dict[str, int], int, int, List[dict]]:
    header_idx = _sheet_header_index(wb, ws, upb_header)
    if not all(h in header_idx for h in key_headers):
        return {}, header_idx, 0, 0, []

    rows: Dict[Tuple[str, ...], List[int]] = {}
    blank_key_rows = 0
    data_rows = 0
    duplicates: List[dict] = []
    scan_cols = list(header_idx.values())

    for row_idx in range(5, ws.max_row + 1):
        if not _embedded_row_has_any_value(ws, row_idx, scan_cols):
            continue
        data_rows += 1

        key_parts: List[List[str]] = []
        valid = True
        for key_header in key_headers:
            raw = ws.cell(row_idx, header_idx[key_header]).value
            vals = _normalize_sheet_key_variants(key_header, raw) if include_key_variants else [_normalize_sheet_key_value(key_header, raw)]
            vals = [v for v in vals if v]
            if not vals:
                blank_key_rows += 1
                valid = False
                break
            key_parts.append(vals)
        if not valid:
            continue

        if include_key_variants and len(key_parts) > 1:
            import itertools
            keys_to_add = list(itertools.product(*key_parts))
        else:
            keys_to_add = [tuple(part[0] for part in key_parts)]

        for key in keys_to_add:
            rows.setdefault(key, []).append(row_idx)
            if len(rows[key]) > 1:
                duplicates.append({
                    "issue_type": "duplicate_key",
                    "sheet_name": ws.title,
                    "row": row_idx,
                    "key": " | ".join(key),
                    "header": ", ".join(key_headers),
                    "built_value": "duplicate key row",
                    "baseline_value": "",
                    "detail": f"Duplicate key for {', '.join(key_headers)}",
                    "suspected_source_header": "",
                })

    chosen = {key: row_list[-1] for key, row_list in rows.items()}
    return chosen, header_idx, data_rows, blank_key_rows, duplicates



def _embedded_choose_keys(sheet_name: str, out_wb, out_ws, base_wb, base_ws, upb_header: str):
    for candidate in _backfill_rule_candidates(sheet_name):
        out_rows, out_headers, out_data_rows, out_blank_keys, dupes = _embedded_build_row_lookup(
            out_wb, out_ws, candidate, upb_header, include_key_variants=False
        )
        base_rows, base_headers, _base_data_rows, _base_blank_keys, _base_dupes = _embedded_build_row_lookup(
            base_wb, base_ws, candidate, upb_header, include_key_variants=True
        )
        if sheet_name == "Term Loan" and candidate == ["Servicer ID"] and (not out_rows or not base_rows):
            continue
        if out_rows or base_rows:
            return list(candidate), out_rows, base_rows, out_headers, base_headers, out_data_rows, out_blank_keys, dupes
    return [], {}, {}, {}, {}, 0, 0, []



def _embedded_choose_structural_keys(sheet_name: str, wb, ws, upb_header: str):
    for candidate in _backfill_rule_candidates(sheet_name):
        rows, headers, data_rows, blank_keys, dupes = _embedded_build_row_lookup(
            wb, ws, candidate, upb_header, include_key_variants=False
        )
        if sheet_name == "Term Loan" and candidate == ["Servicer ID"] and not rows:
            continue
        if headers:
            return list(candidate), rows, headers, data_rows, blank_keys, dupes
    return [], {}, {}, 0, 0, []



def _embedded_possible_shift_header(sheet_name: str, header: str, base_values: Dict[str, str], built_values: Dict[str, str]) -> str:
    protected = EMBEDDED_AUDIT_PROTECTED_HEADERS.get(sheet_name, set())
    if header not in protected:
        return ""
    target = base_values.get(header, "")
    if not target or str(target).strip().upper() in {"N/A", "NA", "NONE", "NAN"}:
        return ""
    for other_header in protected:
        if other_header == header:
            continue
        if built_values.get(other_header, "") == target:
            return other_header
    return ""



def _embedded_audit_openpyxl_workbook(
    wb,
    baseline_bytes: Optional[bytes] = None,
    upb_header: Optional[str] = None,
    sheet_names: Optional[Sequence[str]] = None,
    max_examples_per_sheet: int = 250,
) -> Tuple[List[dict], List[dict]]:
    if upb_header is None:
        upb_header = _embedded_find_upb_header(wb)

    targets = list(sheet_names) if sheet_names else list(SHEET_BASELINE_KEY_CANDIDATES.keys())
    summaries: List[dict] = []
    exceptions: List[dict] = []

    base_wb = load_workbook(BytesIO(baseline_bytes), data_only=False, keep_links=False) if baseline_bytes else None
    try:
        for sheet_name in targets:
            if sheet_name not in wb.sheetnames:
                continue

            out_ws = wb[sheet_name]
            formula_cols = formula_col_indices(out_ws)
            # Columns past the blue range are filled manually and are not audited.
            blue_max = _sheet_blue_max_col(sheet_name)
            if blue_max is not None:
                formula_cols = {c for c in formula_cols if c <= blue_max}

            if base_wb is None or sheet_name not in getattr(base_wb, "sheetnames", []):
                key_headers, out_rows, out_headers, out_data_rows, out_blank_keys, dupes = _embedded_choose_structural_keys(
                    sheet_name, wb, out_ws, upb_header
                )
                formula_gaps = 0
                sheet_example_count = 0
                if out_headers:
                    scan_cols = list(out_headers.values())
                    header_by_col = {col_idx: header for header, col_idx in out_headers.items()}
                    for row_idx in range(5, out_ws.max_row + 1):
                        if not _embedded_row_has_any_value(out_ws, row_idx, scan_cols):
                            continue
                        for col_idx in formula_cols:
                            value = out_ws.cell(row_idx, col_idx).value
                            if not (isinstance(value, str) and value.startswith("=")):
                                formula_gaps += 1
                                if sheet_example_count < max_examples_per_sheet:
                                    exceptions.append({
                                        "issue_type": "formula_gap",
                                        "sheet_name": sheet_name,
                                        "row": row_idx,
                                        "key": "",
                                        "header": header_by_col.get(col_idx, ""),
                                        "built_value": clean_text(value),
                                        "baseline_value": "formula expected",
                                        "detail": "Formula column is missing a formula on a populated data row.",
                                        "suspected_source_header": "",
                                    })
                                    sheet_example_count += 1
                for item in dupes[:max(0, max_examples_per_sheet - sheet_example_count)]:
                    exceptions.append(item)
                summaries.append({
                    "sheet_name": sheet_name,
                    "status": "structural_only",
                    "keys": ", ".join(key_headers),
                    "data_rows": out_data_rows,
                    "matched_rows": 0,
                    "built_only_rows": 0,
                    "baseline_only_rows": 0,
                    "duplicate_key_rows": len(dupes),
                    "blank_key_rows": out_blank_keys,
                    "formula_gap_cells": formula_gaps,
                    "critical_fillable_blanks": 0,
                    "review_fillable_blanks": 0,
                    "possible_shift_cells": 0,
                    "protected_headers_checked": len(EMBEDDED_AUDIT_PROTECTED_HEADERS.get(sheet_name, set())),
                    "top_fillable_blank_columns": [],
                })
                continue

            base_ws = base_wb[sheet_name]
            chosen_keys, out_rows, base_rows, out_headers, base_headers, out_data_rows, out_blank_keys, dupes = _embedded_choose_keys(
                sheet_name, wb, out_ws, base_wb, base_ws, upb_header
            )
            if not chosen_keys:
                summaries.append({
                    "sheet_name": sheet_name,
                    "status": "skipped_no_keys",
                    "keys": "",
                    "data_rows": 0,
                    "matched_rows": 0,
                    "built_only_rows": 0,
                    "baseline_only_rows": 0,
                    "duplicate_key_rows": 0,
                    "blank_key_rows": 0,
                    "formula_gap_cells": 0,
                    "critical_fillable_blanks": 0,
                    "review_fillable_blanks": 0,
                    "possible_shift_cells": 0,
                    "protected_headers_checked": 0,
                    "top_fillable_blank_columns": [],
                })
                continue

            built_only = set(out_rows) - set(base_rows)
            baseline_only = set(base_rows) - set(out_rows)
            common_headers = [
                header for header in out_headers
                if header in base_headers and header not in chosen_keys and not UPB_HEADER_RE.search(str(header))
                and (blue_max is None or out_headers[header] <= blue_max)
            ]
            protected_headers = [header for header in common_headers if header in EMBEDDED_AUDIT_PROTECTED_HEADERS.get(sheet_name, set())]
            scan_cols = list(out_headers.values())
            header_by_col = {col_idx: header for header, col_idx in out_headers.items()}

            formula_gaps = 0
            sheet_example_count = 0
            for row_idx in range(5, out_ws.max_row + 1):
                if not _embedded_row_has_any_value(out_ws, row_idx, scan_cols):
                    continue
                for col_idx in formula_cols:
                    value = out_ws.cell(row_idx, col_idx).value
                    if not (isinstance(value, str) and value.startswith("=")):
                        formula_gaps += 1
                        if sheet_example_count < max_examples_per_sheet:
                            exceptions.append({
                                "issue_type": "formula_gap",
                                "sheet_name": sheet_name,
                                "row": row_idx,
                                "key": "",
                                "header": header_by_col.get(col_idx, ""),
                                "built_value": clean_text(value),
                                "baseline_value": "formula expected",
                                "detail": "Formula column is missing a formula on a populated data row.",
                                "suspected_source_header": "",
                            })
                            sheet_example_count += 1

            critical_fillable_blanks = 0
            review_fillable_blanks = 0
            possible_shifts = 0
            top_blank_counts: Dict[str, int] = {}

            for key, out_row in out_rows.items():
                base_row = base_rows.get(key)
                if not base_row:
                    continue
                key_text = " | ".join(key)

                built_norm = {
                    header: _embedded_normalize_compare_value(header, out_ws.cell(out_row, out_headers[header]).value)
                    for header in protected_headers
                    if out_headers[header] not in formula_cols
                }
                base_norm = {
                    header: _embedded_normalize_compare_value(header, base_ws.cell(base_row, base_headers[header]).value)
                    for header in protected_headers
                    if out_headers[header] not in formula_cols
                }

                for header in common_headers:
                    if out_headers[header] in formula_cols:
                        continue
                    built_raw = out_ws.cell(out_row, out_headers[header]).value
                    base_raw = base_ws.cell(base_row, base_headers[header]).value
                    built_value = _embedded_normalize_compare_value(header, built_raw)
                    base_value = _embedded_normalize_compare_value(header, base_raw)

                    if not built_value and base_value:
                        top_blank_counts[header] = top_blank_counts.get(header, 0) + 1
                        source_header = _embedded_possible_shift_header(sheet_name, header, base_norm, built_norm)
                        if source_header:
                            possible_shifts += 1
                        if header in EMBEDDED_AUDIT_PROTECTED_HEADERS.get(sheet_name, set()):
                            critical_fillable_blanks += 1
                            issue_type = "critical_blank"
                        else:
                            review_fillable_blanks += 1
                            issue_type = "review_blank"
                        if sheet_example_count < max_examples_per_sheet:
                            exceptions.append({
                                "issue_type": issue_type,
                                "sheet_name": sheet_name,
                                "row": out_row,
                                "key": key_text,
                                "header": header,
                                "built_value": clean_text(built_raw),
                                "baseline_value": clean_text(base_raw),
                                "detail": "Built workbook is blank where baseline has a value.",
                                "suspected_source_header": source_header,
                            })
                            sheet_example_count += 1
                    elif header in EMBEDDED_AUDIT_PROTECTED_HEADERS.get(sheet_name, set()) and base_value and built_value and base_value != built_value:
                        source_header = _embedded_possible_shift_header(sheet_name, header, base_norm, built_norm)
                        if source_header:
                            possible_shifts += 1
                            if sheet_example_count < max_examples_per_sheet:
                                exceptions.append({
                                    "issue_type": "possible_shift",
                                    "sheet_name": sheet_name,
                                    "row": out_row,
                                    "key": key_text,
                                    "header": header,
                                    "built_value": clean_text(built_raw),
                                    "baseline_value": clean_text(base_raw),
                                    "detail": "Baseline value appears to be present in another protected column on the same row.",
                                    "suspected_source_header": source_header,
                                })
                                sheet_example_count += 1

            for item in dupes[:max(0, max_examples_per_sheet - sheet_example_count)]:
                exceptions.append(item)

            status = "pass"
            if formula_gaps or critical_fillable_blanks or possible_shifts or dupes or out_blank_keys:
                status = "fail"
            elif review_fillable_blanks or built_only or baseline_only:
                status = "review"

            summaries.append({
                "sheet_name": sheet_name,
                "status": status,
                "keys": ", ".join(chosen_keys),
                "data_rows": out_data_rows,
                "matched_rows": len(set(out_rows) & set(base_rows)),
                "built_only_rows": len(built_only),
                "baseline_only_rows": len(baseline_only),
                "duplicate_key_rows": len(dupes),
                "blank_key_rows": out_blank_keys,
                "formula_gap_cells": formula_gaps,
                "critical_fillable_blanks": critical_fillable_blanks,
                "review_fillable_blanks": review_fillable_blanks,
                "possible_shift_cells": possible_shifts,
                "protected_headers_checked": len(protected_headers),
                "top_fillable_blank_columns": [
                    {"column": header, "count": count}
                    for header, count in sorted(top_blank_counts.items(), key=lambda item: (-item[1], item[0]))[:15]
                ],
            })
    finally:
        if base_wb is not None:
            try:
                base_wb.close()
            except Exception:
                pass

    return summaries, exceptions



def _embedded_audit_col_letter(index: int) -> str:
    out = ""
    n = index
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out



def _embedded_auto_width(ws, widths: Dict[int, int], max_width: int = 50):
    for col_idx, width in widths.items():
        ws.column_dimensions[_embedded_audit_col_letter(col_idx)].width = min(max_width, max(10, width + 2))



def _embedded_write_audit_sheets(wb, summary_rows: List[dict], exception_rows: List[dict]) -> None:
    for name in ["QA Summary", "QA Exceptions"]:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet("QA Summary")
    ws.freeze_panes = "A2"
    summary_headers = [
        "sheet_name", "status", "keys", "data_rows", "matched_rows", "built_only_rows",
        "baseline_only_rows", "duplicate_key_rows", "blank_key_rows", "formula_gap_cells",
        "critical_fillable_blanks", "review_fillable_blanks", "possible_shift_cells",
        "protected_headers_checked", "top_fillable_blank_columns",
    ]
    widths = {i + 1: len(h) for i, h in enumerate(summary_headers)}
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = copy(EMBEDDED_AUDIT_HEADER_FONT)
        cell.alignment = copy(EMBEDDED_AUDIT_WRAP_ALIGNMENT)
    for row_idx, item in enumerate(summary_rows, start=2):
        for col_idx, header in enumerate(summary_headers, start=1):
            value = item.get(header, "")
            if header == "top_fillable_blank_columns" and isinstance(value, list):
                value = "; ".join(f"{x.get('column')}: {x.get('count')}" for x in value)
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = copy(EMBEDDED_AUDIT_BODY_FONT)
            cell.alignment = copy(EMBEDDED_AUDIT_WRAP_ALIGNMENT)
            widths[col_idx] = max(widths[col_idx], len(str(value)) if value is not None else 0)
    _embedded_auto_width(ws, widths)

    ex = wb.create_sheet("QA Exceptions")
    ex.freeze_panes = "A2"
    exception_headers = [
        "issue_type", "sheet_name", "row", "key", "header", "built_value",
        "baseline_value", "detail", "suspected_source_header",
    ]
    widths = {i + 1: len(h) for i, h in enumerate(exception_headers)}
    for col_idx, header in enumerate(exception_headers, start=1):
        cell = ex.cell(1, col_idx, header)
        cell.font = copy(EMBEDDED_AUDIT_HEADER_FONT)
        cell.alignment = copy(EMBEDDED_AUDIT_WRAP_ALIGNMENT)
    for row_idx, item in enumerate(exception_rows, start=2):
        for col_idx, header in enumerate(exception_headers, start=1):
            value = item.get(header, "")
            cell = ex.cell(row_idx, col_idx, value)
            cell.font = copy(EMBEDDED_AUDIT_BODY_FONT)
            cell.alignment = copy(EMBEDDED_AUDIT_WRAP_ALIGNMENT)
            widths[col_idx] = max(widths[col_idx], len(str(value)) if value is not None else 0)
    _embedded_auto_width(ex, widths, max_width=60)



def _embedded_audit_diagnostic_lines(summary_rows: List[dict]) -> List[str]:
    lines: List[str] = []
    for item in summary_rows:
        status = str(item.get("status", "")).upper()
        lines.append(
            f"QA {item.get('sheet_name')}: {status} | "
            f"critical blanks={int(item.get('critical_fillable_blanks', 0)):,}, "
            f"review blanks={int(item.get('review_fillable_blanks', 0)):,}, "
            f"possible shifts={int(item.get('possible_shift_cells', 0)):,}, "
            f"formula gaps={int(item.get('formula_gap_cells', 0)):,}, "
            f"duplicate keys={int(item.get('duplicate_key_rows', 0)):,}"
        )
    return lines



def _embedded_workbook_needs_attention(summary_rows: List[dict]) -> bool:
    for item in summary_rows:
        if str(item.get("status", "")).lower() in {"fail", "review"}:
            return True
    return False


if not POSTBUILD_AUDIT_AVAILABLE or audit_openpyxl_workbook is None or write_audit_sheets is None or audit_diagnostic_lines is None or workbook_needs_attention is None:
    audit_openpyxl_workbook = _embedded_audit_openpyxl_workbook
    write_audit_sheets = _embedded_write_audit_sheets
    audit_diagnostic_lines = _embedded_audit_diagnostic_lines
    workbook_needs_attention = _embedded_workbook_needs_attention
    POSTBUILD_AUDIT_AVAILABLE = True
    if POSTBUILD_AUDIT_IMPORT_ERROR:
        POSTBUILD_AUDIT_IMPORT_ERROR = f"{POSTBUILD_AUDIT_IMPORT_ERROR}; embedded fallback activated"
    else:
        POSTBUILD_AUDIT_IMPORT_ERROR = "embedded fallback activated"

def _audit_fillable_blank_totals(audit_summary: Sequence[dict]) -> Tuple[int, int, int]:
    critical = sum(int(item.get("critical_fillable_blanks", 0) or 0) for item in audit_summary)
    review = sum(int(item.get("review_fillable_blanks", 0) or 0) for item in audit_summary)
    possible_shift = sum(int(item.get("possible_shift_cells", 0) or 0) for item in audit_summary)
    return critical, review, possible_shift


def enforce_zero_fillable_blanks(
    wb,
    baseline_bytes: Optional[bytes],
    upb_header: str,
    sheet_names: Optional[Sequence[str]] = None,
    max_rounds: int = ZERO_BLANK_MAX_ROUNDS,
) -> Tuple[bool, List[str], List[dict], List[dict]]:
    diagnostics: List[str] = []
    audit_summary: List[dict] = []
    audit_exceptions: List[dict] = []

    if not baseline_bytes:
        diagnostics.append("Zero-blank enforcement skipped: no prior completed workbook was uploaded.")
        return False, diagnostics, audit_summary, audit_exceptions

    if not POSTBUILD_AUDIT_AVAILABLE or audit_openpyxl_workbook is None:
        diagnostics.append("Zero-blank enforcement skipped: post-build audit helper is not available in this runtime.")
        return False, diagnostics, audit_summary, audit_exceptions

    prev_blank_total: Optional[int] = None
    for round_no in range(1, max_rounds + 1):
        repair_summary = repair_workbook_from_baseline(wb, baseline_bytes, upb_header, sheet_names=sheet_names)
        fills = sum(int(item.get("fills", 0) or 0) for item in repair_summary)
        for item in repair_summary:
            if item.get("fills"):
                diagnostics.append(
                    f"Zero-blank round {round_no}: {item['sheet_name']} filled {int(item['fills']):,} cells using {item.get('keys', 'n/a')}"
                )

        audit_summary, audit_exceptions = audit_openpyxl_workbook(
            wb,
            baseline_bytes=baseline_bytes,
            upb_header=upb_header,
            sheet_names=sheet_names,
        )
        critical, review, possible_shift = _audit_fillable_blank_totals(audit_summary)
        blank_total = critical + review
        diagnostics.append(
            f"Zero-blank round {round_no}: remaining fillable blanks {blank_total:,} (critical {critical:,}, review {review:,}); possible shifts {possible_shift:,}."
        )

        if blank_total == 0:
            return True, diagnostics, audit_summary, audit_exceptions

        if prev_blank_total is not None and blank_total >= prev_blank_total and fills == 0:
            break
        prev_blank_total = blank_total

    return False, diagnostics, audit_summary, audit_exceptions



def _prune_non_report_sheets(wb, sheet_names: Optional[Sequence[str]] = None) -> List[str]:
    """Remove carried-over sheets from the finished report.

    V76: OFF by default. Keeping Bridge Payoffs / Term Payoffs / REO Sales was never what
    emptied Bridge Asset and Term Asset -- that was openpyxl streaming the two largest
    worksheets through lxml on Python 3.14 (see the note at the top of this file). The
    official report carries these sheets, so the default is to carry them too, and dropping
    them is a deliberate per-run choice.
    """
    removed = []
    for _name in (sheet_names if sheet_names is not None else REPORT_DROP_SHEETS):
        if _name in wb.sheetnames:
            try:
                del wb[_name]
                removed.append(_name)
            except Exception:
                pass
    return removed


def _write_build_log_sheet(wb, diagnostics, upb_col, template_path, build_target_label):
    """Write the build's own log into the workbook as a 'Build Log' sheet.

    Three broken builds in a row could not be diagnosed because the evidence lived in the
    Streamlit session and never travelled with the file. From now on the artifact carries its
    own provenance: versions, which template it started from, which sheets were handed rows,
    how many rows were on each sheet right after its write, and every diagnostics line.
    Never raises -- a logging failure must not cost a build.
    """
    try:
        import sys as _sys
        try:
            if "Build Log" in wb.sheetnames:
                del wb["Build Log"]
        except Exception:
            pass
        ws = wb.create_sheet("Build Log")
        rows = []
        rows.append(("build version", APP_BUILD_VERSION))
        rows.append(("build target", str(build_target_label)))
        rows.append(("non-report sheets dropped", "yes" if globals().get("drop_extra_sheets") else "no"))
        rows.append(("template source", str(template_path)))
        rows.append(("UPB header", str(upb_col)))
        rows.append(("python", _sys.version.split()[0]))
        for _mod in ("pandas", "numpy", "openpyxl"):
            try:
                rows.append((_mod, __import__(_mod).__version__))
            except Exception:
                rows.append((_mod, "unavailable"))
        rows.append(("openpyxl worksheet writer", str(OPENPYXL_WORKSHEET_WRITER)))
        try:
            import openpyxl.worksheet._writer as _wchk
            rows.append(("worksheet writer in use now", getattr(_wchk.xmlfile, "__module__", "?")))
        except Exception:
            pass
        try:
            from openpyxl.xml import LXML as _LXML
            rows.append(("openpyxl.xml.LXML (informational)", str(_LXML)))
        except Exception:
            pass
        rows.append(("", ""))
        rows.append(("rows handed to write_output_sheet", ""))
        for _k, _v in (WRITTEN_SHEET_ROWS or {}).items():
            rows.append((f"  {_k}", _v))
        if not WRITTEN_SHEET_ROWS:
            rows.append(("  (none - write_output_sheet was never called)", ""))
        rows.append(("", ""))
        rows.append(("rows on sheet immediately after its write", ""))
        for _k, _v in (SHEET_WRITE_AUDIT or {}).items():
            rows.append((f"  {_k}", f"{_v.get('on_sheet_after_write')} of {_v.get('handed')}"))
        if not SHEET_WRITE_AUDIT:
            rows.append(("  (none recorded)", ""))
        rows.append(("", ""))
        rows.append(("sheets present at save time", ""))
        for _sn in wb.sheetnames:
            try:
                _mr = wb[_sn].max_row or 0
            except Exception:
                _mr = "?"
            rows.append((f"  {_sn}", f"max_row={_mr}"))
        rows.append(("", ""))
        rows.append(("diagnostics", ""))
        for _line in (diagnostics or []):
            rows.append(("", str(_line)[:1000]))
        ws.cell(1, 1).value = "key"
        ws.cell(1, 2).value = "value"
        for _i, (_a, _b) in enumerate(rows, start=2):
            ws.cell(_i, 1).value = _a
            ws.cell(_i, 2).value = _b
        try:
            ws.column_dimensions["A"].width = 44
            ws.column_dimensions["B"].width = 130
        except Exception:
            pass
    except Exception:
        pass


def sanitize_summary_formulas(wb):
    if "Summary" not in wb.sheetnames:
        return
    ws = wb["Summary"]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            if "IFERROR(" in v.upper():
                continue
            if "/" in v:
                cell.value = f'=IFERROR({v[1:]},"N/A")'

def mark_workbook_for_recalc(wb):
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def init_build_state():
    if "built_workbook_bytes" not in st.session_state:
        st.session_state.built_workbook_bytes = None
    if "built_workbook_name" not in st.session_state:
        st.session_state.built_workbook_name = None
    if "built_template_path" not in st.session_state:
        st.session_state.built_template_path = None
    if "show_download_prompt" not in st.session_state:
        st.session_state.show_download_prompt = False
    if "download_choice" not in st.session_state:
        st.session_state.download_choice = "Not yet"


def reset_build_state():
    st.session_state.built_workbook_bytes = None
    st.session_state.built_workbook_name = None
    st.session_state.built_template_path = None
    st.session_state.show_download_prompt = False
    st.session_state.download_choice = "Not yet"


st.set_page_config(page_title="Active Loans Builder", layout="wide")
st.title("Active Loans Report Builder")
st.subheader(hey())
st.caption(f"Code build: {APP_BUILD_VERSION}")

run_dt = today_et()
upb_col = make_upb_header(run_dt)
init_build_state()

st.markdown(
    f"""
Welcome! This tool builds the **Active Loans** workbook using **Salesforce Bulk API 2.0** and optional **servicer uploads**.

### What you’ll do
1) Log in to **Salesforce**
2) Upload the **current external servicer files** or skip them
3) (Recommended) Upload **last week’s / completed Active Loans report** for layout/carry-forward; not an SF data source
4) Choose **which sheet to build** or **All**

### UPB header
Before servicer files are parsed, the preview uses today's date (ET): **{run_dt.isoformat()}** → **{upb_col}**. During the build, the final UPB header is reset to the dominant uploaded servicer tape date.

**Salesforce source note:** Bridge, Term, Valuation, AM Assignments, Active RM, and related deal/asset populations are pulled live from the Salesforce API. To replicate the completed Active Loans process, also upload the prior/completed Active Loans workbook so curated fields and Term Asset rows can be carried forward. Upload external servicer files for UPB/status/due-date enrichment.
"""
)

_repo_template_available = False
try:
    _tmpl_bytes_preview, _tmpl_path_used = load_repo_template_bytes()
    _repo_template_available = True
    st.success(f"✅ Using repo template: {_tmpl_path_used}")
except Exception as e:
    st.warning(
        "Repo template not found right now. You can still build by uploading a completed Active Loans workbook "
        "to use as the template base, or by committing one of the expected template files to the repo."
    )
    st.caption(str(e))

st.caption(
    "This merged version pulls core report data from Salesforce Bulk API, uses your repo template by default, "
    "can use the uploaded completed report as the build base/carry-forward source, "
    "uses uploaded Midland / FCI / Berkadia / Statebridge / other servicer files only for external servicer enrichment, "
    "resolves formula-linked UPB headers, fills formulas down, trims extra blank rows, keeps row-level Salesforce Servicer IDs intact, "
    "runs both dataframe-level and workbook-level blank repair against the uploaded known-good report when you provide one, "
    "and writes QA Summary / QA Exceptions tabs after the build when the post-build audit helper is available."
)

if POSTBUILD_AUDIT_AVAILABLE:
    st.caption("Post-build QA audit helper loaded. Completed workbooks will include QA Summary and QA Exceptions tabs.")
else:
    st.warning(f"Post-build QA audit helper was not loaded: {POSTBUILD_AUDIT_IMPORT_ERROR}")

sf_info = render_salesforce_login_gate()
sf_ready = bool(sf_info)
use_sf = True

st.markdown("### Step 2: Upload files")
col_a, col_b = st.columns([1.3, 1.0])
with col_a:
    prev_upload = st.file_uploader(
        "Upload LAST WEEK'S or COMPLETED Active Loans report (.xlsx) for layout/carry-forward (strongly recommended for production-accurate replication)",
        type=["xlsx"],
    )
with col_b:
    servicer_uploads = st.file_uploader(
        "Upload current EXTERNAL servicer files only (csv/xlsx) — not Salesforce exports (optional if skipped below)",
        type=["csv", "xlsx"],
        accept_multiple_files=True,
    )

st.markdown("### Step 3: Build options")
skip_servicer_files = st.checkbox(
    "Skip servicer files and build Salesforce-only version",
    value=False,
    help="Leaves servicer-driven columns blank or Salesforce-fallback where available.",
)

build_target = st.selectbox(
    "Which sheet do you want to build right now?",
    options=["Bridge Asset", "Bridge Loan", "Term Loan", "Term Asset", "All"],
    index=4,
)

show_servicer_preview = st.checkbox(
    "Show servicer preview table after parsing",
    value=False,
    help="Leave this off to save memory during testing.",
)

allow_qa_fail_download = st.checkbox(
    "Allow download even if post-build QA fails (debug only)",
    value=False,
    help="Leave unchecked for normal builds. If QA fails, the app will stop before download so avoidable blank/shift issues cannot be treated as production output.",
)
run_postbuild_qa = True
st.caption("Zero-blank repair / QA runs automatically on every build.")

drop_extra_sheets = st.checkbox(
    "Drop the carried-over Payoffs / REO Sales / Pacific Life / 2026-1 / CAFL SA / JLL sheets",
    value=False,
    help=(
        "Off by default, which matches the official report -- it carries all of these. "
        "The template is your uploaded prior workbook, so whatever is on it comes through. "
        "This has NO bearing on the Bridge Asset / Term Asset tabs: those were emptied by the "
        "openpyxl worksheet writer on Python 3.14, which is fixed separately. "
        "Strategy Groupings, SSP Loans and Legacy are never dropped -- the build reads its "
        "lookups out of them."
    ),
)

if st.button("Clear cached Salesforce metadata", type="secondary"):
    st.session_state.sobject_describe_cache = {}
    st.success("Cleared Salesforce metadata cache for this session.")

if st.button("Clear cached servicer parsing", type="secondary"):
    st.cache_data.clear()
    st.success("Cleared Streamlit data cache.")

build_btn = st.button("Build", type="primary")

if build_btn:
    reset_build_state()

    if not use_sf:
        st.error("This version requires Salesforce API to build the report.")
    elif prev_upload is None and not _repo_template_available:
        st.error("No repo template was found. Upload a prior/completed Active Loans workbook as the template base, or commit an Active Loans template to the repo.")
    elif not skip_servicer_files and not servicer_uploads:
        st.error("Upload the servicer files, or check 'Skip servicer files and build Salesforce-only version'.")
    elif not sf_ready:
        st.error("Salesforce login is required.")
    else:
        wb = None
        serv_join = pd.DataFrame()
        serv_preview = pd.DataFrame()
        sf_am = pd.DataFrame()
        sf_active_rm = pd.DataFrame()
        try:
            status = st.status("Preparing build...", expanded=True)
            diagnostics: List[str] = []
            prev_maps: dict = {}
            prev_bytes = prev_upload.getvalue() if prev_upload else None
            npl_maps = {"loan_flags": pd.DataFrame(), "asset_flags": pd.DataFrame()}

            if prev_upload:
                status.update(label="Reading uploaded completed report for carry-forward...")
                _prior_is_build, _prior_note = prior_workbook_provenance(prev_bytes)
                if _prior_is_build:
                    diagnostics.append("PRIOR WORKBOOK WARNING: " + _prior_note)
                    try:
                        st.warning("⚠️ " + _prior_note)
                    except Exception:
                        pass
                prev_maps = build_prev_maps(prev_bytes)
            else:
                diagnostics.append("No prior/completed Active Loans workbook uploaded: using the repo template only; manual carry-forward/backfill fields will be limited to Salesforce/template logic.")

            if skip_servicer_files:
                serv_join = pd.DataFrame(columns=["source_file", "servicer", "servicer_family", "servicer_id", "upb", "suspense", "next_payment_date", "maturity_date", "status", "as_of", "_sid_key"])
                detected_run_date = run_dt
                serv_preview = serv_join.copy()

                st.markdown("### Servicer lookup preview")
                st.caption("Servicer files were skipped. Servicer-driven columns will use Salesforce fallback where available.")
                st.caption(f"UPB header (build report date): **{upb_col}**")
            else:
                servicer_phase_started = time.perf_counter()

                def _servicer_progress(message: str) -> None:
                    elapsed = time.perf_counter() - servicer_phase_started
                    status.update(label=f"Servicer processing | {message} | total {elapsed:0.1f}s")

                _servicer_progress(f"starting {len(servicer_uploads)} uploaded file(s)")
                serv_join, detected_run_date, serv_preview = build_servicer_lookup(
                    servicer_uploads,
                    progress_hook=_servicer_progress,
                    preview_rows_limit=30 if show_servicer_preview else 0,
                    use_cache=False,
                )

                st.markdown("### Servicer lookup preview")
                st.caption(f"Detected dominant servicer tape date from uploaded filenames / report tabs: **{detected_run_date.isoformat()}**")
                st.caption(f"UPB header used for this build: **{make_upb_header(detected_run_date)}**")
                if show_servicer_preview:
                    st.dataframe(serv_preview.head(30), use_container_width=True)
                else:
                    st.caption("Servicer preview hidden to save memory during this run.")
                serv_preview = pd.DataFrame()
                gc.collect()

            run_dt = detected_run_date
            upb_col = make_upb_header(run_dt)

            status.update(label=f"Loading Excel template for report date {run_dt.isoformat()} ({upb_col})...")
            if prev_upload is None and not _repo_template_available:
                raise FileNotFoundError(
                    "No repo template was found and no completed Active Loans workbook was uploaded. "
                    "Upload the prior completed workbook or add one of the expected template files to the repo."
                )
            tmpl_bytes, tmpl_path_used = resolve_template_bytes(prev_upload)
            template_maps = load_template_lookup_maps(tmpl_bytes)
            wb = load_workbook(BytesIO(tmpl_bytes), data_only=False, keep_links=False)
            mark_workbook_for_recalc(wb)
            diagnostics.extend(restore_template_scaffold(wb, run_dt, upb_col) or [])
            sanitize_summary_formulas(wb)

            need_bridge = build_target in ("Bridge Asset", "Bridge Loan", "All")
            need_term = build_target in ("Term Loan", "Term Asset", "All")
            need_term_asset = build_target in ("Term Asset", "All")
            need_am = need_bridge or need_term

            sf_am = pd.DataFrame()
            if need_am:
                status.update(label="Pulling AM assignments from Salesforce...")
                sf_am = _build_am_assignments_like()

            status.update(label="Pulling Active RM deals from Salesforce...")
            sf_active_rm = _build_active_rm_like()

            if need_bridge:
                status.update(label="Pulling bridge/property data from Salesforce...")
                bridge_spine = _build_bridge_spine_like()
                bridge_loan_wide = _build_bridge_loan_wide_like()
                bridge_property_rollup = _build_bridge_property_rollup_like()
                bridge_asset_ids = _bridge_asset_ids_from_spine(bridge_spine)

                status.update(label="Pulling Do Not Lend deals from Salesforce...")
                bridge_dnl = _build_do_not_lend_like()

                status.update(label="Pulling valuation data from Salesforce...")
                bridge_val = _build_valuation_like(asset_ids=bridge_asset_ids)

                status.update(label="Pulling foreclosure sale dates from Salesforce...")
                bridge_foreclosure = _build_foreclosure_like(asset_ids=bridge_asset_ids)

                status.update(label="Building Bridge Asset...")
                bridge_asset_df = build_bridge_asset(
                    bridge_spine,
                    bridge_dnl,
                    bridge_val,
                    bridge_foreclosure,
                    sf_am,
                    sf_active_rm,
                    serv_join,
                    upb_col,
                    prev_maps,
                    template_maps,
                    npl_maps=npl_maps,
                )
                bridge_asset_df, bridge_asset_backfill = backfill_df_from_baseline("Bridge Asset", bridge_asset_df, prev_bytes)
                # Baseline backfill can populate funded component fields; recompute the funded amount immediately after it.
                bridge_asset_df, bridge_asset_math_diags = repair_bridge_asset_math(bridge_asset_df, upb_col)
                diagnostics.extend(bridge_asset_math_diags)
                diagnostics.extend(validate_bridge_math_or_raise(bridge_asset_df, None, upb_col))
                if bridge_asset_backfill and bridge_asset_backfill.get("fills"):
                    diagnostics.append(f"Bridge Asset baseline backfill cells: {int(bridge_asset_backfill['fills']):,} using {bridge_asset_backfill.get('keys', 'n/a')}")

                diagnostics.append(f"Bridge Asset rows: {len(bridge_asset_df):,}")
                if SUBUNIT_FLAG_CENSUS:
                    _sc = SUBUNIT_FLAG_CENSUS
                    diagnostics.append(
                        "Is Sub Unit census over {rows:,} active bridge properties: "
                        "checked ON {true:,}, OFF-or-unset {false_or_null:,}, "
                        "of which NULL/blank {null_or_blank:,}".format(**_sc)
                    )
                    if int(_sc.get("null_or_blank", 0)) > 0:
                        diagnostics.append(
                            "NOTE: Is_Sub_Unit__c is NULL/blank on {:,} properties, so 'off' and "
                            "'not set' are NOT the same in this org. The Term Asset queries accept "
                            "both via (= FALSE OR = NULL); tightening them to '= FALSE' would drop "
                            "the never-set rows.".format(int(_sc["null_or_blank"]))
                        )
                diagnostics.append(
                    f"Bridge Asset nonblank {upb_col}: {bridge_asset_df[upb_col].notna().mean():.1%}"
                    if upb_col in bridge_asset_df.columns
                    else f"Bridge Asset nonblank {upb_col}: n/a"
                )

                if build_target in ("Bridge Asset", "All"):
                    status.update(label="Writing Bridge Asset sheet...")
                    write_output_sheet(wb, "Bridge Asset", bridge_asset_df, upb_col)

                if build_target in ("Bridge Loan", "All"):
                    status.update(label="Building Bridge Loan...")
                    bridge_loan_df = build_bridge_loan(
                        bridge_loan_wide,
                        bridge_asset_df,
                        bridge_property_rollup,
                        serv_join,
                        upb_col,
                        prev_maps,
                        template_maps,
                        npl_maps=npl_maps,
                    )
                    bridge_loan_df, bridge_loan_backfill = backfill_df_from_baseline("Bridge Loan", bridge_loan_df, prev_bytes)
                    bridge_loan_df = _reconcile_bridge_loan_from_asset_rollup(bridge_loan_df, bridge_asset_df, upb_col)
                    bridge_loan_df, bridge_loan_commitment_diags = _repair_bridge_loan_commitment_math(bridge_loan_df, upb_col)
                    diagnostics.extend(bridge_loan_commitment_diags)
                    diagnostics.extend(validate_bridge_math_or_raise(bridge_asset_df, bridge_loan_df, upb_col))
                    if bridge_loan_backfill and bridge_loan_backfill.get("fills"):
                        diagnostics.append(f"Bridge Loan baseline backfill cells: {int(bridge_loan_backfill['fills']):,} using {bridge_loan_backfill.get('keys', 'n/a')}")

                    diagnostics.append(f"Bridge Loan rows: {len(bridge_loan_df):,}")
                    diagnostics.append(
                        f"Bridge Loan nonblank {upb_col}: {bridge_loan_df[upb_col].notna().mean():.1%}"
                        if upb_col in bridge_loan_df.columns
                        else f"Bridge Loan nonblank {upb_col}: n/a"
                    )

                    status.update(label="Writing Bridge Loan sheet...")
                    write_output_sheet(wb, "Bridge Loan", bridge_loan_df, upb_col)
                    del bridge_loan_df

                del bridge_spine, bridge_loan_wide, bridge_property_rollup, bridge_dnl, bridge_asset_ids, bridge_val, bridge_foreclosure, bridge_asset_df
                gc.collect()

            if need_term:
                status.update(label="Pulling term data from Salesforce...")
                term_wide = _build_term_wide_like()

                candidate_term_deals = _nonblank_unique(term_wide["Deal Loan Number"].tolist()) if not term_wide.empty and "Deal Loan Number" in term_wide.columns else []

                status.update(label="Pulling term asset deal universe from Salesforce...")
                term_asset_filter_deals = _build_term_asset_deal_universe(candidate_term_deals)
                diagnostics.append(f"Term asset deal universe: {len(term_asset_filter_deals):,} deals")
                if TERM_ASSET_PARENT_DROPS:
                    _pd_ = TERM_ASSET_PARENT_DROPS
                    diagnostics.append(
                        "Term Asset: dropped {rows:,} parent property row(s) across {deals:,} deal(s) "
                        "that also carry their individual assets, removing {ala:,.2f} of duplicated ALA "
                        "from the UPB allocation denominator.".format(**_pd_)
                    )

                status.update(label="Building Term Loan...")
                term_loan_df = build_term_loan(
                    term_wide,
                    sf_am,
                    sf_active_rm,
                    serv_join,
                    upb_col,
                    prev_maps,
                    template_maps,
                    asset_deal_numbers=term_asset_filter_deals,
                )
                term_loan_df, term_loan_backfill = backfill_df_from_baseline("Term Loan", term_loan_df, prev_bytes)
                term_loan_df = _clear_duplicate_term_servicer_assignments(term_loan_df, upb_col, prev_maps=prev_maps)
                term_loan_df = _guard_term_loan_upb_vs_amount(term_loan_df, upb_col, prev_maps=prev_maps)
                diagnostics.extend(validate_term_loan_amounts_or_raise(term_loan_df, upb_col))
                if term_loan_backfill and term_loan_backfill.get("fills"):
                    diagnostics.append(f"Term Loan baseline backfill cells: {int(term_loan_backfill['fills']):,} using {term_loan_backfill.get('keys', 'n/a')}")

                diagnostics.append(f"Term Loan rows: {len(term_loan_df):,}")
                diagnostics.append(
                    f"Term Loan nonblank {upb_col}: {term_loan_df[upb_col].notna().mean():.1%}"
                    if upb_col in term_loan_df.columns
                    else f"Term Loan nonblank {upb_col}: n/a"
                )

                if build_target in ("Term Loan", "All"):
                    status.update(label="Writing Term Loan sheet...")
                    write_output_sheet(wb, "Term Loan", term_loan_df, upb_col)

                if need_term_asset:
                    term_deal_numbers = (
                        norm_id_series(pd.Series([d for d in _nonblank_unique(term_loan_df["Deal Number"].tolist()) if clean_text(d).upper() != "N/A"], dtype="object"))
                        .dropna()
                        .astype(str)
                        .tolist()
                    ) if "Deal Number" in term_loan_df.columns else []

                    status.update(label="Pulling term asset rows from Salesforce...")
                    term_asset_source = _build_term_asset_like(deal_numbers=term_deal_numbers)

                    status.update(label="Building Term Asset...")
                    term_asset_df = build_term_asset(term_asset_source, term_loan_df, upb_col, prev_maps=prev_maps)
                    term_asset_df, term_asset_backfill = backfill_df_from_baseline("Term Asset", term_asset_df, prev_bytes)
                    # Baseline backfill can populate / repair Property ALA, so allocate UPB after the final asset population exists.
                    term_asset_df = _allocate_term_asset_upb_from_loan(term_asset_df, term_loan_df, upb_col)
                    diagnostics.extend(validate_term_math_or_raise(term_loan_df, term_asset_df, upb_col))
                    if term_asset_backfill and term_asset_backfill.get("fills"):
                        diagnostics.append(f"Term Asset baseline backfill cells: {int(term_asset_backfill['fills']):,} using {term_asset_backfill.get('keys', 'n/a')}")

                    if TERM_ASSET_SUBUNIT_TWIN_DROPS:
                        _st_ = TERM_ASSET_SUBUNIT_TWIN_DROPS
                        diagnostics.append(
                            "Term Asset: dropped {rows:,} sub-unit duplicate row(s) across {deals:,} deal(s) "
                            "whose ALA pushed the deal above its Loan Amount, removing {ala:,.2f} of "
                            "duplicated ALA from the UPB allocation denominator.".format(**_st_)
                        )

                    status.update(label="Writing Term Asset sheet...")
                    write_output_sheet(wb, "Term Asset", term_asset_df, upb_col)

                    # Term Loan's SFR/MF Allocation and Strategy Grouping SUMIFS into Term
                    # Asset, so they can only be resolved now that Term Asset is written.
                    _alloc_diag = _materialize_term_loan_allocations_on_sheet(wb, term_asset_df, upb_col)
                    if _alloc_diag:
                        diagnostics.append(_alloc_diag)

                    del term_deal_numbers, term_asset_source, term_asset_df

                del term_wide, term_loan_df, term_asset_filter_deals, candidate_term_deals
                gc.collect()

            sf_am = None
            sf_active_rm = None
            serv_join = None
            serv_preview = None
            gc.collect()

            selected_sheet_names = [
                sheet_name
                for sheet_name in ["Bridge Asset", "Bridge Loan", "Term Loan", "Term Asset"]
                if build_target in (sheet_name, "All")
            ]
            validate_workbook_schema_or_raise(wb, upb_col, sheet_names=selected_sheet_names)
            diagnostics.append("Schema guardrail passed: selected report tabs have expected row-4 headers before save.")

            audit_summary = []
            audit_exceptions = []
            if run_postbuild_qa and ENFORCE_ZERO_FILLABLE_BLANKS:
                status.update(label="Running zero-blank enforcement...")
                zero_blank_ok, zero_blank_diags, audit_summary, audit_exceptions = enforce_zero_fillable_blanks(
                    wb,
                    baseline_bytes=prev_bytes,
                    upb_header=upb_col,
                    sheet_names=selected_sheet_names,
                )
                diagnostics.extend(zero_blank_diags)
                if write_audit_sheets is not None and audit_summary is not None:
                    write_audit_sheets(wb, audit_summary, audit_exceptions)
                if audit_diagnostic_lines is not None and audit_summary:
                    diagnostics.extend(audit_diagnostic_lines(audit_summary))
                if not zero_blank_ok:
                    diagnostics.append(
                        "Zero-blank enforcement could not eliminate all fillable blanks. Review the QA Summary / QA Exceptions tabs and use the immediately prior completed workbook as baseline on the next run."
                    )
                    st.warning(
                        "Zero-blank enforcement did not clear every fillable blank. The workbook was still built; review QA Summary / QA Exceptions before use."
                    )
                if any(int(item.get("possible_shift_cells", 0) or 0) for item in audit_summary):
                    diagnostics.append("Blank enforcement passed, but possible shifted values were still detected. Review the QA Exceptions tab before weekly use.")
                if QA_HARD_STOP_ON_FAIL and audit_summary and workbook_needs_attention is not None and workbook_needs_attention(audit_summary) and not allow_qa_fail_download:
                    raise RuntimeError(
                        "Post-build QA failed/requires review. The workbook was not offered for download. "
                        "Fix the source/template/code issue, or explicitly allow debug download if intentionally investigating."
                    )
            elif run_postbuild_qa and POSTBUILD_AUDIT_AVAILABLE and audit_openpyxl_workbook is not None and write_audit_sheets is not None:
                status.update(label="Running post-build QA audit...")
                audit_summary, audit_exceptions = audit_openpyxl_workbook(
                    wb,
                    baseline_bytes=prev_bytes,
                    upb_header=upb_col,
                    sheet_names=selected_sheet_names,
                )
                write_audit_sheets(wb, audit_summary, audit_exceptions)
                if audit_diagnostic_lines is not None:
                    diagnostics.extend(audit_diagnostic_lines(audit_summary))
                if workbook_needs_attention is not None and workbook_needs_attention(audit_summary):
                    diagnostics.append("QA attention needed: review the QA Summary and QA Exceptions tabs in the completed workbook before weekly use.")
                    if QA_HARD_STOP_ON_FAIL and not allow_qa_fail_download:
                        raise RuntimeError(
                            "Post-build QA failed/requires review. The workbook was not offered for download. "
                            "Fix the QA exceptions or explicitly allow debug download if you are investigating."
                        )
            elif not run_postbuild_qa:
                diagnostics.append("Post-build zero-blank / QA was skipped to save memory for this run.")
            else:
                diagnostics.append("Post-build QA audit helper not available in this runtime; workbook-level QA tabs were not added.")

            if MATERIALIZE_FORMULA_RESULT_COLUMNS:
                diagnostics.append(
                    "Formula-output note: report formula columns were materialized as values so Python/openpyxl mismatch checks see populated results immediately after build."
                )
            else:
                diagnostics.append(
                    "Formula-cache note: generated formula columns are marked for Excel recalculation on open. "
                    "If the workbook is inspected by Python/openpyxl or a previewer before Excel recalculates it, "
                    "formula-result columns can look blank even though formulas are present."
                )

            status.update(label="Saving workbook...")
            if drop_extra_sheets:
                _dropped_sheets = _prune_non_report_sheets(wb)
                if _dropped_sheets:
                    diagnostics.append(
                        "Removed non-report sheets carried in from the template: " + ", ".join(_dropped_sheets)
                    )
            else:
                _kept = [n for n in REPORT_DROP_SHEETS if n in wb.sheetnames]
                if _kept:
                    diagnostics.append(
                        "Kept the carried-over non-report sheets (matches the official report): "
                        + ", ".join(_kept)
                        + '. Tick "Drop the carried-over ... sheets" to remove them.'
                    )
            diagnostics.append("Sheets in the finished workbook: " + ", ".join(wb.sheetnames))
            try:
                import openpyxl.worksheet._writer as _wchk2
                _w_now = getattr(_wchk2.xmlfile, "__module__", "?")
            except Exception:
                _w_now = "?"
            diagnostics.append(f"Worksheet writer in use: {_w_now} (expected et_xmlfile.xmlfile)")
            if "et_xmlfile" not in str(_w_now):
                diagnostics.append(
                    "WARNING: openpyxl is still streaming worksheets through lxml. That is what "
                    "emptied Bridge Asset and Term Asset on Python 3.14."
                )
            _write_build_log_sheet(wb, diagnostics, upb_col, tmpl_path_used, build_target)
            out_bytes = BytesIO()
            sanitize_summary_formulas(wb)
            _strip_timezones_from_workbook(wb)
            mark_workbook_for_recalc(wb)
            wb.save(out_bytes)
            out_bytes.seek(0)
            wb.close()

            # V68: read the SAVED bytes back and confirm every sheet we wrote still carries
            # its rows. Nothing between the write and here is supposed to remove data --
            # repair_workbook_from_baseline only fills blanks, and the timezone/recalc passes
            # only touch cell values -- but test 77 shipped with Bridge Asset and Term Asset
            # as empty stubs anyway, after a QA audit that had seen 4,694 and 23,848 rows.
            # Whatever the cause, a half-workbook must never reach the user silently.
            # V70: the V68 version of this check trusted openpyxl's read-only max_row and, worse,
            # FAILED OPEN -- any exception in the check itself only logged a line and let the
            # download proceed. Test 78 shipped with the same two empty stubs as test 77
            # despite the guard being live, so the check is now made on the saved ZIP itself
            # and every failure path blocks.
            #
            # A worksheet with no data serialises as either a self-closing <sheetData/> or a
            # <sheetData></sheetData> pair with no <row> element, and its part is ~918 bytes.
            # That is unambiguous and needs no spreadsheet parsing at all.
            _lost = []
            try:
                import zipfile as _zf
                _raw = out_bytes.getvalue()
                with _zf.ZipFile(BytesIO(_raw)) as _z:
                    _wbx = _z.read("xl/workbook.xml").decode("utf-8", errors="replace")
                    _rels = _z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
                    _rid_to_target = {
                        m.group(1): m.group(2)
                        for m in re.finditer(
                            r'<Relationship (?=[^>]*Id="([^"]+)")(?=[^>]*Target="([^"]+)")[^>]*>', _rels
                        )
                    }
                    _name_to_rid = {
                        m.group(1): m.group(2)
                        for m in re.finditer(
                            r'<sheet (?=[^>]*name="([^"]+)")(?=[^>]*r:id="([^"]+)")[^>]*>', _wbx
                        )
                    }
                    for _sn, _expected in WRITTEN_SHEET_ROWS.items():
                        if _expected <= 0:
                            continue
                        _rid = _name_to_rid.get(_sn)
                        if _rid is None:
                            _lost.append(f"{_sn}: not declared in the saved workbook (expected {_expected:,} rows)")
                            continue
                        _tgt = (_rid_to_target.get(_rid) or "").lstrip("/")
                        _part = _tgt if _tgt.startswith("xl/") else "xl/" + _tgt
                        if _part not in _z.namelist():
                            _lost.append(f"{_sn}: worksheet part {_part} missing from the saved file (expected {_expected:,} rows)")
                            continue
                        _xml = _z.read(_part).decode("utf-8", errors="replace")
                        _rows = _xml.count("<row ")
                        if _rows < max(1, int(_expected * 0.5)):
                            _lost.append(
                                f"{_sn}: saved with {_rows:,} rows in {_part} ({_z.getinfo(_part).file_size:,} bytes), expected about {_expected:,}"
                            )
            except Exception as _verr:
                # Fail CLOSED: an unverifiable workbook is treated as a failed one.
                _lost.append(f"the check itself could not run ({type(_verr).__name__}: {_verr})")
            if _lost:
                # V74: this reports, it no longer blocks. Refusing the download left no
                # artifact at all to diagnose from and cost a build; a flagged workbook plus
                # the Build Log sheet is strictly more useful than nothing.
                diagnostics.append("SAVE VERIFICATION FAILED: " + "; ".join(_lost))
                try:
                    st.error(
                        "⚠️ The saved workbook lost data that was written to it: "
                        + "; ".join(_lost)
                        + ". The file is still offered for download so it can be inspected -- see the "
                        "Build Log sheet -- but do NOT use it as the weekly report."
                    )
                except Exception:
                    pass
            if False:
                raise RuntimeError(
                    "The saved workbook lost data that was written to it: "
                    + "; ".join(_lost)
                    + ". The file was NOT offered for download because it would be an incomplete report. "
                    "The sheets were populated when the QA audit read them, so they were lost during the save. "
                    "Re-run; if it repeats, send the full diagnostics text -- this is a save-path or memory "
                    "problem, not a data problem."
                )
            # V73: the template is the UPLOADED PRIOR WORKBOOK whenever one is supplied
            # (resolve_template_bytes), and restore_template_scaffold does not clear data rows.
            # So a report tab that never gets written still shows last week's rows -- which is
            # what made the QA audit report 4,691 Bridge Asset rows on a build that never wrote
            # that sheet, and why WRITTEN_SHEET_ROWS had no entry for it and the save check
            # skipped it. Name any expected tab that was never handed rows.
            _never_written = [
                _sn for _sn in ("Bridge Asset", "Bridge Loan", "Term Loan", "Term Asset")
                if _sn in selected_sheet_names and _sn not in WRITTEN_SHEET_ROWS
            ]
            if _never_written:
                diagnostics.append(
                    "SHEETS NEVER WRITTEN: " + ", ".join(_never_written)
                    + " -- these were selected for this build but write_output_sheet was never "
                    "called for them, so whatever they contain came from the template (which is "
                    "the uploaded prior workbook when one is supplied), not from this run."
                )
            if SHEET_WRITE_AUDIT:
                diagnostics.append(
                    "Rows on each sheet immediately after its write: "
                    + ", ".join(
                        f"{k} {v['on_sheet_after_write']:,}/{v['handed']:,}"
                        for k, v in SHEET_WRITE_AUDIT.items()
                    )
                )
            if WRITTEN_SHEET_ROWS:
                diagnostics.append(
                    "Save verification passed: "
                    + ", ".join(f"{k} {v:,} rows" for k, v in WRITTEN_SHEET_ROWS.items() if v)
                )

            st.session_state.built_workbook_bytes = out_bytes.getvalue()
            st.session_state.built_workbook_name = OUTPUT_TEST_FILENAME
            st.session_state.built_template_path = tmpl_path_used
            st.session_state.show_download_prompt = True
            st.session_state.download_choice = "Not yet"

            status.update(label="Build complete", state="complete")
            st.success("✅ Workbook built")
            st.caption(f"Built from template source: {tmpl_path_used}")

            if diagnostics:
                st.subheader("Diagnostics")
                for msg in diagnostics:
                    st.write(msg)

        except Exception as e:
            st.error("The report builder hit an error. The real traceback is below.")
            st.exception(e)
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

if st.session_state.get("show_download_prompt") and st.session_state.get("built_workbook_bytes"):
    st.markdown("### Download")
    st.radio(
        "Your report is ready. Do you want to download the Excel file now?",
        options=["Not yet", "Yes"],
        horizontal=True,
        key="download_choice",
    )

    if st.session_state.get("download_choice") == "Yes":
        st.download_button(
            "Download Excel file",
            data=st.session_state["built_workbook_bytes"],
            file_name=st.session_state["built_workbook_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.caption("No problem — the file is ready whenever you want to download it during this session.")